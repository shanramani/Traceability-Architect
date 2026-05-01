"""
VALINTEL.AI — DCI (Deviation & CAPA Investigation) Module
============================================================

Per DCI Spec v1.3. 14 rules across 4 engines, keyword-match only, no LLM
in detection path.

Engines:
    A — RCA Recurrence      Rules 1-3
    B — Weak Investigation  Rules 4-8
    C — CAPA Effectiveness  Rules 9-11
    D — SLA / Aging Risk    Rules 12-14

Regulatory coverage:
    21 CFR Part 820.100  |  ICH Q10 §3.2  |  EU GMP Annex 11, Clause 10
    |  FDA CAPA Guidance (2014)  |  FDA CSA Final Guidance (Sep 2021)

This module imports UI helpers from generator.py at load time. Circular
imports are avoided via a deferred-import pattern in generator.py's
dispatcher (it imports from dci_module only when the user clicks the
DCI module button).
"""
import io
import hashlib
import datetime as _dt

import pandas as pd
import streamlit as st

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Helpers from generator.py — LAZY import pattern ───────────────────────
# We MUST NOT do `from generator import ...` at module top level.
# Reason: when Streamlit first imports dci_module (via the deferred import
# in generator.py's dispatcher), generator.py is still mid-execution. A
# top-level `from generator import X` would re-enter generator, which
# re-runs `st.set_page_config()` — but widgets have already rendered, so
# Streamlit raises StreamlitSetPageConfigMustBeFirstCommandError.
#
# Solution: defer the import inside a cached accessor. First call actually
# imports (generator.py is fully loaded by then). Subsequent calls return
# cached references.
_gen_helpers = None

def _gen():
    """
    Lazy-load helpers from generator.py — looking up the ALREADY-RUNNING
    generator module via sys.modules, NEVER using `import generator`.

    Why this matters in Streamlit:
      - Streamlit runs generator.py as the entry script. Python stores it
        under the module name "__main__" (or an ephemeral name), NOT under
        "generator".
      - Doing `import generator` would cause Python to find generator.py
        on the path, see no cached entry under that exact name, and RUN
        the file top-to-bottom AGAIN. That re-executes `st.set_page_config()`
        which Streamlit forbids past the first call.
      - The fix: walk sys.modules to find the module object that contains
        the helpers we need. That object is already loaded and in memory.
    """
    global _gen_helpers
    if _gen_helpers is not None:
        return _gen_helpers

    import sys as _sys

    # Strategy: find the module object that defines the functions we need.
    # Generator.py is guaranteed to be loaded (we're running from it) but
    # its module name depends on how Streamlit invoked it. Search both
    # the obvious candidates and the full module registry.
    _candidate_names = ["__main__", "generator", "streamlit.generator"]
    _gen_mod = None
    for _name in _candidate_names:
        _m = _sys.modules.get(_name)
        if _m is not None and hasattr(_m, "log_audit") and hasattr(_m, "_scroll_top"):
            _gen_mod = _m
            break

    # Fallback: scan all loaded modules for one whose __file__ ends with
    # generator.py. Handles any naming convention Streamlit might use.
    if _gen_mod is None:
        for _name, _m in list(_sys.modules.items()):
            try:
                _f = getattr(_m, "__file__", "") or ""
            except Exception:
                continue
            if _f.endswith("generator.py") and hasattr(_m, "log_audit"):
                _gen_mod = _m
                break

    if _gen_mod is None:
        # sys.modules didn't contain a match. We deliberately do NOT fall back
        # to `import generator` — that would re-execute generator.py and
        # re-trigger st.set_page_config, which is exactly the bug this
        # function exists to avoid. Raise a clear error instead.
        raise RuntimeError(
            "dci_module._gen() could not locate the already-loaded generator "
            "module in sys.modules. Candidates searched: "
            f"{_candidate_names}. Loaded modules count: {len(_sys.modules)}. "
            "This usually means dci_module.py was loaded before generator.py "
            "finished its top-level execution, or the two files are not in "
            "the same folder. Verify deployment layout."
        )

    _gen_helpers = {
        "log_audit":                      _gen_mod.log_audit,
        "_render_validator_verdict":      _gen_mod._render_validator_verdict,
        "_scroll_top":                    _gen_mod._scroll_top,
        "_cols_lower":                    _gen_mod._cols_lower,
        "_matching_cols":                 _gen_mod._matching_cols,
        "_find_col":                      _gen_mod._find_col,
        "_verdict_from_results":          _gen_mod._verdict_from_results,
        "_VALINTEL_SHEET_FINGERPRINTS":   _gen_mod._VALINTEL_SHEET_FINGERPRINTS,
        "_VALINTEL_COLUMN_FINGERPRINTS":  _gen_mod._VALINTEL_COLUMN_FINGERPRINTS,
        "_AT_VOCAB_COLUMNS":              _gen_mod._AT_VOCAB_COLUMNS,
        "_UAR_VOCAB_COLUMNS":             _gen_mod._UAR_VOCAB_COLUMNS,
        "_dim_event_category":            getattr(_gen_mod, "_dim_event_category", None),
        # v96 — cross-module upload invalidation helpers
        "_file_content_hash":             getattr(_gen_mod, "_file_content_hash", None),
        "_check_and_invalidate_on_new_upload": getattr(_gen_mod, "_check_and_invalidate_on_new_upload", None),
        "_record_run_hash":               getattr(_gen_mod, "_record_run_hash", None),
    }
    return _gen_helpers


# ═══════════════════════════════════════════════════════════════════════════
#  Regulatory reference
# ═══════════════════════════════════════════════════════════════════════════
_REG_DCI = ("21 CFR Part 820.100  |  ICH Q10 §3.2  "
            "|  EU GMP Annex 11, Clause 10  |  FDA CAPA Guidance (2014)  "
            "|  FDA CSA Final Guidance (Sep 2021)")


# ═══════════════════════════════════════════════════════════════════════════
#  Required columns + column alias map
# ═══════════════════════════════════════════════════════════════════════════
_DCI_REQUIRED_COLS = {
    "record_id", "record_type", "deviation_category",
    "system_name", "open_date", "close_date",
    "rca_text", "capa_text", "assigned_to",
    "approved_by", "status", "sla_days",
}

_DCI_COLUMN_ALIASES = {
    # record_id
    "record_id": "record_id", "recordid": "record_id", "id": "record_id",
    "deviation_id": "record_id", "deviation id": "record_id",
    "capa_id": "record_id", "capa id": "record_id",
    "nc_id": "record_id", "event_id": "record_id", "case_id": "record_id",
    # record_type
    "record_type": "record_type", "recordtype": "record_type",
    "type": "record_type", "record type": "record_type",
    # deviation_category
    "deviation_category": "deviation_category", "category": "deviation_category",
    "cause_category": "deviation_category", "root_cause_category": "deviation_category",
    "rca_category": "deviation_category",
    # system_name
    "system_name": "system_name", "system": "system_name", "systemname": "system_name",
    "application": "system_name", "gxp_system": "system_name",
    # open_date
    "open_date": "open_date", "opendate": "open_date",
    "date_opened": "open_date", "opened_date": "open_date",
    "occurrence_date": "open_date", "report_date": "open_date",
    # close_date
    "close_date": "close_date", "closedate": "close_date",
    "date_closed": "close_date", "closed_date": "close_date",
    "closure_date": "close_date",
    # rca_text
    "rca_text": "rca_text", "rca": "rca_text", "root_cause": "rca_text",
    "root cause": "rca_text", "root_cause_text": "rca_text",
    "investigation": "rca_text", "investigation_summary": "rca_text",
    # capa_text
    "capa_text": "capa_text", "capa": "capa_text",
    "corrective_action": "capa_text", "corrective action": "capa_text",
    "action_taken": "capa_text", "capa_description": "capa_text",
    # assigned_to
    "assigned_to": "assigned_to", "assignee": "assigned_to",
    "owner": "assigned_to", "assigned": "assigned_to",
    "investigator": "assigned_to", "responsible": "assigned_to",
    # approved_by
    "approved_by": "approved_by", "approver": "approved_by",
    "approved": "approved_by", "closed_by": "approved_by",
    "reviewer": "approved_by",
    # status
    "status": "status", "state": "status", "current_status": "status",
    "record_status": "status",
    # sla_days
    "sla_days": "sla_days", "sla": "sla_days", "due_days": "sla_days",
    "target_days": "sla_days", "closure_sla": "sla_days",
}

# Cross-module rejection detection
_DCI_VOCAB_COLUMNS = {
    "record_id", "recordid", "deviation_id", "deviationid",
    "capa_id", "capaid", "nc_id", "case_id", "caseid",
    "deviation_category", "cause_category", "root_cause_category",
    "rca_text", "rca", "root_cause", "investigation",
    "capa_text", "capa", "corrective_action", "action_taken",
    "open_date", "opendate", "date_opened", "occurrence_date",
    "close_date", "closedate", "date_closed", "closure_date",
    "sla_days", "sla", "due_days", "target_days",
    "assigned_to", "assignee", "investigator", "approved_by", "approver",
}


# ═══════════════════════════════════════════════════════════════════════════
#  DCI vocabularies (per Spec v1.3 §4.1)
# ═══════════════════════════════════════════════════════════════════════════
# Reviewer Round 2 guardrails locked here:
#   - Keep _DCI_SPECIFIC_CAUSE_TERMS SMALL (10 terms). Do NOT expand.
#   - Match semantics: case-insensitive substring via Python `in` operator.
#     No regex, no stemming, no NLP.

_DCI_VAGUE_RCA_TERMS = {
    "human error", "operator error", "mistake", "carelessness",
    "lack of attention", "oversight", "misunderstanding",
    "forgot", "forgotten", "negligence",
}

_DCI_SPECIFIC_CAUSE_TERMS = {
    "instrument",     # "instrument failure", "instrument malfunction"
    "calibration",    # "calibration drift", "out of calibration"
    "sop",            # "SOP deviation", "SOP step missed"
    "method",         # "method error", "method not followed"
    "equipment",      # "equipment failure", "equipment malfunction"
    "software",       # "software glitch", "software bug"
    "procedure",      # "procedure not followed", "procedure outdated"
    "reagent",        # "reagent expired", "reagent contamination"
    "specification",  # "out of specification", "specification misread"
    "component",      # "component failure", "component worn"
}

_DCI_TRAINING_ONLY_TERMS = {
    "retrain", "retraining", "refresher", "reminder",
    "toolbox talk", "additional training", "training session",
    "awareness", "briefing",
}

_DCI_CAPA_ACTION_TERMS = {
    "sop", "procedure", "method", "specification",
    "engineering", "design", "hardware", "software",
    "equipment", "requalification", "revalidation",
    "preventive maintenance",
}


# ═══════════════════════════════════════════════════════════════════════════
#  Input file validator (per Spec v1.3 §3)
# ═══════════════════════════════════════════════════════════════════════════
def _validate_dci_input_file(raw_bytes: bytes, file_name: str, df: pd.DataFrame,
                              all_sheet_names: list = None) -> tuple:
    """
    DCI input file validator. 8 invariants (3 fatal, 5 normal).

    Returns (ok, severity, title, results, evidence).
    severity: "hard_reject" | "warn" | "ok"
    """
    results  = []
    evidence = []

    def _has_dci_col(canonical: str) -> tuple:
        """Returns (has_col, actual_column_name_found)."""
        aliases_for_canonical = {
            alias for alias, canon in _DCI_COLUMN_ALIASES.items()
            if canon == canonical
        }
        cols_norm = _gen()["_cols_lower"](df)
        hit = cols_norm & aliases_for_canonical
        if hit:
            return True, sorted(hit)[0]
        return False, ""

    # ── DCI-F1: Not a VALINTEL output ─────────────────────────────────────
    _sheet_hits = set()
    if all_sheet_names:
        _sheet_hits = ({str(s).strip().lower() for s in all_sheet_names}
                       & _gen()["_VALINTEL_SHEET_FINGERPRINTS"])
    _col_hits = _gen()["_cols_lower"](df) & _gen()["_VALINTEL_COLUMN_FINGERPRINTS"]
    _f1_passed = not _sheet_hits and not _col_hits
    _f1_detail = "Not a VALINTEL output"
    if not _f1_passed:
        if _sheet_hits:
            _d = f"VALINTEL-only sheets detected: {', '.join(sorted(_sheet_hits))}"
            _f1_detail = _d
            evidence.append(_d)
        if _col_hits:
            _d = f"VALINTEL-only columns detected: {', '.join(sorted(_col_hits))}"
            _f1_detail = _d if _f1_passed else _f1_detail + " · " + _d
            evidence.append(_d)
    results.append({
        "id": "DCI-F1", "name": "Not a VALINTEL output",
        "severity_class": "fatal",
        "passed": _f1_passed, "detail": _f1_detail,
    })

    # ── DCI-F2: Has record_id column ──────────────────────────────────────
    _f2_has, _f2_col = _has_dci_col("record_id")
    _f2_detail = (f"record_id column found: '{_f2_col}'"
                  if _f2_has else
                  "No record identifier column found. Required to uniquely "
                  "identify deviations/CAPAs.")
    if not _f2_has:
        evidence.append("Missing: record_id (or alias: deviation_id, capa_id, "
                        "nc_id, case_id, event_id)")
    results.append({
        "id": "DCI-F2", "name": "Has record_id column (or alias)",
        "severity_class": "fatal",
        "passed": _f2_has, "detail": _f2_detail,
    })

    # ── DCI-F3: Has BOTH rca_text AND capa_text ───────────────────────────
    _f3_rca_has, _f3_rca_col   = _has_dci_col("rca_text")
    _f3_capa_has, _f3_capa_col = _has_dci_col("capa_text")
    _f3_passed = _f3_rca_has and _f3_capa_has
    if _f3_passed:
        _f3_detail = (f"Both RCA ('{_f3_rca_col}') and CAPA ('{_f3_capa_col}') "
                      "columns present.")
    else:
        _missing = []
        if not _f3_rca_has:  _missing.append("rca_text")
        if not _f3_capa_has: _missing.append("capa_text")
        _f3_detail = ("Both RCA text and CAPA text columns are required. "
                      f"Missing: {', '.join(_missing)}.")
        evidence.append(_f3_detail)
    results.append({
        "id": "DCI-F3", "name": "Has both rca_text AND capa_text columns",
        "severity_class": "fatal",
        "passed": _f3_passed, "detail": _f3_detail,
    })

    # ── DCI-N1: Has open_date ─────────────────────────────────────────────
    _n1_has, _n1_col = _has_dci_col("open_date")
    _n1_detail = (f"open_date column found: '{_n1_col}'" if _n1_has else
                  "No record open date column found — SLA/aging rules "
                  "(12/13/14) will be unavailable.")
    if not _n1_has:
        evidence.append("Missing: open_date (or alias: opendate, date_opened, "
                        "occurrence_date, report_date)")
    results.append({
        "id": "DCI-N1", "name": "Has open_date column (or alias)",
        "severity_class": "normal",
        "passed": _n1_has, "detail": _n1_detail,
    })

    # ── DCI-N2: Has status column ─────────────────────────────────────────
    _n2_has, _n2_col = _has_dci_col("status")
    _n2_detail = (f"status column found: '{_n2_col}'" if _n2_has else
                  "No status column found — closure-based rules will be "
                  "unavailable.")
    if not _n2_has:
        evidence.append("Missing: status (or alias: state, current_status, "
                        "record_status)")
    results.append({
        "id": "DCI-N2", "name": "Has status column (or alias)",
        "severity_class": "normal",
        "passed": _n2_has, "detail": _n2_detail,
    })

    # ── DCI-N3: AT vocabulary does NOT dominate ───────────────────────────
    _at_cols  = _gen()["_matching_cols"](df, _gen()["_AT_VOCAB_COLUMNS"])
    _dci_cols = _gen()["_matching_cols"](df, _DCI_VOCAB_COLUMNS)
    _n3_dominated = (len(_at_cols) >= 3 and len(_at_cols) > len(_dci_cols))
    _n3_passed = not _n3_dominated
    if _n3_passed:
        _n3_detail = ("No audit-trail vocabulary dominance "
                      f"(AT cols: {len(_at_cols)}, DCI cols: {len(_dci_cols)}).")
    else:
        _n3_detail = ("File appears to be an audit trail log "
                      f"(AT vocabulary columns: {', '.join(_at_cols)}). "
                      "Upload via AT module instead.")
        evidence.append(_n3_detail)
    results.append({
        "id": "DCI-N3", "name": "No AT vocabulary dominance",
        "severity_class": "normal",
        "passed": _n3_passed, "detail": _n3_detail,
    })

    # ── DCI-N4: UAR vocabulary does NOT dominate ──────────────────────────
    _uar_cols = _gen()["_matching_cols"](df, _gen()["_UAR_VOCAB_COLUMNS"])
    _n4_dominated = (len(_uar_cols) >= 3 and len(_uar_cols) > len(_dci_cols))
    _n4_passed = not _n4_dominated
    if _n4_passed:
        _n4_detail = ("No user-access vocabulary dominance "
                      f"(UAR cols: {len(_uar_cols)}, DCI cols: {len(_dci_cols)}).")
    else:
        _n4_detail = ("File appears to be a user access list "
                      f"(UAR vocabulary columns: {', '.join(_uar_cols)}). "
                      "Upload via UAR module instead.")
        evidence.append(_n4_detail)
    results.append({
        "id": "DCI-N4", "name": "No UAR vocabulary dominance",
        "severity_class": "normal",
        "passed": _n4_passed, "detail": _n4_detail,
    })

    # ── DCI-N5: Record-centric structure ──────────────────────────────────
    _n5_passed = True
    _n5_detail = "Record uniqueness check skipped (no record_id column)."
    if _f2_has and _f2_col and len(df) > 0:
        try:
            _rid_col_actual = _gen()["_find_col"](df, {_f2_col})
            _rid_vals = df[_rid_col_actual].astype(str)
            _unique_ratio = len(set(_rid_vals)) / max(len(_rid_vals), 1)
            _n5_passed = _unique_ratio >= 0.50
            if _n5_passed:
                _n5_detail = (f"Record uniqueness: {_unique_ratio:.0%} "
                              f"({len(set(_rid_vals))}/{len(_rid_vals)}).")
            else:
                _n5_detail = (f"Low record_id uniqueness ({_unique_ratio:.0%}) — "
                              "snapshot data assumed. If this is an audit-history "
                              "export, collapse to latest state per record first.")
                evidence.append(_n5_detail)
        except Exception:
            _n5_detail = "Record uniqueness check skipped (column read error)."
    results.append({
        "id": "DCI-N5", "name": "Record-centric structure (>=50% unique)",
        "severity_class": "normal",
        "passed": _n5_passed, "detail": _n5_detail,
    })

    # ── Verdict ───────────────────────────────────────────────────────────
    _severity = _gen()["_verdict_from_results"](results)
    _title = ""
    if _severity == "hard_reject":
        _title = ("File rejected — does not match "
                  "Deviation & CAPA Investigation source pattern")
    elif _severity == "warn":
        _title = "File accepted with a warning"

    return (_severity != "hard_reject", _severity, _title, results, evidence)


# ═══════════════════════════════════════════════════════════════════════════
#  Rule config, metadata, tier priority
# ═══════════════════════════════════════════════════════════════════════════
_DCI_RULE_DEFAULTS = {
    "dci_r1_on":  True,   "dci_r2_on":  True,   "dci_r3_on":  True,
    "dci_r4_on":  True,   "dci_r5_on":  True,   "dci_r6_on":  True,
    "dci_r7_on":  True,   "dci_r8_on":  True,
    "dci_r9_on":  True,   "dci_r10_on": True,   "dci_r11_on": False,
    "dci_r12_on": True,   "dci_r13_on": True,   "dci_r14_on": False,
    "dci_r15_on": True,   # Rule 15 — Training-Only CAPA on Recurrent Failure
}

_DCI_CFG_SCORE_MAP = {
    "dci_r1_on":  ["score_dci_rule1_recurring_category"],
    "dci_r2_on":  ["score_dci_rule2_repeat_category_hivol"],
    "dci_r3_on":  ["score_dci_rule3_repeat_system"],
    "dci_r4_on":  ["score_dci_rule4_short_rca"],
    "dci_r5_on":  ["score_dci_rule5_vague_rca"],
    "dci_r6_on":  ["score_dci_rule6_missing_rca"],
    "dci_r7_on":  ["score_dci_rule7_missing_capa"],
    "dci_r8_on":  ["score_dci_rule8_weak_capa"],
    "dci_r9_on":  ["score_dci_rule9_repeat_post_closure"],
    "dci_r10_on": ["score_dci_rule10_reopened_capa"],
    "dci_r11_on": ["score_dci_rule11_short_close"],
    "dci_r12_on": ["score_dci_rule12_overdue"],
    "dci_r13_on": ["score_dci_rule13_near_breach"],
    "dci_r14_on": ["score_dci_rule14_no_activity"],
}

_DCI_RULE_META = [
    # (rule_num, name, severity, tier, engine, default, cfg_key)
    (1,  "Recurring Category",                "Medium",   "T1", "A", True,  "dci_r1_on"),
    (2,  "Repeat Category High Volume",       "High",     "T1", "A", True,  "dci_r2_on"),
    (3,  "Repeat System",                     "High",     "T1", "A", True,  "dci_r3_on"),
    (4,  "Short RCA Narrative",               "High",     "T1", "B", True,  "dci_r4_on"),
    (5,  "Vague RCA (generic cause only)",    "Critical", "T1", "B", True,  "dci_r5_on"),
    (6,  "Missing RCA",                       "Critical", "T1", "B", True,  "dci_r6_on"),
    (7,  "Missing CAPA",                      "Critical", "T1", "B", True,  "dci_r7_on"),
    (8,  "Weak CAPA (training-only)",         "High",     "T1", "B", True,  "dci_r8_on"),
    (9,  "Repeat Deviation Post-Closure",     "Critical", "T1", "C", True,  "dci_r9_on"),
    (10, "Re-opened CAPA",                    "High",     "T1", "C", True,  "dci_r10_on"),
    (11, "Short Close Cycle",                 "Medium",   "T2", "C", False, "dci_r11_on"),
    (12, "Overdue",                           "High",     "T1", "D", True,  "dci_r12_on"),
    (13, "Near-Breach",                       "Medium",   "T1", "D", True,  "dci_r13_on"),
    (14, "No Activity",                       "Medium",   "T2", "D", False, "dci_r14_on"),
    (15, "Training-Only CAPA on Recurrence",  "High",     "T1", "C", True,  "dci_r15_on"),
]

_DCI_SEVERITY_SCORE = {"Critical": 9.0, "High": 7.0, "Medium": 6.0}

_DCI_RULE_TIER_PRIORITY = [
    ("score_dci_rule6_missing_rca",             9.0, "Critical"),
    ("score_dci_rule7_missing_capa",            9.0, "Critical"),
    ("score_dci_rule5_vague_rca",               9.0, "Critical"),
    ("score_dci_rule9_repeat_post_closure",     9.0, "Critical"),
    ("score_dci_rule2_repeat_category_hivol",   7.0, "High"),
    ("score_dci_rule3_repeat_system",           7.0, "High"),
    ("score_dci_rule4_short_rca",               7.0, "High"),
    ("score_dci_rule8_weak_capa",               7.0, "High"),
    ("score_dci_rule10_reopened_capa",          7.0, "High"),
    ("score_dci_rule12_overdue",                7.0, "High"),
    ("score_dci_rule1_recurring_category",      6.0, "Medium"),
    ("score_dci_rule11_short_close",            6.0, "Medium"),
    ("score_dci_rule13_near_breach",            6.0, "Medium"),
    ("score_dci_rule14_no_activity",            6.0, "Medium"),
]

_DCI_RULE_DISPLAY_NAMES = {
    "score_dci_rule1_recurring_category":     "Rule 1 — Recurring Category",
    "score_dci_rule2_repeat_category_hivol":  "Rule 2 — Repeat Category (High Volume)",
    "score_dci_rule3_repeat_system":          "Rule 3 — Repeat System",
    "score_dci_rule4_short_rca":              "Rule 4 — Short RCA Narrative",
    "score_dci_rule5_vague_rca":              "Rule 5 — Vague RCA (generic cause only)",
    "score_dci_rule6_missing_rca":            "Rule 6 — Missing RCA",
    "score_dci_rule7_missing_capa":           "Rule 7 — Missing CAPA",
    "score_dci_rule8_weak_capa":              "Rule 8 — Weak CAPA (training-only)",
    "score_dci_rule9_repeat_post_closure":    "Rule 9 — Repeat Deviation Post-Closure",
    "score_dci_rule10_reopened_capa":         "Rule 10 — Re-opened CAPA",
    "score_dci_rule11_short_close":           "Rule 11 — Short Close Cycle",
    "score_dci_rule12_overdue":               "Rule 12 — Overdue",
    "score_dci_rule13_near_breach":           "Rule 13 — Near-Breach",
    "score_dci_rule14_no_activity":           "Rule 14 — No Activity",
}


# ═══════════════════════════════════════════════════════════════════════════
#  Helpers
# ═══════════════════════════════════════════════════════════════════════════
def _dci_normalize_status(status) -> str:
    """Canonicalize status to one of: closed | reopened | open | other."""
    if status is None or (isinstance(status, float) and pd.isna(status)):
        return "other"
    s = str(status).strip().lower()
    if s in ("", "nan", "none"):
        return "other"
    if s in ("closed", "complete", "completed"):
        return "closed"
    if s in ("re-opened", "reopened", "re_opened", "reopen"):
        return "reopened"
    if s in ("open", "in-progress", "in progress", "in_progress",
             "investigating", "pending"):
        return "open"
    return "other"


def _dci_parse_date(v):
    """Parse to pd.Timestamp; return pd.NaT on failure."""
    if v is None:
        return pd.NaT
    try:
        return pd.to_datetime(v, errors="coerce")
    except Exception:
        return pd.NaT


# ═══════════════════════════════════════════════════════════════════════════
#  ENGINE A — RCA Recurrence (Rules 1-3)
# ═══════════════════════════════════════════════════════════════════════════
def _dci_rule1_recurring_category(df):
    """Rule 1 — Recurring Category. Same deviation_category in >=3 records
    within 180-day sliding window. Medium severity."""
    n = len(df)
    scores    = [0.0] * n
    rationales = [""]  * n
    if "deviation_category" not in df.columns or "open_date" not in df.columns:
        return pd.Series(scores, index=df.index), pd.Series(rationales, index=df.index)

    cats  = df["deviation_category"].astype(str).str.strip().str.lower()
    dates = pd.to_datetime(df["open_date"], errors="coerce")

    from collections import defaultdict
    cat_index = defaultdict(list)
    for i, (c, d) in enumerate(zip(cats, dates)):
        if c and c not in ("nan", "none", "") and pd.notna(d):
            cat_index[c].append((i, d))

    for cat, entries in cat_index.items():
        if len(entries) < 3:
            continue
        entries.sort(key=lambda x: x[1])
        dates_arr = [e[1] for e in entries]
        for start_i in range(len(entries)):
            end_i = start_i
            while (end_i + 1 < len(entries) and
                   (dates_arr[end_i + 1] - dates_arr[start_i]).days <= 180):
                end_i += 1
            if end_i - start_i + 1 >= 3:
                n_in_window = end_i - start_i + 1
                window_start = dates_arr[start_i].strftime("%Y-%m-%d")
                window_end   = dates_arr[end_i].strftime("%Y-%m-%d")
                for j in range(start_i, end_i + 1):
                    row_idx = entries[j][0]
                    if scores[row_idx] == 0.0:
                        scores[row_idx] = _DCI_SEVERITY_SCORE["Medium"]
                        rationales[row_idx] = (
                            f"Deviation category '{cat}' appeared in "
                            f"{n_in_window} records between {window_start} and "
                            f"{window_end} — possible recurring systemic cause. "
                            "ICH Q10 §3.2 requires effective CAPA to prevent "
                            "recurrence."
                        )
                break
    return (pd.Series(scores, index=df.index),
            pd.Series(rationales, index=df.index))


def _dci_rule2_repeat_category_hivol(df):
    """Rule 2 — Repeat Category High Volume. Same deviation_category in >=5
    records within 90-day window. High severity."""
    n = len(df)
    scores    = [0.0] * n
    rationales = [""]  * n
    if "deviation_category" not in df.columns or "open_date" not in df.columns:
        return pd.Series(scores, index=df.index), pd.Series(rationales, index=df.index)

    cats  = df["deviation_category"].astype(str).str.strip().str.lower()
    dates = pd.to_datetime(df["open_date"], errors="coerce")

    from collections import defaultdict
    cat_index = defaultdict(list)
    for i, (c, d) in enumerate(zip(cats, dates)):
        if c and c not in ("nan", "none", "") and pd.notna(d):
            cat_index[c].append((i, d))

    for cat, entries in cat_index.items():
        if len(entries) < 5:
            continue
        entries.sort(key=lambda x: x[1])
        dates_arr = [e[1] for e in entries]
        for start_i in range(len(entries)):
            end_i = start_i
            while (end_i + 1 < len(entries) and
                   (dates_arr[end_i + 1] - dates_arr[start_i]).days <= 90):
                end_i += 1
            if end_i - start_i + 1 >= 5:
                n_in_window = end_i - start_i + 1
                for j in range(start_i, end_i + 1):
                    row_idx = entries[j][0]
                    if scores[row_idx] == 0.0:
                        scores[row_idx] = _DCI_SEVERITY_SCORE["High"]
                        rationales[row_idx] = (
                            f"Deviation category '{cat}' appeared in "
                            f"{n_in_window} records within 90 days — "
                            "high-volume recurrence indicates prior CAPA "
                            "ineffective. 21 CFR 820.100(a)(2), ICH Q10 §3.2.3."
                        )
                break
    return (pd.Series(scores, index=df.index),
            pd.Series(rationales, index=df.index))


def _dci_rule3_repeat_system(df):
    """Rule 3 — Repeat System. Same system_name has >=3 deviations within
    60-day window. High severity."""
    n = len(df)
    scores    = [0.0] * n
    rationales = [""]  * n
    if "system_name" not in df.columns or "open_date" not in df.columns:
        return pd.Series(scores, index=df.index), pd.Series(rationales, index=df.index)

    systems = df["system_name"].astype(str).str.strip().str.lower()
    dates   = pd.to_datetime(df["open_date"], errors="coerce")

    from collections import defaultdict
    sys_index = defaultdict(list)
    for i, (s, d) in enumerate(zip(systems, dates)):
        if s and s not in ("nan", "none", "") and pd.notna(d):
            sys_index[s].append((i, d))

    for sys_name, entries in sys_index.items():
        if len(entries) < 3:
            continue
        entries.sort(key=lambda x: x[1])
        dates_arr = [e[1] for e in entries]
        for start_i in range(len(entries)):
            end_i = start_i
            while (end_i + 1 < len(entries) and
                   (dates_arr[end_i + 1] - dates_arr[start_i]).days <= 60):
                end_i += 1
            if end_i - start_i + 1 >= 3:
                n_in_window = end_i - start_i + 1
                for j in range(start_i, end_i + 1):
                    row_idx = entries[j][0]
                    if scores[row_idx] == 0.0:
                        scores[row_idx] = _DCI_SEVERITY_SCORE["High"]
                        rationales[row_idx] = (
                            f"System '{sys_name}' had {n_in_window} deviations "
                            "in 60 days — concentrated issue pattern. Annex 11 "
                            "§10 requires change management to address systemic "
                            "deficiencies."
                        )
                break
    return (pd.Series(scores, index=df.index),
            pd.Series(rationales, index=df.index))


# ═══════════════════════════════════════════════════════════════════════════
#  ENGINE B — Weak Investigation (Rules 4-8)
# ═══════════════════════════════════════════════════════════════════════════
def _dci_rule4_short_rca(row):
    """Rule 4 — Short RCA Narrative. Fires on Closed records with non-blank
    RCA text shorter than 50 chars. High severity."""
    status = _dci_normalize_status(row.get("status"))
    if status != "closed":
        return 0.0, ""
    rca = str(row.get("rca_text", "")).strip()
    if not rca or rca.lower() in ("nan", "none"):
        return 0.0, ""
    if len(rca) < 50:
        return _DCI_SEVERITY_SCORE["High"], (
            f"RCA text is only {len(rca)} characters — insufficient detail "
            "to document root cause analysis per 21 CFR 820.100(b)(4) and "
            "ICH Q10 §3.2.2."
        )
    return 0.0, ""


def _dci_rule5_vague_rca(row):
    """Rule 5 — Vague RCA (generic cause only). Per Spec v1.3 §4.3.
    Fires when: status=Closed AND rca not blank AND vague_term_present
    AND NOT specific_cause_present. Substring match, case-insensitive."""
    status = _dci_normalize_status(row.get("status"))
    if status != "closed":
        return 0.0, ""
    rca = str(row.get("rca_text", "")).strip().lower()
    if not rca or rca in ("nan", "none"):
        return 0.0, ""

    vague_present    = any(v in rca for v in _DCI_VAGUE_RCA_TERMS)
    specific_present = any(s in rca for s in _DCI_SPECIFIC_CAUSE_TERMS)

    if vague_present and not specific_present:
        matched_vague = [v for v in _DCI_VAGUE_RCA_TERMS if v in rca]
        return _DCI_SEVERITY_SCORE["Critical"], (
            f"RCA attributes cause to '{matched_vague[0]}' without specific "
            "equipment, procedural, material, or process factor. "
            "FDA CAPA Guidance (2014) requires root-cause identification "
            "beyond human-factor attribution."
        )
    return 0.0, ""


def _dci_rule6_missing_rca(row):
    """Rule 6 — Missing RCA. Closed with blank rca_text. Critical."""
    status = _dci_normalize_status(row.get("status"))
    if status != "closed":
        return 0.0, ""
    rca = row.get("rca_text")
    is_blank = (
        rca is None
        or (isinstance(rca, float) and pd.isna(rca))
        or str(rca).strip().lower() in ("", "nan", "none")
    )
    if is_blank:
        return _DCI_SEVERITY_SCORE["Critical"], (
            "Record closed with no root cause analysis recorded. "
            "21 CFR 820.100(b)(2) requires investigation of the cause "
            "of nonconformities."
        )
    return 0.0, ""


def _dci_rule7_missing_capa(row):
    """Rule 7 — Missing CAPA. Closed with blank capa_text. Critical."""
    status = _dci_normalize_status(row.get("status"))
    if status != "closed":
        return 0.0, ""
    capa = row.get("capa_text")
    is_blank = (
        capa is None
        or (isinstance(capa, float) and pd.isna(capa))
        or str(capa).strip().lower() in ("", "nan", "none")
    )
    if is_blank:
        return _DCI_SEVERITY_SCORE["Critical"], (
            "Record closed with no corrective/preventive action recorded. "
            "21 CFR 820.100(a)(3) requires identification of action needed "
            "to correct and prevent recurrence."
        )
    return 0.0, ""


def _dci_rule8_weak_capa(row):
    """Rule 8 — Weak CAPA (training-only). Per Spec v1.3 §4.3.
    Fires when: status=Closed AND capa not blank AND training term present
    AND NOT action term present."""
    status = _dci_normalize_status(row.get("status"))
    if status != "closed":
        return 0.0, ""
    capa = str(row.get("capa_text", "")).strip().lower()
    if not capa or capa in ("nan", "none"):
        return 0.0, ""

    training_present = any(t in capa for t in _DCI_TRAINING_ONLY_TERMS)
    action_present   = any(a in capa for a in _DCI_CAPA_ACTION_TERMS)

    if training_present and not action_present:
        return _DCI_SEVERITY_SCORE["High"], (
            "CAPA addresses the issue only via training (no procedural, "
            "engineering, or material change). Training-only CAPA for "
            "non-training-root-cause deviations typically indicates weak "
            "corrective action. ICH Q10 §3.2.3."
        )
    return 0.0, ""


# ═══════════════════════════════════════════════════════════════════════════
#  ENGINE C — CAPA Effectiveness (Rules 9-11)
# ═══════════════════════════════════════════════════════════════════════════
def _dci_rule9_repeat_post_closure(df):
    """Rule 9 — Repeat Deviation Post-Closure. Same record_type +
    system_name + deviation_category re-opens within 90 days of prior
    closure. Critical severity."""
    n = len(df)
    scores    = [0.0] * n
    rationales = [""]  * n

    required = {"record_type", "system_name", "deviation_category",
                "open_date", "close_date", "status"}
    if not required.issubset(df.columns):
        return pd.Series(scores, index=df.index), pd.Series(rationales, index=df.index)

    rtypes   = df["record_type"].astype(str).str.strip().str.lower()
    syss     = df["system_name"].astype(str).str.strip().str.lower()
    cats     = df["deviation_category"].astype(str).str.strip().str.lower()
    opens    = pd.to_datetime(df["open_date"],  errors="coerce")
    closes   = pd.to_datetime(df["close_date"], errors="coerce")
    statuses = df["status"].apply(_dci_normalize_status)
    rids     = df.get("record_id", pd.Series([""] * n)).astype(str)

    from collections import defaultdict
    key_index = defaultdict(list)
    for i in range(n):
        if not (rtypes.iloc[i] and syss.iloc[i] and cats.iloc[i]):
            continue
        if rtypes.iloc[i] in ("nan", "none", ""):
            continue
        key = (rtypes.iloc[i], syss.iloc[i], cats.iloc[i])
        key_index[key].append((
            i, opens.iloc[i], closes.iloc[i],
            statuses.iloc[i], rids.iloc[i]
        ))

    for key, entries in key_index.items():
        if len(entries) < 2:
            continue
        entries.sort(key=lambda x: (x[1] if pd.notna(x[1]) else pd.Timestamp.max))
        for i_cur in range(1, len(entries)):
            cur = entries[i_cur]
            cur_idx, cur_open = cur[0], cur[1]
            if pd.isna(cur_open):
                continue
            for i_prev in range(i_cur):
                prev = entries[i_prev]
                prev_idx, prev_open, prev_close, prev_status, prev_rid = prev
                if prev_status != "closed" or pd.isna(prev_close):
                    continue
                delta = (cur_open - prev_close).days
                if 0 <= delta <= 90:
                    if scores[cur_idx] == 0.0:
                        scores[cur_idx] = _DCI_SEVERITY_SCORE["Critical"]
                        rationales[cur_idx] = (
                            f"Same {key[0]} recurred for {key[1]} in "
                            f"category '{key[2]}' {delta} days after previous "
                            f"closure ({prev_rid} closed "
                            f"{prev_close.strftime('%Y-%m-%d')}). Indicates "
                            "prior CAPA was ineffective. 21 CFR 820.100(a)(2), "
                            "ICH Q10 §3.2.3."
                        )
                    break
    return (pd.Series(scores, index=df.index),
            pd.Series(rationales, index=df.index))


def _dci_rule10_reopened_capa(row):
    """Rule 10 — Re-opened CAPA. Status normalizes to 'reopened'. High."""
    status = _dci_normalize_status(row.get("status"))
    if status == "reopened":
        return _DCI_SEVERITY_SCORE["High"], (
            "Record status is 'Re-opened' — prior closure was invalid or "
            "CAPA ineffective. Per 21 CFR 820.100(a)(7), re-opening requires "
            "documented re-investigation."
        )
    return 0.0, ""


def _dci_rule11_short_close(row):
    """Rule 11 — Short Close Cycle. Closed with <3 days between open and
    close. Default OFF. Medium severity."""
    status = _dci_normalize_status(row.get("status"))
    if status != "closed":
        return 0.0, ""
    od = _dci_parse_date(row.get("open_date"))
    cd = _dci_parse_date(row.get("close_date"))
    if pd.isna(od) or pd.isna(cd):
        return 0.0, ""
    days = (cd - od).days
    if days < 3:
        return _DCI_SEVERITY_SCORE["Medium"], (
            f"Record closed within {days} day(s) of opening. Investigations "
            "typically require multi-day evidence gathering; rapid closure "
            "may indicate insufficient review. ICH Q10 §3.2.2."
        )
    return 0.0, ""


# ═══════════════════════════════════════════════════════════════════════════
#  ENGINE D — SLA / Aging Risk (Rules 12-14)
# ═══════════════════════════════════════════════════════════════════════════
def _dci_rule12_overdue(row, today=None):
    """Rule 12 — Overdue. open_date+sla_days < today AND status != Closed.
    High severity."""
    if today is None:
        today = pd.Timestamp.now().normalize()
    status = _dci_normalize_status(row.get("status"))
    if status == "closed":
        return 0.0, ""
    od = _dci_parse_date(row.get("open_date"))
    if pd.isna(od):
        return 0.0, ""
    sla_raw = row.get("sla_days")
    try:
        sla = float(sla_raw)
        if pd.isna(sla):
            return 0.0, ""
    except (TypeError, ValueError):
        return 0.0, ""
    due_date = od + pd.Timedelta(days=sla)
    if today > due_date:
        days_overdue = (today - due_date).days
        return _DCI_SEVERITY_SCORE["High"], (
            f"Record open for {days_overdue} day(s) past SLA "
            f"({int(sla)}-day window). Overdue investigations breach "
            "21 CFR 820.100(b)(5) timeliness and Annex 11 §10 "
            "change-management timing."
        )
    return 0.0, ""


def _dci_rule13_near_breach(row, today=None):
    """Rule 13 — Near-Breach. Within 7 days of SLA AND still open.
    Medium severity."""
    if today is None:
        today = pd.Timestamp.now().normalize()
    status = _dci_normalize_status(row.get("status"))
    if status == "closed":
        return 0.0, ""
    od = _dci_parse_date(row.get("open_date"))
    if pd.isna(od):
        return 0.0, ""
    sla_raw = row.get("sla_days")
    try:
        sla = float(sla_raw)
        if pd.isna(sla):
            return 0.0, ""
    except (TypeError, ValueError):
        return 0.0, ""
    due_date = od + pd.Timedelta(days=sla)
    days_until = (due_date - today).days
    if 0 <= days_until <= 7:
        return _DCI_SEVERITY_SCORE["Medium"], (
            f"Record within {days_until} day(s) of SLA breach and still open. "
            "Near-breach status per FDA CAPA Guidance (2014) — proactive "
            "escalation recommended."
        )
    return 0.0, ""


def _dci_rule14_no_activity(row, today=None):
    """Rule 14 — No Activity (snapshot fallback). Open >=30 days with no
    close_date. Default OFF. Medium severity."""
    if today is None:
        today = pd.Timestamp.now().normalize()
    status = _dci_normalize_status(row.get("status"))
    if status == "closed":
        return 0.0, ""
    od = _dci_parse_date(row.get("open_date"))
    if pd.isna(od):
        return 0.0, ""
    cd = _dci_parse_date(row.get("close_date"))
    if pd.notna(cd):
        return 0.0, ""
    days_open = (today - od).days
    if days_open >= 30:
        return _DCI_SEVERITY_SCORE["Medium"], (
            f"Record open {days_open} days with no closure. Without "
            "status-change history, sustained-inactivity detection uses "
            "duration proxy per ICH Q10 §3.2.2."
        )
    return 0.0, ""


# ═══════════════════════════════════════════════════════════════════════════
#  Orchestrator
# ═══════════════════════════════════════════════════════════════════════════
def dci_score_records(df, rule_config=None):
    """Score every DCI record across 14 rules. Returns sorted DataFrame
    with individual scores, rule flags, Risk_Tier, Risk_Score,
    Primary_Rule, All_Rules_Fired, Detection_Basis."""
    if rule_config is None:
        rule_config = dict(_DCI_RULE_DEFAULTS)

    df = df.copy()
    today = pd.Timestamp.now().normalize()

    # Vectorized rules
    s1, r1 = _dci_rule1_recurring_category(df)
    s2, r2 = _dci_rule2_repeat_category_hivol(df)
    s3, r3 = _dci_rule3_repeat_system(df)
    s9, r9 = _dci_rule9_repeat_post_closure(df)
    df["score_dci_rule1_recurring_category"]    = s1
    df["rule1_rationale"]                        = r1
    df["score_dci_rule2_repeat_category_hivol"] = s2
    df["rule2_rationale"]                        = r2
    df["score_dci_rule3_repeat_system"]         = s3
    df["rule3_rationale"]                        = r3
    df["score_dci_rule9_repeat_post_closure"]   = s9
    df["rule9_rationale"]                        = r9

    # Per-row rules
    per_row_rules = [
        ("score_dci_rule4_short_rca",       "rule4_rationale",   _dci_rule4_short_rca,      False),
        ("score_dci_rule5_vague_rca",       "rule5_rationale",   _dci_rule5_vague_rca,      False),
        ("score_dci_rule6_missing_rca",     "rule6_rationale",   _dci_rule6_missing_rca,    False),
        ("score_dci_rule7_missing_capa",    "rule7_rationale",   _dci_rule7_missing_capa,   False),
        ("score_dci_rule8_weak_capa",       "rule8_rationale",   _dci_rule8_weak_capa,      False),
        ("score_dci_rule10_reopened_capa",  "rule10_rationale",  _dci_rule10_reopened_capa, False),
        ("score_dci_rule11_short_close",    "rule11_rationale",  _dci_rule11_short_close,   False),
        ("score_dci_rule12_overdue",        "rule12_rationale",  _dci_rule12_overdue,       True),
        ("score_dci_rule13_near_breach",    "rule13_rationale",  _dci_rule13_near_breach,   True),
        ("score_dci_rule14_no_activity",    "rule14_rationale",  _dci_rule14_no_activity,   True),
    ]
    for score_col, rat_col, fn, needs_today in per_row_rules:
        scores, rats = [], []
        for _, row in df.iterrows():
            try:
                s, r = fn(row, today=today) if needs_today else fn(row)
            except Exception:
                s, r = 0.0, ""
            scores.append(s)
            rats.append(r)
        df[score_col] = scores
        df[rat_col]   = rats

    # Zero out disabled rules
    for cfg_key, score_cols in _DCI_CFG_SCORE_MAP.items():
        if not rule_config.get(cfg_key, True):
            for sc in score_cols:
                if sc in df.columns:
                    df[sc] = 0.0

    # Aggregate to Risk_Tier
    _TIER_RANK = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}

    def _priority_tier_and_rule(row):
        best_tier  = "Low"
        best_score = 0.0
        best_col   = ""
        for score_col, threshold, base_tier in _DCI_RULE_TIER_PRIORITY:
            val = float(row.get(score_col, 0))
            if val >= threshold:
                if _TIER_RANK.get(base_tier, 3) < _TIER_RANK.get(best_tier, 3):
                    best_tier  = base_tier
                    best_score = val
                    best_col   = score_col
                elif (_TIER_RANK.get(base_tier, 3) == _TIER_RANK.get(best_tier, 3)
                      and val > best_score):
                    best_score = val
                    best_col   = score_col
        return best_tier, round(best_score, 2), best_col

    _triples = [_priority_tier_and_rule(row) for _, row in df.iterrows()]
    df["Risk_Tier"]    = [t[0] for t in _triples]
    df["Risk_Score"]   = [t[1] for t in _triples]
    df["Primary_Rule"] = [_DCI_RULE_DISPLAY_NAMES.get(t[2], "—") for t in _triples]

    def _rules_fired(row):
        fired = []
        for score_col, threshold, _ in _DCI_RULE_TIER_PRIORITY:
            if float(row.get(score_col, 0)) >= threshold:
                fired.append(_DCI_RULE_DISPLAY_NAMES.get(score_col, score_col))
        return "; ".join(fired) if fired else ""

    def _detection_basis(row):
        parts = []
        for rn in range(1, 15):
            rat = str(row.get(f"rule{rn}_rationale", "")).strip()
            if rat:
                parts.append(rat)
        return "  ||  ".join(parts) if parts else ""

    df["All_Rules_Fired"] = df.apply(_rules_fired, axis=1)
    df["Detection_Basis"] = df.apply(_detection_basis, axis=1)

    # ── IQI — Investigation Quality Index ─────────────────────────────────────
    # IQI = 100 × (1 − total_fired_score / max_possible_score)
    # max_possible = 9 × number of active rules (default 12 ON → 108)
    # All score columns for active rules
    _active_score_cols = [
        sc for cfg_key, score_cols in _DCI_CFG_SCORE_MAP.items()
        if rule_config.get(cfg_key, True)
        for sc in score_cols
        if sc in df.columns
    ]
    _n_active = len(_active_score_cols)
    _max_possible = 9.0 * _n_active if _n_active > 0 else 108.0

    def _compute_iqi(row):
        total = sum(float(row.get(sc, 0)) for sc in _active_score_cols)
        raw = 100.0 * (1.0 - total / _max_possible)
        return int(max(0, min(100, round(raw))))

    def _iqi_band(iqi_val):
        if iqi_val >= 85: return "Strong"
        if iqi_val >= 65: return "Acceptable"
        if iqi_val >= 40: return "Weak"
        return "Poor"

    def _iqi_drivers(row):
        """One-line explanation: which rules fired and how much each cost."""
        hits = []
        for sc, threshold, _ in _DCI_RULE_TIER_PRIORITY:
            val = float(row.get(sc, 0))
            if val >= threshold:
                name = _DCI_RULE_DISPLAY_NAMES.get(sc, sc)
                hits.append(f"{name} (−{int(val)})")
        if not hits:
            return "No issues detected — investigation meets all quality checks."
        return "Deductions: " + " · ".join(hits)

    df["IQI"]         = df.apply(_compute_iqi, axis=1)
    df["IQI_Band"]    = df["IQI"].apply(_iqi_band)
    df["IQI_Drivers"] = df.apply(_iqi_drivers, axis=1)

    # ── CAPA Type Classification ───────────────────────────────────────────────
    # Deterministic keyword classifier — tags each record's CAPA by control type.
    # Four types in effectiveness hierarchy: Engineering > Systemic > Procedural > Training
    _CAPA_TYPE_KEYWORDS = {
        "Engineering":  ["engineer", "redesign", "hardware", "automation", "interlock",
                         "sensor", "software patch", "firmware", "equipment mod",
                         "physical control", "poka-yoke", "alarm"],
        "Systemic":     ["process change", "system change", "workflow", "redesigned",
                         "restructur", "reorgan", "root cause system", "process redesign",
                         "system redesign", "fundamental", "overhaul"],
        "Procedural":   ["sop", "procedure", "protocol", "work instruction", "checklist",
                         "policy", "revised", "updated procedure", "amended", "guideline"],
        "Training":     ["train", "retrain", "awareness", "remind", "coached",
                         "refresher", "education", "briefing", "inform", "instruct"],
    }
    # Order matters — first match wins (engineering > systemic > procedural > training)
    _CAPA_TYPE_ORDER = ["Engineering", "Systemic", "Procedural", "Training"]

    def _classify_capa_type(row):
        capa_text = str(row.get("capa_text", "")).lower().strip()
        if not capa_text or capa_text in ("nan", "none", ""):
            return "None"
        for ctype in _CAPA_TYPE_ORDER:
            if any(kw in capa_text for kw in _CAPA_TYPE_KEYWORDS[ctype]):
                return ctype
        return "Other"

    df["CAPA_Type"] = df.apply(_classify_capa_type, axis=1)

    # ── Rule 15 — Training-Only CAPA on Recurrent Failure (CAPA Effectiveness) ──
    # Fires when: CAPA_Type = Training AND (Rule 1 OR Rule 9 already fired)
    # Effect: escalates Risk_Tier one level (Medium→High, High→Critical)
    # Does NOT change Risk_Score (score stays deterministic from rules 1-14)
    # Records the escalation reason in a new column.
    _TIER_ESCALATE = {"Low": "Medium", "Medium": "High", "High": "Critical", "Critical": "Critical"}

    def _apply_rule15(row):
        if row["CAPA_Type"] != "Training":
            return row["Risk_Tier"], ""
        r1_fired = float(row.get("score_dci_rule1_recurring_category", 0)) >= 6.0
        r9_fired = float(row.get("score_dci_rule9_repeat_post_closure", 0)) >= 9.0
        if r1_fired or r9_fired:
            trigger = "Rule 1 (Recurring Category)" if r1_fired else "Rule 9 (Repeat Post-Closure)"
            new_tier = _TIER_ESCALATE.get(row["Risk_Tier"], row["Risk_Tier"])
            if new_tier != row["Risk_Tier"]:
                return new_tier, (
                    f"Rule 15 escalation: Training-only CAPA on recurrent failure "
                    f"({trigger} fired). Training CAPAs do not prevent recurrence — "
                    f"systemic action required per ICH Q10 §3.2.3."
                )
        return row["Risk_Tier"], ""

    r15_results = df.apply(_apply_rule15, axis=1, result_type="expand")
    df["Rule15_Tier"]   = r15_results[0]
    df["Rule15_Reason"] = r15_results[1]

    # Apply Rule 15 escalation to Risk_Tier (original preserved in Risk_Tier_Base)
    df["Risk_Tier_Base"] = df["Risk_Tier"]
    df["Risk_Tier"]      = df["Rule15_Tier"]
    df = df.drop(columns=["Rule15_Tier"])

    # Recount after Rule 15 escalation for sort
    df["_tier_rank"] = df["Risk_Tier"].map(_TIER_RANK).fillna(3).astype(int)
    df = df.sort_values(
        ["_tier_rank", "IQI"], ascending=[True, True]   # worst IQI first within tier
    ).drop(columns=["_tier_rank"]).reset_index(drop=True)

    return df


# ═══════════════════════════════════════════════════════════════════════════
#  Excel builder — 6 sheets per Spec v1.3 §6
# ═══════════════════════════════════════════════════════════════════════════
def dci_build_excel(scored_df, system_name, r_start, r_end, fname,
                     config_hash="", operator_user="", model_used="",
                     rule_config=None):
    """Build 6-sheet GxP evidence workbook for DCI findings."""
    if rule_config is None:
        rule_config = dict(_DCI_RULE_DEFAULTS)

    output = io.BytesIO()
    wb     = Workbook()

    C_NAVY      = "1E3A5F"
    C_WHITE     = "FFFFFF"
    C_AMBER     = "D97706"
    C_RED       = "C0392B"
    C_ORANGE    = "EA580C"
    C_GREY      = "F1F5F9"
    C_MID       = "94A3B8"
    C_LIGHT     = "EFF6FF"
    C_DARK_TEXT = "2C3E50"

    _TIER_COLORS = {
        "Critical": (C_RED,    C_WHITE),
        "High":     (C_ORANGE, C_WHITE),
        "Medium":   (C_AMBER,  C_WHITE),
        "Low":      (C_GREY,   C_DARK_TEXT),
    }

    bdr = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _hdr(ws, row, col, val, width=None, bg=C_NAVY, fg=C_WHITE, size=9, wrap=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", bold=True, size=size, color=fg)
        c.fill = _fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
        c.border = bdr
        if width:
            ws.column_dimensions[get_column_letter(col)].width = width
        return c

    def _cell(ws, row, col, val, bold=False, bg=None, fg=C_DARK_TEXT,
              size=9, wrap=False, align="left"):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Calibri", bold=bold, size=size, color=fg)
        c.fill = _fill(bg) if bg else PatternFill()
        c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
        c.border = bdr
        return c

    n_total    = len(scored_df)
    n_critical = int((scored_df["Risk_Tier"] == "Critical").sum()) if n_total else 0
    n_high     = int((scored_df["Risk_Tier"] == "High").sum())     if n_total else 0
    n_medium   = int((scored_df["Risk_Tier"] == "Medium").sum())   if n_total else 0
    n_hc       = n_critical + n_high

    # IQI summary stats
    _iqi_vals   = scored_df["IQI"].dropna() if "IQI" in scored_df.columns and n_total else []
    _iqi_mean   = round(float(_iqi_vals.mean()), 1) if len(_iqi_vals) else "—"
    _iqi_min    = int(_iqi_vals.min()) if len(_iqi_vals) else "—"
    _iqi_poor   = int((_iqi_vals < 40).sum()) if len(_iqi_vals) else 0
    _iqi_strong = int((_iqi_vals >= 85).sum()) if len(_iqi_vals) else 0

    # Rule 15 escalations
    _r15_count  = int((scored_df.get("Rule15_Reason", "") != "").sum()) if n_total else 0

    # CAPA type breakdown
    _capa_types = scored_df["CAPA_Type"].value_counts().to_dict() if "CAPA_Type" in scored_df.columns and n_total else {}

    # ── Sheet 1 — Summary ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False

    ws1.cell(row=1, column=1, value="VALINTEL.AI — Deviation & CAPA Investigation"
    ).font = Font(name="Calibri", bold=True, size=14, color=C_NAVY)
    ws1.cell(row=2, column=1,
             value=f"System: {system_name}   ·   Review Period: {r_start} → {r_end}"
    ).font = Font(name="Calibri", size=10, color=C_DARK_TEXT)
    ws1.cell(row=3, column=1,
             value=f"Source File: {fname}"
    ).font = Font(name="Calibri", size=9, italic=True, color=C_MID)

    ws1.row_dimensions[1].height = 22
    ws1.column_dimensions["A"].width = 26
    ws1.column_dimensions["B"].width = 60

    _hdr(ws1, 5, 1, "KPI", width=26)
    _hdr(ws1, 5, 2, "Value", width=60)

    kpi_rows = [
        ("Records Analyzed",         str(n_total)),
        ("Critical Findings",         str(n_critical)),
        ("High Findings",             str(n_high)),
        ("Medium Findings",           str(n_medium)),
        ("Records Requiring Review",  f"{n_hc} (Critical + High)"),
        ("— — —",                     ""),
        ("Period IQI (avg)",          f"{_iqi_mean}" + (" / 100" if _iqi_mean != "—" else "")),
        ("Lowest Record IQI",         str(_iqi_min) + (" / 100" if _iqi_min != "—" else "")),
        ("Records IQI < 40 (Poor)",   str(_iqi_poor)),
        ("Records IQI ≥ 85 (Strong)", str(_iqi_strong)),
        ("Rule 15 Escalations",       str(_r15_count) + (" (training CAPA on recurrence)" if _r15_count else "")),
    ]
    # Add CAPA type breakdown if data present
    if _capa_types:
        kpi_rows.append(("— — —", ""))
        for ctype in ["Engineering", "Systemic", "Procedural", "Training", "Other", "None"]:
            cnt = _capa_types.get(ctype, 0)
            if cnt:
                kpi_rows.append((f"CAPA Type: {ctype}", str(cnt)))
    for i, (k, v) in enumerate(kpi_rows, 6):
        _cell(ws1, i, 1, k, bold=True, bg=C_LIGHT)
        _cell(ws1, i, 2, v)
        ws1.row_dimensions[i].height = 16

    _hdr(ws1, 12, 1, "Top Rules Fired")
    _hdr(ws1, 12, 2, "Count")
    if n_total:
        rule_counts = {}
        for score_col, threshold, _ in _DCI_RULE_TIER_PRIORITY:
            if score_col in scored_df.columns:
                cnt = int((scored_df[score_col] >= threshold).sum())
                if cnt > 0:
                    disp = _DCI_RULE_DISPLAY_NAMES.get(score_col, score_col)
                    rule_counts[disp] = cnt
        top3 = sorted(rule_counts.items(), key=lambda x: -x[1])[:3]
        for i, (rname, cnt) in enumerate(top3, 13):
            _cell(ws1, i, 1, rname)
            _cell(ws1, i, 2, str(cnt), align="center")
            ws1.row_dimensions[i].height = 15
    else:
        _cell(ws1, 13, 1, "No findings", fg=C_MID)

    _hdr(ws1, 18, 1, "Regulatory References")
    _cell(ws1, 19, 1, _REG_DCI, wrap=True)
    ws1.merge_cells(start_row=19, start_column=1, end_row=19, end_column=2)
    ws1.row_dimensions[19].height = 48

    _cell(ws1, 22, 1, f"Config Hash: {config_hash or 'n/a'}",
          fg=C_MID, size=8)

    # ── Sheet 2 — Records for Review ────────────────────────────────────
    ws2 = wb.create_sheet("Records for Review")
    ws2.sheet_view.showGridLines = False

    review_df = scored_df[scored_df["Risk_Tier"].isin(
        ["Critical", "High", "Medium"])].copy()

    review_cols = [
        ("Record ID",       "record_id",           18),
        ("Type",            "record_type",         14),
        ("Category",        "deviation_category",  20),
        ("System",          "system_name",         18),
        ("Open Date",       "open_date",           14),
        ("Close Date",      "close_date",          14),
        ("Status",          "status",              12),
        ("Assigned To",     "assigned_to",         18),
        ("Approved By",     "approved_by",         18),
        ("Risk Tier",       "Risk_Tier",           12),
        ("Risk Score",      "Risk_Score",          11),
        ("IQI",             "IQI",                  8),
        ("IQI Band",        "IQI_Band",            14),
        ("CAPA Type",       "CAPA_Type",           14),
        ("Primary Rule",    "Primary_Rule",        36),
        ("IQI Drivers",     "IQI_Drivers",         60),
        ("Rule 15 Note",    "Rule15_Reason",       50),
        ("Detection Basis", "Detection_Basis",     80),
    ]

    ws2.cell(row=1, column=1,
             value=f"Records for Review — {len(review_df)} record(s) "
                   f"at Medium or higher risk"
    ).font = Font(name="Calibri", bold=True, size=12, color=C_NAVY)
    ws2.row_dimensions[1].height = 20

    for ci, (hdr, _, width) in enumerate(review_cols, 1):
        _hdr(ws2, 3, ci, hdr, width=width)
    ws2.row_dimensions[3].height = 18

    for ri, (_, row_data) in enumerate(review_df.iterrows(), 4):
        tier = str(row_data.get("Risk_Tier", "Low"))
        tier_bg, tier_fg = _TIER_COLORS.get(tier, (C_GREY, C_DARK_TEXT))
        for ci, (_, col_key, _) in enumerate(review_cols, 1):
            val = row_data.get(col_key, "")
            if pd.isna(val):
                val = ""
            if col_key in ("open_date", "close_date"):
                try:
                    dt = pd.to_datetime(val, errors="coerce")
                    val = dt.strftime("%Y-%m-%d") if pd.notna(dt) else ""
                except Exception:
                    val = str(val) if val else ""
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Calibri", size=9, color=C_DARK_TEXT)
            c.alignment = Alignment(
                horizontal="center" if ci > 1 else "left",
                vertical="top",
                wrap_text=col_key in ("Detection_Basis", "All_Rules_Fired",
                                      "Primary_Rule", "IQI_Drivers", "Rule15_Reason"))
            c.border = bdr
            if col_key == "Risk_Tier":
                c.font = Font(name="Calibri", bold=True, size=9, color=tier_fg)
                c.fill = _fill(tier_bg)
            elif col_key == "IQI":
                # Colour-code IQI: green (strong) → red (poor)
                try:
                    iq = int(val) if val != "" else -1
                except (ValueError, TypeError):
                    iq = -1
                if iq >= 85:
                    c.fill = _fill("D1FAE5"); c.font = Font(name="Calibri", bold=True, size=9, color="065F46")
                elif iq >= 65:
                    c.fill = _fill("FEF9C3"); c.font = Font(name="Calibri", bold=True, size=9, color="713F12")
                elif iq >= 40:
                    c.fill = _fill("FED7AA"); c.font = Font(name="Calibri", bold=True, size=9, color="7C2D12")
                elif iq >= 0:
                    c.fill = _fill("FEE2E2"); c.font = Font(name="Calibri", bold=True, size=9, color="7F1D1D")
            elif col_key == "Rule15_Reason" and val:
                c.fill = _fill("FFF7ED")  # amber tint for escalation notes
            elif ri % 2 == 0:
                c.fill = _fill("F8FAFC")
        ws2.row_dimensions[ri].height = 36

    ws2.freeze_panes = "A4"

    # ── Sheet 3 — Full Log ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Full Log")
    ws3.sheet_view.showGridLines = False
    ws3.cell(row=1, column=1,
             value=f"Full Record Log — {n_total} record(s), all tiers"
    ).font = Font(name="Calibri", bold=True, size=12, color=C_NAVY)
    ws3.row_dimensions[1].height = 20

    full_cols = [
        ("Record ID",     "record_id",           16),
        ("Type",          "record_type",         14),
        ("Category",      "deviation_category",  18),
        ("System",        "system_name",         16),
        ("Open Date",     "open_date",           12),
        ("Close Date",    "close_date",          12),
        ("Status",        "status",              12),
        ("SLA Days",      "sla_days",            10),
        ("Assigned To",   "assigned_to",         16),
        ("Approved By",   "approved_by",         16),
        ("RCA Text",      "rca_text",            40),
        ("CAPA Text",     "capa_text",           40),
        ("CAPA Type",     "CAPA_Type",           14),
        ("Risk Tier",     "Risk_Tier",           10),
        ("Risk Score",    "Risk_Score",          10),
        ("IQI",           "IQI",                  8),
        ("IQI Band",      "IQI_Band",            14),
        ("IQI Drivers",   "IQI_Drivers",         55),
        ("Rule 15 Note",  "Rule15_Reason",       45),
        ("Primary Rule",  "Primary_Rule",        30),
    ]
    for sc, _, _ in _DCI_RULE_TIER_PRIORITY:
        rn = sc.split("rule")[1].split("_")[0]
        full_cols.append((f"R{rn} Score", sc, 9))

    for ci, (hdr, _, width) in enumerate(full_cols, 1):
        _hdr(ws3, 3, ci, hdr, width=width)
    ws3.row_dimensions[3].height = 18

    for ri, (_, row_data) in enumerate(scored_df.iterrows(), 4):
        tier = str(row_data.get("Risk_Tier", "Low"))
        tier_bg, tier_fg = _TIER_COLORS.get(tier, (C_GREY, C_DARK_TEXT))
        for ci, (_, col_key, _) in enumerate(full_cols, 1):
            val = row_data.get(col_key, "")
            if pd.isna(val):
                val = ""
            if col_key in ("open_date", "close_date"):
                try:
                    dt = pd.to_datetime(val, errors="coerce")
                    val = dt.strftime("%Y-%m-%d") if pd.notna(dt) else ""
                except Exception:
                    val = str(val) if val else ""
            c = ws3.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Calibri", size=8.5, color=C_DARK_TEXT)
            c.alignment = Alignment(
                horizontal="center" if ci > 1 else "left",
                vertical="top",
                wrap_text=col_key in ("rca_text", "capa_text", "IQI_Drivers", "Rule15_Reason"))
            c.border = bdr
            if col_key == "Risk_Tier":
                c.font = Font(name="Calibri", bold=True, size=9, color=tier_fg)
                c.fill = _fill(tier_bg)
            elif col_key == "IQI":
                try:
                    iq = int(val) if val != "" else -1
                except (ValueError, TypeError):
                    iq = -1
                if iq >= 85:
                    c.fill = _fill("D1FAE5"); c.font = Font(name="Calibri", bold=True, size=8.5, color="065F46")
                elif iq >= 65:
                    c.fill = _fill("FEF9C3"); c.font = Font(name="Calibri", bold=True, size=8.5, color="713F12")
                elif iq >= 40:
                    c.fill = _fill("FED7AA"); c.font = Font(name="Calibri", bold=True, size=8.5, color="7C2D12")
                elif iq >= 0:
                    c.fill = _fill("FEE2E2"); c.font = Font(name="Calibri", bold=True, size=8.5, color="7F1D1D")
            elif ri % 2 == 0:
                c.fill = _fill("F8FAFC")
        ws3.row_dimensions[ri].height = 22

    ws3.freeze_panes = "A4"

    # ── Sheet 4 — Detection Logic ───────────────────────────────────────
    ws4 = wb.create_sheet("Detection Logic")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 6
    ws4.column_dimensions["B"].width = 42
    ws4.column_dimensions["C"].width = 12
    ws4.column_dimensions["D"].width = 10
    ws4.column_dimensions["E"].width = 12
    ws4.column_dimensions["F"].width = 70

    ws4.cell(row=1, column=1, value="DCI Detection Rules — 14 Rules Across 4 Engines"
    ).font = Font(name="Calibri", bold=True, size=12, color=C_NAVY)
    ws4.row_dimensions[1].height = 20

    for ci, hdr in enumerate(
        ["#", "Rule Name", "Severity", "Tier", "Engine", "Fire Condition"], 1
    ):
        _hdr(ws4, 3, ci, hdr)
    ws4.row_dimensions[3].height = 18

    engine_labels = {
        "A": "RCA Recurrence",
        "B": "Weak Investigation",
        "C": "CAPA Effectiveness",
        "D": "SLA / Aging",
    }
    fire_conditions = {
        1:  "Same deviation_category in >=3 records within 180-day window.",
        2:  "Same deviation_category in >=5 records within 90-day window.",
        3:  "Same system_name has >=3 deviations within 60-day window.",
        4:  "Closed record with len(rca_text.strip()) < 50 chars.",
        5:  "Closed record, RCA contains vague term AND no specific-cause term. "
            "Substring match, case-insensitive. Non-blank RCA only.",
        6:  "Closed record with blank/null rca_text.",
        7:  "Closed record with blank/null capa_text.",
        8:  "Closed record, CAPA contains training term AND no action term. "
            "Substring match. Non-blank CAPA only.",
        9:  "Same record_type + system_name + deviation_category re-opens within "
            "90 days of a prior closure.",
        10: "Status normalizes to 're-opened' / 'reopened'.",
        11: "Closed record, close_date - open_date < threshold days (default 7). "
            "Both dates required.",
        12: "open_date + sla_days < today AND status != Closed. "
            "Requires numeric sla_days.",
        13: "open_date + sla_days - today <= near_breach_pct% elapsed AND status != Closed.",
        14: "Open >= 30 days, no close_date, status != Closed.",
        15: "CAPA_Type = Training AND (Rule 1 OR Rule 9 fired on same record). "
            "Escalates Risk_Tier one level. Does not change Risk_Score. "
            "Rationale: training-only CAPAs on recurrent failures indicate "
            "CAPA ineffectiveness per ICH Q10 §3.2.3.",
    }

    for ri, (num, name, sev, tier, eng, dflt, cfg_key) in enumerate(
        _DCI_RULE_META, 4
    ):
        _cell(ws4, ri, 1, str(num), bold=True, align="center")
        _cell(ws4, ri, 2, name)
        sev_bg, sev_fg = _TIER_COLORS.get(sev, (C_GREY, C_DARK_TEXT))
        c = _cell(ws4, ri, 3, sev, bold=True, align="center")
        c.font = Font(name="Calibri", bold=True, size=9, color=sev_fg)
        c.fill = _fill(sev_bg)
        _cell(ws4, ri, 4, tier, align="center")
        _cell(ws4, ri, 5, f"{eng} — {engine_labels.get(eng, eng)}", align="center")
        _cell(ws4, ri, 6, fire_conditions.get(num, ""), wrap=True)
        ws4.row_dimensions[ri].height = 30

    ri_ai = 4 + len(_DCI_RULE_META) + 2
    _hdr(ws4, ri_ai, 1, "AI Use in Scoring")
    ws4.merge_cells(start_row=ri_ai, start_column=1, end_row=ri_ai, end_column=6)
    _cell(ws4, ri_ai + 1, 1,
          "No AI is used in DCI rule scoring or tier assignment. All 15 rules "
          "are deterministic Python logic operating on keyword/date checks. "
          "AI (if present in narrative generation for Summary) is advisory text "
          "only — Risk_Score, Risk_Tier, and rule firing are fully reproducible "
          "from config hash + input file hash. Per ISPE GAMP® Guide: Artificial "
          "Intelligence (July 2025).",
          wrap=True)
    ws4.merge_cells(start_row=ri_ai+1, start_column=1,
                    end_row=ri_ai+1, end_column=6)
    ws4.row_dimensions[ri_ai + 1].height = 60

    # IQI explanation block
    ri_iqi = ri_ai + 4
    _hdr(ws4, ri_iqi, 1, "Investigation Quality Index (IQI) — How It Works")
    ws4.merge_cells(start_row=ri_iqi, start_column=1, end_row=ri_iqi, end_column=6)
    _cell(ws4, ri_iqi + 1, 1,
          "IQI = 100 × (1 − total_fired_score ÷ max_possible_score)  "
          "where max_possible = 9 × number of active rules (default 12 ON → 108 max).  "
          "IQI 100 = perfect investigation, no issues found.  "
          "IQI 0 = every active rule fired.  "
          "Bands: 85–100 Strong · 65–84 Acceptable · 40–64 Weak · 0–39 Poor.  "
          "IQI is a parallel quality lens — it does NOT replace Risk_Tier. "
          "Sort the Full Log by IQI ascending to find the weakest investigations first.",
          wrap=True)
    ws4.merge_cells(start_row=ri_iqi+1, start_column=1,
                    end_row=ri_iqi+1, end_column=6)
    ws4.row_dimensions[ri_iqi + 1].height = 70

    # CAPA Type explanation block
    ri_ct = ri_iqi + 4
    _hdr(ws4, ri_ct, 1, "CAPA Type Classification — Deterministic Keyword Matching")
    ws4.merge_cells(start_row=ri_ct, start_column=1, end_row=ri_ct, end_column=6)
    _cell(ws4, ri_ct + 1, 1,
          "Each CAPA is classified by control type using keyword matching on capa_text.  "
          "Types (effectiveness hierarchy, strongest first):  "
          "Engineering (hardware/automation/interlock) → "
          "Systemic (process redesign/workflow change) → "
          "Procedural (SOP/protocol/checklist update) → "
          "Training (retrain/awareness/reminder).  "
          "First match wins. Rule 15 escalates records where CAPA_Type = Training "
          "AND the same failure has already recurred (Rule 1 or Rule 9 fired).",
          wrap=True)
    ws4.merge_cells(start_row=ri_ct+1, start_column=1,
                    end_row=ri_ct+1, end_column=6)
    ws4.row_dimensions[ri_ct + 1].height = 80

    # ── Sheet 5 — Integrity Audit ────────────────────────────────────────
    ws5 = wb.create_sheet("Integrity Audit")
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions["A"].width = 28
    ws5.column_dimensions["B"].width = 70

    ws5.cell(row=1, column=1, value="DCI Integrity Audit"
    ).font = Font(name="Calibri", bold=True, size=12, color=C_NAVY)
    ws5.row_dimensions[1].height = 20

    try:
        file_hash = hashlib.sha256(str(fname).encode()).hexdigest()[:16]
    except Exception:
        file_hash = "—"

    audit_rows = [
        ("Module",                 "Deviation & CAPA Investigation"),
        ("Source File",            fname),
        ("Source File Hash",       file_hash),
        ("System Name",            system_name),
        ("Review Period",          f"{r_start} → {r_end}"),
        ("Analysis Timestamp",     pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Operator",               operator_user or "—"),
        ("Config Hash",            config_hash or "—"),
        ("Model Used (narrative)", model_used or "n/a — no AI in scoring"),
        ("Records Analyzed",       str(n_total)),
        ("Critical Findings",      str(n_critical)),
        ("High Findings",          str(n_high)),
        ("Medium Findings",        str(n_medium)),
        ("Regulatory References",  _REG_DCI),
    ]
    for ri, (k, v) in enumerate(audit_rows, 3):
        _cell(ws5, ri, 1, k, bold=True, bg=C_LIGHT)
        _cell(ws5, ri, 2, v, wrap=True)
        ws5.row_dimensions[ri].height = 18 if len(str(v)) < 60 else 32

    # ── Sheet 6 — Compliance Checklist ───────────────────────────────────
    ws6 = wb.create_sheet("Compliance Checklist")
    ws6.sheet_view.showGridLines = False
    ws6.column_dimensions["A"].width = 6
    ws6.column_dimensions["B"].width = 40
    ws6.column_dimensions["C"].width = 42
    ws6.column_dimensions["D"].width = 14

    ws6.cell(row=1, column=1,
             value="DCI Compliance Checklist — Fired Rules Mapped to Regulatory Clauses"
    ).font = Font(name="Calibri", bold=True, size=12, color=C_NAVY)
    ws6.row_dimensions[1].height = 20

    for ci, hdr in enumerate(["#", "Rule", "Regulatory Clause", "Count"], 1):
        _hdr(ws6, 3, ci, hdr)
    ws6.row_dimensions[3].height = 18

    reg_map = {
        1:  "ICH Q10 §3.2 · 21 CFR 820.100(a)(2)",
        2:  "21 CFR 820.100(a)(2) · ICH Q10 §3.2.3",
        3:  "EU Annex 11 §10 · ICH Q10 §3.2",
        4:  "21 CFR 820.100(b)(4) · ICH Q10 §3.2.2",
        5:  "FDA CAPA Guidance (2014) · 21 CFR 820.100(b)(2)",
        6:  "21 CFR 820.100(b)(2) · 21 CFR 211.192",
        7:  "21 CFR 820.100(a)(3) · 21 CFR 211.192",
        8:  "ICH Q10 §3.2.3 · FDA CAPA Guidance (2014)",
        9:  "21 CFR 820.100(a)(2) · ICH Q10 §3.2.3",
        10: "21 CFR 820.100(a)(7)",
        11: "ICH Q10 §3.2.2",
        12: "21 CFR 820.100(b)(5) · EU Annex 11 §10",
        13: "FDA CAPA Guidance (2014)",
        14: "ICH Q10 §3.2.2",
    }

    rule_fire_counts = {}
    for sc, thr, _ in _DCI_RULE_TIER_PRIORITY:
        if sc in scored_df.columns and n_total:
            cnt = int((scored_df[sc] >= thr).sum())
            rn = int(sc.split("rule")[1].split("_")[0])
            rule_fire_counts[rn] = cnt

    for ri, (num, name, _, _, _, _, _) in enumerate(_DCI_RULE_META, 4):
        _cell(ws6, ri, 1, str(num), bold=True, align="center")
        _cell(ws6, ri, 2, name)
        _cell(ws6, ri, 3, reg_map.get(num, ""), wrap=True)
        cnt = rule_fire_counts.get(num, 0)
        cell_cnt = _cell(ws6, ri, 4, str(cnt), align="center", bold=(cnt > 0))
        if cnt > 0:
            cell_cnt.fill = _fill("FEF2F2" if cnt >= 5 else "FFF7ED")
        ws6.row_dimensions[ri].height = 24

    wb.save(output)
    return output.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
#  DIM banking hook
# ═══════════════════════════════════════════════════════════════════════════
def _dci_bank_to_dim(scored_df, period_label, system_name, file_name,
                     event_category_fn=None):
    """Bank H/C DCI findings to st.session_state.dim_accumulated_rows.

    event_category_fn: callable that maps Rule_Triggered -> Event_Category.
    When called from generator.py, pass the _dim_event_category function so
    DCI rule names get properly classified to Investigation / Change Control.
    """
    if event_category_fn is None:
        # Fallback — use "Investigation" as default category for DCI findings
        event_category_fn = lambda rt: "Investigation"

    hc_df = scored_df[
        scored_df["Risk_Tier"].isin(["Critical", "High"])
    ] if not scored_df.empty else pd.DataFrame()

    dci_dim_rows = []
    for _, row in hc_df.iterrows():
        status_norm = _dci_normalize_status(row.get("status"))
        cd = _dci_parse_date(row.get("close_date"))
        od = _dci_parse_date(row.get("open_date"))
        if status_norm == "closed" and pd.notna(cd):
            event_ts = str(cd)
        elif pd.notna(od):
            event_ts = str(od)
        else:
            event_ts = ""

        rule_str = str(row.get("Primary_Rule", "DCI finding"))[:120]
        dci_dim_rows.append({
            "Review_Period":   period_label,
            "Username":        str(row.get("assigned_to", "unknown")),
            "Risk_Level":      str(row.get("Risk_Tier", "High")),
            "Rule_Triggered":  rule_str,
            "Event_Category":  event_category_fn(rule_str),
            "System_Name":     str(row.get("system_name", system_name)),
            "Event_Type":      str(row.get("record_type", "DEVIATION")),
            "Event_Timestamp": event_ts,
            "Source_File":     file_name,
            "Source_Module":   "DCI",
            "IQI":             int(row.get("IQI", 0)) if "IQI" in row.index else 0,
        })

    existing = [
        r for r in st.session_state.get("dim_accumulated_rows", [])
        if not (r.get("Review_Period") == period_label
                and r.get("Source_Module") == "DCI")
    ]

    if dci_dim_rows:
        existing.extend(dci_dim_rows)
        banked = len(dci_dim_rows)
    else:
        sentinel = {
            "Review_Period":   period_label,
            "Username":        "(no escalations)",
            "Risk_Level":      "Low",
            "Rule_Triggered":  "No named rules triggered",
            "Event_Category":  "Other",
            "System_Name":     system_name,
            "Event_Type":      "DCI_REVIEW",
            "Event_Timestamp": "",
            "Source_File":     file_name,
            "Source_Module":   "DCI",
        }
        existing.append(sentinel)
        banked = 1

    st.session_state["dim_accumulated_rows"] = existing
    st.session_state["dim_periods_banked"] = len(
        set(r["Review_Period"] for r in existing))
    st.session_state["dim_analysis_done"] = False
    st.session_state["dim_result"] = None

    return banked


# ═══════════════════════════════════════════════════════════════════════════
#  Event_Category classifier — exported for generator.py to use
# ═══════════════════════════════════════════════════════════════════════════
def dci_classify_event_category(rule_triggered):
    """Map a DCI Rule_Triggered label to Event_Category.
    Rules 9,10,11 -> Change Control; Rules 1-8, 12-14 -> Investigation.
    Returns None if the rule is not a DCI rule (caller falls back to
    existing AT/UAR logic)."""
    r = str(rule_triggered).lower()
    if any(x in r for x in (
        "repeat deviation post-closure", "re-opened capa",
        "reopened capa", "short close cycle"
    )):
        return "Change Control"
    if any(x in r for x in (
        "recurring category", "repeat category", "repeat system",
        "short rca", "vague rca", "missing rca", "missing capa",
        "weak capa", "overdue", "near-breach", "no activity"
    )):
        return "Investigation"
    return None  # Not a DCI rule — let caller handle


# ═══════════════════════════════════════════════════════════════════════════
#  UI — main entry point
# ═══════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════
#  UI helper — render results from session state (no upload required)
# ═══════════════════════════════════════════════════════════════════════════
def _render_dci_results_from_session(user, model_id):
    """Render previously-computed DCI results without requiring re-upload.

    Called from show_dci_review when user navigates back to DCI after
    visiting another module — the file_uploader has been reset to None
    but the analysis is still in session state. Mirrors the active-flow
    Results + Download + Bank sections.
    """
    scored_df = st.session_state.get("dci_scored_df")
    if scored_df is None or scored_df.empty:
        return
    sys_name  = st.session_state.get("dci_system_name", "System")
    file_name = st.session_state.get("dci_file_name", "")
    cfg_hash  = st.session_state.get("dci_config_hash", "")
    cfg       = st.session_state.get("dci_rule_config", dict(_DCI_RULE_DEFAULTS))

    # Reconstruct review-period strings from derived keys
    dci_r_start = st.session_state.get("dci_r_start_derived", "")
    dci_r_end   = st.session_state.get("dci_r_end_derived",   "")

    n_crit = int((scored_df["Risk_Tier"] == "Critical").sum())
    n_high = int((scored_df["Risk_Tier"] == "High").sum())
    n_med  = int((scored_df["Risk_Tier"] == "Medium").sum())
    n_low  = int((scored_df["Risk_Tier"] == "Low").sum())
    n_total_p = len(scored_df)

    st.markdown("---")
    st.markdown("### Results")

    # Color-coded risk tier cards — matches AT/UAR theme
    _tier_cards_p = [
        ("#fee2e2", "#dc2626", "#7f1d1d", "🔴 Critical", n_crit),
        ("#fef3c7", "#d97706", "#78350f", "🟠 High",     n_high),
        ("#dbeafe", "#2563eb", "#1e3a8a", "🔵 Medium",   n_med),
        ("#f0fdf4", "#16a34a", "#14532d", "🟢 Low",      n_low),
    ]
    _pcard_html = "<div style='display:flex;gap:12px;margin-bottom:6px;'>"
    for _bg, _accent, _text, _label, _val in _tier_cards_p:
        _pcard_html += (
            f"<div style='flex:1;background:{_bg};border:1.5px solid {_accent};"
            f"border-radius:10px;padding:14px 16px;text-align:center;'>"
            f"<div style='font-size:0.82rem;color:{_accent};font-weight:600;"
            f"letter-spacing:0.04em;margin-bottom:4px;'>{_label}</div>"
            f"<div style='font-size:2.1rem;font-weight:800;color:{_text};'>{_val}</div>"
            f"</div>"
        )
    _pcard_html += "</div>"
    st.markdown(_pcard_html, unsafe_allow_html=True)

    # Context explanation under tier cards
    st.markdown(
        "<div style='background:#f8fafc;border-left:3px solid #94a3b8;"
        "padding:8px 14px;border-radius:0 6px 6px 0;margin-bottom:14px;'>"
        "<span style='font-size:0.82rem;color:#475569;'>"
        "<b>🔴 Critical</b> — multiple serious gaps (missing CAPA, repeat post-closure, vague RCA). "
        "Requires immediate remediation and re-investigation. &nbsp;|&nbsp; "
        "<b>🟠 High</b> — significant issue present (short RCA, repeat system, overdue). "
        "Review and action required before next period. &nbsp;|&nbsp; "
        "<b>🔵 Medium</b> — pattern detected but lower severity (recurring category, near-breach). "
        "Monitor and address in current cycle. &nbsp;|&nbsp; "
        "<b>🟢 Low</b> — no rules fired. Investigation meets quality checks."
        "</span></div>",
        unsafe_allow_html=True,
    )

    if "IQI" in scored_df.columns and len(scored_df):
        _ivals = scored_df["IQI"].dropna()
        _imean = round(float(_ivals.mean()), 1)
        _ipoor = int((_ivals < 40).sum())

        def _iqi_color_p(val):
            if val >= 85: return ("#f0fdf4", "#16a34a", "#14532d")
            if val >= 65: return ("#dbeafe", "#2563eb", "#1e3a8a")
            if val >= 40: return ("#fef3c7", "#d97706", "#78350f")
            return ("#fee2e2", "#dc2626", "#7f1d1d")

        _ibg1p = _iqi_color_p(_imean)
        _ibg3p = ("#fee2e2", "#dc2626", "#7f1d1d") if _ipoor > 0 else ("#f0fdf4", "#16a34a", "#14532d")

        _iqi_html_p = (
            "<div style='display:flex;gap:12px;margin-bottom:6px;'>"
            f"<div style='flex:1;background:{_ibg1p[0]};border:1.5px solid {_ibg1p[1]};"
            f"border-radius:10px;padding:12px 16px;text-align:center;'>"
            f"<div style='font-size:0.78rem;color:{_ibg1p[1]};font-weight:600;'>📊 Investigation Quality Index (avg)</div>"
            f"<div style='font-size:1.8rem;font-weight:800;color:{_ibg1p[2]};'>{_imean}"
            f"<span style='font-size:0.9rem;font-weight:400;'>/100</span></div>"
            f"</div>"
            f"<div style='flex:1;background:{_ibg3p[0]};border:1.5px solid {_ibg3p[1]};"
            f"border-radius:10px;padding:12px 16px;text-align:center;'>"
            f"<div style='font-size:0.78rem;color:{_ibg3p[1]};font-weight:600;'>⚠️ Records with Poor IQI (&lt;40)</div>"
            f"<div style='font-size:1.8rem;font-weight:800;color:{_ibg3p[2]};'>{_ipoor}"
            f"<span style='font-size:0.9rem;font-weight:400;'> of {n_total_p}</span></div>"
            f"</div>"
            "</div>"
        )
        st.markdown(_iqi_html_p, unsafe_allow_html=True)
        st.markdown(
            "<div style='background:#f8fafc;border-left:3px solid #94a3b8;"
            "padding:8px 14px;border-radius:0 6px 6px 0;margin-bottom:14px;'>"
            "<span style='font-size:0.82rem;color:#475569;'>"
            "<b>IQI (Investigation Quality Index)</b> measures investigation completeness across all "
            "active rules. 100 = no issues found. Each rule that fires deducts points proportional "
            "to its severity. Bands: <b style='color:#15803d;'>85–100 Strong</b> · "
            "<b style='color:#1d4ed8;'>65–84 Acceptable</b> · "
            "<b style='color:#b45309;'>40–64 Weak</b> · "
            "<b style='color:#dc2626;'>0–39 Poor</b>. "
            "Check the IQI Drivers column in the evidence package to see exactly which rules "
            "fired and how many points each cost."
            "</span></div>",
            unsafe_allow_html=True,
        )

    review_df = scored_df[scored_df["Risk_Tier"].isin(
        ["Critical", "High", "Medium"])]
    if not review_df.empty:
        st.markdown("**Records for Review** (Critical + High + Medium):")
        display_cols = ["record_id", "record_type", "deviation_category",
                        "system_name", "status", "Risk_Tier", "Risk_Score",
                        "IQI", "IQI_Band", "CAPA_Type", "Primary_Rule"]
        display_cols = [c for c in display_cols if c in review_df.columns]
        st.dataframe(
            review_df[display_cols].head(50),
            use_container_width=True, hide_index=True,
        )
        if len(review_df) > 50:
            st.caption(f"Showing 50 of {len(review_df)} review records. "
                       "Download Excel for full set.")
    else:
        st.success("✅ No records flagged at Medium or higher.")

    # Download + Open DIM
    st.markdown("---")
    st.markdown("### Download Evidence Package")

    # Row 1: Download + Start New Analysis side by side
    dc1, dc2 = st.columns(2)
    with dc1:
        try:
            xlsx_bytes = dci_build_excel(
                scored_df,
                system_name=sys_name,
                r_start=dci_r_start,
                r_end=dci_r_end,
                fname=file_name,
                config_hash=cfg_hash,
                operator_user=user,
                model_used=model_id,
                rule_config=cfg,
            )
            st.download_button(
                "📥 Download DCI Evidence Package (xlsx)",
                data=xlsx_bytes,
                file_name=f"DCI_{sys_name}_{dci_r_end}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dci_download_btn_persistent",
                on_click=lambda: _gen()["log_audit"](
                    user, "DCI_DOWNLOAD", "REPORT",
                    new_value=f"DCI_{sys_name}",
                    reason=f"System: {sys_name} (re-download from session)",
                ),
            )
        except Exception as e:
            st.error(f"Excel build failed: {e}")

    with dc2:
        if st.button("🔄 Start New Analysis", key="dci_reset_btn_persistent",
                     use_container_width=True):
            for k in ["dci_scored_df", "dci_analysis_done",
                      "dci_last_run_fingerprint", "dci_config_hash",
                      "dci_file_name", "dci_pending_hash",
                      "dci_last_run_hash", "dci_last_run_filename",
                      "dci_invalidation_msg",
                      "dci_r_start_derived", "dci_r_end_derived",
                      "dci_dim_banked_period", "dci_dim_banked_count"]:
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()

    # Row 2: Open DIM full width
    st.markdown("<div style='margin-top:8px;'></div>", unsafe_allow_html=True)
    if st.button("📊 Open Data Integrity Monitor →", use_container_width=True,
                 key="dci_open_dim_persistent", type="primary"):
        import streamlit as _st2
        _st2.session_state["main_view"] = "dim"
        _st2.rerun()

    # DIM banked confirmation — below Open DIM
    _bp = st.session_state.get("dci_dim_banked_period")
    _bc = st.session_state.get("dci_dim_banked_count", 0)
    if _bp:
        _msg = (
            f"✅ **Banked to DIM automatically** — clean period ({_bp})."
            if (_bc <= 1 and n_crit == 0 and n_high == 0)
            else f"✅ **Banked to DIM automatically** — {_bc} finding(s) in period *{_bp}*. "
                 "Open DIM to run cross-module convergence analysis."
        )
        st.success(_msg)


def show_dci_review(user, role, model_id):
    """Render Periodic Review — Module 3: Deviation & CAPA Investigation.
    Mirrors show_user_access_review structure."""
    _gen()["_scroll_top"]()
    st.title("🔍 Deviation & CAPA Investigation")
    st.markdown(
        "<p style='color:#94a3b8;margin-top:-12px;'>"
        "Upload your deviation/CAPA log to run the deterministic investigation "
        "quality engine — 14 scoring rules across 4 engines (RCA recurrence, weak "
        "investigation, CAPA effectiveness, SLA aging). Produces a GxP evidence "
        "package for 21 CFR 820.100, ICH Q10 §3.2, and EU Annex 11 §10.</p>",
        unsafe_allow_html=True,
    )

    # v96 — invalidation banner: shown when a new-file upload cleared prior results
    _dci_inv_msg = st.session_state.get("dci_invalidation_msg", "")
    if _dci_inv_msg:
        st.warning(_dci_inv_msg)

    st.markdown("---")

    # System metadata — system name only (no date inputs, period derived from data like AT)
    _sn_col, _ = st.columns([2, 4])
    with _sn_col:
        st.session_state["dci_system_name"] = st.text_input(
            "System Name",
            value=st.session_state.get("dci_system_name", ""),
            placeholder="e.g. TrackWise, MasterControl, Veeva QMS",
            key="dci_sysname",
        )

    # Period strings derived from data (set after scoring, like AT)
    dci_r_start = st.session_state.get("dci_r_start_derived", "")
    dci_r_end   = st.session_state.get("dci_r_end_derived",   "")

    st.markdown("---")

    # Upload
    st.markdown("### 1. Upload Deviation/CAPA Log")

    # ── Collapsible "What columns do you need" — mandatory on top, optional below ──
    with st.expander("📋 What columns do you need? (click to expand)", expanded=False):
        st.markdown(
            "<p style='font-size:0.85rem;color:#94a3b8;margin-bottom:6px;'>"
            "Column aliases are auto-detected (e.g. <code>deviation_id</code> → "
            "<code>record_id</code>). Mandatory columns must be present to run analysis.</p>",
            unsafe_allow_html=True,
        )

        # ── Mandatory columns ──
        st.markdown(
            "<div style='font-weight:700;font-size:0.93rem;color:#1e3a5f;"
            "margin-bottom:4px;'>⬛ Mandatory columns</div>",
            unsafe_allow_html=True,
        )
        _mand_rows = [
            ("`record_id`",          "Unique deviation/CAPA identifier",         "DEV-2024-001"),
            ("`record_type`",        "Deviation, CAPA, NC, Investigation",        "Deviation"),
            ("`deviation_category`", "Cause category / classification",           "Equipment Failure"),
            ("`system_name`",        "Source GxP system",                        "LIMS, MES, EQMS"),
            ("`open_date`",          "When the record was opened",               "2024-03-15"),
            ("`close_date`",         "When the record was closed (blank if open)","2024-04-22"),
            ("`rca_text`",           "Root cause analysis narrative",             "Free-text narrative"),
            ("`capa_text`",          "CAPA description / actions taken",         "Free-text actions"),
            ("`assigned_to`",        "Investigator assigned",                     "jdoe"),
            ("`approved_by`",        "QA approver",                              "jsmith"),
            ("`status`",             "Open, Closed, Cancelled, Re-opened",        "Closed"),
            ("`sla_days`",           "Target close-by SLA in days",              "30"),
        ]
        _mand_table = (
            "| Column | Purpose | Example |\n|---|---|---|\n"
            + "\n".join(f"| {c} | {p} | {e} |" for c, p, e in _mand_rows)
        )
        st.markdown(_mand_table)

        st.markdown("<div style='margin-top:10px;'></div>", unsafe_allow_html=True)

        # ── Optional columns ──
        st.markdown(
            "<div style='font-weight:700;font-size:0.93rem;color:#15803d;"
            "margin-bottom:4px;'>🟩 Optional columns — enhance analysis when present</div>",
            unsafe_allow_html=True,
        )
        _opt_rows = [
            ("`training_expiry_date`", "Adds training expiry narrative to CAPA assessment",  "2025-06-30"),
            ("`reopen_date`",          "Explicit re-open date for Rule 10 (re-opened CAPA)", "2024-07-01"),
            ("`last_activity_date`",   "Last update date; used by Rule 14 (no activity)",    "2024-08-15"),
            ("`risk_level`",           "Pre-assigned risk — shown in evidence package",       "High"),
            ("`product`",              "Product or batch reference",                          "Batch 24B-01"),
        ]
        _opt_html = "".join(
            f"<div style='display:flex;gap:8px;padding:4px 0;border-bottom:1px solid #f0fdf4;'>"
            f"<span style='color:#15803d;font-family:monospace;min-width:160px;'>{c}</span>"
            f"<span style='color:#334155;font-size:0.88rem;flex:1;'>{p}</span>"
            f"<span style='color:#64748b;font-size:0.83rem;min-width:110px;'>{e}</span>"
            f"</div>"
            for c, p, e in _opt_rows
        )
        st.markdown(
            f"<div style='background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;"
            f"padding:10px 14px;'>{_opt_html}</div>",
            unsafe_allow_html=True,
        )

        st.markdown(
            "<p style='font-size:0.82rem;color:#94a3b8;margin-top:10px;'>"
            "14 detection rules across 4 engines — RCA Recurrence (1–3) · "
            "Weak Investigation (4–7) · CAPA Effectiveness (8–11) · SLA Aging (12–15)</p>",
            unsafe_allow_html=True,
        )

    # ── v96 — Sample template download ────────────────────────────────────────
    @st.cache_data
    def _build_dci_sample_xlsx() -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        wb.remove(ws)

        # Sheet 1 — Usage Instructions
        ws_use = wb.create_sheet("Usage Instructions")
        ws_use.sheet_view.showGridLines = False
        ws_use["A1"] = "Sample Deviation/CAPA Log Template — Usage Instructions"
        ws_use["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws_use["A1"].fill = PatternFill("solid", fgColor="1E3A5F")
        ws_use.merge_cells("A1:B1")
        ws_use.row_dimensions[1].height = 32

        instructions = [
            ("How to use this template",
             "Replace the sample rows on the 'Deviation Log' sheet with your "
             "own deviation/CAPA register export. Mandatory columns (★) must be "
             "present. Optional columns (◈) enhance the analysis when available. "
             "Save the file and upload it to VALINTEL's DCI module."),
            ("★ record_id",
             "Unique identifier for the deviation, CAPA, or NC record. "
             "Aliases accepted: deviation_id, capa_id, nc_id, event_id."),
            ("★ record_type",
             "Deviation / CAPA / NC / Investigation — drives rule applicability."),
            ("★ deviation_category",
             "Root-cause category or classification (Equipment Failure, "
             "Procedural Non-Compliance, Human Error, etc.). Required for "
             "Rules 1–3 (recurrence detection)."),
            ("★ system_name",
             "Source GxP system or process owner. Required for Rule 3 "
             "(repeat-system detection)."),
            ("★ open_date / close_date",
             "Dates in any common format. close_date can be blank for open "
             "records. Required for Rules 11–14 (SLA aging)."),
            ("★ rca_text",
             "Root cause analysis narrative. Length and content are scored "
             "against Rule 4 (Short RCA) and Rule 5 (Vague RCA — generic "
             "human-error / training-issue language)."),
            ("★ capa_text",
             "CAPA actions taken. Used by Rule 7 (Missing CAPA) and Rule 8 "
             "(Weak training-only CAPA)."),
            ("★ assigned_to / approved_by",
             "Investigator and QA approver user IDs."),
            ("★ status",
             "Open, Closed, Cancelled, Re-opened. Used by Rule 9 (repeat "
             "deviation post-closure) and Rule 10 (re-opened CAPA)."),
            ("★ sla_days",
             "Target close-by SLA in days. Used by Rules 11–14 to measure "
             "compliance against your own internal SLA."),
        ]
        # Mandatory section header styling
        ws_use.column_dimensions["A"].width = 38
        ws_use.column_dimensions["B"].width = 100

        # Section label: Mandatory
        _mand_hdr_row = 3
        _mand_label = ws_use.cell(row=_mand_hdr_row, column=1,
                                   value="MANDATORY COLUMNS (★)")
        _mand_label.font = Font(bold=True, color="FFFFFF", size=10)
        _mand_label.fill = PatternFill("solid", fgColor="1E3A5F")
        _mand_label.alignment = Alignment(vertical="center", indent=1)
        ws_use.merge_cells(f"A{_mand_hdr_row}:B{_mand_hdr_row}")
        ws_use.row_dimensions[_mand_hdr_row].height = 22

        for r, (k, v) in enumerate(instructions, _mand_hdr_row + 1):
            cell_k = ws_use.cell(row=r, column=1, value=k)
            cell_k.font = Font(bold=True, color="1E3A5F", size=11)
            cell_k.alignment = Alignment(vertical="top", wrap_text=True)
            cell_v = ws_use.cell(row=r, column=2, value=v)
            cell_v.font = Font(color="334155", size=11)
            cell_v.alignment = Alignment(vertical="top", wrap_text=True)
            ws_use.row_dimensions[r].height = 38

        # Section label: Optional columns
        _opt_start_row = _mand_hdr_row + 1 + len(instructions)
        _opt_label = ws_use.cell(row=_opt_start_row, column=1,
                                  value="OPTIONAL COLUMNS (◈)  — enhance analysis when present")
        _opt_label.font = Font(bold=True, color="FFFFFF", size=10)
        _opt_label.fill = PatternFill("solid", fgColor="166534")
        _opt_label.alignment = Alignment(vertical="center", indent=1)
        ws_use.merge_cells(f"A{_opt_start_row}:B{_opt_start_row}")
        ws_use.row_dimensions[_opt_start_row].height = 22

        optional_instructions = [
            ("◈ training_expiry_date",
             "Date the assigned investigator's GxP training expires. When present "
             "and expired, CAPA assessment narrative is enhanced with expiry date "
             "and regulatory citation. No score impact — narrative only."),
            ("◈ reopen_date",
             "Explicit date the record was re-opened. Strengthens Rule 10 "
             "(Re-opened CAPA) rationale when populated."),
            ("◈ last_activity_date",
             "Date of the most recent update or activity on the record. "
             "Used by Rule 14 (No Activity) to flag stale open records. "
             "If absent, Rule 14 is automatically skipped."),
            ("◈ risk_level",
             "Pre-assigned risk level from your QMS (Low / Medium / High / Critical). "
             "Included in the evidence package for cross-reference; not used in scoring."),
            ("◈ product",
             "Product name or batch reference. Included in evidence package "
             "for traceability; not used in scoring."),
        ]
        for r, (k, v) in enumerate(optional_instructions, _opt_start_row + 1):
            cell_k = ws_use.cell(row=r, column=1, value=k)
            cell_k.font = Font(bold=True, color="166534", size=11)
            cell_k.fill = PatternFill("solid", fgColor="F0FDF4")
            cell_k.alignment = Alignment(vertical="top", wrap_text=True)
            cell_v = ws_use.cell(row=r, column=2, value=v)
            cell_v.font = Font(color="166534", size=11)
            cell_v.fill = PatternFill("solid", fgColor="F0FDF4")
            cell_v.alignment = Alignment(vertical="top", wrap_text=True)
            ws_use.row_dimensions[r].height = 48

        # Run DCI row — after optional
        _run_row = _opt_start_row + 1 + len(optional_instructions)
        _rk = ws_use.cell(row=_run_row, column=1, value="Run DCI")
        _rk.font = Font(bold=True, color="334155", size=11)
        _rk.alignment = Alignment(vertical="top", wrap_text=True)
        _rv = ws_use.cell(row=_run_row, column=2,
                           value="Once uploaded, DCI will validate column presence, score every "
                                 "record across 15 rules, surface pattern findings (recurrence, "
                                 "weak investigations, SLA breaches), and produce a 5-sheet Excel "
                                 "evidence package.")
        _rv.font = Font(color="475569", size=11)
        _rv.alignment = Alignment(vertical="top", wrap_text=True)
        ws_use.row_dimensions[_run_row].height = 38

        # Sheet 2 — Deviation Log (sample data)
        ws_data = wb.create_sheet("Deviation Log")
        ws_data.sheet_view.showGridLines = False
        DCI_HEADERS_MANDATORY = [
            "record_id", "record_type", "deviation_category", "system_name",
            "open_date", "close_date", "rca_text", "capa_text",
            "assigned_to", "approved_by", "status", "sla_days",
        ]
        DCI_HEADERS_OPTIONAL = [
            "training_expiry_date", "reopen_date", "last_activity_date",
            "risk_level", "product",
        ]
        DCI_HEADERS = DCI_HEADERS_MANDATORY + DCI_HEADERS_OPTIONAL
        for ci, h in enumerate(DCI_HEADERS, 1):
            c = ws_data.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF", size=11)
            if h in DCI_HEADERS_OPTIONAL:
                c.fill = PatternFill("solid", fgColor="166534")  # green for optional
            else:
                c.fill = PatternFill("solid", fgColor="1E3A5F")  # navy for mandatory
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # 25 sample rows demonstrating all 14 detection rules
        SAMPLE_ROWS = [
            ("DEV-2024-001","Deviation","Equipment Failure","LIMS","2024-01-08","2024-02-15","Probe drift identified during weekly verification; calibration was within tolerance at last verification but had drifted over the 7-day period due to ambient temperature swings affecting the probe housing.","Probe replaced with temperature-compensated unit (P/N 47821); revised calibration interval from 7 to 3 days for this lab area; updated SOP-LIMS-014 §4.2.3.","jdoe","jsmith","Closed","30"),
            ("DEV-2024-002","Deviation","Equipment Failure","LIMS","2024-01-22","2024-02-25","Same probe drift issue.","Recalibrate.","jdoe","jsmith","Closed","30"),  # Rule 1 + Rule 4 (short RCA)
            ("DEV-2024-003","Deviation","Equipment Failure","LIMS","2024-02-14","2024-03-08","Probe drift again.","Replaced probe.","mclark","jsmith","Closed","30"),  # Rule 2 (repeat HIGH volume)
            ("DEV-2024-004","Deviation","Procedural Non-Compliance","MES","2024-02-05","2024-03-22","Operator failed to record batch start time within required 5 minutes — recorded 18 minutes late. Investigation traced delay to operator dealing with a concurrent equipment alarm on adjacent line.","Re-trained operator on SOP-MFG-007 §3.1; added secondary confirmation step requiring supervisor sign-off when alarm-handling delays primary documentation.","ptan","bjones","Closed","30"),
            ("DEV-2024-005","Deviation","Procedural Non-Compliance","MES","2024-02-19","2024-03-30","Same procedural issue, different operator.","Re-trained.","asmith","bjones","Closed","30"),  # Rule 1 again, Rule 8 (training-only CAPA)
            ("DEV-2024-006","CAPA","Human Error","LIMS","2024-03-01","2024-04-15","Operator entered wrong sample ID","Trained operator","jdoe","jsmith","Closed","30"),  # Rule 5 (vague human error) + Rule 8
            ("CAPA-2024-007","CAPA","Procedure Gap","QMS","2024-03-12","2024-05-08","SOP-QMS-022 §6 required interpretation by reviewer regarding which records constituted 'related batches'. Multiple investigators interpreted differently leading to inconsistent classifications.","Revised SOP-QMS-022 §6 with explicit decision tree; trained 12 QA reviewers; added quarterly calibration check on classification consistency.","ghoward","bjones","Closed","45"),
            ("DEV-2024-008","Deviation","Out of Specification","LIMS","2024-03-20","","","","","","Open","30"),  # Rule 6 (missing RCA), Rule 7 (missing CAPA), Rule 12 (overdue, currently)
            ("CAPA-2024-009","CAPA","Equipment Failure","LIMS","2024-04-02","2024-04-09","Probe drift.","Calibrate.","jdoe","jsmith","Closed","30"),  # Rule 11 (short close cycle)
            ("DEV-2024-010","Deviation","Documentation Gap","DMS","2024-04-08","2024-05-10","Approval signature missing on validation protocol VAL-2024-005 prior to execution. Discovered during in-process review when QA noted ambiguity in protocol §4.","Approval secured retrospectively with QA risk assessment documenting no impact; revised protocol issuance checklist requiring electronic confirmation of all approvals before assignment to executor.","ghoward","jsmith","Closed","30"),
            ("DEV-2024-011","Deviation","Documentation Gap","DMS","2024-04-22","2024-06-15","Same issue.","Same fix.","ghoward","jsmith","Closed","30"),  # Rule 1 (repeat)
            ("DEV-2024-012","Deviation","Documentation Gap","DMS","2024-05-08","2024-07-01","Documentation gap identified.","Approval secured retrospectively with QA documentation.","tnovak","jsmith","Closed","30"),
            ("CAPA-2024-013","CAPA","Procedural Non-Compliance","MES","2024-05-15","","Procedure non-compliance recurring across multiple operators despite 2024-Q1 retraining initiative. Root cause investigation identified that the training material did not address the alarm-handling delay scenario, which was the most common contributor.","Revised training material to include alarm-handling scenarios; added a competency-assessment quiz before unsupervised work; engineering controls being evaluated to reduce alarm frequency.","bjones","ghoward","Open","45"),  # Rule 13 (near-breach if SLA approaching)
            ("DEV-2024-014","Deviation","Equipment Failure","MES","2024-06-01","2024-08-20","Mixer interlock failure resulting in unplanned hold.","","jdoe","jsmith","Closed","30"),  # Rule 7 (missing CAPA), overdue
            ("DEV-2024-015","Deviation","Out of Specification","LIMS","2024-06-12","2024-07-18","Sample preparation step skipped.","Operator retrained.","ptan","bjones","Closed","30"),  # Rule 8
            ("CAPA-2024-016","CAPA","Equipment Failure","LIMS","2024-07-02","","Probe failure investigation continues.","","jdoe","jsmith","Open","30"),  # Rule 14 (no activity if old)
            ("DEV-2024-017","Deviation","Software Bug","EQMS","2024-07-15","2024-09-10","Software bug in EQMS preventing proper rendering of CAPA timelines on dashboard. Patch validated and deployed.","Patch v2.4.1 deployed; validation testing performed per VAL-2024-018; users notified of fix.","fchen","ghoward","Closed","60"),
            ("DEV-2024-018","Deviation","Procedural Non-Compliance","MES","2024-08-04","2024-09-22","Operator skipped step.","Retrained.","hrivera","bjones","Closed","30"),  # Rule 5 + Rule 8
            ("DEV-2024-019","Deviation","Out of Specification","LIMS","2024-08-25","2024-09-30","OOS investigation; assignable cause identified as instrument calibration drift exceeding action limit. Hardware replaced and calibration program revised.","Hardware replacement; calibration interval reduced from monthly to bi-weekly; trend monitoring added to weekly QC review.","jdoe","jsmith","Closed","30"),
            ("CAPA-2024-020","CAPA","Equipment Failure","LIMS","2024-09-12","2024-09-15","Calibration issue.","Calibrated.","jdoe","jsmith","Closed","30"),  # Rule 11 (short close, 3 days)
            ("DEV-2024-021","Deviation","Equipment Failure","MES","2024-10-08","","Mixer failure being investigated.","","ikim","jsmith","Open","30"),  # Rule 12 overdue
            ("CAPA-2024-022","CAPA","Equipment Failure","MES","2024-10-15","2024-11-20","Mixer interlock recurrence; deeper root cause analysis revealed PLC firmware regression on controller v3.2.1 that was not caught in pre-deployment testing.","Reverted to PLC firmware v3.1.8; extended pre-deployment testing protocol to include interlock-stress scenarios; documented in CC-2024-089.","ikim","jsmith","Closed","45"),
            ("DEV-2024-023","Deviation","Out of Specification","LIMS","2024-11-02","","Investigation in progress","","jdoe","jsmith","Open","30"),  # Rule 4 (short RCA), Rule 12 (overdue)
            ("DEV-2024-024","Deviation","Equipment Failure","MES","2024-11-15","2024-12-10","Recurring mixer failure post-CAPA closure on CAPA-2024-022. Suggests CAPA effectiveness not yet realized; new investigation underway.","Re-opened CAPA-2024-022 (now CAPA-2024-022-R1); engineering escalation to vendor for design review.","ikim","jsmith","Closed","30"),  # Rule 9 (repeat post-closure), Rule 10 (re-opened)
            ("CAPA-2024-022-R1","CAPA","Equipment Failure","MES","2024-12-12","","Re-opened from CAPA-2024-022. Vendor engaged for design review; interim controls in place.","","ikim","jsmith","Re-opened","45"),  # Rule 10
        ]
        for ri, row in enumerate(SAMPLE_ROWS, 2):
            for ci, val in enumerate(row, 1):
                cell = ws_data.cell(row=ri, column=ci, value=val)
                cell.font = Font(color="1E293B", size=11)
                cell.alignment = Alignment(horizontal="left", vertical="top",
                                            wrap_text=True, indent=1)
            ws_data.row_dimensions[ri].height = 60
        col_widths = [16, 12, 22, 14, 12, 12, 50, 50, 12, 12, 11, 9,
                      18, 14, 18, 12, 18]
        for ci, w in enumerate(col_widths, 1):
            ws_data.column_dimensions[get_column_letter(ci)].width = w
        ws_data.freeze_panes = "A2"
        ws_data.auto_filter.ref = f"A1:{get_column_letter(len(DCI_HEADERS))}1"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    st.download_button(
        label="📥 Download Sample Deviation/CAPA Template",
        data=_build_dci_sample_xlsx(),
        file_name="valintel_sample_dci_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dci_sample_download",
        help=(
            "25-row sample deviation/CAPA log covering all 14 detection rules. "
            "Sheet 1 (Usage) explains every column; Sheet 2 (Deviation Log) is "
            "the data — replace with your own export."
        ),
    )
    st.markdown("<div style='margin-top:10px;'></div>", unsafe_allow_html=True)

    dci_file = st.file_uploader(
        "Excel or CSV file containing your deviation and CAPA records",
        type=["xlsx", "xls", "csv"],
        key="dci_uploader",
    )

    # ── Persistent-results path ──────────────────────────────────────────────
    # Streamlit's file_uploader resets to None when the user navigates away
    # and comes back. If we have analysis results from a prior run, surface
    # them instead of forcing a fresh upload. User can hit "Clear results"
    # to start over.
    _prior_scored = st.session_state.get("dci_scored_df")
    _prior_done   = st.session_state.get("dci_analysis_done", False)
    _prior_fname  = st.session_state.get("dci_file_name", "")

    if dci_file is None and _prior_done and _prior_scored is not None:
        st.info(
            f"📊 Showing results from previous analysis of "
            f"**{_prior_fname or 'last uploaded file'}**. "
            "Upload a new file or click **Clear results** to start over."
        )
        # Clear button — single click to reset the whole DCI session
        _cc1, _cc2 = st.columns([1, 4])
        with _cc1:
            if st.button("🗑️ Clear results", key="dci_clear_results_btn",
                         use_container_width=True):
                for _k in ("dci_scored_df", "dci_analysis_done",
                           "dci_last_run_fingerprint", "dci_config_hash",
                           "dci_file_name"):
                    if _k in st.session_state:
                        del st.session_state[_k]
                st.rerun()

        # Render results from session state — bypass upload/validate/run flow
        _render_dci_results_from_session(user, model_id)
        return

    if dci_file is None:
        st.caption("Upload your deviation/CAPA log above to begin. "
                   "Expand 'What columns do you need?' for column guidance.")
        return

    try:
        raw_bytes = dci_file.getvalue()
        file_name = dci_file.name

        # v96 — upload invalidation: detect content change vs last successful run
        _dci_check = _gen().get("_check_and_invalidate_on_new_upload")
        _dci_hash_fn = _gen().get("_file_content_hash")
        if _dci_check and _dci_hash_fn:
            _dci_new_hash = _dci_hash_fn(raw_bytes)
            _dci_check(
                module="dci",
                new_hash=_dci_new_hash,
                new_filename=file_name,
                invalidate_keys=[
                    "dci_scored_df", "dci_analysis_done",
                    "dci_last_run_fingerprint", "dci_config_hash",
                ],
            )
            st.session_state["dci_pending_hash"] = _dci_new_hash

        st.session_state["dci_file_name"] = file_name
        if file_name.lower().endswith(".csv"):
            dci_df_raw = pd.read_csv(io.BytesIO(raw_bytes), dtype=str).fillna("")
            all_sheets = [file_name]
        else:
            xl = pd.ExcelFile(io.BytesIO(raw_bytes))
            all_sheets = xl.sheet_names
            preferred = None
            for s in all_sheets:
                sn = str(s).lower()
                if any(k in sn for k in ("deviation", "capa", "incident",
                                          "investigation", "nc", "record")):
                    preferred = s; break
            sheet_to_use = preferred or all_sheets[0]
            dci_df_raw = pd.read_excel(
                io.BytesIO(raw_bytes), sheet_name=sheet_to_use, dtype=str
            ).fillna("")
    except Exception as e:
        st.error(f"Failed to read file: {e}")
        return

    # Validator
    v_ok, v_sev, v_title, v_results, v_evidence = _validate_dci_input_file(
        raw_bytes, file_name, dci_df_raw, all_sheets
    )

    override_key = f"validator_override__DCI__{file_name}"
    already_overridden = st.session_state.get(override_key, False)

    if not (_gen()["_render_validator_verdict"](
        v_sev, v_title, v_results, v_evidence,
        file_name, user, module="DCI"
    ) or already_overridden):
        return

    # Apply column aliases (auto-detection from _DCI_COLUMN_ALIASES)
    dci_df = dci_df_raw.copy()
    col_rename = {}
    for col in dci_df.columns:
        norm = str(col).strip().lower().replace(" ", "_")
        if norm in _DCI_COLUMN_ALIASES:
            canonical = _DCI_COLUMN_ALIASES[norm]
            if canonical != col:
                col_rename[col] = canonical
    if col_rename:
        dci_df = dci_df.rename(columns=col_rename)

    # ── Column Mapping — always shown, all 12 required columns ──────────────
    # Show before the hard-stop so users can fix missing columns in-UI.
    # Green tick = auto-resolved. Dropdown = needs manual selection.
    _raw_file_cols = list(dci_df_raw.columns)
    _file_col_options = ["— Skip —"] + _raw_file_cols
    _map_override = {}
    _col_map_needed = False  # True if any required col still needs manual mapping

    # Build mapping state: for each required canonical col, what did we resolve to?
    _resolved = {}   # canon → actual column name in dci_df (auto or already renamed)
    for _canon in sorted(_DCI_REQUIRED_COLS):
        if _canon in dci_df.columns:
            _resolved[_canon] = _canon  # already there (direct match or alias rename)
        else:
            _col_map_needed = True

    with st.expander(
        "🗂️ Column Mapping" + (" — ⚠️ some columns need mapping" if _col_map_needed else " — all columns auto-detected ✓"),
        expanded=_col_map_needed
    ):
        st.markdown(
            "<p style='font-size:0.85rem;color:#94a3b8;margin-bottom:10px;'>"
            "Common column name aliases are auto-detected. "
            "Green rows are confirmed. Use dropdowns to map any undetected columns, "
            "or leave as <em>— Skip —</em> to exclude that rule engine from scoring.</p>",
            unsafe_allow_html=True,
        )
        _map_cols_ui = st.columns(3)
        for _mi, _canon in enumerate(sorted(_DCI_REQUIRED_COLS)):
            with _map_cols_ui[_mi % 3]:
                if _canon in _resolved:
                    # Auto-resolved — show as green confirmed label
                    _matched = _resolved[_canon]
                    st.markdown(
                        f"<div style='background:#f0fdf4;border:1px solid #bbf7d0;"
                        f"border-radius:6px;padding:6px 10px;margin-bottom:6px;'>"
                        f"<div style='font-size:0.72rem;color:#15803d;font-weight:600;'>"
                        f"✓ `{_canon}`</div>"
                        f"<div style='font-size:0.8rem;color:#166534;'>"
                        f"← <code>{_matched}</code></div>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
                else:
                    # Not resolved — show selectbox
                    _sel = st.selectbox(
                        f"⚠️ `{_canon}`",
                        options=_file_col_options,
                        index=0,
                        key=f"dci_colmap_{_canon}",
                        help=f"Select the column in your file that contains '{_canon}' data",
                    )
                    if _sel != "— Skip —":
                        _map_override[_sel] = _canon

    # Apply manual mappings
    if _map_override:
        dci_df = dci_df.rename(columns=_map_override)

    # Hard stop — only if still missing after manual mapping
    missing_required = _DCI_REQUIRED_COLS - set(dci_df.columns)
    if missing_required:
        st.error(
            f"❌ {len(missing_required)} required column(s) still unmapped: "
            f"{', '.join(sorted(missing_required))}. "
            "Use the Column Mapping section above to assign them, or add these "
            "columns to your file and re-upload."
        )
        return

    st.success(f"✓ File loaded: **{len(dci_df)} record(s)** · {len(dci_df.columns)} columns")

    # ── Data preview — top 10 rows, expandable ────────────────────────────────
    with st.expander("🔍 Data Preview — first 10 rows (click to expand)", expanded=False):
        st.dataframe(dci_df_raw.head(10), use_container_width=True, hide_index=True)

    # Rule display — read-only, all rules always active, shown chronologically
    st.markdown("---")
    st.markdown("### 2. Detection Rules")
    st.caption(
        "15 rules across 4 engines — all active. Rules are deterministic: "
        "keyword-match and threshold-based only, no AI in the scoring path."
    )

    # Always use all defaults — no user config
    cfg = dict(_DCI_RULE_DEFAULTS)
    # Force all rules ON (Rules 11 and 14 were optional — now all active)
    for k in cfg:
        cfg[k] = True
    st.session_state["dci_rule_config"] = cfg

    engine_labels = {
        "A": ("🟡", "RCA Recurrence",     "Rules 1–3  ·  Recurring categories, repeat systems"),
        "B": ("🔴", "Weak Investigation",  "Rules 4–7  ·  Short RCA, vague cause, missing CAPA"),
        "C": ("🔵", "CAPA Effectiveness",  "Rules 8–11 ·  Training-only CAPA, re-opened, repeat post-closure"),
        "D": ("🟣", "SLA / Aging",         "Rules 12–13 · Overdue, near-breach"),
    }

    active_count = len(_DCI_RULE_META)  # Always all rules
    st.markdown(
        f"<div style='background:#f0fdf4;border:1px solid #bbf7d0;border-radius:7px;"
        f"padding:7px 14px;margin-bottom:8px;display:inline-block;'>"
        f"<span style='color:#15803d;font-weight:600;font-size:0.88rem;'>"
        f"✓ All {active_count} rules active — no configuration required</span></div>",
        unsafe_allow_html=True,
    )

    # Regulation mapped to each rule — shown in UI expander
    _DCI_RULE_REGS = {
        1:  "ICH Q10 §3.2",
        2:  "ICH Q10 §3.2.3 · 21 CFR 820.100(a)(2)",
        3:  "EU Annex 11 §10 · ICH Q10 §3.2",
        4:  "21 CFR 820.100(b)(4) · ICH Q10 §3.2.2",
        5:  "FDA CAPA Guidance (2014) · 21 CFR 820.100(b)(2)",
        6:  "21 CFR 820.100(b)(2) · 21 CFR 211.192",
        7:  "21 CFR 820.100(a)(3) · 21 CFR 211.192",
        8:  "ICH Q10 §3.2.3 · FDA CAPA Guidance (2014)",
        9:  "21 CFR 820.100(a)(2) · ICH Q10 §3.2.3",
        10: "21 CFR 820.100(a)(7)",
        11: "ICH Q10 §3.2.2",
        12: "21 CFR 820.100(b)(5) · EU Annex 11 §10",
        13: "FDA CAPA Guidance (2014)",
        14: "ICH Q10 §3.2.2",
        15: "ICH Q10 §3.2.3",
    }

    with st.expander("View all 15 detection rules", expanded=False):
        # Sort rules chronologically by rule number
        sorted_rules = sorted(_DCI_RULE_META, key=lambda m: m[0])
        current_engine = None
        for num, name, sev, tier, engine_code, _, cfg_key in sorted_rules:
            if engine_code != current_engine:
                current_engine = engine_code
                em, en, edesc = engine_labels[engine_code]
                st.markdown(
                    f"<div style='background:#f8fafc;border-left:3px solid #334155;"
                    f"padding:6px 12px;margin:10px 0 4px 0;'>"
                    f"<span style='font-weight:700;font-size:0.9rem;'>{em} {en}</span>"
                    f"<span style='color:#64748b;font-size:0.82rem;margin-left:10px;'>{edesc}</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            _sev_colors = {
                "Critical": ("#fee2e2", "#dc2626"),
                "High":     ("#fef3c7", "#d97706"),
                "Medium":   ("#dbeafe", "#2563eb"),
            }
            _sbg, _sc = _sev_colors.get(sev, ("#f1f5f9", "#475569"))
            _reg = _DCI_RULE_REGS.get(num, "")
            st.markdown(
                f"<div style='display:flex;align-items:center;gap:10px;"
                f"padding:5px 8px;margin-bottom:2px;border-bottom:1px solid #f1f5f9;'>"
                f"<span style='font-size:0.82rem;color:#334155;min-width:210px;'>"
                f"<b>Rule {num}</b> — {name}</span>"
                f"<span style='background:{_sbg};color:{_sc};border-radius:4px;"
                f"padding:1px 7px;font-size:0.75rem;font-weight:600;min-width:64px;text-align:center;'>{sev}</span>"
                f"<span style='font-size:0.74rem;color:#94a3b8;font-family:monospace;'>{_reg}</span>"
                f"</div>",
                unsafe_allow_html=True,
            )

    # Run Analysis
    st.markdown("---")
    st.markdown("### 3. Run Analysis")

    dci_running = st.session_state.get("dci_running", False)

    # ── Fingerprint of current inputs ────────────────────────────────────────
    # If file + config haven't changed since last successful run, the Run
    # button is disabled with a helpful message. User must change the file
    # OR toggle a rule config to enable a re-run.
    import json as _json_fp, hashlib as _hash_fp
    _input_fp = _hash_fp.sha256(
        _json_fp.dumps(
            {
                "file":  file_name,
                "rows":  len(dci_df),
                "cols":  sorted(dci_df.columns.tolist()),
                "cfg":   {k: bool(v) for k, v in cfg.items()},
            },
            sort_keys=True,
        ).encode()
    ).hexdigest()[:16]
    _last_run_fp = st.session_state.get("dci_last_run_fingerprint")
    _already_analyzed_this_input = (
        _last_run_fp == _input_fp
        and st.session_state.get("dci_analysis_done", False)
    )

    run_col1, _run_col2 = st.columns([1, 3])
    with run_col1:
        if dci_running:
            _btn_label    = "⏳ Running…"
            _btn_disabled = True
        elif _already_analyzed_this_input:
            _btn_label    = "✓ Analyzed — change file or rules to re-run"
            _btn_disabled = True
        else:
            _btn_label    = "🚀 Run DCI Analysis"
            _btn_disabled = False
        run_clicked = st.button(
            _btn_label,
            disabled=_btn_disabled,
            key="dci_run_btn",
            use_container_width=True,
        )

    if run_clicked:
        st.session_state["dci_running"] = True
        try:
            with st.spinner("Scoring records across 14 rules…"):
                scored_df = dci_score_records(dci_df, rule_config=cfg)
            st.session_state["dci_scored_df"] = scored_df

            # Derive period dates from data (like AT — no manual date pickers)
            try:
                if "open_date" in scored_df.columns:
                    _dates = pd.to_datetime(scored_df["open_date"], errors="coerce").dropna()
                    if len(_dates):
                        st.session_state["dci_r_start_derived"] = _dates.min().strftime("%d-%b-%Y")
                        st.session_state["dci_r_end_derived"]   = _dates.max().strftime("%d-%b-%Y")
                    else:
                        st.session_state["dci_r_start_derived"] = ""
                        st.session_state["dci_r_end_derived"]   = ""
            except Exception:
                st.session_state["dci_r_start_derived"] = ""
                st.session_state["dci_r_end_derived"]   = ""
            st.session_state["dci_analysis_done"] = True

            # v96 — record run hash for upload invalidation on next file
            _dci_record = _gen().get("_record_run_hash")
            if _dci_record:
                _dci_record(
                    module="dci",
                    file_hash=st.session_state.get("dci_pending_hash", ""),
                    filename=st.session_state.get("dci_file_name", ""),
                )
            st.session_state["dci_last_run_fingerprint"] = _input_fp
            try:
                cfg_str = _json_fp.dumps(cfg, sort_keys=True)
                st.session_state["dci_config_hash"] = \
                    _hash_fp.sha256(cfg_str.encode()).hexdigest()[:16]
            except Exception:
                st.session_state["dci_config_hash"] = ""

            # ── Auto-bank to DIM immediately after scoring ────────────────────
            # No button required — DCI findings always bank on run completion,
            # same as AT and UAR. Re-banking the same period is idempotent
            # (existing rows for this period+module are replaced, not duplicated).
            try:
                _ec_fn = _gen().get("_dim_event_category")
                # Period label mirrors UAR pattern: "date_range (filename_stem)"
                # so two different CAPA exports for the same date range get
                # distinct DIM period keys and don't overwrite each other.
                _dci_fname_stem = file_name.rsplit(".", 1)[0][:30] if file_name else ""
                _dci_r_start_d = st.session_state.get("dci_r_start_derived", "")
                _dci_r_end_d   = st.session_state.get("dci_r_end_derived", "")
                _dci_date_range = (
                    f"{_dci_r_start_d} → {_dci_r_end_d}"
                    if (_dci_r_start_d and _dci_r_end_d) else ""
                )
                if _dci_date_range and _dci_fname_stem:
                    _auto_period = f"{_dci_date_range} ({_dci_fname_stem})"
                elif _dci_date_range:
                    _auto_period = _dci_date_range
                else:
                    _auto_period = f"DCI Period {st.session_state.get('dim_periods_banked', 0) + 1}"
                _auto_banked = _dci_bank_to_dim(
                    scored_df,
                    period_label=_auto_period,
                    system_name=st.session_state.get("dci_system_name", "System"),
                    file_name=file_name,
                    event_category_fn=_ec_fn,
                )
                st.session_state["dci_dim_banked_period"] = _auto_period
                st.session_state["dci_dim_banked_count"]  = _auto_banked
            except Exception as _be:
                # Banking failure is non-fatal — results are still shown
                st.session_state["dci_dim_banked_period"] = None
                st.session_state["dci_dim_banked_count"]  = 0

            _gen()["log_audit"](
                user, "DCI_ANALYSIS_RUN", "DATASET",
                new_value=f"DCI_{st.session_state.get('dci_system_name','')}",
                reason=(
                    f"Records: {len(scored_df)} · "
                    f"Rules active: {active_count} · "
                    f"File: {file_name}"
                ),
            )
        except Exception as e:
            st.error(f"Analysis failed: {e}")
            st.session_state["dci_analysis_done"] = False
            st.session_state["dci_last_run_fingerprint"] = None
        finally:
            st.session_state["dci_running"] = False
            st.rerun()

    if not st.session_state.get("dci_analysis_done"):
        return

    # Results
    scored_df = st.session_state.get("dci_scored_df")
    if scored_df is None or scored_df.empty:
        st.warning("No records scored.")
        return

    st.markdown("---")
    _hr1, _hr2 = st.columns([4, 1])
    with _hr1:
        st.markdown("### 4. Results")
    with _hr2:
        if st.button("🗑️ Clear results", key="dci_clear_results_btn_inflow",
                     use_container_width=True):
            for _k in ("dci_scored_df", "dci_analysis_done",
                       "dci_last_run_fingerprint", "dci_config_hash",
                       "dci_file_name"):
                if _k in st.session_state:
                    del st.session_state[_k]
            st.rerun()

    n_crit = int((scored_df["Risk_Tier"] == "Critical").sum())
    n_high = int((scored_df["Risk_Tier"] == "High").sum())
    n_med  = int((scored_df["Risk_Tier"] == "Medium").sum())
    n_low  = int((scored_df["Risk_Tier"] == "Low").sum())
    n_total = len(scored_df)

    # Color-coded risk tier cards — matches AT/UAR theme
    _tier_cards = [
        ("#fee2e2", "#dc2626", "#7f1d1d", "🔴 Critical", n_crit),
        ("#fef3c7", "#d97706", "#78350f", "🟠 High",     n_high),
        ("#dbeafe", "#2563eb", "#1e3a8a", "🔵 Medium",   n_med),
        ("#f0fdf4", "#16a34a", "#14532d", "🟢 Low",      n_low),
    ]
    _card_html = "<div style='display:flex;gap:12px;margin-bottom:6px;'>"
    for _bg, _accent, _text, _label, _val in _tier_cards:
        _card_html += (
            f"<div style='flex:1;background:{_bg};border:1.5px solid {_accent};"
            f"border-radius:10px;padding:14px 16px;text-align:center;'>"
            f"<div style='font-size:0.82rem;color:{_accent};font-weight:600;"
            f"letter-spacing:0.04em;margin-bottom:4px;'>{_label}</div>"
            f"<div style='font-size:2.1rem;font-weight:800;color:{_text};'>{_val}</div>"
            f"</div>"
        )
    _card_html += "</div>"
    st.markdown(_card_html, unsafe_allow_html=True)

    # Context explanation under tier cards
    st.markdown(
        "<div style='background:#f8fafc;border-left:3px solid #94a3b8;"
        "padding:8px 14px;border-radius:0 6px 6px 0;margin-bottom:14px;'>"
        "<span style='font-size:0.82rem;color:#475569;'>"
        "<b>🔴 Critical</b> — multiple serious gaps (missing CAPA, repeat post-closure, vague RCA). "
        "Requires immediate remediation and re-investigation. &nbsp;|&nbsp; "
        "<b>🟠 High</b> — significant issue present (short RCA, repeat system, overdue). "
        "Review and action required before next period. &nbsp;|&nbsp; "
        "<b>🔵 Medium</b> — pattern detected but lower severity (recurring category, near-breach). "
        "Monitor and address in current cycle. &nbsp;|&nbsp; "
        "<b>🟢 Low</b> — no rules fired. Investigation meets quality checks."
        "</span></div>",
        unsafe_allow_html=True,
    )

    # IQI summary — 2 cards only (Period avg + Poor count); Lowest Record removed
    if "IQI" in scored_df.columns and len(scored_df):
        _ivals = scored_df["IQI"].dropna()
        _imean = round(float(_ivals.mean()), 1)
        _ipoor = int((_ivals < 40).sum())

        def _iqi_color(val):
            if val >= 85: return ("#f0fdf4", "#16a34a", "#14532d")
            if val >= 65: return ("#dbeafe", "#2563eb", "#1e3a8a")
            if val >= 40: return ("#fef3c7", "#d97706", "#78350f")
            return ("#fee2e2", "#dc2626", "#7f1d1d")

        _ibg1 = _iqi_color(_imean)
        _ibg3 = ("#fee2e2", "#dc2626", "#7f1d1d") if _ipoor > 0 else ("#f0fdf4", "#16a34a", "#14532d")

        _iqi_html = (
            "<div style='display:flex;gap:12px;margin-bottom:6px;'>"
            f"<div style='flex:1;background:{_ibg1[0]};border:1.5px solid {_ibg1[1]};"
            f"border-radius:10px;padding:12px 16px;text-align:center;'>"
            f"<div style='font-size:0.78rem;color:{_ibg1[1]};font-weight:600;'>📊 Investigation Quality Index (avg)</div>"
            f"<div style='font-size:1.8rem;font-weight:800;color:{_ibg1[2]};'>{_imean}"
            f"<span style='font-size:0.9rem;font-weight:400;'>/100</span></div>"
            f"</div>"
            f"<div style='flex:1;background:{_ibg3[0]};border:1.5px solid {_ibg3[1]};"
            f"border-radius:10px;padding:12px 16px;text-align:center;'>"
            f"<div style='font-size:0.78rem;color:{_ibg3[1]};font-weight:600;'>⚠️ Records with Poor IQI (&lt;40)</div>"
            f"<div style='font-size:1.8rem;font-weight:800;color:{_ibg3[2]};'>{_ipoor}"
            f"<span style='font-size:0.9rem;font-weight:400;'> of {n_total}</span></div>"
            f"</div>"
            "</div>"
        )
        st.markdown(_iqi_html, unsafe_allow_html=True)
        # IQI context explanation
        st.markdown(
            "<div style='background:#f8fafc;border-left:3px solid #94a3b8;"
            "padding:8px 14px;border-radius:0 6px 6px 0;margin-bottom:14px;'>"
            "<span style='font-size:0.82rem;color:#475569;'>"
            "<b>IQI (Investigation Quality Index)</b> measures investigation completeness across all "
            "active rules. 100 = no issues found. Each rule that fires deducts points proportional "
            "to its severity. Bands: <b style='color:#15803d;'>85–100 Strong</b> · "
            "<b style='color:#1d4ed8;'>65–84 Acceptable</b> · "
            "<b style='color:#b45309;'>40–64 Weak</b> · "
            "<b style='color:#dc2626;'>0–39 Poor</b>. "
            "Check the IQI Drivers column in the evidence package to see exactly which rules "
            "fired and how many points each cost."
            "</span></div>",
            unsafe_allow_html=True,
        )

    review_df = scored_df[scored_df["Risk_Tier"].isin(
        ["Critical", "High", "Medium"])]
    if not review_df.empty:
        st.markdown("**Records for Review** (Critical + High + Medium):")
        display_cols = ["record_id", "record_type", "deviation_category",
                        "system_name", "status", "Risk_Tier", "Risk_Score",
                        "IQI", "IQI_Band", "CAPA_Type", "Primary_Rule"]
        display_cols = [c for c in display_cols if c in review_df.columns]
        st.dataframe(
            review_df[display_cols].head(50),
            use_container_width=True, hide_index=True,
        )
        if len(review_df) > 50:
            st.caption(f"Showing 50 of {len(review_df)} review records. "
                       "Download Excel for full set.")
    else:
        st.success("✅ No records flagged at Medium or higher.")

    # Download + Open DIM
    st.markdown("---")
    st.markdown("### 5. Download Evidence Package")

    sys_name = st.session_state.get("dci_system_name", "System")
    cfg_hash = st.session_state.get("dci_config_hash", "")
    dci_r_start = st.session_state.get("dci_r_start_derived", "")
    dci_r_end   = st.session_state.get("dci_r_end_derived", "")

    # Row 1: Download + Start New Analysis side by side
    dc1, dc2 = st.columns(2)
    with dc1:
        try:
            xlsx_bytes = dci_build_excel(
                scored_df,
                system_name=sys_name,
                r_start=dci_r_start,
                r_end=dci_r_end,
                fname=file_name,
                config_hash=cfg_hash,
                operator_user=user,
                model_used=model_id,
                rule_config=cfg,
            )
            st.download_button(
                "📥 Download DCI Evidence Package (xlsx)",
                data=xlsx_bytes,
                file_name=f"DCI_{sys_name}_{dci_r_end}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                on_click=lambda: _gen()["log_audit"](
                    user, "DCI_DOWNLOAD", "REPORT",
                    new_value=f"DCI_{sys_name}",
                    reason=f"System: {sys_name}",
                ),
            )
        except Exception as e:
            st.error(f"Excel build failed: {e}")

    with dc2:
        if st.button("🔄 Start New Analysis", key="dci_reset_btn",
                     use_container_width=True):
            for k in ["dci_scored_df", "dci_analysis_done",
                      "dci_last_run_fingerprint", "dci_config_hash",
                      "dci_file_name", "dci_pending_hash",
                      "dci_last_run_hash", "dci_last_run_filename",
                      "dci_invalidation_msg",
                      "dci_r_start_derived", "dci_r_end_derived",
                      "dci_dim_banked_period", "dci_dim_banked_count"]:
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()

    # Row 2: Open DIM (full width)
    st.markdown("<div style='margin-top:8px;'></div>", unsafe_allow_html=True)
    if st.button("📊 Open Data Integrity Monitor →", key="dci_open_dim_btn",
                 use_container_width=True, type="primary"):
        st.session_state["main_view"] = "dim"
        st.rerun()

    # DIM banked confirmation — below Open DIM
    _banked_period = st.session_state.get("dci_dim_banked_period")
    _banked_count  = st.session_state.get("dci_dim_banked_count", 0)
    if _banked_period:
        if _banked_count <= 1 and n_crit == 0 and n_high == 0:
            st.success(
                f"✅ **Banked to DIM automatically** — clean period "
                f"({_banked_period}). DCI ran with zero High/Critical findings."
            )
        else:
            st.success(
                f"✅ **Banked to DIM automatically** — {_banked_count} "
                f"High/Critical finding(s) added to period *{_banked_period}*. "
                f"Open DIM to run cross-module convergence analysis."
            )

    # Compliance evidence confirmation
    if st.session_state.get("dci_analysis_done"):
        st.markdown(
            "<div style='background:#f0fdf4;border:1.5px solid #16a34a;"
            "border-radius:10px;padding:10px 18px;margin-top:10px;'>"
            "<span style='color:#15803d;font-size:0.92rem;'>"
            "✓ <b>This run is now part of your DIM evidence package.</b> "
            "Results banked to DIM. Open DIM to run cross-module convergence."
            "</span></div>",
            unsafe_allow_html=True,
        )
