"""
Validation Doc Assist — v20.0
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Changes over v19.0:
  ROOT CAUSE FIX — Phantom Test_ID in Traceability:
  1. PYTHON-OWNED TRACEABILITY  — LLM no longer generates the
                                   Traceability dataset at all.
                                   _build_traceability() rebuilds
                                   it from actual OQ rows only.
                                   If OQ-004 doesn't exist in the
                                   OQ sheet, it CANNOT appear as
                                   "Covered" in Traceability.
  2. PASS 2 NOW 3 DATASETS      — FRS, OQ, Gap_Analysis only.
                                   Traceability dataset removed
                                   from prompt entirely.
  3. COVERAGE_STATUS GUARANTEED — Covered / Partial / Not Covered
                                   set by Python cross-reference,
                                   never by LLM claim.
  4. DASHBOARD + METRICS        — All coverage stats now read from
                                   Python-built Traceability sheet,
                                   not from OQ.Requirement_Link.
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

import streamlit as st
import streamlit.components.v1 as _st_components
import os
import datetime
import pandas as pd
from litellm import completion
import tempfile
import io
import sqlite3
import re
import hashlib
import secrets
import html as _html_lib

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

from langchain_community.document_loaders import PyPDFLoader
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import bcrypt
    BCRYPT_AVAILABLE = True
except ImportError:
    BCRYPT_AVAILABLE = False

# =============================================================================
# 1. CONFIG
# =============================================================================
VERSION        = "37.0"
PROMPT_VERSION = "v19.0-esignature-test-type-r3c"
TEMPERATURE    = 0.2
CHUNK_SIZE     = 8
DB_PATH        = os.path.join(os.path.dirname(os.path.abspath(__file__)), "validation_app.db")

SESSION_TIMEOUT_MINUTES = 15   # 21 CFR Part 11 — 15-minute inactivity timeout
MAX_FAILED_ATTEMPTS     = 5
LOCKOUT_MINUTES         = 15
MAX_UPLOAD_BYTES        = 10 * 1024 * 1024   # 10 MB hard limit per uploaded file

ROLES = ["Admin", "QA", "Validator"]

# =============================================================================
# 1b. PROMPT LOADER
# =============================================================================
# All prompts live in ./prompts/*.md — loaded once at startup.
# Separating prompts from code lets domain experts edit clinical/regulatory
# language without touching Python, and gives prompt changes their own git history.

def _load_prompt(filename: str) -> str:
    """
    Load a prompt template from the prompts/ directory next to this file.
    Supports subdirectory paths e.g. 'change_impact/pass3_justification.md'.
    Skills are organised by mode:
      prompts/                        ← shared gateway prompts
      prompts/new_validation/         ← New Validation mode
      prompts/change_impact/          ← Change Impact Analysis mode
    """
    prompt_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "prompts")
    path       = os.path.join(prompt_dir, filename)
    try:
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        # Graceful fallback — log warning but don't crash the app
        import warnings
        warnings.warn(f"Prompt file not found: {path}. Using empty string fallback.")
        return ""

# Load all prompt templates at module level — single I/O hit at startup
#
# Shared gateway prompts (root level — used across all modes)
_PROMPT_SYSTEM_RAW     = _load_prompt("system_prompt.md")
_PROMPT_PREFLIGHT_RAW  = _load_prompt("preflight_classifier.md")
#
# New Validation mode prompts
_PROMPT_PASS1_RAW      = _load_prompt("new_validation/pass1_urs_extraction.md")
_PROMPT_PASS2_RAW      = _load_prompt("new_validation/pass2_frs_oq_gap.md")
#
# Change Impact Analysis mode prompts
_PROMPT_CIA_PASS1_RAW  = _load_prompt("change_impact/pass1_change_extraction.md")
_PROMPT_CIA_PASS2_RAW  = _load_prompt("change_impact/pass2_impact_mapping.md")
_PROMPT_CIA_PASS3_RAW  = _load_prompt("change_impact/pass3_justification.md")

# =============================================================================
# 1b. SECURITY HELPERS
# =============================================================================

# In-memory rate-limiter (keyed by username, clears on server restart intentionally)
_LOGIN_RATE: dict = {}
_RATE_WINDOW_SEC   = 60
_RATE_MAX_ATTEMPTS = 15   # per 60-second window (above DB-level account lockout)


def _rate_allowed(key: str) -> bool:
    """True = request is within rate limit. False = reject immediately."""
    now      = datetime.datetime.utcnow().timestamp()
    attempts = [t for t in _LOGIN_RATE.get(key, []) if now - t < _RATE_WINDOW_SEC]
    _LOGIN_RATE[key] = attempts
    return len(attempts) < _RATE_MAX_ATTEMPTS


def _rate_record(key: str):
    _LOGIN_RATE.setdefault(key, []).append(datetime.datetime.utcnow().timestamp())


def sanitize_input(value: str, max_length: int = 128) -> str:
    """
    Strip null bytes, ASCII control characters, and HTML tags.
    Limits string length to max_length. Safe for use before any DB write.
    """
    if not isinstance(value, str):
        value = str(value)
    # Remove null bytes and non-printable control characters (keep tab/newline for text areas)
    value = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', value)
    # Remove HTML tags
    value = re.sub(r'<[^>]+>', '', value)
    return value.strip()[:max_length]


def _inject_password_security():
    """
    Inject JavaScript (via iframe parent access) that sets autocomplete='new-password'
    on every password input in the Streamlit page. This prevents browsers and password
    managers from offering to save or auto-fill credentials.

    Also sets data-lpignore (LastPass) and data-1p-ignore (1Password) opt-out attributes.
    A MutationObserver re-applies the attributes after every Streamlit rerender.
    """
    _st_components.html("""
    <script>
    (function() {
      var ATTRS = {
        'autocomplete': 'new-password',
        'data-form-type': 'other',
        'data-lpignore': 'true',
        'data-1p-ignore': 'true',
        'aria-autocomplete': 'none'
      };
      function patch() {
        try {
          var inputs = window.parent.document.querySelectorAll('input[type="password"]');
          inputs.forEach(function(el) {
            for (var k in ATTRS) { el.setAttribute(k, ATTRS[k]); }
            // Prevent browser from reading back cached value
            if (el._pwPatched) return;
            el._pwPatched = true;
            el.addEventListener('focus', function() {
              this.removeAttribute('value');
            });
          });
        } catch(e) {}
      }
      patch();
      try {
        new MutationObserver(patch).observe(
          window.parent.document.body, {childList: true, subtree: true}
        );
      } catch(e) {}
    })();
    </script>
    """, height=0, scrolling=False)

st.set_page_config(page_title="VALINTEL.AI — Validation Intelligence", layout="wide")

# =============================================================================
# 2. DATABASE
# =============================================================================

def db_connect():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False, timeout=10)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA busy_timeout=5000")
    # Deny any attempt to write to the audit_log table via DDL at runtime
    # (defence-in-depth — the Python layer never calls DELETE/UPDATE on audit_log)
    return conn

def db_migrate():
    try:
        conn = db_connect()

        conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                username        TEXT    UNIQUE NOT NULL,
                password_hash   TEXT    NOT NULL,
                role            TEXT    DEFAULT 'Validator',
                failed_attempts INTEGER DEFAULT 0,
                locked_until    TEXT    DEFAULT NULL,
                created_at      TEXT
            )
        """)

        # INSERT-ONLY audit trail — never UPDATE or DELETE this table
        conn.execute("""
            CREATE TABLE IF NOT EXISTS audit_log (
                event_id       INTEGER PRIMARY KEY AUTOINCREMENT,
                user           TEXT    NOT NULL,
                timestamp      TEXT    NOT NULL,
                action         TEXT    NOT NULL,
                object_changed TEXT,
                old_value      TEXT,
                new_value      TEXT,
                reason         TEXT,
                location       TEXT,
                user_ip        TEXT
            )
        """)

        conn.execute("""
            CREATE TABLE IF NOT EXISTS documents (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                name        TEXT    NOT NULL,
                type        TEXT    NOT NULL,
                version     INTEGER NOT NULL,
                uploaded_by TEXT,
                timestamp   TEXT,
                file_path   TEXT,
                status      TEXT    DEFAULT 'Active',
                content     TEXT,
                project_ref TEXT
            )
        """)

        conn.execute("""
            CREATE TABLE IF NOT EXISTS ai_gen_log (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                model            TEXT,
                prompt_version   TEXT,
                temperature      REAL,
                timestamp        TEXT,
                generated_by     TEXT,
                project_ref      TEXT,
                input_file       TEXT,
                document_ids_used TEXT
            )
        """)

        # ── Electronic Signature log — 21 CFR Part 11 §11.50 / §11.200 ────────
        # INSERT-ONLY. Append-only triggers added below alongside audit_log.
        conn.execute("""
            CREATE TABLE IF NOT EXISTS signature_log (
                signature_id      INTEGER PRIMARY KEY AUTOINCREMENT,
                user              TEXT    NOT NULL,
                role              TEXT,
                timestamp         TEXT    NOT NULL,
                action            TEXT    NOT NULL,
                signature_meaning TEXT    NOT NULL,
                document_hash     TEXT    NOT NULL,
                document_name     TEXT,
                model_used        TEXT,
                prompt_version    TEXT,
                ip_address        TEXT,
                doc_ids           TEXT
            )
        """)

        # Safe migrations for pre-existing databases
        user_cols = [r[1] for r in conn.execute("PRAGMA table_info(users)").fetchall()]
        for col, defn in [
            ("failed_attempts", "INTEGER DEFAULT 0"),
            ("locked_until",    "TEXT DEFAULT NULL"),
            ("created_at",      "TEXT"),
        ]:
            if col not in user_cols:
                conn.execute(f"ALTER TABLE users ADD COLUMN {col} {defn}")

        audit_cols = [r[1] for r in conn.execute("PRAGMA table_info(audit_log)").fetchall()]
        for col, defn in [
            ("object_changed", "TEXT"),
            ("old_value",      "TEXT"),
            ("new_value",      "TEXT"),
            ("reason",         "TEXT"),
            ("location",       "TEXT"),
            ("user_ip",        "TEXT"),   # v29 — separate column for faster audit search
        ]:
            if col not in audit_cols:
                try:
                    conn.execute(f"ALTER TABLE audit_log ADD COLUMN {col} {defn}")
                except Exception:
                    pass

        doc_cols = [r[1] for r in conn.execute("PRAGMA table_info(documents)").fetchall()]
        for col, defn in [
            ("name",        "TEXT"),
            ("type",        "TEXT"),
            ("file_path",   "TEXT"),
            ("status",      "TEXT DEFAULT 'Active'"),
            ("project_ref", "TEXT"),
            ("content",     "TEXT"),
            ("uploaded_by", "TEXT"),
            ("timestamp",   "TEXT"),
        ]:
            if col not in doc_cols:
                try:
                    conn.execute(f"ALTER TABLE documents ADD COLUMN {col} {defn}")
                except Exception:
                    pass

        ai_cols = [r[1] for r in conn.execute("PRAGMA table_info(ai_gen_log)").fetchall()]
        for col, defn in [
            ("temperature",       "REAL"),
            ("project_ref",       "TEXT"),
            ("input_file",        "TEXT"),
            ("document_ids_used", "TEXT"),
        ]:
            if col not in ai_cols:
                conn.execute(f"ALTER TABLE ai_gen_log ADD COLUMN {col} {defn}")

        # ── Async job queue table ────────────────────────────────────────────
        conn.execute("""
            CREATE TABLE IF NOT EXISTS jobs (
                job_id      TEXT    PRIMARY KEY,
                user        TEXT    NOT NULL,
                status      TEXT    NOT NULL DEFAULT 'queued',
                file_name   TEXT,
                model_id    TEXT,
                created_at  TEXT,
                started_at  TEXT,
                completed_at TEXT,
                progress    INTEGER DEFAULT 0,
                progress_msg TEXT   DEFAULT '',
                result_urs  TEXT,
                result_frs  TEXT,
                result_oq   TEXT,
                result_trace TEXT,
                result_gap  TEXT,
                result_xlsx BLOB,
                error_msg   TEXT,
                sys_ctx_name TEXT
            )
        """)

        conn.commit()
        conn.close()

        # ── Append-only enforcement: block UPDATE + DELETE on audit_log ──────
        # These triggers make the audit trail structurally immutable at the
        # database level — a defence-in-depth layer on top of the Python
        # "never call UPDATE/DELETE" convention.
        conn2 = db_connect()
        conn2.execute("""
            CREATE TRIGGER IF NOT EXISTS trg_audit_no_update
            BEFORE UPDATE ON audit_log
            BEGIN
                SELECT RAISE(ABORT, '21CFR11: audit_log rows are immutable — UPDATE denied');
            END
        """)
        conn2.execute("""
            CREATE TRIGGER IF NOT EXISTS trg_audit_no_delete
            BEFORE DELETE ON audit_log
            BEGIN
                SELECT RAISE(ABORT, '21CFR11: audit_log rows are immutable — DELETE denied');
            END
        """)
        # signature_log is also append-only — e-signatures cannot be retracted
        conn2.execute("""
            CREATE TRIGGER IF NOT EXISTS trg_esig_no_update
            BEFORE UPDATE ON signature_log
            BEGIN
                SELECT RAISE(ABORT, '21CFR11: signature_log rows are immutable — UPDATE denied');
            END
        """)
        conn2.execute("""
            CREATE TRIGGER IF NOT EXISTS trg_esig_no_delete
            BEFORE DELETE ON signature_log
            BEGIN
                SELECT RAISE(ABORT, '21CFR11: signature_log rows are immutable — DELETE denied');
            END
        """)
        conn2.commit()
        conn2.close()

    except Exception as e:
        st.warning(f"DB migration warning: {e}")


def db_diagnostics() -> dict:
    try:
        conn   = db_connect()
        result = {t: conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
                  for t in ["users", "audit_log", "documents", "ai_gen_log", "signature_log"]}
        conn.close()
        return result
    except Exception as e:
        return {"error": str(e)}



# =============================================================================
# ASYNC JOB QUEUE
# Background thread processes validation jobs independently of Streamlit.
# UI submits a job_id and polls status — no blocking, no timeouts.
# =============================================================================

import threading as _threading
import uuid      as _uuid
import time      as _time_mod

# Global worker state — one worker thread per process
_worker_lock    = _threading.Lock()
_worker_thread  = None
_worker_running = False


def _job_update(job_id: str, **kwargs):
    """Update job fields atomically."""
    if not kwargs:
        return
    fields = ", ".join(f"{k} = ?" for k in kwargs)
    vals   = list(kwargs.values()) + [job_id]
    try:
        conn = db_connect()
        conn.execute(f"UPDATE jobs SET {fields} WHERE job_id = ?", vals)
        conn.commit()
        conn.close()
    except Exception:
        pass


def _job_get(job_id: str) -> dict:
    """Fetch a single job row as a dict."""
    try:
        conn = db_connect()
        row  = conn.execute(
            "SELECT * FROM jobs WHERE job_id = ?", (job_id,)
        ).fetchone()
        conn.close()
        if row:
            cols = ["job_id","user","status","file_name","model_id",
                    "created_at","started_at","completed_at",
                    "progress","progress_msg",
                    "result_urs","result_frs","result_oq",
                    "result_trace","result_gap","result_xlsx",
                    "error_msg","sys_ctx_name"]
            return dict(zip(cols, row))
    except Exception:
        pass
    return {}


def _run_job(job_id: str, file_bytes: bytes, sys_ctx_bytes,
             model_id: str, user: str):
    """
    Execute the full validation pipeline for one job.
    Runs in a background thread — never touches Streamlit state directly.
    Updates the jobs table with progress and results.
    """
    import io as _io
    import datetime as _dt

    _job_update(job_id,
                status="running",
                started_at=_dt.datetime.utcnow().isoformat())

    # Minimal progress_bar / status_text shims so run_segmented_analysis
    # can call them without hitting Streamlit from a background thread.
    class _FakeProgress:
        def progress(self, v): pass
        def empty(self): pass

    class _FakeText:
        def __init__(self, job_id):
            self._jid = job_id
            self._last = ""
        def text(self, msg):
            if msg != self._last:
                self._last = msg
                _job_update(self._jid, progress_msg=str(msg)[:500])
        def empty(self): pass

    fake_bar  = _FakeProgress()
    fake_text = _FakeText(job_id)

    try:
        urs_df, frs_df, oq_df, trace_df, gap_df = run_segmented_analysis(
            file_bytes, model_id, fake_bar, fake_text, sys_ctx_bytes
        )

        if urs_df.empty and frs_df.empty:
            _job_update(job_id,
                        status="failed",
                        error_msg="Pipeline returned empty output. Check API quota.",
                        completed_at=_dt.datetime.utcnow().isoformat())
            return

        # ── Post-processing (mirrors synchronous pipeline) ────────────────
        _job_update(job_id, progress=70, progress_msg="Running deterministic checks...")
        gap_df, det_df = run_deterministic_validation(frs_df, oq_df, gap_df, urs_df)
        for _df in [gap_df, det_df, trace_df]:
            _df.fillna("N/A", inplace=True)
            _df.replace("", "N/A", inplace=True)

        _job_update(job_id, progress=80, progress_msg="Saving documents...")
        save_document("URS_Extraction", urs_df.to_csv(index=False),  user, "async_job")
        save_document("FRS",            frs_df.to_csv(index=False),  user, "async_job")
        save_document("OQ",             oq_df.to_csv(index=False),   user, "async_job")
        save_document("Traceability",   trace_df.to_csv(index=False),user, "async_job")
        save_document("Gap_Analysis",   gap_df.to_csv(index=False),  user, "async_job")
        save_document("Det_Validation", det_df.to_csv(index=False),  user, "async_job")

        _job_update(job_id, progress=88, progress_msg="Building validation workbook...")
        audit_df     = build_audit_log_sheet(
            user, "async_job", model_id,
            frs_df, oq_df, gap_df, det_df, 1, 1, ""
        )
        dashboard_df = build_dashboard_sheet(
            frs_df, oq_df, gap_df, det_df, trace_df, "async_job", model_id
        )
        dataframes = {
            "Dashboard":      dashboard_df,
            "URS_Extraction": urs_df,
            "FRS":            frs_df,
            "OQ":             oq_df,
            "Traceability":   trace_df,
            "Gap_Analysis":   gap_df,
            "Det_Validation": det_df,
            "Audit_Log":      audit_df,
        }
        xlsx_bytes = build_styled_excel(
            dataframes, user=user, file_name="async_job",
            model_name=model_id, sys_context_name="",
            dashboard_df=dashboard_df
        )

        _job_update(job_id,
                    status="complete",
                    progress=100,
                    progress_msg=(
                        f"✅ Done — {len(urs_df)} requirements, "
                        f"{len(frs_df)} FRS rows, {len(oq_df)} OQ tests"
                    ),
                    result_urs   = urs_df.to_csv(index=False),
                    result_frs   = frs_df.to_csv(index=False),
                    result_oq    = oq_df.to_csv(index=False),
                    result_trace = trace_df.to_csv(index=False),
                    result_gap   = gap_df.to_csv(index=False),
                    result_xlsx  = xlsx_bytes,
                    completed_at = _dt.datetime.utcnow().isoformat())

    except Exception as exc:
        _job_update(job_id,
                    status="failed",
                    error_msg=str(exc)[:1000],
                    completed_at=_dt.datetime.utcnow().isoformat())


def _worker_loop():
    """
    Continuously poll for queued jobs and process them one at a time.
    Runs as a daemon thread started once at app boot.
    """
    global _worker_running
    _worker_running = True
    try:
        while True:
            try:
                conn = db_connect()
                row  = conn.execute(
                    "SELECT job_id, user, model_id FROM jobs "
                    "WHERE status = 'queued' ORDER BY created_at LIMIT 1"
                ).fetchone()
                conn.close()
            except Exception:
                row = None

            if row:
                job_id, user, model_id = row
                # Fetch file bytes from the job blobs table
                try:
                    conn = db_connect()
                    blob_row = conn.execute(
                        "SELECT file_bytes, sys_ctx_bytes FROM job_blobs "
                        "WHERE job_id = ?", (job_id,)
                    ).fetchone()
                    conn.close()
                    if blob_row:
                        file_bytes   = blob_row[0]
                        sys_ctx_bytes = blob_row[1]
                        _run_job(job_id, file_bytes, sys_ctx_bytes,
                                 model_id, user)
                except Exception as exc:
                    _job_update(job_id, status="failed",
                                error_msg=f"Worker fetch error: {exc}")
            else:
                _time_mod.sleep(3)  # poll every 3 seconds when idle
    finally:
        _worker_running = False


def ensure_worker_running():
    """Start the background worker thread if not already running."""
    global _worker_thread, _worker_running
    with _worker_lock:
        if _worker_thread is None or not _worker_thread.is_alive():
            _worker_thread = _threading.Thread(
                target=_worker_loop, daemon=True, name="valintel-worker"
            )
            _worker_thread.start()


def submit_job(user: str, file_bytes: bytes, file_name: str,
               model_id: str, sys_ctx_bytes=None,
               sys_ctx_name: str = "") -> str:
    """
    Queue a new validation job. Returns the job_id immediately.
    File bytes are stored in a separate job_blobs table to keep jobs table lean.
    """
    import datetime as _dt
    job_id = str(_uuid.uuid4())[:12].upper()

    conn = db_connect()
    # Ensure job_blobs table exists
    conn.execute("""
        CREATE TABLE IF NOT EXISTS job_blobs (
            job_id       TEXT PRIMARY KEY,
            file_bytes   BLOB,
            sys_ctx_bytes BLOB
        )
    """)
    conn.execute(
        "INSERT INTO job_blobs (job_id, file_bytes, sys_ctx_bytes) VALUES (?,?,?)",
        (job_id, file_bytes, sys_ctx_bytes)
    )
    conn.execute(
        """INSERT INTO jobs
           (job_id, user, status, file_name, model_id, created_at, sys_ctx_name)
           VALUES (?,?,?,?,?,?,?)""",
        (job_id, user, "queued", file_name, model_id,
         _dt.datetime.utcnow().isoformat(), sys_ctx_name)
    )
    conn.commit()
    conn.close()

    ensure_worker_running()
    return job_id


def log_audit(user: str, action: str, object_changed: str = "",
              old_value: str = "", new_value: str = "", reason: str = ""):
    """
    Append-only audit write. This function must never UPDATE or DELETE rows.
    user_ip is pulled from st.session_state and written as a separate indexed
    column (not embedded in Reason) to support fast audit searches per 21 CFR Part 11.
    """
    try:
        user_ip = st.session_state.get("user_ip", "")
    except Exception:
        user_ip = ""
    try:
        conn = db_connect()
        conn.execute(
            """INSERT INTO audit_log
               (user, timestamp, action, object_changed, old_value, new_value,
                reason, location, user_ip)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (user,
             datetime.datetime.utcnow().isoformat(),
             action,
             str(object_changed)[:500],
             str(old_value)[:2000]  if old_value  else "",
             str(new_value)[:2000]  if new_value  else "",
             str(reason)[:1000]     if reason     else "",
             "",          # location intentionally blank — removed from UI in v27
             str(user_ip)[:100])
        )
        conn.commit()
        conn.close()
    except Exception as e:
        st.warning(f"Audit log write failed: {e}")


# =============================================================================
# 21 CFR PART 11 — ELECTRONIC SIGNATURE CONSTANTS & WRITER
# =============================================================================
# Controlled vocabulary for signature meaning per §11.50(a)(1).
# Free text is NOT allowed — auditors expect a finite, pre-approved list.
# ── MANUAL EDIT v29-custom — DO NOT OVERWRITE ──────────────
ESIG_MEANINGS = [
    "I executed this validation package",
    "I reviewed this validation package",
    "I approved this validation package",
    ]
ESIG_DEFAULT_MEANING = ESIG_MEANINGS[0]
# ── END MANUAL EDIT ────────────────────────────────────────

def log_esignature(user: str, role: str, action: str, meaning: str,
                   document_hash: str, document_name: str = "",
                   model_used: str = "", prompt_ver: str = "",
                   ip_address: str = "", doc_ids: str = "") -> int:
    """
    Insert one row into signature_log (append-only).

    21 CFR Part 11 compliance:
      §11.50(a)(1) — printed name, date/time, meaning recorded
      §11.50(a)(2) — linked to associated record via document_hash
      §11.100(a)   — unique to one individual (username)
      §11.200(b)   — two distinct components: username + password re-entry
                     (verification happens in calling code before this runs)

    document_hash covers all workbook sheets EXCEPT the Signature sheet.
    This is standard practice (same as PDF digital signatures) and is
    documented explicitly on the Signature sheet.

    Returns the new signature_id (or -1 on failure).
    """
    try:
        conn = db_connect()
        cur  = conn.execute(
            """INSERT INTO signature_log
               (user, role, timestamp, action, signature_meaning,
                document_hash, document_name, model_used, prompt_version,
                ip_address, doc_ids)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (user, role,
             datetime.datetime.utcnow().isoformat(),
             action, meaning, document_hash, document_name,
             model_used, prompt_ver, ip_address, doc_ids)
        )
        sig_id = cur.lastrowid
        conn.commit()
        conn.close()
        return sig_id
    except Exception as e:
        st.warning(f"E-signature log write failed: {e}")
        return -1


def log_ai_generation(user: str, model: str, prompt_version: str,
                      temperature: float, input_file: str = "",
                      project_ref: str = "", document_ids_used: str = ""):
    try:
        conn = db_connect()
        conn.execute(
            """INSERT INTO ai_gen_log
               (model, prompt_version, temperature, timestamp, generated_by,
                project_ref, input_file, document_ids_used)
               VALUES (?,?,?,?,?,?,?,?)""",
            (model, prompt_version, temperature,
             datetime.datetime.utcnow().isoformat(),
             user, project_ref, input_file, document_ids_used)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        st.warning(f"AI gen log write failed: {e}")


def get_next_doc_version(doc_type: str) -> int:
    try:
        conn = db_connect()
        row  = conn.execute(
            "SELECT MAX(version) FROM documents WHERE type=?", (doc_type,)
        ).fetchone()
        conn.close()
        return (row[0] or 0) + 1
    except Exception:
        return 1


def save_document(doc_type: str, content: str, created_by: str,
                  project_ref: str = "", file_path: str = "") -> int:
    """Always inserts a new version — never overwrites. Returns new doc ID."""
    version = get_next_doc_version(doc_type)
    name    = f"{doc_type}_v{version}.0_{datetime.date.today()}"
    try:
        conn = db_connect()
        cur  = conn.execute(
            """INSERT INTO documents
               (name, type, version, uploaded_by, timestamp, file_path,
                status, content, project_ref)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (name, doc_type, version, created_by,
             datetime.datetime.utcnow().isoformat(),
             file_path, "Active", content[:10000], project_ref)
        )
        doc_id = cur.lastrowid
        conn.commit()
        conn.close()
        return doc_id
    except Exception as e:
        st.warning(f"Document save failed: {e}")
        return -1


# =============================================================================
# 3. AUTHENTICATION
# =============================================================================

def hash_password(plain: str) -> str:
    """bcrypt only — no plaintext fallback. Auditor-safe."""
    if not BCRYPT_AVAILABLE:
        raise RuntimeError(
            "bcrypt is required but not installed. "
            "Run: pip install bcrypt  — SHA-256 plaintext fallback is disabled for GxP compliance."
        )
    return bcrypt.hashpw(plain.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, stored_hash: str) -> bool:
    """bcrypt only — rejects any non-bcrypt hash rather than falling back to SHA-256."""
    if not BCRYPT_AVAILABLE:
        return False   # cannot verify without bcrypt — fail closed
    try:
        if not stored_hash.startswith("$2"):
            # Hash is not a bcrypt hash — refuse to compare
            return False
        return bcrypt.checkpw(plain.encode("utf-8"), stored_hash.encode("utf-8"))
    except Exception:
        return False


def create_user(username: str, plain_password: str, role: str = "Validator"):
    pw_hash = hash_password(plain_password)
    conn    = db_connect()
    try:
        conn.execute(
            "INSERT INTO users (username, password_hash, role, created_at) VALUES (?,?,?,?)",
            (username, pw_hash, role, datetime.datetime.utcnow().isoformat())
        )
        conn.commit()
    except sqlite3.IntegrityError:
        pass
    finally:
        conn.close()


def _is_account_locked(username: str) -> tuple:
    conn = db_connect()
    row  = conn.execute(
        "SELECT failed_attempts, locked_until FROM users WHERE username=?", (username,)
    ).fetchone()
    conn.close()
    if not row:
        return False, ""
    _, locked_until = row
    if locked_until:
        try:
            unlock_time = datetime.datetime.fromisoformat(locked_until)
            if datetime.datetime.utcnow() < unlock_time:
                remaining = int((unlock_time - datetime.datetime.utcnow()).total_seconds() / 60)
                return True, f"Account locked. Try again in {remaining} minute(s)."
            else:
                conn2 = db_connect()
                conn2.execute(
                    "UPDATE users SET failed_attempts=0, locked_until=NULL WHERE username=?",
                    (username,)
                )
                conn2.commit()
                conn2.close()
        except Exception:
            pass
    return False, ""


def _record_failed_attempt(username: str):
    conn = db_connect()
    conn.execute(
        "UPDATE users SET failed_attempts = failed_attempts + 1 WHERE username=?", (username,)
    )
    row = conn.execute(
        "SELECT failed_attempts FROM users WHERE username=?", (username,)
    ).fetchone()
    if row and row[0] >= MAX_FAILED_ATTEMPTS:
        lock_until = (datetime.datetime.utcnow() +
                      datetime.timedelta(minutes=LOCKOUT_MINUTES)).isoformat()
        conn.execute(
            "UPDATE users SET locked_until=? WHERE username=?", (lock_until, username)
        )
        log_audit(username, "ACCOUNT_LOCKED", "USER",
                  reason=f"Exceeded {MAX_FAILED_ATTEMPTS} failed attempts")
    conn.commit()
    conn.close()


def _reset_failed_attempts(username: str):
    conn = db_connect()
    conn.execute(
        "UPDATE users SET failed_attempts=0, locked_until=NULL WHERE username=?", (username,)
    )
    conn.commit()
    conn.close()


def authenticate_user(username: str, password: str) -> tuple:
    """Returns (success: bool, error_message: str)."""
    # Sanitize inputs before any processing
    username = sanitize_input(username, max_length=64)
    password = sanitize_input(password, max_length=256)

    if not username:
        return False, "Username is required."

    # In-memory rate limit (complements DB-level account lockout)
    if not _rate_allowed(username):
        log_audit(username, "LOGIN_RATE_LIMITED", "SESSION",
                  reason="Exceeded in-memory rate limit")
        return False, "Too many login attempts. Please wait 60 seconds."
    _rate_record(username)

    try:
        conn  = db_connect()
        row   = conn.execute(
            "SELECT password_hash, role FROM users WHERE username=?", (username,)
        ).fetchone()
        count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        conn.close()

        if count == 0:
            create_user(username, password, role="Admin")
            log_audit(username, "FIRST_RUN_ADMIN_CREATED", "USER",
                      new_value=username, reason="First-run bootstrap")
            _reset_failed_attempts(username)
            return True, ""

        if not row:
            return False, "Invalid credentials."

        is_locked, lock_msg = _is_account_locked(username)
        if is_locked:
            log_audit(username, "LOGIN_BLOCKED_LOCKED", "SESSION")
            return False, lock_msg

        if verify_password(password, row[0]):
            _reset_failed_attempts(username)
            return True, ""

        _record_failed_attempt(username)
        conn2    = db_connect()
        attempts = conn2.execute(
            "SELECT failed_attempts FROM users WHERE username=?", (username,)
        ).fetchone()
        conn2.close()
        remaining = MAX_FAILED_ATTEMPTS - (attempts[0] if attempts else 1)
        log_audit(username, "LOGIN_FAILED", "SESSION",
                  reason=f"Bad password, {max(remaining,0)} attempt(s) remaining")
        if remaining <= 0:
            return False, f"Account locked for {LOCKOUT_MINUTES} minutes."
        return False, f"Invalid credentials. {remaining} attempt(s) remaining."

    except Exception as ex:
        return False, f"Authentication error: {ex}"


def get_user_role(username: str) -> str:
    try:
        conn = db_connect()
        row  = conn.execute("SELECT role FROM users WHERE username=?", (username,)).fetchone()
        conn.close()
        return row[0] if row else "Validator"
    except Exception:
        return "Validator"


def check_session_timeout() -> bool:
    last = st.session_state.get("last_activity")
    if not last:
        return True
    elapsed = (datetime.datetime.utcnow() - last).total_seconds() / 60
    return elapsed <= SESSION_TIMEOUT_MINUTES


def touch_session():
    st.session_state["last_activity"] = datetime.datetime.utcnow()


# =============================================================================
# 4. PDF EXTRACTION
# =============================================================================

# PDF magic bytes — all valid PDF files start with %PDF
_PDF_MAGIC = b'%PDF'


def validate_upload(file_obj) -> tuple:
    """
    Two-gate upload security check.

    Gate 1 — File size: reject anything over MAX_UPLOAD_BYTES (10 MB).
      Prevents memory exhaustion and DoS via oversized uploads.

    Gate 2 — MIME / magic bytes: read the first 4 bytes and confirm the
      file starts with %PDF regardless of the .pdf extension chosen.
      Extension-only checks are trivially bypassed by renaming a file.

    Returns (is_valid: bool, error_message: str).
    """
    if file_obj is None:
        return False, "No file provided."

    # Gate 1 — size
    size = file_obj.size
    if size > MAX_UPLOAD_BYTES:
        mb = size / (1024 * 1024)
        return False, (
            f"⛔ File rejected: {mb:.1f} MB exceeds the 10 MB limit. "
            "Split the document or remove embedded images before uploading."
        )

    # Gate 2 — MIME / magic bytes
    raw = file_obj.getvalue()
    if not raw or raw[:4] != _PDF_MAGIC:
        return False, (
            "⛔ File rejected: content does not match PDF format. "
            "Renaming a non-PDF file to .pdf does not make it a valid PDF. "
            "Upload a genuine PDF document."
        )

    return True, ""


def extract_pages(file_bytes: bytes) -> list:
    pages_text = []
    if PDFPLUMBER_AVAILABLE:
        try:
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page_num, page in enumerate(pdf.pages, start=1):
                    parts = []
                    prose = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
                    if prose.strip():
                        parts.append(prose.strip())
                    for t_idx, table in enumerate(page.extract_tables() or []):
                        if not table:
                            continue
                        rows_md = []
                        for r_idx, row in enumerate(table):
                            cells = [str(c).replace("\n", " ").strip() if c else "" for c in row]
                            rows_md.append(" | ".join(cells))
                            if r_idx == 0:
                                rows_md.append(" | ".join(["---"] * len(row)))
                        parts.append(
                            f"\n[TABLE {t_idx+1} — Page {page_num}]\n"
                            + "\n".join(rows_md) + "\n[/TABLE]\n"
                        )
                    pages_text.append(f"--- Page {page_num} ---\n" + "\n".join(parts))
            if sum(len(p) for p in pages_text) >= 50:
                return pages_text
        except Exception:
            pages_text = []

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        lc_pages   = PyPDFLoader(tmp_path).load()
        pages_text = [f"--- Page {i+1} ---\n{p.page_content}"
                      for i, p in enumerate(lc_pages)]
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    return pages_text



# =============================================================================
# 4b. URS DOCUMENT VALIDATION — Two-stage gate
#     Stage 1: Fast heuristic (free, instant) — rejects obvious non-URS docs
#     Stage 2: LLM pre-flight (one cheap call) — catches edge cases
# =============================================================================

# Positive signals — language expected in a URS / SOP
_URS_POSITIVE = [
    r'\bshall\b', r'\bmust\b', r'\brequirement[s]?\b', r'\buser requirement\b',
    r'\bsystem shall\b', r'\bthe system\b', r'\burs\b', r'\bsop\b',
    r'\bfunctional requirement\b', r'\buse case\b', r'\bstakeholder\b',
    r'\bscope\b', r'\bpurpose\b', r'\bspecification\b',
    r'\bvalidation\b', r'\bcompliance\b', r'\baudit trail\b',
    r'\breq[-_\s]?\d+\b', r'\burs[-_\s]?\d+\b',   # REQ-001, URS-001 style IDs
    r'\d+\.\d+[\s]+\w',                              # 1.1 Section style numbering
]

# Negative signals — language that identifies clearly wrong document types
_URS_NEGATIVE = [
    r'\bdate of birth\b', r'\blicense number\b', r'\bdriver.?s license\b',
    r'\bpassport\b', r'\bstate id\b', r'\bidentification card\b',
    r'\bexpir(?:es|ation)\b.*\b\d{4}\b',   # "Expires 2027" — ID card pattern
    r'\bsocial security\b', r'\bssn\b',
    r'\binvoice\b', r'\bpurchase order\b', r'\breceipt\b',
    r'\btotal due\b', r'\bamount due\b', r'\bremit payment\b',
    r'\bdear\b.*\bsincerely\b',            # letter pattern
    r'\bresume\b', r'\bcurriculum vitae\b',
    r'\bmenu\b.*\bprice\b',               # restaurant menu
]

# LLM pre-flight prompt — binary YES/NO, one sentence reasoning
_PREFLIGHT_PROMPT = _PROMPT_PREFLIGHT_RAW


def validate_urs_document(
    file_bytes: bytes,
    model_id: str,
) -> tuple[bool, str]:
    """
    Three-stage URS document gate. Returns (is_valid: bool, message: str).

    Stage 0 — Structural pre-check (free, <5ms):
        Minimum extractable text length and page count. Rejects empty or
        image-only PDFs before any pattern matching.

    Stage 1 — Heuristic scoring (free, <10ms):
        Score the first ~3000 chars against positive/negative patterns.
        Also checks for minimum "shall"/"must" density and minimum
        extractable requirement count.

    Stage 2 — LLM pre-flight (one cheap API call, only if Stage 1 passes):
        Send the first page to the model with a binary YES/NO prompt.

    New checks added in v26:
      - Minimum text length gate (< 300 chars = not a document)
      - Minimum "shall"/"must" count (< 2 = insufficient requirement density)
      - Minimum positive signal score raised to 3
      - Structural section check: warns if no section headings found
    """
    # ── Extract text ─────────────────────────────────────────────────────────
    try:
        pages = extract_pages(file_bytes)
    except Exception as e:
        return False, f"Could not extract text from PDF: {e}"

    if not pages:
        return False, "The uploaded PDF appears to be empty or image-only (no extractable text)."

    # ── Stage 0: Structural pre-check ─────────────────────────────────────────
    full_text    = "\n".join(pages)
    full_lower   = full_text.lower()
    sample_text  = "\n\n".join(pages[:5])  # extended from 2 to 5 pages
    sample_lower = sample_text.lower()

    if len(full_text.strip()) < 300:
        return False, (
            "⛔ Document too short to be a valid URS.\n\n"
            f"Only {len(full_text.strip())} characters of text were extracted. "
            "A User Requirements Specification must contain substantive requirement "
            "statements. Check that the PDF is not an image-only scan."
        )

    # Minimum requirement-statement density
    shall_must_count = len(re.findall(r'\b(shall|must)\b', full_lower, re.IGNORECASE))
    if shall_must_count < 2:
        return False, (
            f"⛔ Insufficient requirement language detected.\n\n"
            f"Found only {shall_must_count} statement(s) using 'shall' or 'must'. "
            f"A valid URS must contain at least 2 requirement statements written in "
            f"prescriptive language (shall/must). This document does not appear to be "
            f"a User Requirements Specification."
        )

    # ── Stage 1: Heuristic scoring ────────────────────────────────────────────
    pos_hits = [p for p in _URS_POSITIVE if re.search(p, sample_lower, re.IGNORECASE)]
    neg_hits = [p for p in _URS_NEGATIVE if re.search(p, sample_lower, re.IGNORECASE)]

    score = len(pos_hits) - (3 * len(neg_hits))

    if neg_hits:
        matched = [p.replace(r'\b', '').replace('\\', '') for p in neg_hits[:3]]
        return False, (
            f"⛔ Document rejected at content screening.\n\n"
            f"This does not appear to be a URS, SRS, or SOP. "
            f"Detected non-URS content: **{', '.join(matched)}**.\n\n"
            f"Please upload a User Requirements Specification, System Requirements "
            f"Specification, or Standard Operating Procedure."
        )

    if score < 3:
        # Raised threshold from 2 → 3 for stricter gate
        return False, (
            f"⛔ Document rejected: insufficient URS content detected.\n\n"
            f"Only {len(pos_hits)} URS indicator(s) found in the document "
            f"(minimum 3 required). The document may not be a URS, SRS, or SOP.\n\n"
            f"Expected content: requirement statements using 'shall'/'must', "
            f"numbered requirements (URS-001, REQ-001), section headings like "
            f"'Scope', 'Purpose', 'Functional Requirements'."
        )

    # ── Structural section check (soft warning — does not block) ─────────────
    section_patterns = [
        r'\bscope\b', r'\bpurpose\b', r'\bintroduction\b', r'\boverview\b',
        r'\bfunctional requirement', r'\bnon-functional', r'\bsecurity requirement',
        r'\bperformance requirement', r'\bsystem requirement', r'\buser requirement',
        r'\bversion history\b', r'\bdocument control\b', r'\bapproval\b',
    ]
    section_hits = sum(1 for p in section_patterns
                       if re.search(p, sample_lower, re.IGNORECASE))
    structural_warning = "" if section_hits >= 2 else (
        " Note: fewer than 2 standard URS section headings detected — "
        "document may be informal but has been accepted on requirement density."
    )

    # ── Stage 2: LLM pre-flight ───────────────────────────────────────────────
    try:
        preflight_text = sample_text[:3000]
        response = completion(
            model=model_id,
            stream=False,
            temperature=0.0,
            max_tokens=100,
            messages=[
                {"role": "user", "content": _PREFLIGHT_PROMPT.format(text=preflight_text)}
            ]
        )
        reply = (response.choices[0].message.content or "").strip()

        verdict_match = re.search(r'VERDICT:\s*(YES|NO)', reply, re.IGNORECASE)
        reason_match  = re.search(r'REASON:\s*(.+)',      reply, re.IGNORECASE)

        verdict = verdict_match.group(1).upper() if verdict_match else None
        reason  = reason_match.group(1).strip()  if reason_match  else reply[:200]

        if verdict == "YES":
            return True, (
                f"Document validated ({len(pos_hits)} URS indicators, "
                f"{shall_must_count} shall/must statements). "
                f"LLM: {reason}{structural_warning}"
            )
        elif verdict == "NO":
            return False, (
                f"⛔ Document rejected by AI content classifier.\n\n"
                f"**AI assessment:** {reason}\n\n"
                f"Please upload a User Requirements Specification, System Requirements "
                f"Specification, or Standard Operating Procedure."
            )
        else:
            return True, (
                f"Document accepted (LLM response ambiguous; Stage 1 score={score}). "
                f"{structural_warning}"
            )

    except Exception as e:
        if score >= 5:
            return True, (
                f"LLM pre-flight failed ({e}); accepted on strong Stage 1 signal "
                f"(score={score}, {shall_must_count} shall/must statements). "
                f"{structural_warning}"
            )
        return False, (
            f"⛔ Document validation incomplete — LLM pre-flight failed: {e}\n\n"
            f"Stage 1 score was {score} (threshold: 5 for auto-accept without LLM). "
            f"Please try again or use a different model."
        )



#    Pass 1: Structured URS extraction (deterministic input table)
#    Pass 2: FRS / OQ / Traceability / Gap derived from Pass-1 table
#            FRS descriptions are ALWAYS written in engineering/implementation
#            language — never a copy of the URS wording.
# =============================================================================

SYSTEM_PROMPT = _PROMPT_SYSTEM_RAW.strip()


def _make_system_prompt(sys_context: str = "") -> str:
    """
    Build the system prompt for a completion call.

    When a SysContext (product user guide / system manual) has been uploaded,
    it is appended as Reference Material so the LLM can use real screen names,
    field names, and module terminology in BOTH Pass 1 URS extraction AND Pass 2
    FRS/OQ generation — completing the URS → FRS chain the Instructions require.

    The sys_context is capped at 4000 chars to stay within token budgets while
    still providing enough product vocabulary to produce credible engineering FRS.
    """
    if not sys_context:
        return SYSTEM_PROMPT
    return (
        SYSTEM_PROMPT
        + "\n\nREFERENCE MATERIAL — TARGET SYSTEM USER GUIDE:\n"
          "Use the following product documentation to inform ALL outputs with accurate "
          "screen names, field names, module names, and workflow terminology. "
          "Do NOT copy this text verbatim; use it to ground your engineering descriptions.\n\n"
        + sys_context[:4000]
    )

# ── PASS 1 PROMPT: extract a clean, structured URS table ─────────────────────
def build_pass1_prompt(chunk_text: str, chunk_index: int, total_chunks: int) -> str:
    return _PROMPT_PASS1_RAW.format(
        chunk_index  = chunk_index + 1,
        total_chunks = total_chunks,
        chunk_text   = chunk_text,
    )

# ── PASS 2 PROMPT: generate FRS / OQ / Traceability / Gap from URS table ─────
def build_pass2_prompt(urs_csv: str, sys_context: str = "") -> str:

    if sys_context:
        context_block = (
            f"SYSTEM USER GUIDE (product manual uploaded by user — use this to shape "
            f"implementation details in FRS descriptions):\n{sys_context[:3000]}\n\n"
        )
        system_guidance = (
            "Use the System User Guide above to determine the specific screens, fields, "
            "modules, and workflows that the system uses to implement each URS requirement. "
            "FRS descriptions must reference the actual product terminology from that guide. "
            "For OQ test steps: apply RULE A for infrastructure/non-functional requirements "
            "(availability, uptime, SLA, failover) — write technical verification procedures, "
            "NOT UI navigation. Apply RULE B for application features described in the guide "
            "using exact screen and field names. Apply RULE C (prefix [SCREEN UNVERIFIED], "
            "Confidence=0.60) for application features NOT described in the guide."
        )
    else:
        context_block = ""
        system_guidance = (
            "NO system user guide was provided. "
            "Infer the system type from the URS content (e.g. LIMS, SAP, Veeva Vault, ERP, "
            "MES, QMS, CTMS, eTMF, or similar GxP platform). "
            "Write FRS descriptions as best-practice implementation for that system type using "
            "plausible but generic screen/module names. Set Source_Section = 'URS-derived' for "
            "all FRS rows. "
            "For OQ test steps: apply RULE A for infrastructure/non-functional requirements "
            "(write technical verification procedures, not UI navigation). "
            "Apply RULE C for all application feature requirements — prefix every Test_Step "
            "with [SCREEN UNVERIFIED] and set Confidence = 0.60. "
            "This is required because screen names cannot be verified without a guide."
        )

    return _PROMPT_PASS2_RAW.format(
        context_block   = context_block,
        urs_csv         = urs_csv,
        system_guidance = system_guidance,
    )


def _summarise_sys_context(sys_context: str) -> str:
    """
    Phase 2: condense the system guide to ~500 chars of key vocabulary.
    Called ONCE before the per-requirement loop. Reused across all calls
    to avoid sending the full guide with every individual requirement.
    """
    if not sys_context:
        return ""
    # Take first 3000 chars, extract unique noun phrases (screens, modules)
    # Simple heuristic: lines containing capitalised compound words are likely
    # screen/module names — include them verbatim, truncate everything else.
    lines = sys_context[:3000].split("\n")
    key_lines = []
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        # Include lines that look like navigation paths, module names, or
        # screen names (contain capitals + LIMS-specific keywords)
        if any(kw in stripped for kw in [
            "Tramline", "Tramstop", "SDC", "SDI", "Parameter List",
            "Data Entry", "Sample", "Stability", "Instrument", "Audit",
            "navigate", "screen", "page", "module", "tab", "button",
        ]):
            key_lines.append(stripped)
        if len("\n".join(key_lines)) > 500:
            break
    summary = "\n".join(key_lines)[:500]
    return summary if summary else sys_context[:500]


def build_pass2_single_prompt(req_row: str, header: str,
                               sys_summary: str = "") -> str:
    """
    Phase 2: generate FRS + OQ for ONE requirement row.
    Sends only a 500-char system guide summary instead of full guide.
    Returns same CSV format as batch prompt so existing parsers work unchanged.
    """
    if sys_summary:
        context_block = (
            f"SYSTEM GUIDE (key terminology only):\n{sys_summary}\n\n"
        )
        system_guidance = (
            "Use the system guide terminology above for screen names and "
            "navigation paths. Apply RULE B for features described in the "
            "guide. Apply RULE C ([SCREEN UNVERIFIED], Confidence=0.60) for "
            "features not described."
        )
    else:
        context_block = ""
        system_guidance = (
            "No system guide provided. Use best-practice GxP LIMS terminology. "
            "Prefix all OQ Test_Steps with [SCREEN UNVERIFIED] and set "
            "Confidence=0.60."
        )

    # Single-requirement CSV: header + one data row
    single_csv = header + "\n" + req_row

    return _PROMPT_PASS2_RAW.format(
        context_block   = context_block,
        urs_csv         = single_csv,
        system_guidance = system_guidance,
    )


# =============================================================================
# 6. TWO-PASS AI ANALYSIS ENGINE
# =============================================================================

# Known header signatures for each of the 4 datasets in Pass-2 output.
# Used by the robust splitter to locate dataset boundaries even when the
# LLM embeds stray ||| tokens inside quoted field values.
_PASS2_HEADERS = [
    # Dataset 1 — FRS (now includes Source_Section)
    r"^ID[,\t]Requirement_Description",
    # Dataset 2 — OQ (now includes Suggested_Evidence)
    r"^Test_ID[,\t]Requirement_Link",
    # Dataset 3 — Gap_Analysis (LLM-detected URS-level gaps only)
    r"^Req_ID[,\t]Gap_Type",
]

_PASS1_HEADER = r"^Req_ID[,\t]Requirement_Description"


def _strip_fences(raw: str) -> str:
    """Remove markdown code-fences that LLMs sometimes wrap output in."""
    raw = re.sub(r'^```[a-zA-Z]*\n?', '', raw, flags=re.MULTILINE)
    raw = re.sub(r'```\s*$',          '', raw, flags=re.MULTILINE)
    return raw.strip()


def _strip_preamble(text: str) -> str:
    """
    Strip any non-CSV prose the LLM prepends before the actual header row.

    Handles patterns like:
      "Here is the CSV:"            — narrative intro
      "Dataset 1 (FRS):"            — section label
      "Dataset 2:"                  — numbered label
      "## FRS Requirements"         — markdown heading
      "Sure! Here are the results:" — chatty preamble

    Strategy: scan lines top-to-bottom; return from the first line that
    looks like a real CSV row — i.e. it contains at least one comma AND
    starts with an alphanumeric or quote character AND does NOT start with
    a known preamble keyword.
    """
    PREAMBLE_RE = re.compile(
        r'^(here|dataset\s*\d|csv|the\s+following|below|output|result|sure|note|'
        r'i\s+have|please\s+find|as\s+requested|```|#)',
        re.IGNORECASE
    )
    lines = text.splitlines()
    for i, line in enumerate(lines):
        s = line.strip()
        if not s:
            continue
        # Must contain a comma (CSV requirement)
        if ',' not in s:
            continue
        # Must start with alphanumeric or quote — not a sentence
        if not re.match(r'^[A-Za-z0-9"\']', s):
            continue
        # Must not match known preamble starters
        if PREAMBLE_RE.match(s):
            continue
        return '\n'.join(lines[i:])
    return text  # fallback — return as-is if nothing matched


def _robust_split_datasets(raw: str, headers: list) -> list:
    """
    Split LLM output into N CSV blocks by finding each known header line.
    Immune to:
      - stray ||| tokens inside quoted cell values
      - "Here is the CSV:" / "Dataset 1 (FRS):" label lines before the header
      - markdown fences wrapping each section

    Strategy:
      1. Strip fences and leading preamble from the whole block.
      2. For each header pattern, find the first matching line.
      3. Extract text from that line to the next header (or end).
      4. Apply _strip_preamble to each extracted section for safety.
      5. Return a list of N strings (empty string if a section is missing).
    """
    raw    = _strip_fences(raw)
    lines  = raw.splitlines()
    n      = len(headers)
    starts = [None] * n

    for i, pat in enumerate(headers):
        for idx, line in enumerate(lines):
            if re.match(pat, line.strip(), re.IGNORECASE):
                starts[i] = idx
                break

    # Build sections
    sections = []
    for i in range(n):
        if starts[i] is None:
            sections.append("")
            continue
        end = len(lines)
        for j in range(i + 1, n):
            if starts[j] is not None and starts[j] > starts[i]:
                end = starts[j]
                break
        section_lines = lines[starts[i]:end]
        # Strip: empty lines, ||| separators, ---, and section-label lines.
        # A section label is any line that ends with ":" and contains no commas
        # (e.g. "Dataset 2 (OQ):", "Here is the CSV:") — never a CSV data row.
        def _is_noise(line: str) -> bool:
            s = line.strip()
            if not s:                        return True
            if s in ("|||", "---"):          return True
            if s.endswith(":") and "," not in s:  return True   # label line
            return False
        cleaned = [l for l in section_lines if not _is_noise(l)]
        section_text = "\n".join(cleaned)
        # Final safety: strip any residual preamble within the section
        sections.append(_strip_preamble(section_text))

    while len(sections) < n:
        sections.append("")
    return sections


def _csv_to_df(csv_text: str) -> pd.DataFrame:
    if not csv_text or not csv_text.strip():
        return pd.DataFrame()
    try:
        df = pd.read_csv(
            io.StringIO(csv_text),
            quotechar='"',
            on_bad_lines='skip',
            skipinitialspace=True,
            dtype=str,           # keep everything as str; avoids float/int confusion
        )
        # Drop rows where every cell is NaN
        df.dropna(how='all', inplace=True)
        return df
    except Exception:
        return pd.DataFrame()


def _remove_duplicate_headers(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or len(df.columns) == 0:
        return df
    return df[df.iloc[:, 0].astype(str).str.strip() != df.columns[0]].reset_index(drop=True)


def _renumber_frs_ids(df: pd.DataFrame) -> pd.DataFrame:
    """
    Guarantee FRS ID column contains clean FRS-NNN codes.

    Problem: LLMs sometimes shift columns when a long description that
    contains commas is not properly quoted, causing the description text
    to land in the ID column.

    Fix: detect any row where ID looks like a sentence (> 10 chars or
    does not start with 'FRS') and rebuild the entire ID column as a
    clean sequence.  Also resets the 'Requirement_Link' in OQ if passed.
    """
    if df.empty or "ID" not in df.columns:
        return df
    df = df.copy()

    def _looks_like_id(val: str) -> bool:
        v = str(val).strip()
        return bool(re.match(r'^FRS-?\d+$', v, re.IGNORECASE)) and len(v) <= 10

    bad_ids = df["ID"].apply(lambda v: not _looks_like_id(str(v)))
    if bad_ids.any():
        # Renumber all rows unconditionally for consistency
        df["ID"] = [f"FRS-{i+1:03d}" for i in range(len(df))]
    else:
        # Normalise formatting: FRS001 → FRS-001
        def _normalise(v):
            v = str(v).strip().upper()
            m = re.match(r'FRS-?(\d+)', v)
            if m:
                return f"FRS-{int(m.group(1)):03d}"
            return v
        df["ID"] = df["ID"].apply(_normalise)
    return df


def _fill_missing_frs(urs_df: pd.DataFrame, frs_df: pd.DataFrame) -> pd.DataFrame:
    """
    Detect any URS requirement that has no FRS row and insert a clearly-flagged
    placeholder so nothing silently disappears from the output.
    The placeholder has Confidence=0.50 and Confidence_Flag='⚠️ Review Required'
    so it is immediately visible as needing manual completion.
    """
    if urs_df.empty or "Req_ID" not in urs_df.columns:
        return frs_df

    frs_urs_refs = set()
    if not frs_df.empty and "Source_URS_Ref" in frs_df.columns:
        frs_urs_refs = set(frs_df["Source_URS_Ref"].dropna().astype(str).str.strip())

    placeholders = []
    for _, row in urs_df.iterrows():
        uid  = str(row.get("Req_ID", "")).strip()
        desc = str(row.get("Requirement_Description", "")).strip()
        if uid and uid not in frs_urs_refs:
            # Determine next FRS number
            if not frs_df.empty and "ID" in frs_df.columns:
                existing_nums = []
                for fid in frs_df["ID"].dropna().astype(str):
                    m = re.match(r'FRS-(\d+)', fid.strip(), re.IGNORECASE)
                    if m: existing_nums.append(int(m.group(1)))
                next_n = max(existing_nums, default=0) + len(placeholders) + 1
            else:
                next_n = 1 + len(placeholders)
            frs_id = f"FRS-{next_n:03d}"
            placeholders.append({
                "ID":                      frs_id,
                "Requirement_Description": f"[HUMAN-IN-THE-LOOP SAFEGUARD — MANUAL REVIEW REQUIRED] "
                                           f"No FRS was generated for URS requirement: '{desc}'. "
                                           f"Please define the engineering implementation.",
                "Priority":                "N/A",
                "Risk":                    "High",
                "GxP_Impact":              "Direct",
                "Source_URS_Ref":          uid,
                "Source_Text":             str(row.get("Source_Text", "N/A")).strip() or "N/A",
                "Source_Page":             str(row.get("Source_Page", "N/A")).strip() or "N/A",
                "Confidence":              "0.50",
                "Confidence_Flag":         "⚠️ Review Required",
            })

    if not placeholders:
        return frs_df

    placeholder_df = pd.DataFrame(placeholders)
    result = pd.concat([frs_df, placeholder_df], ignore_index=True)
    result.fillna("N/A", inplace=True)
    return result


def _clean_frs_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Strip column-name prefixes that LLMs sometimes echo into values.
    e.g. "Risk: High" → "High", "Priority: Critical" → "Critical",
         "GxP_Impact: Direct" → "Direct"
    Also normalises capitalisation so downstream comparisons are reliable.
    """
    if df.empty:
        return df
    df = df.copy()
    _prefixes = {
        "Risk":       ["risk:", "risk :", "Risk:", "Risk :"],
        "Priority":   ["priority:", "priority :", "Priority:", "Priority :"],
        "GxP_Impact": ["gxp_impact:", "gxp_impact :", "GxP_Impact:", "GxP Impact:"],
    }
    for col, prefixes in _prefixes.items():
        if col in df.columns:
            for pfx in prefixes:
                df[col] = df[col].astype(str).str.replace(
                    pfx, "", case=False, regex=False
                ).str.strip()
            # Normalise to title-case so "high" → "High", "MEDIUM" → "Medium"
            # but keep the exact strings the rest of the code expects
            _map = {"high": "High", "medium": "Medium", "low": "Low",
                    "critical": "Critical", "direct": "Direct",
                    "indirect": "Indirect", "none": "None"}
            df[col] = df[col].str.lower().map(
                lambda v: _map.get(v, v.title()) if isinstance(v, str) else v
            )
    return df


def _renumber_oq_ids(df: pd.DataFrame) -> pd.DataFrame:
    """
    Always renumber OQ Test_IDs as a clean sequential OQ-001, OQ-002, OQ-003...
    This is unconditional — the LLM frequently skips numbers (e.g. OQ-001, 002, 003, 005)
    when it skips an FRS, leaving gaps. We never preserve gap-filled sequences.
    """
    if df.empty or "Test_ID" not in df.columns:
        return df
    df = df.copy()
    df["Test_ID"] = [f"OQ-{i+1:03d}" for i in range(len(df))]
    return df


def _fill_missing_oq(frs_df: pd.DataFrame, oq_df: pd.DataFrame) -> pd.DataFrame:
    """
    Detect any FRS requirement that has no OQ test case and insert a clearly-flagged
    placeholder so nothing silently disappears from the test matrix.

    Mirrors _fill_missing_frs — same philosophy: the AI skips quietly, Python catches it.
    The placeholder is stamped [AI SKIPPED — MANUAL REVIEW REQUIRED] so a validator
    immediately knows it needs a real test case written. Confidence = 0.50.

    Called AFTER _fill_missing_frs so that FRS placeholder rows (themselves inserted
    because the AI skipped a URS→FRS) are also caught here and get a matching OQ
    placeholder. After inserting, _renumber_oq_ids is called again by the caller
    to assign clean sequential IDs to the expanded set.
    """
    if frs_df.empty or "ID" not in frs_df.columns:
        return oq_df

    # Build set of FRS IDs that already have at least one OQ test
    linked_frs = set()
    if not oq_df.empty and "Requirement_Link" in oq_df.columns:
        linked_frs = set(
            oq_df["Requirement_Link"].dropna().astype(str).str.strip()
        )

    placeholders = []
    for _, row in frs_df.iterrows():
        fid  = str(row.get("ID", "")).strip()
        desc = str(row.get("Requirement_Description", "")).strip()[:120]
        urs  = str(row.get("Source_URS_Ref", "N/A")).strip()
        if fid and fid not in linked_frs:
            placeholders.append({
                "Test_ID":               "OQ-PLACEHOLDER",   # renumbered below
                "Requirement_Link":      fid,
                "Requirement_Link_Type": "FRS",
                "Test_Step":             (
                    f"[HUMAN-IN-THE-LOOP SAFEGUARD — MANUAL REVIEW REQUIRED] "
                    f"No OQ test was generated for {fid}: '{desc}'. "
                    f"Write executable test steps for this requirement."
                ),
                "Expected_Result":       f"[MANUAL ENTRY REQUIRED] Expected outcome for {fid}",
                "Pass_Fail_Criteria":    "[MANUAL ENTRY REQUIRED] Define pass/fail criteria",
                "Source":                f"Derived from {urs}",
                "Confidence":            "0.50",
                "Confidence_Flag":       "⚠️ Review Required",
            })

    if not placeholders:
        return oq_df

    result = pd.concat([oq_df, pd.DataFrame(placeholders)], ignore_index=True)
    result.fillna("N/A", inplace=True)
    return result


def _apply_confidence_flags(df: pd.DataFrame) -> pd.DataFrame:
    """
    Guarantee Confidence_Flag is set correctly.
    - Confidence < 0.70  → "⚠️ Review Required"
    - Confidence >= 0.70 → ""   (blank — never None)
    - Missing/unparseable → ""
    Python enforces this independently of whatever the LLM wrote.
    """
    if df.empty:
        return df
    df = df.copy()

    if "Confidence" not in df.columns:
        df["Confidence"]      = "1.00"
        df["Confidence_Flag"] = ""
        return df

    if "Confidence_Flag" not in df.columns:
        df["Confidence_Flag"] = ""

    def _flag(conf_val):
        try:
            c = float(str(conf_val).strip())
            return "⚠️ Review Required" if c < 0.70 else ""
        except (ValueError, TypeError):
            return ""

    df["Confidence_Flag"] = df["Confidence"].apply(_flag)

    # Replace any NaN/None in ALL columns with empty string for clean Excel output
    df = df.fillna("")
    return df


def run_cross_source_analysis(
    urs_text: str,
    sys_context_text: str,
    model_id: str,
    sys_context_name: str = "User Guide"
) -> tuple:
    """
    Cross-Source Gap Analysis (v29):
    Compares the URS against the User Guide to find:
      1. Features in the User Guide NOT mentioned in the URS
         → generate an FRS flagged [GAP-SOURCE: User Guide Only]
         → these represent system capabilities the company failed to validate
      2. URS requirements with NO corresponding feature in the User Guide
         → flag as [GAP-SOURCE: URS Only — Not in User Guide]
         → these suggest the system may not support the requirement at all

    Returns: (cross_frs_rows: list[dict], cross_gap_rows: list[dict])
    Both are appended to the main FRS and Gap_Analysis tables.
    """
    CROSS_SOURCE_PROMPT = f"""
You are a GxP validation engineer performing a bidirectional gap analysis.

DOCUMENT A — USER REQUIREMENTS SPECIFICATION (URS):
{urs_text[:4000]}

DOCUMENT B — SYSTEM USER GUIDE / PRODUCT MANUAL ('{sys_context_name}'):
{sys_context_text[:4000]}

TASK: Perform a bidirectional gap analysis between the two documents.

PART 1 — Features in User Guide NOT in URS:
Identify system features or capabilities described in the User Guide that have NO
corresponding requirement in the URS. These represent missed validation scope.
For each, generate one FRS row with Source_URS_Ref = "[GAP-SOURCE: User Guide Only]".

PART 2 — URS Requirements NOT supported by User Guide:
Identify URS requirements that describe functionality NOT described anywhere in the
User Guide. These suggest the system may not support the requirement.
Generate one gap row for each with Gap_Type = "URS_Not_In_UserGuide".

Output EXACTLY 2 CSV datasets separated by |||.
Include the header row in EACH dataset. Wrap comma-containing values in double-quotes.
Use N/A for any field that is not applicable.

Dataset 1 (cross-source FRS rows):
ID,Requirement_Description,Priority,Risk,GxP_Impact,Source_URS_Ref,Source_Text,Source_Page,Confidence,Confidence_Flag
- ID: XFRS-001, XFRS-002, etc.
- Requirement_Description: engineering/technical description of the feature from the User Guide
- Source_URS_Ref: EXACTLY the string "[GAP-SOURCE: User Guide Only]"
- Confidence_Flag: "Cross-Source Gap — Review Required"

Dataset 2 (cross-source gap rows):
Req_ID,Gap_Type,Description,Recommendation,Severity
- Req_ID: use the URS Req_ID if known, otherwise "URS-UNMATCHED-NNN"
- Gap_Type: URS_Not_In_UserGuide
- Description: what the URS requires and why the User Guide doesn't cover it
- Recommendation: specific action (e.g. contact vendor, raise a change request, or add to URS)
- Severity: Critical / High / Medium

If no gaps exist in either direction, output two CSV headers with no data rows.
"""
    try:
        response = completion(
            model=model_id,
            stream=False,
            temperature=0.1,   # low temp for deterministic gap comparison
            messages=[
                {"role": "system", "content": (
                    "You are a senior GxP validation specialist. "
                    "You produce only structured CSV output — no prose, no markdown fences. "
                    "Your gap analysis findings will be incorporated into a regulated validation package."
                )},
                {"role": "user", "content": CROSS_SOURCE_PROMPT}
            ]
        )
        raw = response.choices[0].message.content or ""
        raw = re.sub(r'^```[a-zA-Z]*\n?', '', raw, flags=re.MULTILINE)
        raw = re.sub(r'```\s*$',          '', raw, flags=re.MULTILINE)

        parts = raw.split("|||")
        xfrs_csv = parts[0].strip() if len(parts) > 0 else ""
        xgap_csv = parts[1].strip() if len(parts) > 1 else ""

        xfrs_df = _csv_to_df(_strip_preamble(xfrs_csv))
        xgap_df = _csv_to_df(_strip_preamble(xgap_csv))

        # Filter out "no gaps" prose rows
        if not xfrs_df.empty and "ID" in xfrs_df.columns:
            xfrs_df = xfrs_df[xfrs_df["ID"].astype(str).str.startswith("XFRS")].reset_index(drop=True)
        if not xgap_df.empty and "Req_ID" in xgap_df.columns:
            no_gap_mask = xgap_df["Req_ID"].astype(str).str.lower().str.contains(
                r'no gap|none|n/a|not applicable', na=False, regex=True
            )
            xgap_df = xgap_df[~no_gap_mask].reset_index(drop=True)

        return xfrs_df, xgap_df

    except Exception as e:
        st.warning(f"⚠️ Cross-source analysis failed ({e}) — proceeding without it.")
        return pd.DataFrame(), pd.DataFrame()


def run_segmented_analysis(
    file_bytes: bytes,
    model_id: str,
    progress_bar,
    status_text,
    sys_context_bytes: bytes = None
) -> tuple:
    """
    Two-pass analysis with Fail-Stop Protocol (v27).

    Pass 1 — per-chunk URS extraction: produces a clean structured URS table.
    Pass 2 — single call with full URS table: produces FRS / OQ / Gap.
    Returns: (urs_df, frs_df, oq_df, trace_df, gap_df)

    Fail-Stop Protocol (21 CFR Part 11 / GxP compliance):
      If ANY Pass-1 segment fails, the entire analysis is aborted and a
      SegmentFailureError is raised. A validation package with missing
      pages would fail a regulatory audit — 100% coverage or nothing.
      The exception is caught in show_app() which logs the failure and
      shows a compliance-grade error message.

    SysContext (User Guide) injection:
      If sys_context_bytes is provided, the first 6 pages of the guide are
      extracted and injected into BOTH Pass-1 and Pass-2 system prompts so
      the LLM can reference actual screen names, module names, and field
      names when writing FRS descriptions and OQ test steps.

    Cross-chunk ID resequencing:
      After combining chunks, FRS IDs and OQ IDs are globally resequenced
      so FRS-001…FRS-N and OQ-001…OQ-M are always unique and sequential
      regardless of how many chunks the document was split into.
    """
    class SegmentFailureError(RuntimeError):
        pass

    all_pages = extract_pages(file_bytes)
    if not all_pages:
        raise SegmentFailureError(
            "No pages could be extracted from the uploaded PDF. "
            "The file may be image-only (scanned) or corrupt. "
            "Per ALCOA+ standards, non-searchable PDFs cannot be AI-validated."
        )

    # OCR / searchability gate: require minimum text density
    joined_text = "\n".join(all_pages)
    if len(joined_text.strip()) < 100:
        raise SegmentFailureError(
            "⛔ Compliance Warning: Document is not OCR-searchable.\n"
            "Non-searchable PDFs cannot be validated by the AI engine per ALCOA+ standards.\n"
            "Please convert the document to a text-based PDF using OCR software before uploading."
        )

    chunks = [all_pages[i:i + CHUNK_SIZE] for i in range(0, len(all_pages), CHUNK_SIZE)]
    total  = len(chunks)

    # ── SysContext (User Guide) extraction ───────────────────────────────────
    sys_context = ""
    if sys_context_bytes:
        try:
            sys_pages   = extract_pages(sys_context_bytes)
            sys_context = "\n\n".join(sys_pages[:6])   # first 6 pages ≈ 4-8k chars
            status_text.text("📖 User Guide loaded — injecting into analysis context...")
        except Exception as e:
            st.warning(f"⚠️ Could not extract User Guide context: {e} — proceeding without it.")

    # ── PASS 1: Extract structured URS table from each chunk ─────────────────
    urs_frames = []
    for idx, chunk_pages in enumerate(chunks):
        chunk_text = "\n\n".join(chunk_pages)
        status_text.text(f"📄 Pass 1 — Extracting URS: segment {idx + 1} of {total}...")
        progress_bar.progress((idx) / (total * 2))

        try:
            # Phase 1: stream=True prevents silent 600s hang on Pass 1 segments
            stream_resp_p1 = completion(
                model=model_id,
                stream=True,
                temperature=TEMPERATURE,
                timeout=900,
                messages=[
                    {"role": "system", "content": _make_system_prompt(sys_context)},
                    {"role": "user",   "content": build_pass1_prompt(chunk_text, idx, total)}
                ]
            )
            raw_urs = ""
            for chunk in stream_resp_p1:
                delta = (chunk.choices[0].delta.content or "") if chunk.choices else ""
                raw_urs += delta
                if len(raw_urs) % 600 < len(delta) + 1:
                    status_text.text(
                        f"📄 Pass 1 — segment {idx+1}/{total}: "
                        f"extracting... ({len(raw_urs):,} chars)"
                    )
            raw_urs = re.sub(r'^```[a-zA-Z]*\n?', '', raw_urs, flags=re.MULTILINE)
            raw_urs = re.sub(r'```\s*$',          '', raw_urs, flags=re.MULTILINE)
            raw_urs = _strip_preamble(raw_urs.strip())
            df_urs  = _csv_to_df(raw_urs)
            if not df_urs.empty:
                urs_frames.append(df_urs)
        except Exception as e:
            # FAIL-STOP: any segment failure aborts the entire run
            raise SegmentFailureError(
                f"Pass 1 segment {idx + 1}/{total} failed: {e}\n\n"
                f"Analysis aborted. Per GxP Fail-Stop Protocol, an incomplete analysis "
                f"(missing pages {idx * CHUNK_SIZE + 1}–{min((idx + 1) * CHUNK_SIZE, len(all_pages))}) "
                f"cannot be used as a validation artifact. Please retry."
            ) from e

    def _combine(frames):
        if not frames:
            return pd.DataFrame()
        c = pd.concat(frames, ignore_index=True)
        c = _remove_duplicate_headers(c)
        c.dropna(how='all', inplace=True)
        return c

    urs_final = _combine(urs_frames)

    # ── Cross-chunk URS ID resequencing ──────────────────────────────────────
    # Each chunk's LLM may restart numbering at URS-001. Renumber globally.
    if not urs_final.empty and "Req_ID" in urs_final.columns:
        urs_final = urs_final.copy()
        urs_final["Req_ID"] = [f"URS-{i+1:03d}" for i in range(len(urs_final))]

    urs_final = _apply_confidence_flags(urs_final)

    # ── Filter fabricated rows from empty table shells ─────────────────────
    # Pass 1 sometimes extracts section headers and empty table rows from
    # sparse URS documents as fake requirements. Remove rows where:
    #   (a) Requirement_Description is very short (< 15 chars) — section header
    #   (b) Requirement_Description matches known header patterns
    #   (c) Req_ID is empty, N/A, or non-URS format
    if not urs_final.empty and "Requirement_Description" in urs_final.columns:
        _before = len(urs_final)
        _hdr_patterns = [
            r"^(single sample|bulk sample|parent.child|aliquot|barcode|bi.direct"
            r"|data integrity|audit trail|computer.gen|electronic sig|access level"
            r"|reagent.*registry|study protocol|pull sched|oot detect|system.*arch"
            r"|login proc|role def|instrument.*reg|note:|n/?a)$"
        ]
        import re as _re
        def _is_fabricated(row):
            desc = str(row.get("Requirement_Description", "")).strip()
            req_id = str(row.get("Req_ID", "")).strip()
            # Too short to be a real requirement
            if len(desc) < 15:
                return True
            # Looks like a section sub-heading (no "shall" or "must")
            if not _re.search(r"(shall|must|will)", desc, _re.IGNORECASE):
                if len(desc) < 60:
                    return True
            # Matches known empty-shell header patterns
            for pat in _hdr_patterns:
                if _re.match(pat, desc.lower()):
                    return True
            return False

        mask = urs_final.apply(_is_fabricated, axis=1)
        urs_final = urs_final[~mask].reset_index(drop=True)
        # Re-apply sequential IDs after filtering
        urs_final["Req_ID"] = [f"URS-{i+1:03d}" for i in range(len(urs_final))]
        _after = len(urs_final)
        if _before != _after:
            st.caption(f"ℹ️ Pass 1 extracted {_before} rows — filtered {_before - _after} "
                       f"section headers/empty rows → {_after} real requirements.")

    progress_bar.progress(0.5)
    status_text.text(f"✅ Pass 1 complete — {len(urs_final)} requirements found. Running Pass 2...")

    # ── PASS 2: Generate FRS / OQ / Gap from URS table ────────────────────────
    frs_frames, oq_frames, gap_frames = [], [], []

    if urs_final.empty:
        raise SegmentFailureError(
            "Pass 1 extracted zero requirements. The document may be empty, "
            "image-only, or contain no recognisable requirement statements. "
            "Analysis cannot continue."
        )

    urs_csv_str = urs_final.to_csv(index=False)
    urs_lines   = urs_csv_str.split("\n")
    header_line = urs_lines[0]
    data_lines  = [l for l in urs_lines[1:] if l.strip()]  # skip blank rows
    p2_total    = len(data_lines)
    sys_summary = _summarise_sys_context(sys_context)
    _failed_reqs = []

    # ── Fast-path: small documents (≤ 15 requirements) ───────────────────────
    # Send all requirements in a single streaming batch call.
    # Avoids the per-req overhead (15 × 30s = 7.5 min) for small docs.
    # Large documents (> 15 requirements) use the per-req loop below.
    if p2_total <= 15:
        status_text.text(
            f"🔬 Pass 2 — small document ({p2_total} requirements): "
            f"single batch call..."
        )
        progress_bar.progress(0.52)
        try:
            _full_csv = header_line + "\n" + "\n".join(data_lines)
            _stream_fp = completion(
                model=model_id,
                stream=True,
                temperature=TEMPERATURE,
                timeout=300,
                messages=[
                    {"role": "system", "content": _make_system_prompt(sys_summary)},
                    {"role": "user",   "content": build_pass2_prompt(_full_csv, sys_summary)}
                ]
            )
            _raw_fp = ""
            for _chunk in _stream_fp:
                _delta = (_chunk.choices[0].delta.content or "") if _chunk.choices else ""
                _raw_fp += _delta
                if len(_raw_fp) % 800 < len(_delta) + 1:
                    status_text.text(
                        f"🔬 Pass 2 — generating... ({len(_raw_fp):,} chars)"
                    )
            _sections = _robust_split_datasets(_raw_fp, _PASS2_HEADERS)
            for _frames, _csv_text in [
                (frs_frames, _sections[0]),
                (oq_frames,  _sections[1]),
                (gap_frames, _sections[2]),
            ]:
                _df = _csv_to_df(_csv_text)
                if not _df.empty:
                    _frames.append(_df)
            progress_bar.progress(0.95)
            status_text.text(
                f"✅ Pass 2 complete — "
                f"FRS: {sum(len(f) for f in frs_frames)} rows  |  "
                f"OQ: {sum(len(f) for f in oq_frames)} tests"
            )
        except Exception as _e:
            raise SegmentFailureError(
                f"Pass 2 fast-path failed: {_e}\n\n"
                f"Analysis aborted. Please retry."
            ) from _e
    else:
        # ── Phase 2: per-requirement processing (large documents > 15 reqs) ──
        for p2_idx, req_row in enumerate(data_lines):
            if not req_row.strip():
                continue

            pct = 0.50 + (p2_idx / max(p2_total, 1)) * 0.44
            progress_bar.progress(min(pct, 0.94))
            status_text.text(
                f"🔬 Pass 2 — requirement {p2_idx+1}/{p2_total}  |  "
                f"FRS: {sum(len(f) for f in frs_frames)} rows  |  "
                f"OQ: {sum(len(f) for f in oq_frames)} tests"
            )

            try:
                stream_resp = completion(
                    model=model_id,
                    stream=True,
                    temperature=TEMPERATURE,
                    timeout=120,
                    messages=[
                        {"role": "system", "content": _make_system_prompt(sys_summary)},
                        {"role": "user",   "content": build_pass2_single_prompt(
                            req_row, header_line, sys_summary)}
                    ]
                )
                raw_p2 = ""
                for chunk in stream_resp:
                    delta = (chunk.choices[0].delta.content or "") if chunk.choices else ""
                    raw_p2 += delta

            except Exception as e:
                # Phase 2: per-requirement retry once before skipping
                try:
                    import time as _time
                    _time.sleep(8)
                    stream_resp2 = completion(
                        model=model_id,
                        stream=True,
                        temperature=TEMPERATURE,
                        timeout=120,
                        messages=[
                            {"role": "system", "content": _make_system_prompt(sys_summary)},
                            {"role": "user",   "content": build_pass2_single_prompt(
                                req_row, header_line, sys_summary)}
                        ]
                    )
                    raw_p2 = ""
                    for chunk in stream_resp2:
                        delta = (chunk.choices[0].delta.content or "") if chunk.choices else ""
                        raw_p2 += delta
                except Exception as e2:
                    # Log the skip — do not abort the whole run
                    _failed_reqs.append(f"req {p2_idx+1}: {str(e2)[:80]}")
                    continue  # move to next requirement

            sections = _robust_split_datasets(raw_p2, _PASS2_HEADERS)
            frs_csv, oq_csv, gap_csv = sections[0], sections[1], sections[2]
            for frames, csv_text in [
                (frs_frames,  frs_csv),
                (oq_frames,   oq_csv),
                (gap_frames,  gap_csv),
            ]:
                df = _csv_to_df(csv_text)
                if not df.empty:
                    frames.append(df)

    # Surface any skipped requirements as a soft warning (not a hard abort)
    if _failed_reqs:
        st.warning(
            f"⚠️ {len(_failed_reqs)} requirement(s) skipped after retry failure "
            f"and excluded from this package. Re-run to recover: "
            f"{'; '.join(_failed_reqs[:5])}"
        )

    progress_bar.progress(0.95)
    status_text.text("✅ Both passes complete — running deterministic checks...")

    frs_final = _combine(frs_frames)
    oq_final  = _combine(oq_frames)
    gap_final = _combine(gap_frames)

    # ── Post-processing: global ID resequencing (cross-chunk dedup) ──────────
    # Each Pass-2 batch restarts FRS/OQ numbering. Renumber globally BEFORE
    # any cross-reference so FRS-001…FRS-N are unique across all batches.
    frs_final = _renumber_frs_ids(frs_final)
    oq_final  = _renumber_oq_ids(oq_final)

    # Patch OQ Requirement_Link to the renumbered FRS IDs using Source_URS_Ref mapping.
    # Build a map: Source_URS_Ref → new FRS ID (after renumber)
    if not frs_final.empty and "ID" in frs_final.columns and "Source_URS_Ref" in frs_final.columns:
        urs_to_frs = {}
        for _, r in frs_final.iterrows():
            u = str(r.get("Source_URS_Ref", "")).strip()
            f = str(r.get("ID", "")).strip()
            if u and f:
                urs_to_frs[u] = f
        # Update OQ Requirement_Link using Source column ("Derived from URS-NNN")
        if not oq_final.empty and "Requirement_Link" in oq_final.columns and "Source" in oq_final.columns:
            def _remap_link(row):
                link = str(row.get("Requirement_Link", "")).strip()
                src  = str(row.get("Source", "")).strip()
                # If link looks like a FRS ID already and it exists, keep it
                if re.match(r'^FRS-\d+$', link, re.IGNORECASE) and link in set(frs_final["ID"]):
                    return link
                # Try to extract URS ref from Source field
                m = re.search(r'URS-(\d+)', src, re.IGNORECASE)
                if m:
                    urs_ref = f"URS-{int(m.group(1)):03d}"
                    return urs_to_frs.get(urs_ref, link)
                return link
            oq_final["Requirement_Link"] = oq_final.apply(_remap_link, axis=1)

    frs_final = _clean_frs_columns(frs_final)
    frs_final = _fill_missing_frs(urs_final, frs_final)
    oq_final  = _fill_missing_oq(frs_final, oq_final)
    oq_final  = _renumber_oq_ids(oq_final)

    frs_final = _apply_confidence_flags(frs_final)
    oq_final  = _apply_confidence_flags(oq_final)
    urs_final = _apply_confidence_flags(urs_final)

    for df in [frs_final, oq_final, urs_final]:
        df.fillna("N/A", inplace=True)
        df.replace("", "N/A", inplace=True)

    gap_final   = _clean_gap_analysis(gap_final)
    trace_final = _build_traceability(urs_final, frs_final, oq_final)

    # ── Pass 3 (optional): Cross-Source Gap Analysis ─────────────────────────
    # Only runs when a User Guide was uploaded. Compares the URS against the
    # guide to find: (a) features in the guide not in the URS, (b) URS reqs
    # not supported by the guide. Results are appended to FRS and Gap tables.
    if sys_context and len(sys_context.strip()) > 200:
        status_text.text("🔀 Pass 3 — Cross-source URS ↔ User Guide gap analysis...")
        urs_text_for_cross = "\n".join(all_pages[:8])   # first 8 pages of URS
        xfrs_df, xgap_df = run_cross_source_analysis(
            urs_text    = urs_text_for_cross,
            sys_context_text = sys_context,
            model_id    = model_id,
            sys_context_name = "User Guide"
        )
        # Append cross-source FRS rows to main FRS table
        if not xfrs_df.empty:
            frs_final = pd.concat([frs_final, xfrs_df], ignore_index=True)
            frs_final.fillna("N/A", inplace=True)
        # Append cross-source gap rows to main gap table
        if not xgap_df.empty:
            # Ensure matching columns
            for col in ["Req_ID", "Gap_Type", "Description", "Recommendation", "Severity"]:
                if col not in xgap_df.columns:
                    xgap_df[col] = "N/A"
            gap_final = pd.concat([gap_final, xgap_df], ignore_index=True)
            gap_final.fillna("N/A", inplace=True)

        # Rebuild traceability to include cross-source FRS rows
        trace_final = _build_traceability(urs_final, frs_final, oq_final)

    progress_bar.progress(1.0)
    status_text.text("✅ All segments processed — compiling workbook...")

    return urs_final, frs_final, oq_final, trace_final, gap_final


def _build_traceability(urs_df: pd.DataFrame,
                        frs_df: pd.DataFrame,
                        oq_df:  pd.DataFrame) -> pd.DataFrame:
    """
    Rebuild Traceability Matrix from URS as the primary key.

    Every URS requirement gets a row regardless of whether the LLM generated
    an FRS for it. This makes skipped/missing FRS rows visible.

    Columns: URS_Req_ID, URS_Description, FRS_Ref, Test_IDs, Test_Count,
             Coverage_Status, Gap_Analysis

    Coverage logic per FRS row (risk-aware):
      - No FRS generated  → Missing_FRS + [GAP]
      - FRS exists, 0 OQ  → Not Covered  + [GAP]
      - FRS exists, < min → Partial       + [PARTIAL GAP]
      - FRS exists, ≥ min → Covered
    """
    MIN_TESTS = {"high": 3, "medium": 2, "low": 1}

    # Build URS description lookup
    urs_desc: dict = {}
    if not urs_df.empty and "Req_ID" in urs_df.columns:
        for _, r in urs_df.iterrows():
            uid  = str(r.get("Req_ID", "")).strip()
            desc = str(r.get("Requirement_Description", "")).strip()
            if uid:
                urs_desc[uid] = desc

    # Build FRS lookup keyed by Source_URS_Ref → list of FRS rows
    frs_by_urs: dict = {}
    if not frs_df.empty and "Source_URS_Ref" in frs_df.columns:
        for _, r in frs_df.iterrows():
            ref = str(r.get("Source_URS_Ref", "")).strip()
            if ref:
                frs_by_urs.setdefault(ref, []).append(r)

    # Build OQ lookup keyed by Requirement_Link (FRS_ID) → list of OQ Test_IDs
    oq_map: dict = {}
    if not oq_df.empty and "Requirement_Link" in oq_df.columns and "Test_ID" in oq_df.columns:
        for _, r in oq_df.iterrows():
            link    = str(r.get("Requirement_Link", "")).strip()
            test_id = str(r.get("Test_ID", "")).strip()
            if link and test_id:
                oq_map.setdefault(link, []).append(test_id)

    # Determine URS ID list — prefer urs_df order, fallback to FRS refs
    if not urs_df.empty and "Req_ID" in urs_df.columns:
        urs_ids = [str(v).strip() for v in urs_df["Req_ID"].dropna() if str(v).strip()]
    else:
        urs_ids = sorted(frs_by_urs.keys())

    rows = []
    for urs_id in urs_ids:
        urs_description = urs_desc.get(urs_id, "")
        frs_rows = frs_by_urs.get(urs_id, [])

        if not frs_rows:
            # LLM skipped this URS — critical gap
            rows.append({
                "URS_Req_ID":       urs_id,
                "URS_Description":  urs_description,
                "FRS_Ref":          "—",
                "Test_IDs":         "",
                "Test_Count":       "0",
                "Coverage_Status":  "Missing FRS",
                "Gap_Analysis":     f"[GAP] No FRS requirement was generated for {urs_id}. "
                                    "AI may have skipped this requirement.",
            })
            continue

        for frs_row in frs_rows:
            frs_id  = str(frs_row.get("ID", "")).strip()
            risk    = str(frs_row.get("Risk", "low")).strip().lower()
            min_req = MIN_TESTS.get(risk, 1)

            real_tests = oq_map.get(frs_id, [])
            count      = len(real_tests)
            test_str   = "; ".join(sorted(real_tests)) if real_tests else ""

            if count == 0:
                status = "Not Covered"
                gap    = (f"[GAP] {frs_id} has no OQ test case. "
                          f"0/{min_req} required for {risk.title()} risk.")
            elif count < min_req:
                status = "Partial"
                gap    = (f"[PARTIAL GAP] {frs_id} has {count}/{min_req} tests "
                          f"required for {risk.title()} risk.")
            else:
                status = "Covered"
                gap    = ""

            rows.append({
                "URS_Req_ID":       urs_id,
                "URS_Description":  urs_description,
                "FRS_Ref":          frs_id,
                "Test_IDs":         test_str,
                "Test_Count":       str(count),
                "Coverage_Status":  status,
                "Gap_Analysis":     gap,
            })

    result = pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        "URS_Req_ID", "URS_Description", "FRS_Ref", "Test_IDs",
        "Test_Count", "Coverage_Status", "Gap_Analysis"
    ])
    result.fillna("", inplace=True)
    return result


# =============================================================================
# 6b. DETERMINISTIC GAP VALIDATION ENGINE  (R1–R5)
# =============================================================================

NON_TESTABLE_KEYWORDS = [
    "user friendly", "user-friendly", "easy to use", "easy-to-use",
    "intuitive", "fast", "quickly", "seamlessly", "simple", "straightforward",
    "efficient", "smooth", "pleasant", "elegant", "modern", "robust",
    "flexible", "scalable", "reliable", "stable", "responsive",
    "convenient", "accessible", "appealing",
]

AMBIGUOUS_KEYWORDS = [
    "appropriate", "adequate", "sufficient", "reasonable", "as needed",
    "if necessary", "where applicable", "etc", "and/or", "various",
    "many", "several", "some", "other", "etc.", "normal", "standard",
    "should consider", "may need", "could be",
]


def _token_overlap(a: str, b: str) -> float:
    """Simple Jaccard token overlap for duplicate detection."""
    ta = set(re.findall(r'\w+', a.lower()))
    tb = set(re.findall(r'\w+', b.lower()))
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


VALID_GAP_TYPES = {"Untestable", "No_Test_Coverage", "Orphan_Test",
                   "Ambiguous", "Duplicate", "Missing_FRS",
                   "Non_Functional", "Missing_Test", "Non_Testable_Requirement"}

def _clean_gap_analysis(gap_df: pd.DataFrame) -> pd.DataFrame:
    """
    Post-process LLM-generated Gap_Analysis to fix two known corruption patterns:

    1. LLM writes "No gaps identified" (or similar) as the Req_ID value while still
       populating other fields with real or nonsensical values.
       → Remove these rows entirely. The absence of real gap rows IS the signal.

    2. LLM writes a severity value (Critical/High/Medium/Low) in Gap_Type column.
       → Correct by validating Gap_Type against allowed enum values; replace
         invalid values with "Ambiguous".

    3. Replace all remaining None / blank cells with "N/A".
    """
    if gap_df.empty:
        return gap_df

    df = gap_df.copy()

    # Drop rows where Req_ID signals "no gap" prose
    if "Req_ID" in df.columns:
        no_gap_mask = df["Req_ID"].astype(str).str.lower().str.contains(
            r'no gap|no issue|none|nothing|not applicable|n/a', na=False, regex=True
        )
        df = df[~no_gap_mask].reset_index(drop=True)

    # Fix Gap_Type values that contain severity words instead of gap type enum
    if "Gap_Type" in df.columns:
        severity_words = {"critical", "high", "medium", "low"}
        def _fix_gap_type(v):
            v_str = str(v).strip()
            if v_str in VALID_GAP_TYPES:
                return v_str
            if v_str.lower() in severity_words:
                return "Ambiguous"   # best guess when LLM put severity in gap type col
            return v_str if v_str else "N/A"
        df["Gap_Type"] = df["Gap_Type"].apply(_fix_gap_type)

    # Replace all blank / None with N/A
    df = df.fillna("N/A")
    df = df.replace("", "N/A")

    return df


def run_deterministic_validation(
    frs_df: pd.DataFrame,
    oq_df: pd.DataFrame,
    gap_df: pd.DataFrame,
    urs_df: pd.DataFrame = None,
) -> tuple:
    """
    Rules:
      R0 — URS req with no FRS generated           → Gap_Type: Missing_FRS
      R1 — FRS req without any OQ test             → Gap_Type: No_Test_Coverage
      R2 — OQ test with no matching FRS req         → Gap_Type: Orphan_Test
      R3 — Non-testable keywords in description     → Gap_Type: Untestable
      R4 — Risk-tier req with < required OQ tests   → Gap_Type: No_Test_Coverage
      R5 — Near-duplicate FRS descriptions          → Gap_Type: Duplicate

    Traceability is NOT touched here — it is Python-built by _build_traceability().
    Returns: (enriched_gap_df, det_issues_df)
    """
    issues = []

    frs_ids = set()
    if not frs_df.empty and "ID" in frs_df.columns:
        frs_ids = set(frs_df["ID"].dropna().astype(str).str.strip())

    # FRS lookup by Source_URS_Ref for R0
    frs_urs_refs = set()
    if not frs_df.empty and "Source_URS_Ref" in frs_df.columns:
        frs_urs_refs = set(frs_df["Source_URS_Ref"].dropna().astype(str).str.strip())

    oq_req_links = set()
    oq_test_ids  = set()
    if not oq_df.empty:
        if "Requirement_Link" in oq_df.columns:
            oq_req_links = set(oq_df["Requirement_Link"].dropna().astype(str).str.strip())
        if "Test_ID" in oq_df.columns:
            oq_test_ids = set(oq_df["Test_ID"].dropna().astype(str).str.strip())

    # ── R0: URS requirement with no FRS generated ────────────────────────────
    if urs_df is not None and not urs_df.empty and "Req_ID" in urs_df.columns:
        for _, row in urs_df.iterrows():
            uid = str(row.get("Req_ID", "")).strip()
            if uid and uid not in frs_urs_refs:
                issues.append({
                    "Rule":           "R0",
                    "Req_ID":         uid,
                    "Gap_Type":       "Missing_FRS",
                    "Description":    f"{uid} has no FRS requirement generated. "
                                      "The AI may have skipped this requirement.",
                    "Recommendation": "Re-run analysis or manually create an FRS "
                                      "requirement for this URS item.",
                    "Severity":       "Critical",
                })

    # ── R1: FRS req without any OQ test ──────────────────────────────────────
    # Source of truth is the actual OQ rows — never the traceability matrix.
    for frs_id in sorted(frs_ids):
        if frs_id not in oq_req_links:
            issues.append({
                "Rule":            "R1",
                "Req_ID":          frs_id,
                "Gap_Type":        "No_Test_Coverage",
                "Description":     f"{frs_id} has no OQ test case linked to it.",
                "Recommendation":  "Create at least one OQ test case for this FRS requirement.",
                "Severity":        "High",
            })
            # Traceability is Python-built; Coverage_Status is already "Not Covered".
            # No mutation needed here — _build_traceability already set [GAP].

    # ── R2: Orphan OQ tests ───────────────────────────────────────────────────
    if not oq_df.empty and "Requirement_Link" in oq_df.columns:
        for _, row in oq_df.iterrows():
            link    = str(row.get("Requirement_Link", "")).strip()
            test_id = str(row.get("Test_ID", "")).strip()
            if link and link not in frs_ids:
                issues.append({
                    "Rule":            "R2",
                    "Req_ID":          test_id,
                    "Gap_Type":        "Orphan_Test",
                    "Description":     f"OQ test {test_id} links to '{link}' which has no FRS entry.",
                    "Recommendation":  "Verify the requirement reference or remove/reassign this test.",
                    "Severity":        "Medium",
                })

    # ── R3: Non-testable + Ambiguous keyword detection ────────────────────────
    desc_col = "Requirement_Description" if not frs_df.empty and "Requirement_Description" in frs_df.columns else None
    if desc_col:
        for _, row in frs_df.iterrows():
            desc = str(row.get(desc_col, "")).lower()
            fid  = str(row.get("ID", "")).strip()

            nt_found  = [kw for kw in NON_TESTABLE_KEYWORDS if kw in desc]
            amb_found = [kw for kw in AMBIGUOUS_KEYWORDS      if kw in desc]

            if nt_found:
                issues.append({
                    "Rule":            "R3",
                    "Req_ID":          fid,
                    "Gap_Type":        "Untestable",
                    "Description":     f"Non-testable language detected: {', '.join(nt_found)}",
                    "Recommendation":  "Rewrite as specific, measurable requirement (e.g. add numeric criteria).",
                    "Severity":        "High",
                })
                gap_df = pd.concat([gap_df, pd.DataFrame([{
                    "Req_ID":          fid,
                    "Gap_Type":        "Untestable",
                    "Description":     f"Non-testable keywords: {', '.join(nt_found)}",
                    "Recommendation":  "Rewrite as measurable, specific requirement.",
                    "Severity":        "High",
                }])], ignore_index=True)

            elif amb_found:
                issues.append({
                    "Rule":            "R3b",
                    "Req_ID":          fid,
                    "Gap_Type":        "Ambiguous",
                    "Description":     f"Ambiguous language detected: {', '.join(amb_found)}",
                    "Recommendation":  "Clarify intent with specific, unambiguous wording.",
                    "Severity":        "Medium",
                })
                gap_df = pd.concat([gap_df, pd.DataFrame([{
                    "Req_ID":          fid,
                    "Gap_Type":        "Ambiguous",
                    "Description":     f"Ambiguous keywords: {', '.join(amb_found)}",
                    "Recommendation":  "Clarify requirement with precise, unambiguous language.",
                    "Severity":        "Medium",
                }])], ignore_index=True)

    # ── R3c: Non-functional requirement detection ─────────────────────────────
    # Non-functional requirements (availability, SLA, performance targets) need
    # a different test strategy — not testable via standard OQ steps.
    # Flagged so teams define a separate load/stress/performance protocol.
    NON_FUNCTIONAL_KEYWORDS = [
        "availability", "uptime", "sla", "response time", "throughput",
        "maintainability", "portability", "interoperability", "disaster recovery",
        "backup", "recovery time", "rto", "rpo", "capacity", "concurrency",
        "load", "stress", "fault tolerance",
    ]
    if desc_col:
        for _, row in frs_df.iterrows():
            desc = str(row.get(desc_col, "")).lower()
            fid  = str(row.get("ID", "")).strip()
            nf_found = [kw for kw in NON_FUNCTIONAL_KEYWORDS if kw in desc]
            if nf_found:
                issues.append({
                    "Rule":            "R3c",
                    "Req_ID":          fid,
                    "Gap_Type":        "Non_Functional",
                    "Description":     (f"Non-functional requirement detected: {', '.join(nf_found)}. "
                                        f"Standard OQ steps are insufficient for this requirement type."),
                    "Recommendation":  ("Define a separate non-functional test protocol: load/stress test, "
                                        "SLA monitoring, or performance benchmark. Document in IQ/PQ."),
                    "Severity":        "Medium",
                })

    # ── R3d: Non-Testable Requirement detection (scans URS, not FRS) ─────────
    # Flags URS requirements marked Testable=No or containing weak/ambiguous
    # verbs with no quantitative criteria. These are a direct 21 CFR Part 11
    # compliance risk — a requirement that cannot be tested cannot be validated.
    # Maps each weak term to a concrete remediation recommendation.
    WEAK_VERB_REMEDIATION = {
        "user-friendly":  "Define measurable usability criteria (e.g., task completion rate ≥95%, error rate <5%)",
        "user friendly":  "Define measurable usability criteria (e.g., task completion rate ≥95%, error rate <5%)",
        "easy":           "Specify what 'easy' means quantitatively (e.g., onboarding ≤3 clicks, help requests <2/session)",
        "intuitive":      "Define learnability criteria (e.g., new user completes primary task without help within 5 minutes)",
        "fast":           "Specify maximum response time (e.g., page load ≤2 seconds under 100 concurrent users)",
        "quickly":        "Specify maximum response time (e.g., query returns results within 1 second)",
        "seamless":       "Define integration acceptance criteria (e.g., zero data loss, end-to-end transaction ≤3 seconds)",
        "simple":         "Specify complexity metric (e.g., ISO 9241 SUS score ≥80)",
        "flexible":       "Define the specific configuration options required (list each configurable parameter)",
        "robust":         "Specify fault-tolerance criteria (e.g., system recovers from single component failure within 30 seconds)",
        "scalable":       "Define quantitative capacity targets (e.g., supports 500 concurrent users with <5% response degradation)",
        "reliable":       "Specify uptime/availability target (e.g., 99.9% uptime excluding planned maintenance)",
        "convenient":     "Define task-time criteria (e.g., common workflows completable in ≤3 steps)",
        "smooth":         "Specify latency/throughput targets for the identified workflow",
        "modern":         "Remove subjective aesthetic term; specify functional requirements instead",
        "efficient":      "Specify time-on-task or resource consumption target (e.g., batch process ≤10 minutes for 10,000 records)",
        "should":         "Replace 'should' with 'shall' to make this a mandatory requirement, or remove if optional",
        "may":            "Clarify whether this is mandatory (use 'shall') or optional — ambiguous modal verb",
        "appropriate":    "Define specific acceptance criteria for what constitutes 'appropriate'",
        "adequate":       "Define minimum quantitative threshold for adequacy",
        "sufficient":     "Define minimum quantitative threshold for sufficiency",
    }
    if not urs_df.empty:
        urs_req_col  = "Requirement_Description" if "Requirement_Description" in urs_df.columns else None
        urs_test_col = "Testable" if "Testable" in urs_df.columns else None
        urs_id_col   = "Req_ID"   if "Req_ID"   in urs_df.columns else None
        if urs_req_col and urs_id_col:
            for _, row in urs_df.iterrows():
                uid  = str(row.get(urs_id_col, "")).strip()
                desc = str(row.get(urs_req_col, "")).lower()
                testable = str(row.get(urs_test_col, "")).strip().lower() if urs_test_col else "yes"

                # Find the first matching weak verb/phrase
                matched_term = next(
                    (term for term in WEAK_VERB_REMEDIATION if term in desc),
                    None
                )
                if matched_term or testable == "no":
                    term_label = matched_term or "non-testable language"
                    recommendation = WEAK_VERB_REMEDIATION.get(
                        matched_term,
                        "Replace vague language with specific, measurable acceptance criteria."
                    )
                    issues.append({
                        "Rule":           "R3d",
                        "Req_ID":         uid,
                        "Gap_Type":       "Non_Testable_Requirement",
                        "Description":    (f"Contains ambiguous/non-testable term: '{term_label}'. "
                                           f"This requirement cannot be objectively validated — "
                                           f"a direct 21 CFR Part 11 compliance risk."),
                        "Recommendation": recommendation,
                        "Severity":       "High",
                    })
                    gap_df = pd.concat([gap_df, pd.DataFrame([{
                        "Req_ID":         uid,
                        "Gap_Type":       "Non_Testable_Requirement",
                        "Description":    f"Ambiguous term: '{term_label}'",
                        "Recommendation": recommendation,
                        "Severity":       "High",
                    }])], ignore_index=True)

    # ── R4: High-risk reqs with insufficient OQ test count ───────────────────
    if not frs_df.empty and "Risk" in frs_df.columns and not oq_df.empty:
        req_link_col = "Requirement_Link" if "Requirement_Link" in oq_df.columns else None
        if req_link_col:
            for _, row in frs_df.iterrows():
                fid       = str(row.get("ID", "")).strip()
                risk      = str(row.get("Risk", "")).strip().lower()
                min_tests = {"high": 3, "medium": 2, "low": 1}.get(risk, 1)
                test_cnt  = oq_df[oq_df[req_link_col].astype(str).str.strip() == fid].shape[0]
                if test_cnt < min_tests:
                    issues.append({
                        "Rule":            "R4",
                        "Req_ID":          fid,
                        "Gap_Type":        "No_Test_Coverage",
                        "Description":     (f"Risk={risk.title()}: {fid} has {test_cnt} test(s) "
                                            f"but requires ≥{min_tests} for this risk level."),
                        "Recommendation":  f"Add {min_tests - test_cnt} more OQ test case(s).",
                        "Severity":        "High" if risk == "high" else "Medium",
                    })

    # ── R5: Duplicate detection (Jaccard token overlap > 0.80) ───────────────
    if desc_col and len(frs_df) > 1:
        frs_list = frs_df[["ID", desc_col]].dropna().reset_index(drop=True)
        seen_pairs = set()
        for i in range(len(frs_list)):
            for j in range(i + 1, len(frs_list)):
                id_a   = str(frs_list.loc[i, "ID"]).strip()
                id_b   = str(frs_list.loc[j, "ID"]).strip()
                desc_a = str(frs_list.loc[i, desc_col])
                desc_b = str(frs_list.loc[j, desc_col])
                pair   = tuple(sorted([id_a, id_b]))
                if pair in seen_pairs:
                    continue
                overlap = _token_overlap(desc_a, desc_b)
                if overlap >= 0.80:
                    seen_pairs.add(pair)
                    issues.append({
                        "Rule":            "R5",
                        "Req_ID":          f"{id_a} / {id_b}",
                        "Gap_Type":        "Duplicate",
                        "Description":     f"{id_a} and {id_b} have {overlap:.0%} token overlap.",
                        "Recommendation":  "Review and consolidate or differentiate these requirements.",
                        "Severity":        "Medium",
                    })
                    gap_df = pd.concat([gap_df, pd.DataFrame([{
                        "Req_ID":          f"{id_a} / {id_b}",
                        "Gap_Type":        "Duplicate",
                        "Description":     f"{overlap:.0%} overlap between {id_a} and {id_b}",
                        "Recommendation":  "Consolidate or clearly differentiate these requirements.",
                        "Severity":        "Medium",
                    }])], ignore_index=True)

    # ── R6: Human-in-the-Loop safeguard rows ─────────────────────────────────
    # OQ rows with the HITL placeholder text require manual completion.
    # These are not errors but represent mandatory human actions before sign-off.
    if not oq_df.empty and "Test_Step" in oq_df.columns:
        for _, row in oq_df.iterrows():
            step    = str(row.get("Test_Step", ""))
            test_id = str(row.get("Test_ID", "")).strip()
            if "HUMAN-IN-THE-LOOP SAFEGUARD" in step or "MANUAL REVIEW REQUIRED" in step:
                req_link = str(row.get("Requirement_Link", "")).strip()
                issues.append({
                    "Rule":            "R6",
                    "Req_ID":          test_id,
                    "Gap_Type":        "Manual_Action_Required",
                    "Description":     (f"{test_id} (linked to {req_link}) was not generated "
                                        f"by the AI and requires manual test steps to be written "
                                        f"before this validation package can be signed off."),
                    "Recommendation":  (f"Open the OQ sheet, locate {test_id}, and write executable "
                                        f"test steps, expected results, and pass/fail criteria. "
                                        f"This row must be completed before IQ/OQ execution."),
                    "Severity":        "Critical",
                })

    det_issues_df = pd.DataFrame(issues) if issues else pd.DataFrame(
        columns=["Rule", "Req_ID", "Gap_Type", "Description", "Recommendation", "Severity"]
    )

    return gap_df, det_issues_df

def build_audit_log_sheet(user: str, file_name: str, model_name: str,
                          frs_df: pd.DataFrame, oq_df: pd.DataFrame,
                          gap_df: pd.DataFrame, det_df: pd.DataFrame,
                          version_frs: int, version_oq: int,
                          doc_ids: str = "",
                          sys_context_name: str = "") -> pd.DataFrame:
    now_str   = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    role      = get_user_role(user)
    gap_count = len(gap_df) if not gap_df.empty else 0
    det_count = len(det_df) if not det_df.empty else 0

    urs_entry = {
        "Event":            "DOCUMENT_UPLOADED",
        "User":             user,
        "Role":             role,
        "Timestamp":        now_str,
        "Object_Changed":   "URS/SOP",
        "Old_Value":        "",
        "New_Value":        file_name,
        "Reason":           "URS file submitted for analysis",
        "AI_Metadata":      "",
    }

    rows = [
        {
            "Event":            "SESSION_LOGIN",
            "User":             user,
            "Role":             role,
            "Timestamp":        now_str,
            "Object_Changed":   "SESSION",
            "Old_Value":        "",
            "New_Value":        "AUTHENTICATED",
            "Reason":           "User authenticated successfully",
            "AI_Metadata":      "",
        },
        urs_entry,
    ]

    # Chain-of-Custody: record User Guide if one was provided
    if sys_context_name:
        rows.append({
            "Event":            "SYSCONTEXT_UPLOADED",
            "User":             user,
            "Role":             role,
            "Timestamp":        now_str,
            "Object_Changed":   "USER_GUIDE",
            "Old_Value":        "",
            "New_Value":        sys_context_name,
            "Reason":           "System User Guide injected as Reference Material for FRS/OQ generation",
            "AI_Metadata":      f"guide_file={sys_context_name}",
        })

    rows += [
        {
            "Event":            "AI_ANALYSIS_INITIATED",
            "User":             user,
            "Role":             role,
            "Timestamp":        now_str,
            "Object_Changed":   "ANALYSIS_ENGINE",
            "Old_Value":        "",
            "New_Value":        f"Model: {model_name} | Prompt: {PROMPT_VERSION} | Temp: {TEMPERATURE}",
            "Reason":           "GAMP-5 segmented analysis started"
                                + (f" with User Guide: {sys_context_name}" if sys_context_name else ""),
            "AI_Metadata":      f"prompt_version={PROMPT_VERSION} | model={model_name} | "
                                f"temperature={TEMPERATURE} | doc_ids={doc_ids}"
                                + (f" | guide={sys_context_name}" if sys_context_name else ""),
        },
        {
            "Event":            "FRS_GENERATED",
            "User":             user,
            "Role":             role,
            "Timestamp":        now_str,
            "Object_Changed":   f"FRS v{version_frs}.0",
            "Old_Value":        f"v{version_frs - 1}.0" if version_frs > 1 else "N/A",
            "New_Value":        f"v{version_frs}.0 — {len(frs_df)} requirements",
            "Reason":           "Functional requirements derived from URS"
                                + (f" + {sys_context_name}" if sys_context_name else ""),
            "AI_Metadata":      f"model={model_name} | prompt={PROMPT_VERSION} | temp={TEMPERATURE}",
        },
        {
            "Event":            "OQ_GENERATED",
            "User":             user,
            "Role":             role,
            "Timestamp":        now_str,
            "Object_Changed":   f"OQ v{version_oq}.0",
            "Old_Value":        f"v{version_oq - 1}.0" if version_oq > 1 else "N/A",
            "New_Value":        f"v{version_oq}.0 — {len(oq_df)} test cases",
            "Reason":           "OQ test cases generated; High-risk reqs → ≥3 tests",
            "AI_Metadata":      f"model={model_name} | prompt={PROMPT_VERSION} | temp={TEMPERATURE}",
        },
        {
            "Event":            "GAP_ANALYSIS_COMPLETED",
            "User":             user,
            "Role":             role,
            "Timestamp":        now_str,
            "Object_Changed":   "TRACEABILITY_MATRIX",
            "Old_Value":        "",
            "New_Value":        f"AI gaps: {gap_count} | Deterministic issues: {det_count}",
            "Reason":           "RTM compiled; deterministic rules R1-R4 enforced",
            "AI_Metadata":      "",
        },
        {
            "Event":            "WORKBOOK_EXPORTED",
            "User":             user,
            "Role":             role,
            "Timestamp":        now_str,
            "Object_Changed":   "VALIDATION_PACKAGE",
            "Old_Value":        "",
            "New_Value":        f"Validation_Package_{datetime.date.today()}.xlsx",
            "Reason":           "Full package with dashboard downloaded by user",
            "AI_Metadata":      f"doc_ids_used={doc_ids}",
        },
    ]
    return pd.DataFrame(rows)


# =============================================================================
# 7b. VALIDATION DASHBOARD BUILDER
# =============================================================================

def build_dashboard_sheet(frs_df: pd.DataFrame, oq_df: pd.DataFrame,
                           gap_df: pd.DataFrame, det_df: pd.DataFrame,
                           trace_df: pd.DataFrame,
                           file_name: str, model_name: str) -> pd.DataFrame:
    """Build a KPI summary table for the Dashboard sheet."""
    total_reqs  = len(frs_df) if not frs_df.empty else 0
    total_tests = len(oq_df)  if not oq_df.empty  else 0

    # Coverage: from Python-built traceability — the only reliable source
    covered  = 0
    partial  = 0
    missing  = 0
    if not trace_df.empty and "Coverage_Status" in trace_df.columns:
        covered  = int((trace_df["Coverage_Status"] == "Covered").sum())
        partial  = int((trace_df["Coverage_Status"] == "Partial").sum())
        missing  = int((trace_df["Coverage_Status"].isin(
                        ["Not Covered", "Missing FRS"])).sum())

    # Coverage % = (Covered + Partial) / total  — reflects that tests exist even if incomplete
    has_tests = covered + partial
    coverage_pct = round((has_tests / total_reqs * 100), 1) if total_reqs > 0 else 0.0
    fully_covered_pct = round((covered / total_reqs * 100), 1) if total_reqs > 0 else 0.0

    # Gap counts
    ai_gaps  = len(gap_df) if not gap_df.empty else 0
    det_gaps = len(det_df) if not det_df.empty else 0

    # Risk breakdown
    high_risk = med_risk = low_risk = 0
    if not frs_df.empty and "Risk" in frs_df.columns:
        rc = frs_df["Risk"].str.strip().str.lower().value_counts()
        high_risk = int(rc.get("high",   0))
        med_risk  = int(rc.get("medium", 0))
        low_risk  = int(rc.get("low",    0))

    # Non-testable count from det_df
    non_testable = 0
    if not det_df.empty and "Rule" in det_df.columns:
        non_testable = int((det_df["Rule"] == "R3").sum())

    orphan_tests = 0
    if not det_df.empty and "Rule" in det_df.columns:
        orphan_tests = int((det_df["Rule"] == "R2").sum())

    missing_frs = 0
    if not det_df.empty and "Rule" in det_df.columns:
        missing_frs = int((det_df["Rule"] == "R0").sum())

    rows = [
        {"KPI": "📋 Total FRS Requirements",          "Value": total_reqs,            "Status": "N/A"},
        {"KPI": "🧪 Total OQ Test Cases",              "Value": total_tests,           "Status": "N/A"},
        {"KPI": "✅ Fully Covered (all tests met)",    "Value": covered,               "Status": "N/A"},
        {"KPI": "🔶 Partially Covered (some tests)",   "Value": partial,               "Status": "See Traceability sheet"},
        {"KPI": "❌ Not Covered / Missing FRS",         "Value": missing,               "Status": "Immediate action required"},
        {"KPI": "📊 Basic Traceability % (Any Test Exists)",   "Value": f"{coverage_pct}%",
         "Status": "✅ PASS" if coverage_pct >= 80 else ("⚠️ REVIEW" if coverage_pct >= 60 else "❌ FAIL")},
        {"KPI": "🎯 Risk-Adjusted Compliance % (Fully Covered)", "Value": f"{fully_covered_pct}%",
         "Status": "✅ PASS" if fully_covered_pct >= 80 else ("⚠️ REVIEW" if fully_covered_pct >= 60 else "❌ FAIL")},
        {"KPI": "🔴 High Risk Requirements",            "Value": high_risk,             "Status": "Requires ≥3 OQ tests each"},
        {"KPI": "🟡 Medium Risk Requirements",          "Value": med_risk,              "Status": "Requires ≥2 OQ tests each"},
        {"KPI": "🟢 Low Risk Requirements",             "Value": low_risk,              "Status": "Requires ≥1 OQ test each"},
        {"KPI": "🚨 Missing FRS (AI skipped URS)",     "Value": missing_frs,           "Status": "Critical — re-run or add manually"},
        {"KPI": "⚠️ AI-Detected Gaps",                 "Value": ai_gaps,
         "Status": "See Gap_Analysis sheet" if ai_gaps > 0 else "✅ None — see Traceability sheet"},
        {"KPI": "🔍 Deterministic Issues (R0–R5)",     "Value": det_gaps,              "Status": "See Det_Validation sheet"},
        {"KPI": "🚫 Non-Testable Requirements",         "Value": non_testable,          "Status": "Rewrite required"},
        {"KPI": "👻 Orphan Tests (No FRS link)",        "Value": orphan_tests,          "Status": "Investigate"},
        {"KPI": "📁 Source Document",                   "Value": file_name,             "Status": "N/A"},
        {"KPI": "🤖 AI Model Used",                     "Value": model_name,            "Status": "N/A"},
        {"KPI": "🏷️ Prompt Version",                   "Value": PROMPT_VERSION,        "Status": "N/A"},
        {"KPI": "🌡️ Temperature",                      "Value": TEMPERATURE,           "Status": "N/A"},
        {"KPI": "📅 Generated",
         "Value": datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),           "Status": "N/A"},
    ]
    return pd.DataFrame(rows)


def _write_dashboard_chart(wb, ws_dash):
    """Write bar chart to dashboard sheet using openpyxl BarChart."""
    try:
        from openpyxl.chart import BarChart, Reference
        # Chart data: rows 1-4 are the core coverage KPIs (row index = Excel row)
        # KPI col=A(1), Value col=B(2)
        chart        = BarChart()
        chart.type   = "col"
        chart.title  = "Validation Coverage Overview"
        chart.y_axis.title = "Count / %"
        chart.x_axis.title = "Metric"
        chart.style  = 10
        chart.width  = 22
        chart.height = 14

        # Use rows 2-11 (KPI rows 1-10 in the dataframe, Excel rows 2-11)
        data   = Reference(ws_dash, min_col=2, min_row=1, max_row=11)
        cats   = Reference(ws_dash, min_col=1, min_row=2, max_row=11)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.solidFill = "2563EB"
        ws_dash.add_chart(chart, "E2")
    except Exception:
        pass  # Chart is a bonus — never crash on it




SHEET_COLORS = {
    "Summary":          {"header_fill": "0F172A", "tab_color": "1E293B"},
    "Signature":        {"header_fill": "1E3A5F", "tab_color": "1E3A5F"},
    "Dashboard":        {"header_fill": "0F172A", "tab_color": "0F172A"},
    "URS_Extraction":   {"header_fill": "1D4ED8", "tab_color": "1D4ED8"},
    "FRS":              {"header_fill": "2563EB", "tab_color": "2563EB"},
    "OQ":               {"header_fill": "059669", "tab_color": "059669"},
    "Traceability":     {"header_fill": "7C3AED", "tab_color": "7C3AED"},
    "Gap_Analysis":     {"header_fill": "DC2626", "tab_color": "DC2626"},
    "Det_Validation":   {"header_fill": "EA580C", "tab_color": "EA580C"},
    "Audit_Log":        {"header_fill": "B45309", "tab_color": "B45309"},
}


def build_pdf_bytes(r: dict, sig_id: int, sig_meaning: str,
                    sig_timestamp: str, user: str, role: str) -> bytes:
    """
    Generate a signed PDF summary of the validation package.

    Includes: KPI summary table, electronic signature block, SHA-256 hash
    with scope statement, and regulatory disclaimer.

    The document_hash embedded here is identical to the one in the Excel
    Signature sheet — both are produced by the same signing event (sig_id).
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=18*mm, bottomMargin=18*mm)
    styles = getSampleStyleSheet()

    C_NAVY   = colors.HexColor("#0F172A")
    C_DKBLUE = colors.HexColor("#1E3A5F")
    C_TEAL   = colors.HexColor("#2563EB")
    C_GREEN  = colors.HexColor("#059669")
    C_RED    = colors.HexColor("#DC2626")
    C_AMBER  = colors.HexColor("#D97706")
    C_LGREY  = colors.HexColor("#F8FAFC")
    C_MGREY  = colors.HexColor("#94A3B8")
    C_DGREY  = colors.HexColor("#374151")
    C_GRID   = colors.HexColor("#CBD5E1")

    def _style(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    h1   = _style("h1",   fontSize=16, textColor=colors.white, backColor=C_NAVY,
                           alignment=TA_CENTER, fontName="Helvetica-Bold",
                           spaceAfter=0, topPadding=10, bottomPadding=10)
    h2   = _style("h2",   fontSize=9,  textColor=colors.white, backColor=C_TEAL,
                           alignment=TA_CENTER, fontName="Helvetica-BoldOblique",
                           spaceAfter=0, topPadding=6, bottomPadding=6)
    sec  = _style("sec",  fontSize=9,  textColor=colors.white, backColor=C_DKBLUE,
                           alignment=TA_CENTER, fontName="Helvetica-Bold",
                           spaceAfter=0, spaceBefore=6, topPadding=5, bottomPadding=5)
    body = _style("body", fontSize=8.5, textColor=C_DGREY, spaceAfter=2)
    sm   = _style("sm",   fontSize=7,  textColor=C_MGREY,
                           fontName="Helvetica-Oblique", spaceAfter=3)

    cov_pct   = r.get("cov_pct", 0)
    gap_total = r.get("gap_count", 0) + r.get("det_count", 0)
    doc_hash  = r.get("doc_hash", "N/A")
    file_name = r.get("file_name", "")

    story = []

    # ── Title ─────────────────────────────────────────────────────────────────
    story.append(Paragraph("VALIDATION PACKAGE — SIGNED SUMMARY", h1))
    story.append(Paragraph(
        f"21 CFR Part 11 Compliant &nbsp;|&nbsp; App v{VERSION} &nbsp;|&nbsp; "
        f"Generated {datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC", h2))
    story.append(Spacer(1, 5*mm))

    # ── KPI Summary ───────────────────────────────────────────────────────────
    story.append(Paragraph("VALIDATION PACKAGE SUMMARY", sec))
    story.append(Spacer(1, 1*mm))
    kpi_rows = [
        ["Metric", "Value", "Status"],
        ["Requirements extracted (URS)", str(r.get("total_urs", 0)), ""],
        ["FRS requirements generated",   str(r.get("total_reqs", 0)), ""],
        ["OQ test cases generated",      str(r.get("total_tests", 0)), ""],
        ["Fully covered requirements",   str(r.get("covered", 0)), ""],
        ["Traceability coverage",        f"{cov_pct}%",
         "PASS" if cov_pct >= 80 else ("REVIEW" if cov_pct >= 60 else "FAIL")],
        ["Gaps detected (AI + Det.)",    str(gap_total),
         "PASS" if gap_total == 0 else "REVIEW"],
    ]
    kpi_tbl = Table(kpi_rows, colWidths=[90*mm, 30*mm, 30*mm])
    kpi_sty = [
        ("BACKGROUND",    (0, 0), (-1, 0), C_TEAL),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1,-1), 8.5),
        ("ROWBACKGROUNDS",(0, 1), (-1,-1), [C_LGREY, colors.white]),
        ("GRID",          (0, 0), (-1,-1), 0.3, C_GRID),
        ("TOPPADDING",    (0, 0), (-1,-1), 4),
        ("BOTTOMPADDING", (0, 0), (-1,-1), 4),
        ("LEFTPADDING",   (0, 0), (-1,-1), 6),
    ]
    for ri, row in enumerate(kpi_rows[1:], start=1):
        s = row[2]
        c = C_GREEN if s == "PASS" else (C_AMBER if s == "REVIEW" else (C_RED if s == "FAIL" else C_DGREY))
        kpi_sty += [("TEXTCOLOR", (2, ri), (2, ri), c),
                    ("FONTNAME",  (2, ri), (2, ri), "Helvetica-Bold")]
    kpi_tbl.setStyle(TableStyle(kpi_sty))
    story.append(kpi_tbl)
    story.append(Spacer(1, 5*mm))

    # ── Electronic Signature ──────────────────────────────────────────────────
    story.append(Paragraph("ELECTRONIC SIGNATURE RECORD", sec))
    story.append(Spacer(1, 1*mm))
    story.append(Paragraph(
        "21 CFR Part 11 §11.50 / §11.200 — Non-Biometric Electronic Signature", sm))
    sig_rows = [
        ["Field",             "Value",                      "Regulatory Reference"],
        ["Signer Name",       user,                          "§11.50(a)(1) — printed name"],
        ["Role",              role,                          "Role at time of signing"],
        ["Signature ID",      f"SIG-{sig_id:06d}",          "Unique signature identifier"],
        ["Date / Time (UTC)", sig_timestamp[:19],            "§11.50(a)(1) — date and time"],
        ["Action Signed",     "Generated Validation Package","Specific act being signed"],
        ["Meaning",           sig_meaning,                   "§11.50(a)(1) — meaning"],
        ["System Version",    VERSION,                       "App version at signing"],
        ["Prompt Version",    PROMPT_VERSION,                "AI prompt version"],
        ["Source Document",   file_name,                     "URS file analysed"],
    ]
    sig_tbl = Table(sig_rows, colWidths=[40*mm, 68*mm, 42*mm])
    sig_sty = [
        ("BACKGROUND",    (0, 0), (-1, 0), C_DKBLUE),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1,-1), 8),
        ("FONTNAME",      (0, 1), (0, -1), "Helvetica-Bold"),
        ("ROWBACKGROUNDS",(0, 1), (-1,-1), [C_LGREY, colors.white]),
        ("GRID",          (0, 0), (-1,-1), 0.3, C_GRID),
        ("TOPPADDING",    (0, 0), (-1,-1), 4),
        ("BOTTOMPADDING", (0, 0), (-1,-1), 4),
        ("LEFTPADDING",   (0, 0), (-1,-1), 6),
        ("TEXTCOLOR",     (2, 1), (2, -1), C_MGREY),
        ("FONTNAME",      (2, 1), (2, -1), "Helvetica-Oblique"),
    ]
    sig_tbl.setStyle(TableStyle(sig_sty))
    story.append(sig_tbl)
    story.append(Spacer(1, 5*mm))

    # ── Document Hash ─────────────────────────────────────────────────────────
    story.append(Paragraph("DOCUMENT INTEGRITY — SHA-256 HASH", sec))
    story.append(Spacer(1, 1*mm))
    hash_rows = [
        ["Hash Algorithm", "SHA-256 (hashlib.sha256)"],
        ["Document Hash",  doc_hash],
        ["Hash Scope",
         "Covers all Excel workbook sheets EXCEPT the Signature sheet "
         "(standard practice — same as PDF digital signatures). "
         "To verify: re-export the workbook without the Signature sheet "
         "and recompute SHA-256."],
    ]
    hash_tbl = Table(hash_rows, colWidths=[35*mm, 115*mm])
    hash_sty = [
        ("FONTNAME",      (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1,-1), 8),
        ("ROWBACKGROUNDS",(0, 0), (-1,-1), [C_LGREY, colors.HexColor("#D1FAE5"), C_LGREY]),
        ("GRID",          (0, 0), (-1,-1), 0.3, C_GRID),
        ("TOPPADDING",    (0, 0), (-1,-1), 4),
        ("BOTTOMPADDING", (0, 0), (-1,-1), 4),
        ("LEFTPADDING",   (0, 0), (-1,-1), 6),
        ("FONTNAME",      (1, 1), (1,  1), "Courier"),
        ("FONTSIZE",      (1, 1), (1,  1), 7),
        ("TEXTCOLOR",     (1, 1), (1,  1), C_DKBLUE),
        ("VALIGN",        (0, 0), (-1,-1), "TOP"),
        ("FONTSIZE",      (1, 2), (1,  2), 7.5),
    ]
    hash_tbl.setStyle(TableStyle(hash_sty))
    story.append(hash_tbl)
    story.append(Spacer(1, 4*mm))

    # ── Regulatory disclaimer ─────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_GRID))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        "<b>REGULATORY STATEMENT:</b> This electronic signature was applied in accordance with "
        "21 CFR Part 11. The signer's identity was verified at the time of signing by re-entry "
        "of their system password (two-component non-biometric e-signature per §11.200(b)(1)). "
        "This record is stored in an append-only signature log and cannot be altered or deleted.",
        _style("disc", fontSize=7, textColor=C_DGREY, fontName="Helvetica-Oblique")))

    doc.build(story)
    return buf.getvalue()


def build_signature_sheet(wb,
                          user: str,
                          role: str,
                          meaning: str,
                          document_hash: str,
                          document_name: str,
                          model_used: str,
                          signature_id: int,
                          timestamp: str):
    """
    Create the Signature sheet as the LAST tab in the workbook.

    Satisfies 21 CFR Part 11 §11.50:
      - Printed name of signer, date/time, meaning of signature
      - SHA-256 hash links signature to the exact workbook bytes
      - Explicit hash scope statement (excludes this sheet — standard practice)
      - Regulatory statement citing §11.200(b)(1)
    """
    ws = wb.create_sheet("Signature")
    ws.sheet_properties.tabColor = "1E3A5F"

    navy    = PatternFill("solid", fgColor="0F172A")
    dkblue  = PatternFill("solid", fgColor="1E3A5F")
    white   = PatternFill("solid", fgColor="FFFFFF")
    lgrey   = PatternFill("solid", fgColor="F8FAFC")
    green   = PatternFill("solid", fgColor="D1FAE5")
    thin    = Side(style="thin",   color="CBD5E1")
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _cell(row, col, value, bold=False, size=11, color="000000",
              fill=None, align="left", wrap=False, italic=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(bold=bold, size=size, color=color, italic=italic)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        c.fill      = fill or white
        c.border    = border
        return c

    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 44
    _cell(1, 1, "ELECTRONIC SIGNATURE RECORD",
          bold=True, size=16, color="FFFFFF", fill=navy, align="center")

    ws.merge_cells("A2:C2")
    ws.row_dimensions[2].height = 20
    _cell(2, 1,
          "21 CFR Part 11 §11.50 / §11.200 — Non-Biometric Electronic Signature",
          bold=False, size=9, color="FFFFFF", fill=dkblue, align="center", italic=True)

    ws.row_dimensions[3].height = 10

    fields = [
        ("Signer Name",        user,                      "§11.50(a)(1) — printed name"),
        ("Role",               role,                      "User role at time of signing"),
        ("Signature ID",       f"SIG-{signature_id:06d}", "Unique signature record identifier"),
        ("Date / Time (UTC)",  timestamp,                 "§11.50(a)(1) — date and time"),
        ("Action Signed",      "Generated Validation Package",
                                                           "The specific act being signed"),
        ("Meaning",            meaning,                   "§11.50(a)(1) — meaning of signature"),
        ("System Version",     VERSION,                   "Application version at signing"),
        ("AI Model Used",      model_used,                "Model that generated the package"),
        ("Prompt Version",     PROMPT_VERSION,            "AI prompt version used"),
        ("Source Document",    document_name,             "URS/SOP file analysed"),
    ]

    row = 4
    for label, value, note in fields:
        ws.row_dimensions[row].height = 24
        _cell(row, 1, label, bold=True,  size=10, fill=lgrey, align="left")
        _cell(row, 2, value, bold=False, size=10, fill=white, align="left")
        _cell(row, 3, note,  bold=False, size=8,
              color="64748B", fill=lgrey, align="left", italic=True)
        row += 1

    ws.row_dimensions[row].height = 10
    row += 1

    ws.merge_cells(f"A{row}:C{row}")
    ws.row_dimensions[row].height = 22
    _cell(row, 1, "DOCUMENT INTEGRITY — SHA-256 HASH",
          bold=True, size=11, color="FFFFFF", fill=dkblue, align="center")
    row += 1

    ws.row_dimensions[row].height = 22
    _cell(row, 1, "Document Hash (SHA-256)", bold=True, size=10, fill=lgrey)
    ws.merge_cells(f"B{row}:C{row}")
    c = ws.cell(row=row, column=2, value=document_hash)
    c.font      = Font(name="Courier New", size=9, bold=True, color="1E3A5F")
    c.fill      = green
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    c.border    = border
    row += 1

    ws.row_dimensions[row].height = 22
    _cell(row, 1, "Hash Algorithm", bold=True, size=10, fill=lgrey)
    ws.merge_cells(f"B{row}:C{row}")
    _cell(row, 2, "SHA-256 (hashlib.sha256)", bold=False, size=10, fill=white)
    row += 1

    ws.row_dimensions[row].height = 36
    _cell(row, 1, "Hash Scope", bold=True, size=10, fill=lgrey)
    ws.merge_cells(f"B{row}:C{row}")
    _cell(row, 2,
          "Hash covers all workbook sheets EXCEPT this Signature sheet. "
          "This is standard practice (identical to PDF digital signatures). "
          "To verify: re-export the workbook without this sheet and recompute SHA-256.",
          bold=False, size=8, color="374151", fill=lgrey, wrap=True)
    row += 1

    ws.row_dimensions[row].height = 10
    row += 1

    ws.merge_cells(f"A{row}:C{row}")
    ws.row_dimensions[row].height = 48
    _cell(row, 1,
          "REGULATORY STATEMENT: This electronic signature was applied in accordance with "
          "21 CFR Part 11. The signer's identity was verified at the time of signing by "
          "re-entry of their system password (two-component non-biometric e-signature per "
          "§11.200(b)(1)). This record is stored in an append-only signature log and "
          "cannot be altered or deleted.",
          bold=False, size=8, color="374151", fill=lgrey, align="left", wrap=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 38


def build_cover_sheet(wb, user: str, file_name: str, model_name: str,
                      dashboard_df: pd.DataFrame,
                      sys_context_name: str = ""):
    """
    Create an executive-ready Summary tab as the first sheet.
    Displays KPIs in large font, includes Digital Signature lines.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Summary", 0)   # insert at position 0 (first tab)
    ws.sheet_properties.tabColor = "1E293B"

    # ── Freeze first row ──
    ws.freeze_panes = "A2"

    navy   = PatternFill("solid", fgColor="0F172A")
    teal   = PatternFill("solid", fgColor="0E7490")
    white  = PatternFill("solid", fgColor="FFFFFF")
    lgrey  = PatternFill("solid", fgColor="F8FAFC")
    green  = PatternFill("solid", fgColor="D1FAE5")
    yellow = PatternFill("solid", fgColor="FEF9C3")
    red    = PatternFill("solid", fgColor="FEE2E2")
    thin   = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _set(row, col, value, bold=False, size=11, color="000000",
             fill=None, align="left", wrap=False):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font      = Font(bold=bold, size=size, color=color)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if fill:
            cell.fill = fill
        cell.border = border
        return cell

    # ── Row 1: Title banner ──
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 40
    _set(1, 1, "GxP Validation Package — Executive Summary",
         bold=True, size=18, color="FFFFFF", fill=navy, align="center")

    # ── Row 2: Sub-header ──
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 22
    gen_time = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    _set(2, 1, f"Generated: {gen_time}  |  Source: {file_name}  |  Model: {model_name}",
         bold=False, size=10, color="FFFFFF", fill=teal, align="center")

    if sys_context_name:
        ws.merge_cells("A3:F3")
        ws.row_dimensions[3].height = 18
        _set(3, 1, f"User Guide Reference: {sys_context_name}",
             bold=False, size=9, color="334155", fill=lgrey, align="center")
        next_row = 4
    else:
        next_row = 3

    # ── Blank spacer ──
    ws.row_dimensions[next_row].height = 10
    next_row += 1

    # ── KPI section header ──
    ws.merge_cells(f"A{next_row}:F{next_row}")
    ws.row_dimensions[next_row].height = 24
    _set(next_row, 1, "KEY PERFORMANCE INDICATORS",
         bold=True, size=13, color="FFFFFF", fill=navy, align="center")
    next_row += 1

    # ── KPI rows from dashboard_df ──
    kpi_fill_map = {}   # detect pass/fail for colouring
    if not dashboard_df.empty and "KPI" in dashboard_df.columns:
        ws.row_dimensions[next_row].height = 20
        _set(next_row, 1, "Metric",   bold=True, size=10, color="FFFFFF", fill=teal, align="center")
        _set(next_row, 2, "Value",    bold=True, size=10, color="FFFFFF", fill=teal, align="center")
        _set(next_row, 4, "Status",   bold=True, size=10, color="FFFFFF", fill=teal, align="center")
        # Merge B + C for value, E+F for status
        ws.merge_cells(f"B{next_row}:C{next_row}")
        ws.merge_cells(f"D{next_row}:F{next_row}")
        ws.cell(row=next_row, column=4).value = "Status"
        ws.cell(row=next_row, column=4).font  = Font(bold=True, size=10, color="FFFFFF")
        ws.cell(row=next_row, column=4).fill  = teal
        ws.cell(row=next_row, column=4).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=next_row, column=4).border = border
        next_row += 1

        for _, kpi_row in dashboard_df.iterrows():
            kpi_name   = str(kpi_row.get("KPI",    ""))
            kpi_val    = str(kpi_row.get("Value",  ""))
            kpi_status = str(kpi_row.get("Status", ""))
            row_fill   = green if "PASS" in kpi_status else (
                         red   if "FAIL" in kpi_status else (
                         yellow if "REVIEW" in kpi_status else lgrey))
            ws.row_dimensions[next_row].height = 22
            ws.merge_cells(f"B{next_row}:C{next_row}")
            ws.merge_cells(f"D{next_row}:F{next_row}")
            _set(next_row, 1, kpi_name, bold=False, size=10, fill=row_fill, align="left",  wrap=True)
            _set(next_row, 2, kpi_val,  bold=True,  size=10, fill=row_fill, align="center")
            ws.cell(row=next_row, column=4).value     = kpi_status
            ws.cell(row=next_row, column=4).font      = Font(bold=False, size=10)
            ws.cell(row=next_row, column=4).fill      = row_fill
            ws.cell(row=next_row, column=4).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=next_row, column=4).border    = border
            next_row += 1

    # ── Blank spacer ──
    ws.row_dimensions[next_row].height = 16
    next_row += 1

    # ── Digital Signatures section ──
    ws.merge_cells(f"A{next_row}:F{next_row}")
    ws.row_dimensions[next_row].height = 24
    _set(next_row, 1, "ELECTRONIC SIGNATURE — 21 CFR PART 11",
         bold=True, size=13, color="FFFFFF", fill=navy, align="center")
    next_row += 1

    sig_fields = [
        ("Prepared By (Validation Engineer)",  user,        "Signature / Initials"),
        ("Prepared Date",                       gen_time,    "Date (UTC)"),
        ("AI Model Used",                       model_name,  "System"),
        ("Prompt Version",                      PROMPT_VERSION, "Audit Reference"),
        ("Reviewed By (QA Manager)",            "________________________", "Signature"),
        ("Review Date",                         "________________________", "Date"),
        ("Approved By (QA Director / Sponsor)", "________________________", "Signature"),
        ("Approval Date",                       "________________________", "Date"),
    ]
    for sig_label, sig_val, sig_type in sig_fields:
        ws.row_dimensions[next_row].height = 24
        ws.merge_cells(f"C{next_row}:D{next_row}")
        ws.merge_cells(f"E{next_row}:F{next_row}")
        _set(next_row, 1, sig_label,  bold=True,  size=10, fill=lgrey, align="left")
        ws.merge_cells(f"B{next_row}:B{next_row}")
        _set(next_row, 2, sig_type,   bold=False, size=9,  fill=lgrey, color="64748B", align="center")
        ws.cell(row=next_row, column=3).value     = sig_val
        ws.cell(row=next_row, column=3).font      = Font(bold=True, size=10, color="1E40AF")
        ws.cell(row=next_row, column=3).fill      = white
        ws.cell(row=next_row, column=3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=next_row, column=3).border    = border
        next_row += 1

    # ── Disclaimer footer ──
    ws.row_dimensions[next_row].height = 10
    next_row += 1
    ws.merge_cells(f"A{next_row}:F{next_row}")
    ws.row_dimensions[next_row].height = 30
    _set(next_row, 1,
         "DISCLAIMER: This document was AI-generated as a draft validation artefact. "
         "All content must be reviewed and approved by a qualified GxP professional "
         "before use in a regulated submission. AI outputs do not constitute a validated "
         "system qualification without human review per GAMP 5 and ICH Q10.",
         bold=False, size=8, color="64748B", fill=lgrey, align="center", wrap=True)

    # ── Column widths ──
    for col_letter, width in [("A", 38), ("B", 18), ("C", 28), ("D", 18), ("E", 22), ("F", 18)]:
        ws.column_dimensions[col_letter].width = width




def style_worksheet(ws, sheet_name: str):
    colors       = SHEET_COLORS.get(sheet_name, {"header_fill": "334155", "tab_color": "334155"})
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor=colors["header_fill"])
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_fill     = PatternFill("solid", fgColor="F1F5F9")
    thin         = Side(style="thin", color="CBD5E1")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    max_col      = ws.max_column
    max_row      = ws.max_row

    for col in range(1, max_col + 1):
        cell           = ws.cell(row=1, column=col)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border

    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell           = ws.cell(row=row, column=col)
            cell.border    = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row % 2 == 0:
                cell.fill = alt_fill

    ws.auto_filter.ref           = ws.dimensions
    ws.freeze_panes              = "A2"
    ws.row_dimensions[1].height  = 30
    ws.sheet_properties.tabColor = colors["tab_color"]

    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        max_len    = 12
        for row in range(1, max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                cell_len = max(
                    len(str(val).split("\n")[0]),
                    min(len(str(val)) // 2, 40)
                )
                max_len = max(max_len, cell_len)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 80)

    # ── Conditional row colouring ─────────────────────────────────────────────
    gap_fill    = PatternFill("solid", fgColor="FEE2E2")   # light red — gaps
    hitl_fill   = PatternFill("solid", fgColor="FEF9C3")   # light yellow — HITL
    xsrc_fill   = PatternFill("solid", fgColor="EDE9FE")   # light purple — cross-source
    pass_fill   = PatternFill("solid", fgColor="D1FAE5")   # light green — covered/pass
    warn_fill   = PatternFill("solid", fgColor="FFF7ED")   # light orange — partial/review

    # Identify key column indices by header name for this sheet
    header_vals = {ws.cell(row=1, column=c).value: c for c in range(1, max_col + 1)}

    def _colour_row(row_idx, fill_obj):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = fill_obj

    for row_idx in range(2, max_row + 1):
        if sheet_name == "Traceability":
            cov_col = header_vals.get("Coverage_Status")
            if cov_col:
                val = str(ws.cell(row=row_idx, column=cov_col).value or "").strip()
                if val in ("Not Covered", "Missing FRS", "[GAP]"):
                    _colour_row(row_idx, gap_fill)
                elif val == "Partial":
                    _colour_row(row_idx, warn_fill)
                elif val == "Covered":
                    _colour_row(row_idx, pass_fill)

        elif sheet_name == "Gap_Analysis":
            sev_col = header_vals.get("Severity")
            if sev_col:
                sev = str(ws.cell(row=row_idx, column=sev_col).value or "").strip().lower()
                if sev == "critical":
                    _colour_row(row_idx, gap_fill)
                elif sev == "high":
                    _colour_row(row_idx, warn_fill)

        elif sheet_name == "Det_Validation":
            rule_col = header_vals.get("Rule")
            sev_col  = header_vals.get("Severity")
            if rule_col:
                rule = str(ws.cell(row=row_idx, column=rule_col).value or "").strip()
                if rule == "R6":
                    _colour_row(row_idx, hitl_fill)
                elif rule in ("R0", "R1"):
                    _colour_row(row_idx, gap_fill)

        elif sheet_name == "FRS":
            src_col = header_vals.get("Source_URS_Ref")
            cf_col  = header_vals.get("Confidence_Flag")
            if src_col:
                src = str(ws.cell(row=row_idx, column=src_col).value or "")
                if "User Guide Only" in src:
                    _colour_row(row_idx, xsrc_fill)
                elif "Cross-Source Gap" in str(ws.cell(row=row_idx, column=cf_col).value if cf_col else ""):
                    _colour_row(row_idx, xsrc_fill)

        elif sheet_name == "OQ":
            step_col = header_vals.get("Test_Step")
            if step_col:
                step = str(ws.cell(row=row_idx, column=step_col).value or "")
                if "HUMAN-IN-THE-LOOP" in step:
                    _colour_row(row_idx, hitl_fill)


def build_styled_excel(dataframes: dict, user: str = "", file_name: str = "",
                       model_name: str = "", sys_context_name: str = "",
                       dashboard_df=None) -> bytes:
    # ── Inject "DRAFT – AI Generated | Pending Human Review" status column ────
    # Satisfies document control lifecycle: AI output is never presented as Final.
    # Every data sheet gets AI_Review_Status = "DRAFT – AI Generated | Pending Review"
    # so reviewers and auditors can immediately see the document state.
    DRAFT_SHEETS = {"URS_Extraction", "FRS", "OQ", "Traceability", "Gap_Analysis", "Det_Validation"}
    stamped = {}
    for sheet_name, df in dataframes.items():
        if sheet_name in DRAFT_SHEETS and not df.empty:
            df = df.copy()
            df.insert(0, "AI_Review_Status", "DRAFT – AI Generated | Pending Review")
        stamped[sheet_name] = df

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in stamped.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        wb = writer.book
        for sheet_name in stamped:
            if sheet_name in wb.sheetnames:
                style_worksheet(wb[sheet_name], sheet_name)
                # ── Colour the AI_Review_Status column amber to draw the eye ──
                if sheet_name in DRAFT_SHEETS:
                    ws = wb[sheet_name]
                    hdr_vals = {ws.cell(row=1, column=c).value: c
                                for c in range(1, ws.max_column + 1)}
                    status_col = hdr_vals.get("AI_Review_Status")
                    if status_col:
                        amber_fill = PatternFill("solid", fgColor="FEF3C7")
                        amber_font = Font(bold=True, color="92400E", size=9)
                        for row_i in range(1, ws.max_row + 1):
                            cell = ws.cell(row=row_i, column=status_col)
                            cell.fill = amber_fill
                            if row_i > 1:
                                cell.font = amber_font
                        ws.column_dimensions[
                            get_column_letter(status_col)
                        ].width = 36
        # Add bar chart to Dashboard sheet
        if "Dashboard" in wb.sheetnames:
            _write_dashboard_chart(wb, wb["Dashboard"])
        # Add executive Summary cover sheet (inserted at position 0)
        _dash = dashboard_df if dashboard_df is not None else (
            dataframes.get("Dashboard", pd.DataFrame()))
        build_cover_sheet(wb, user=user, file_name=file_name,
                          model_name=model_name, dashboard_df=_dash,
                          sys_context_name=sys_context_name)
    return output.getvalue()


# =============================================================================
# 9. SESSION STATE
# =============================================================================

def get_auto_location():
    """
    IP-based geolocation is intentionally disabled.

    When Streamlit runs on any cloud host (Streamlit Cloud, AWS, GCP, Azure, etc.)
    ip-api.com resolves the *server's* datacenter IP — not the user's browser IP.
    This reliably returns the wrong region (e.g. Oregon when the user is in California).

    For GxP 21 CFR Part 11 compliance, a validator should DECLARE their location
    explicitly rather than have it guessed. Manual entry in the sidebar replaces
    auto-detection entirely.
    """
    return ""


_defaults = {
    "authenticated":      False,
    "selected_model":     "Gemini 1.5 Pro",
    "user_name":          "",
    "user_role":          "",
    "last_activity":      None,
    "sop_file_bytes":     None,
    "sop_file_name":      None,
    "sys_context_bytes":  None,
    "sys_context_name":   None,
    "uploader_key_n":     0,       # incremented to reset the URS file-uploader widget
    "sys_uploader_key_n": 0,       # incremented to reset the sidebar sys-context uploader
    "user_ip":            "",      # client IP stored as separate audit column (v29)
    "esig_pending":       None,    # holds completed analysis awaiting e-signature
    "pass2_chunk_size":   8,       # Phase 1: reduced from 40→8 to prevent 600s timeout
    "show_esig_form":     False,   # True when user clicked a download button → show inline form
    "esig_target":        None,    # "xlsx" or "pdf" — which format triggered the e-sig form
    "app_mode":           "New Validation",   # sidebar mode selector
    # ── Change Impact Analysis slots ────────────────────────────────────────
    "cia_change_spec_bytes": None,
    "cia_change_spec_name":  None,
    "cia_frs_bytes":         None,
    "cia_frs_name":          None,
    "cia_oq_bytes":          None,
    "cia_oq_name":           None,
    "cia_trace_bytes":       None,
    "cia_trace_name":        None,
    "cia_result":            None,   # completed CIA output dict
    "cia_key_n":             0,      # increment to reset all CIA uploaders
    # ── Audit Trail Intelligence (Periodic Review Module 1) ──────────────────
    "at_raw_df":            None,
    "at_mapped_df":         None,
    "at_scored_df":         None,
    "at_top20_df":          None,
    "at_column_map":        {},
    "at_file_name":         "",
    "at_mapping_done":      False,
    "at_analysis_done":     False,
    "at_total_events":      0,
    "at_system_name":       "",
    "at_review_start":      "",
    "at_review_end":        "",
    "at_key_n":             0,      # increment to reset uploaders
    "pr_active_module":     None,   # which PR sub-module is open (None = landing)
    "at_detection_logic_text": "",  # cached detection logic for Excel sheet 4
    # ── Audit Trail risk tier thresholds (GAMP 5 calibrated defaults) ────────
    "at_thresh_critical":   7.0,    # score >= this = Critical
    "at_thresh_high":       5.0,    # score >= this = High
    "at_thresh_medium":     3.0,    # score >= this = Medium
}
for _k, _v in _defaults.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

db_migrate()

# =============================================================================
# 10. CSS  — All v15 branding preserved unchanged
# =============================================================================
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    html, body, [class*="st-"] { font-family: 'Inter', sans-serif; }
    .stApp { background-color: #fcfcfd; }

    /* ── Top Banner ── */
    .top-banner {
        background-color: white; border: 1px solid #eef2f6; border-radius: 10px;
        padding: 12px 0px; text-align: center; margin-bottom: 5px;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
    }
    .banner-text-inner {
        color: #475569; font-weight: 400; letter-spacing: 4px;
        text-transform: uppercase; font-size: 0.85rem; margin: 0;
    }

    /* ── Login inputs — same width as "Initialize Secure Session" button (40%) ── */
    [data-testid="stTextInput"] {
        width: 40% !important;
        min-width: 220px !important;
        margin: 0 auto !important;
    }

    /* ── Button container ── */
    div.stButton {
        width: 100% !important;
        display: flex !important;
        justify-content: center !important;
    }

    /* ── All buttons: base + universal hover ── */
    div.stButton > button {
        border: 1px solid #cbd5e1 !important;
        border-radius: 8px !important;
        font-weight: 500 !important;
        transition: background 0.18s ease, color 0.18s ease,
                    box-shadow 0.18s ease, transform 0.15s ease,
                    border-color 0.18s ease !important;
    }
    div.stButton > button:hover:not(:disabled) {
        background: #eff6ff !important;
        border-color: #3b82f6 !important;
        color: #1d4ed8 !important;
        box-shadow: 0 4px 14px rgba(59, 130, 246, 0.25) !important;
        transform: translateY(-1px) !important;
    }

    /* ── Login button ── */
    div.stButton > button[key="login_btn"] {
        width: 40% !important; margin: 0 auto !important; display: block !important;
        background: linear-gradient(135deg, #3b82f6, #2563eb) !important;
        color: white !important; height: 3.2rem !important;
        border: none !important; font-weight: 600 !important;
        box-shadow: 0 4px 12px rgba(37, 99, 235, 0.3) !important;
    }
    div.stButton > button[key="login_btn"]:hover:not(:disabled) {
        background: linear-gradient(135deg, #60a5fa, #3b82f6) !important;
        color: white !important; border-color: transparent !important;
        box-shadow: 0 6px 18px rgba(37, 99, 235, 0.45) !important;
    }

    /* ── Pull "Auto-Generate" section up by one line ── */
    section[data-testid="stMain"] h1:first-of-type {
        margin-top: -1rem !important;
        padding-top: 0 !important;
    }

    /* ── Main URS file uploader — polished card ── */
    section[data-testid="stMain"] div[data-testid="stFileUploader"] {
        max-width: calc(100% - 54px) !important;
    }
    /* Outer drop-zone card */
    section[data-testid="stMain"] div[data-testid="stFileUploaderDropzone"] {
        background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%) !important;
        border: 1.5px dashed #2563eb !important;
        border-radius: 14px !important;
        padding: 28px 32px !important;
        transition: border-color 0.2s, box-shadow 0.2s !important;
        box-shadow: 0 2px 16px rgba(37, 99, 235, 0.08) !important;
    }
    section[data-testid="stMain"] div[data-testid="stFileUploaderDropzone"]:hover {
        border-color: #3b82f6 !important;
        box-shadow: 0 4px 24px rgba(37, 99, 235, 0.18) !important;
    }
    /* "Drag and drop file here" primary text */
    section[data-testid="stMain"] div[data-testid="stFileUploaderDropzone"] span[data-testid="stFileUploaderDropzoneInstructions"] > div > span:first-child {
        color: #e2e8f0 !important;
        font-size: 0.95rem !important;
        font-weight: 600 !important;
        letter-spacing: 0.01em !important;
    }
    /* "Limit 200MB per file • PDF" sub-text */
    section[data-testid="stMain"] div[data-testid="stFileUploaderDropzone"] small,
    section[data-testid="stMain"] div[data-testid="stFileUploaderDropzone"] span[data-testid="stFileUploaderDropzoneInstructions"] > div > small {
        color: #64748b !important;
        font-size: 0.78rem !important;
    }
    /* Upload cloud icon */
    section[data-testid="stMain"] div[data-testid="stFileUploaderDropzone"] svg {
        fill: #2563eb !important;
        opacity: 0.85 !important;
    }
    /* "Browse files" button */
    section[data-testid="stMain"] div[data-testid="stFileUploaderDropzone"] button[data-testid="baseButton-secondary"] {
        background: #1e40af !important;
        color: #e2e8f0 !important;
        border: 1px solid #2563eb !important;
        border-radius: 8px !important;
        font-size: 0.82rem !important;
        font-weight: 600 !important;
        padding: 5px 18px !important;
        transition: background 0.15s !important;
    }
    section[data-testid="stMain"] div[data-testid="stFileUploaderDropzone"] button[data-testid="baseButton-secondary"]:hover {
        background: #2563eb !important;
        color: white !important;
    }

    /* ── Run Analysis — iOS-inspired ── */
    div.stButton > button[key="run_analysis_btn"] {
        background-color: #007AFF !important;
        color: white !important;
        padding: 0.75rem 2.5rem !important;
        font-size: 1.05rem !important;
        font-weight: 500 !important;
        border-radius: 12px !important;
        border: none !important;
        box-shadow: 0 2px 8px rgba(0, 122, 255, 0.15) !important;
        transition: all 0.3s cubic-bezier(0.25, 0.8, 0.25, 1) !important;
    }
    div.stButton > button[key="run_analysis_btn"]:hover:not(:disabled) {
        background-color: #0063CC !important;
        transform: translateY(-1px) scale(1.02) !important;
        box-shadow: 0 5px 15px rgba(0, 122, 255, 0.25) !important;
        filter: none !important;
        cursor: pointer !important;
    }
    div.stButton > button[key="run_analysis_btn"]:active {
        transform: scale(0.96) !important;
        background-color: #0051A8 !important;
        box-shadow: 0 1px 4px rgba(0, 0, 0, 0.1) !important;
        transition: all 0.1s ease !important;
    }
    div.stButton > button[key="run_analysis_btn"]:disabled {
        background-color: #E9E9EB !important;
        color: #AEAEB2 !important;
        cursor: not-allowed !important;
        transform: none !important;
        box-shadow: none !important;
    }

    /* ── General disabled fallback ── */
    div.stButton > button:disabled {
        background: #e2e8f0 !important; color: #94a3b8 !important;
        cursor: not-allowed !important; transform: none !important;
        box-shadow: none !important; border-color: #e2e8f0 !important;
    }

    /* ── Sidebar ── */
    [data-testid="stSidebar"] { background-color: #0f172a; border-right: 1px solid #1e293b; }
    [data-testid="stSidebar"] [data-testid="stHeader"],
    [data-testid="stSidebarCollapseButton"],
    [title="keyboard_double_arrow_left"] { display: none !important; }

    .sb-title           { color: white !important; font-weight: 700 !important; font-size: 1.1rem; }
    .sb-sub             { color: white !important; font-weight: 700 !important; font-size: 0.95rem; }
    .sb-filename        { color: #94d2f5 !important; font-weight: 400 !important; font-size: 0.80rem;
                          margin: 4px 0 0 0; word-break: break-word; }
    .system-spacer      { margin-top: 80px; }
    .sys-context-spacer { margin-top: 2.4rem; }
    .sidebar-stats      { color: white !important; font-weight: 400 !important; font-size: 0.85rem; margin-bottom: 5px; }

    /* ── Radio button labels white on dark sidebar ── */
    section[data-testid="stSidebar"] .stRadio label,
    section[data-testid="stSidebar"] .stRadio span,
    section[data-testid="stSidebar"] .stRadio p,
    section[data-testid="stSidebar"] .stRadio div,
    section[data-testid="stSidebar"] div[data-testid="stWidgetLabel"] p,
    section[data-testid="stSidebar"] div[data-baseweb="radio"] label,
    section[data-testid="stSidebar"] div[data-baseweb="radio"] span {
        color: white !important;
    }

    div.stButton > button[key="terminate_sidebar"] { width: 100% !important; }

    /* ── New Analysis button — neutral grey, right-aligned with download ── */
    div.stButton > button[key="clear_results_btn"] {
        background: #334155 !important; color: #e2e8f0 !important;
        border: 1px solid #475569 !important; border-radius: 8px !important;
        height: 2.6rem !important; font-size: 0.88rem !important;
    }
    div.stButton > button[key="clear_results_btn"]:hover:not(:disabled) {
        background: #475569 !important; color: white !important;
        border-color: #64748b !important;
    }

    /* ── Back to Periodic Review — top bar, no text wrap ── */
    div.stButton > button[key="pr_back_btn"] {
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        width: 100% !important;
    }

    /* ── Sticky top-right terminate button overlay ── */
    #sticky-terminate-overlay {
        position: fixed !important;
        top: 10px !important;
        right: 20px !important;
        z-index: 999999 !important;
    }
    #sticky-terminate-overlay button {
        background: #dc2626 !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 6px 16px !important;
        font-size: 0.82rem !important;
        font-weight: 600 !important;
        cursor: pointer !important;
        box-shadow: 0 2px 8px rgba(220,38,38,0.4) !important;
        transition: background 0.15s !important;
    }
    #sticky-terminate-overlay button:hover {
        background: #b91c1c !important;
    }

    /* ── Completely hide the invisible trigger button and its wrapper ── */
    /* Streamlit Cloud doesn't expose key= as an HTML attribute, so target
       by multiple selectors to ensure it's invisible across all environments */
    button[data-testid="stButton"][key="terminate_hidden_trigger"],
    div[data-testid="stButton"]:has(button[key="terminate_hidden_trigger"]) {
        display: none !important;
        visibility: hidden !important;
        height: 0 !important;
        width: 0 !important;
        overflow: hidden !important;
        position: absolute !important;
        pointer-events: none !important;
    }

    /* ── E-Signature modal ── */
    .esig-container {
        background: #0f172a;
        border: 2px solid #2563eb;
        border-radius: 12px;
        padding: 28px 32px;
        margin: 24px 0;
        font-family: 'Inter', sans-serif;
    }
    .esig-title {
        color: #e2e8f0;
        font-size: 1.15rem;
        font-weight: 700;
        letter-spacing: 1px;
        margin: 0 0 4px 0;
    }
    .esig-subtitle {
        color: #64748b;
        font-size: 0.78rem;
        margin: 0 0 20px 0;
        font-style: italic;
    }
    .esig-field-label {
        color: #94a3b8;
        font-size: 0.80rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 1px;
        margin: 14px 0 4px 0;
    }
    .esig-user-display {
        background: #1e293b;
        border: 1px solid #334155;
        border-radius: 6px;
        padding: 8px 12px;
        color: #e2e8f0;
        font-size: 0.92rem;
        font-weight: 600;
    }
    .esig-warning {
        color: #fbbf24;
        font-size: 0.78rem;
        margin-top: 8px;
    }
    </style>
    """, unsafe_allow_html=True)

MODELS = {
    "Gemini 1.5 Pro":    "gemini/gemini-1.5-pro",
    "Claude Sonnet 4.5":  "anthropic/claude-sonnet-4-5",
    "GPT-4o":            "openai/gpt-4o",
    "Groq (Llama 3.3)":  "groq/llama-3.3-70b-versatile"
}


def _capture_client_ip():
    """
    Capture client IP via JS → query_params on first page load.
    Stores in st.session_state['user_ip'] which is then written as a
    separate column in audit_log (v29). This satisfies 21 CFR Part 11
    electronic-signature traceability without relying on cloud server IPs.
    """
    if st.session_state.get("user_ip"):
        return  # already captured this session

    # Check if JS already posted the IP via query param
    qp = st.query_params
    if "uip" in qp:
        st.session_state["user_ip"] = str(qp["uip"])[:100]
        return

    # Inject JS: fetch client IP from cloudflare trace (no CORS issues), write to query param
    _st_components.html("""
    <script>
    (function() {
      try {
        fetch('https://www.cloudflare.com/cdn-cgi/trace')
          .then(r => r.text())
          .then(txt => {
            const m = txt.match(/ip=([\\d\\.a-fA-F:]+)/);
            if (m) {
              const url = new URL(window.parent.location.href);
              url.searchParams.set('uip', m[1]);
              window.parent.history.replaceState({}, '', url.toString());
            }
          }).catch(() => {});
      } catch(e) {}
    })();
    </script>
    """, height=0)


# =============================================================================
# 10b. CHANGE IMPACT ANALYSIS — BACKEND
# =============================================================================

def detect_tabular_doc_type(df: pd.DataFrame) -> str:
    """
    Fingerprint an uploaded tabular document (xlsx/csv) by column headers
    and ID patterns.  Returns: 'FRS' | 'OQ' | 'Traceability' | 'URS' | 'Unknown'
    """
    if df is None or df.empty:
        return "Unknown"
    cols      = [str(c).strip().lower() for c in df.columns]
    first_col = df.iloc[:, 0].astype(str).str.strip()

    # Count ID pattern matches in first column
    # OQ pattern is broad — matches OQ-001, OQ-LAB-01, OQ-SYS-002 etc.
    frs_ids = first_col.str.match(r'^FRS-',  case=False).sum()
    oq_ids  = first_col.str.match(r'^OQ-',   case=False).sum()
    urs_ids = first_col.str.match(r'^URS-',  case=False).sum()

    # Also scan ALL columns for OQ-style IDs (traceability has OQ_ID as a non-first column)
    all_vals_oq = sum(
        df[col].astype(str).str.match(r'^OQ-', case=False).sum()
        for col in df.columns
    )

    has_test_step   = any(c in cols for c in ['test_step', 'expected_result', 'pass_fail_criteria'])
    has_req_desc    = any('requirement_description' in c or 'req_desc' in c for c in cols)
    has_coverage    = any('coverage_status' in c or 'coverage' in c for c in cols)
    has_frs_ref     = any('frs_ref' in c or 'frs_id' in c for c in cols)
    has_urs_ref     = any('urs_req_id' in c or 'urs_ref' in c or 'urs_id' in c for c in cols)

    # Test/OQ column detection — handles Test_ID, OQ_ID, and any column starting with "oq"
    has_oq_col      = any('test_id' in c or c == 'oq_id' or c.startswith('oq') for c in cols)

    # Traceability — unique signature: has URS ref + FRS ref + OQ/Test column together
    # OR has coverage_status alongside FRS reference
    # OR has all three ID types (URS, FRS, OQ) present across any columns
    has_all_three_id_types = (
        df.apply(lambda col: col.astype(str).str.match(r'^URS-', case=False).any()).any() and
        df.apply(lambda col: col.astype(str).str.match(r'^FRS-', case=False).any()).any() and
        df.apply(lambda col: col.astype(str).str.match(r'^OQ-',  case=False).any()).any()
    )

    if has_all_three_id_types:
        return "Traceability"
    if (has_urs_ref or urs_ids > 0) and (has_frs_ref or frs_ids > 0) and has_oq_col:
        return "Traceability"
    if has_coverage and has_frs_ref:
        return "Traceability"

    # OQ — test steps present, or first column is OQ-NNN IDs
    if has_test_step or oq_ids > 3 or all_vals_oq > 3:
        return "OQ"

    # FRS
    if (has_req_desc or frs_ids > 3) and not has_test_step:
        return "FRS"

    # URS — first column is URS-NNN IDs but NOT a traceability doc
    if urs_ids > 3:
        return "URS"

    return "Unknown"


def _load_tabular(file_bytes: bytes, file_name: str) -> pd.DataFrame:
    """Load xlsx or csv into a DataFrame regardless of extension."""
    try:
        if file_name.lower().endswith(".csv"):
            return pd.read_csv(io.BytesIO(file_bytes), dtype=str).fillna("")
        else:
            return pd.read_excel(io.BytesIO(file_bytes), dtype=str).fillna("")
    except Exception as e:
        st.error(f"⛔ Could not read {file_name}: {e}")
        return pd.DataFrame()


def _validate_cia_slot(file_bytes, file_name, expected_type: str, slot_label: str):
    """
    Run type detection on an uploaded tabular file and return (df, ok, message).
    For PDFs (change spec + FRS), skip tabular detection.
    """
    if file_bytes is None:
        return None, False, ""

    # PDF slots — change spec and FRS are PDFs, skip column fingerprinting
    if file_name.lower().endswith(".pdf"):
        return file_bytes, True, f"✅ **{slot_label}** — PDF loaded ({len(file_bytes)//1024} KB)"

    df = _load_tabular(file_bytes, file_name)
    if df.empty:
        return None, False, f"⛔ **{slot_label}** — file is empty or could not be read."

    detected = detect_tabular_doc_type(df)

    if detected == expected_type:
        row_count = len(df)
        first_id  = str(df.iloc[0, 0]) if not df.empty else "?"
        last_id   = str(df.iloc[-1, 0]) if not df.empty else "?"
        return df, True, (
            f"✅ **{slot_label}** — {detected} detected, "
            f"{row_count} rows ({first_id} → {last_id})"
        )
    elif detected == "Unknown":
        return df, True, (
            f"⚠️ **{slot_label}** — document type unclear. "
            f"Expected {expected_type}. Verify columns match before running."
        )
    else:
        return None, False, (
            f"⛔ **{slot_label}** — wrong document. "
            f"Detected **{detected}** but this slot expects **{expected_type}**. "
            f"Please upload your {expected_type} file here."
        )


def build_cia_pass1_prompt(change_spec_text: str) -> str:
    """Extract structured change table from change specification PDF."""
    return _PROMPT_CIA_PASS1_RAW.format(
        change_spec_text=change_spec_text[:6000]
    )


def build_cia_pass2_prompt(
    chg_csv: str,
    frs_text: str,
    oq_df: pd.DataFrame,
    trace_df: pd.DataFrame
) -> str:
    """
    Map each change to impact on existing FRS rows, OQ rows, and trace links.
    Uses the trace matrix to build the relationship graph first.
    """
    oq_summary  = oq_df[oq_df.columns[:5]].head(80).to_csv(index=False) if not oq_df.empty else "No OQ provided"
    trc_summary = trace_df.head(80).to_csv(index=False) if not trace_df.empty else "No traceability matrix provided"

    return _PROMPT_CIA_PASS2_RAW.format(
        chg_csv     = chg_csv,
        frs_text    = frs_text[:4000],
        oq_summary  = oq_summary,
        trc_summary = trc_summary,
    )


def build_cia_pass3_prompt(
    frs_impact_df: pd.DataFrame,
    oq_impact_df:  pd.DataFrame,
    change_spec_text: str,
) -> str:
    """
    Build the Pass 3 justification prompt.
    Only feeds Must_Update, New_Required, and Obsolete rows — the three statuses
    that require Change Control justification strings. Needs_Review rows are
    excluded because they do not trigger Change Control actions.
    """
    action_statuses = {"Must_Update", "New_Required", "Obsolete"}

    frs_rows = pd.DataFrame()
    if not frs_impact_df.empty and "Impact_Status" in frs_impact_df.columns:
        frs_rows = frs_impact_df[
            frs_impact_df["Impact_Status"].astype(str).isin(action_statuses)
        ].copy()
        frs_rows["Document_Type"] = "FRS"
        # Rename FRS_ID → Document_ID for the unified prompt
        if "FRS_ID" in frs_rows.columns:
            frs_rows = frs_rows.rename(columns={"FRS_ID": "Document_ID"})

    oq_rows = pd.DataFrame()
    if not oq_impact_df.empty and "Impact_Status" in oq_impact_df.columns:
        oq_rows = oq_impact_df[
            oq_impact_df["Impact_Status"].astype(str).isin(action_statuses)
        ].copy()
        oq_rows["Document_Type"] = "OQ"
        if "OQ_ID" in oq_rows.columns:
            oq_rows = oq_rows.rename(columns={"OQ_ID": "Document_ID"})

    # Combine and select only the columns the prompt needs
    keep_cols = ["Document_ID", "Document_Type", "Impact_Status",
                 "Change_Driver", "Rationale"]
    combined  = pd.concat([frs_rows, oq_rows], ignore_index=True)
    avail     = [c for c in keep_cols if c in combined.columns]
    impacted_csv = combined[avail].to_csv(index=False) if not combined.empty else "No actionable rows."

    return _PROMPT_CIA_PASS3_RAW.format(
        impacted_csv      = impacted_csv,
        change_spec_text  = change_spec_text[:4000],
    )



def run_cia_analysis(
    change_spec_bytes: bytes,
    frs_bytes: bytes,
    oq_df: pd.DataFrame,
    trace_df: pd.DataFrame,
    model_id: str,
    status_widget,
    progress_widget
) -> dict:
    """
    Full Change Impact Analysis pipeline.
    Returns dict with keys: chg_df, frs_impact_df, oq_impact_df,
    justification_df, cia_gap_df, summary
    """
    from litellm import completion as _completion

    # Extract text from PDFs
    status_widget.text("📄 Extracting change specification text...")
    progress_widget.progress(0.1)
    chg_pages  = extract_pages(change_spec_bytes)
    chg_text   = "\n".join(chg_pages)

    frs_pages  = extract_pages(frs_bytes)
    frs_text   = "\n".join(frs_pages)

    # Pass 1 — extract structured change table
    status_widget.text("🔍 Pass 1 — Extracting structured change table from spec...")
    progress_widget.progress(0.25)
    p1_resp = _completion(
        model=model_id, stream=False, temperature=TEMPERATURE,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": build_cia_pass1_prompt(chg_text)}
        ]
    )
    raw_chg = p1_resp.choices[0].message.content or ""
    raw_chg = re.sub(r'^```[a-zA-Z]*\n?', '', raw_chg, flags=re.MULTILINE)
    raw_chg = re.sub(r'```\s*$', '', raw_chg, flags=re.MULTILINE).strip()
    chg_df  = _csv_to_df(raw_chg)

    if chg_df.empty:
        raise RuntimeError(
            "Pass 1 extracted zero changes from the change specification. "
            "Ensure the document describes specific system changes."
        )

    status_widget.text(f"✅ {len(chg_df)} changes extracted. Running impact mapping...")
    progress_widget.progress(0.5)

    # Pass 2 — impact mapping
    status_widget.text("🗺️ Pass 2 — Mapping changes to existing FRS and OQ rows...")
    p2_resp = _completion(
        model=model_id, stream=False, temperature=TEMPERATURE,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": build_cia_pass2_prompt(
                raw_chg, frs_text, oq_df, trace_df
            )}
        ]
    )
    raw_p2 = p2_resp.choices[0].message.content or ""
    raw_p2 = re.sub(r'^```[a-zA-Z]*\n?', '', raw_p2, flags=re.MULTILINE)
    raw_p2 = re.sub(r'```\s*$', '', raw_p2, flags=re.MULTILINE).strip()

    parts = [p.strip() for p in raw_p2.split("|||")]
    frs_impact_df = _csv_to_df(parts[0]) if len(parts) > 0 else pd.DataFrame()
    oq_impact_df  = _csv_to_df(parts[1]) if len(parts) > 1 else pd.DataFrame()

    # ── Trace-Propagated Impact — pandas merge approach ──────────────────────
    # Guarantees 100% compliance: even if the AI missed a linked OQ test,
    # Python will catch it by walking the trace matrix.
    #
    # Rules:
    #   FRS status Must_Update → linked OQ gets Needs_Review (unless already Must_Update)
    #   FRS status Obsolete    → linked OQ gets Needs_Review (unless already Must_Update)
    #   Already Must_Update OQ rows are never downgraded.
    # ─────────────────────────────────────────────────────────────────────────
    if not trace_df.empty and not frs_impact_df.empty:
        status_widget.text("🔗 Propagating impact through traceability matrix...")
        progress_widget.progress(0.75)

        # Step 1 — Detect column names flexibly (handles varied export formats)
        frs_col = next((c for c in trace_df.columns
                        if "frs" in c.lower() and "ref" not in c.lower().replace("frs_ref","x")), None)
        frs_col = frs_col or next((c for c in trace_df.columns if "frs" in c.lower()), None)
        oq_col  = next((c for c in trace_df.columns
                        if "test_id" in c.lower() or c.lower().startswith("oq")), None)

        if frs_col and oq_col:
            # Step 2 — Build a clean bridge: trace rows where FRS col is populated
            trace_bridge = (
                trace_df[[frs_col, oq_col]]
                .copy()
                .rename(columns={frs_col: "FRS_ID", oq_col: "OQ_ID"})
                .assign(
                    FRS_ID=lambda d: d["FRS_ID"].astype(str).str.strip(),
                    OQ_ID =lambda d: d["OQ_ID"].astype(str).str.strip(),
                )
                .query("FRS_ID != '' and OQ_ID != '' and FRS_ID != 'nan' and OQ_ID != 'nan'")
                .drop_duplicates()
            )

            # Step 3 — Find FRS rows that trigger propagation
            trigger_statuses = {"Must_Update", "Obsolete"}
            if "Impact_Status" in frs_impact_df.columns and "FRS_ID" in frs_impact_df.columns:
                triggered_frs = (
                    frs_impact_df[
                        frs_impact_df["Impact_Status"].astype(str).isin(trigger_statuses)
                    ][["FRS_ID", "Impact_Status"]]
                    .assign(FRS_ID=lambda d: d["FRS_ID"].astype(str).str.strip())
                    .drop_duplicates("FRS_ID")
                )
            else:
                triggered_frs = pd.DataFrame(columns=["FRS_ID", "Impact_Status"])

            if not triggered_frs.empty:
                # Step 4 — Merge: triggered FRS → trace bridge → OQ IDs
                # Result: every OQ linked to a triggered FRS with the FRS status attached
                propagated = (
                    triggered_frs
                    .merge(trace_bridge, on="FRS_ID", how="inner")
                    .rename(columns={"Impact_Status": "FRS_Status"})
                    [["OQ_ID", "FRS_ID", "FRS_Status"]]
                    .query("OQ_ID != 'NEW'")
                    .drop_duplicates("OQ_ID")   # one row per OQ (take first FRS if multiple)
                )

                if not propagated.empty:
                    # Step 5 — Determine which OQ IDs are already flagged by the AI
                    already_flagged_must = set()
                    already_flagged_any  = set()
                    if not oq_impact_df.empty and "OQ_ID" in oq_impact_df.columns:
                        already_flagged_any  = set(oq_impact_df["OQ_ID"].astype(str).str.strip())
                        already_flagged_must = set(
                            oq_impact_df[
                                oq_impact_df["Impact_Status"].astype(str) == "Must_Update"
                            ]["OQ_ID"].astype(str).str.strip()
                        )

                    # Step 6 — Build propagated rows, inheriting status from FRS
                    # Rule: Obsolete FRS → OQ becomes Obsolete (test no longer needed)
                    #        Must_Update FRS → OQ becomes Needs_Review (steps may be wrong)
                    #        Already Must_Update OQ → never downgraded
                    new_rows = []
                    for _, pr in propagated.iterrows():
                        oid      = pr["OQ_ID"]
                        frs_id   = pr["FRS_ID"]
                        frs_stat = pr["FRS_Status"]

                        # Derive the OQ status from the FRS status
                        propagated_status = (
                            "Obsolete"     if frs_stat == "Obsolete"    else
                            "Needs_Review"                               # Must_Update → Needs_Review
                        )

                        if oid in already_flagged_must:
                            # Already Must_Update — never downgrade, skip
                            continue

                        if oid in already_flagged_any:
                            # AI already flagged at some level — upgrade/set in place
                            mask = oq_impact_df["OQ_ID"].astype(str).str.strip() == oid
                            existing_status = oq_impact_df.loc[mask, "Impact_Status"].values
                            # Only upgrade, never downgrade
                            if len(existing_status) > 0 and existing_status[0] != "Obsolete":
                                oq_impact_df.loc[mask, "Impact_Status"]    = propagated_status
                            oq_impact_df.loc[mask, "Confidence_Level"] = "High"
                            oq_impact_df.loc[mask, "Change_Driver"]    = (
                                oq_impact_df.loc[mask, "Change_Driver"].astype(str)
                                + " + Trace-propagated"
                            )
                            oq_impact_df.loc[mask, "Rationale"] = (
                                f"System-flagged: Linked FRS [{frs_id}] is {frs_stat}. "
                                + (
                                    "This test case is no longer required for execution."
                                    if frs_stat == "Obsolete" else
                                    "Review test steps for continued validity against updated FRS."
                                )
                            )
                        else:
                            # Brand new row — AI did not flag this OQ at all
                            new_rows.append({
                                "OQ_ID":            oid,
                                "Change_Driver":    "Trace-propagated",
                                "Impact_Status":    propagated_status,
                                "Confidence_Level": "High",
                                "Risk_Category":    "GxP_Critical",
                                "Rationale":        (
                                    f"System-flagged: Linked FRS [{frs_id}] is {frs_stat}. "
                                    + (
                                        "This test case is no longer required for execution."
                                        if frs_stat == "Obsolete" else
                                        "Review test steps and pass/fail criteria for continued validity."
                                    )
                                ),
                                "Action_Required":  (
                                    "Retire this test case from the active test suite."
                                    if frs_stat == "Obsolete" else
                                    "Verify test steps and pass/fail criteria remain valid "
                                    "given the updated FRS requirement. Re-execute if affected."
                                ),
                            })

                    if new_rows:
                        oq_impact_df = pd.concat(
                            [oq_impact_df, pd.DataFrame(new_rows)],
                            ignore_index=True
                        )

                    status_widget.text(
                        f"🔗 Trace propagation complete — "
                        f"{len(new_rows)} new OQ rows added, "
                        f"{len(propagated) - len(new_rows)} existing rows upgraded."
                    )

    # ── Pass 3 — GxP Justification String Generation ─────────────────────────
    # Runs after Python trace propagation so the complete, final impact tables
    # (including all trace-propagated rows) feed into the justification prompt.
    # Only Must_Update, New_Required, and Obsolete rows are processed —
    # these are the only statuses that require Change Control justifications.
    justification_df = pd.DataFrame(
        columns=["Document_ID", "Document_Type", "Impact_Status", "Justification_String"]
    )
    try:
        status_widget.text("✍️ Pass 3 — Generating GxP justification strings for Change Control...")
        progress_widget.progress(0.85)

        p3_resp = _completion(
            model=model_id, stream=False, temperature=0.1,  # lower temp for deterministic phrasing
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user",   "content": build_cia_pass3_prompt(
                    frs_impact_df, oq_impact_df, chg_text
                )}
            ]
        )
        raw_p3 = p3_resp.choices[0].message.content or ""
        raw_p3 = re.sub(r'^```[a-zA-Z]*\n?', '', raw_p3, flags=re.MULTILINE)
        raw_p3 = re.sub(r'```\s*$', '', raw_p3, flags=re.MULTILINE).strip()
        justification_df = _csv_to_df(raw_p3)

        # Validate expected columns came back — if not, create an empty shell
        expected = {"Document_ID", "Document_Type", "Impact_Status", "Justification_String"}
        if not expected.issubset(set(justification_df.columns)):
            justification_df = pd.DataFrame(columns=list(expected))

        status_widget.text(
            f"✅ Pass 3 complete — "
            f"{len(justification_df)} justification string(s) generated."
        )
    except Exception as p3_err:
        # Pass 3 failure is non-fatal — triage output still stands
        import warnings
        warnings.warn(f"Pass 3 justification generation failed: {p3_err}")

    # ── Trace Coverage Verification ───────────────────────────────────────────
    # Compute what % of OQ tests in the uploaded OQ file appear in the trace matrix.
    # Orphan OQ tests (in OQ file but not in trace) indicate broken traceability.
    trace_coverage_pct  = 0
    orphan_oq_count     = 0
    trace_coverage_ok   = False
    if not oq_df.empty and not trace_df.empty:
        oq_col_trace = next((c for c in trace_df.columns if "test_id" in c.lower() or
                             c.lower().startswith("oq")), None)
        if oq_col_trace:
            all_oq_ids    = set(oq_df.iloc[:, 0].astype(str).str.strip())
            traced_oq_ids = set(trace_df[oq_col_trace].astype(str).str.strip())
            linked_count  = len(all_oq_ids & traced_oq_ids)
            orphan_oq_count = len(all_oq_ids - traced_oq_ids)
            trace_coverage_pct = round(linked_count / len(all_oq_ids) * 100, 1) if all_oq_ids else 0
            trace_coverage_ok  = trace_coverage_pct >= 90  # 90%+ = coverage intact

    # ── CIA Gap Detection — New_Required items with no test coverage ─────────
    # Any FRS or OQ row marked New_Required has no existing test coverage.
    # Flag these as Missing_Test gaps so the validator knows new tests are needed.
    cia_gap_rows = []
    if not frs_impact_df.empty and "Impact_Status" in frs_impact_df.columns:
        new_frs = frs_impact_df[
            frs_impact_df["Impact_Status"].astype(str) == "New_Required"
        ]
        for _, row in new_frs.iterrows():
            drv = str(row.get("Change_Driver", ""))
            cia_gap_rows.append({
                "Req_ID":         row.get("FRS_ID", "NEW"),
                "Gap_Type":       "Missing_Test",
                "Description":    (
                    f"New requirement from {drv} has no existing OQ test coverage. "
                    "This item did not exist in the previous validated baseline."
                ),
                "Recommendation": (
                    "Generate new OQ test cases covering positive, negative, and boundary "
                    "conditions before the next validation cycle sign-off."
                ),
                "Severity":       "High",
                "Change_Driver":  drv,
            })

    cia_gap_df = pd.DataFrame(cia_gap_rows) if cia_gap_rows else pd.DataFrame(
        columns=["Req_ID", "Gap_Type", "Description", "Recommendation",
                 "Severity", "Change_Driver"]
    )

    progress_widget.progress(1.0)
    status_widget.text("✅ Change impact analysis complete.")

    # Summary counts
    def _count(df, col, val):
        if df.empty or col not in df.columns:
            return 0
        return int((df[col].astype(str) == val).sum())

    summary = {
        "total_changes":       len(chg_df),
        "frs_must_update":     _count(frs_impact_df, "Impact_Status", "Must_Update"),
        "frs_needs_review":    _count(frs_impact_df, "Impact_Status", "Needs_Review"),
        "frs_obsolete":        _count(frs_impact_df, "Impact_Status", "Obsolete"),
        "frs_new":             _count(frs_impact_df, "Impact_Status", "New_Required"),
        "oq_must_update":      _count(oq_impact_df,  "Impact_Status", "Must_Update"),
        "oq_needs_review":     _count(oq_impact_df,  "Impact_Status", "Needs_Review"),
        "oq_obsolete":         _count(oq_impact_df,  "Impact_Status", "Obsolete"),
        "oq_new":              _count(oq_impact_df,  "Impact_Status", "New_Required"),
        "trace_coverage_pct":  trace_coverage_pct,
        "orphan_oq_count":     orphan_oq_count,
        "trace_coverage_ok":   trace_coverage_ok,
    }

    return {
        "chg_df":           chg_df,
        "frs_impact_df":    frs_impact_df,
        "oq_impact_df":     oq_impact_df,
        "cia_gap_df":       cia_gap_df,
        "justification_df": justification_df,
        "summary":          summary,
    }


def build_cia_excel(result: dict, user: str, file_name: str, model_name: str) -> bytes:
    """Build the Change Impact Analysis Excel workbook."""
    output = io.BytesIO()

    STATUS_COLORS = {
        "Must_Update":   "FEE2E2",   # red
        "Needs_Review":  "FEF9C3",   # yellow
        "Obsolete":      "E5E7EB",   # grey
        "New_Required":  "DBEAFE",   # blue
    }

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Sheet 1 — Change Table
        result["chg_df"].to_excel(writer, sheet_name="Changes", index=False)
        # Sheet 2 — FRS Impact
        result["frs_impact_df"].to_excel(writer, sheet_name="FRS_Impact", index=False)
        # Sheet 3 — OQ Impact
        result["oq_impact_df"].to_excel(writer, sheet_name="OQ_Impact", index=False)
        # Sheet 4 — GxP Justification Strings (Change Control ready)
        just_df = result.get("justification_df", pd.DataFrame())
        if not just_df.empty:
            # Add watermark column — regulatory artefact notice
            just_df = just_df.copy()
            just_df["Review_Status"] = "AI-PROPOSED — Human review and attestation required"
            just_df.to_excel(writer, sheet_name="Justifications", index=False)
        else:
            pd.DataFrame({
                "Note": ["Pass 3 justification generation produced no output. "
                         "Re-run the analysis or draft justifications manually."]
            }).to_excel(writer, sheet_name="Justifications", index=False)
        # Sheet 5 — Gaps (New_Required items with no test coverage)
        cia_gap_df = result.get("cia_gap_df", pd.DataFrame())
        if not cia_gap_df.empty:
            cia_gap_df.to_excel(writer, sheet_name="Gaps", index=False)
        else:
            pd.DataFrame({"Note": ["No gaps detected — all changes have existing test coverage."]
                         }).to_excel(writer, sheet_name="Gaps", index=False)

        wb = writer.book

        for sheet_name in ["Changes", "FRS_Impact", "OQ_Impact", "Justifications", "Gaps"]:
            if sheet_name not in wb.sheetnames:
                continue
            ws  = wb[sheet_name]
            hdr = Font(bold=True, color="FFFFFF", size=10)
            hf  = PatternFill("solid", fgColor="1E293B")
            thin = Side(style="thin", color="CBD5E1")
            bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col in range(1, ws.max_column + 1):
                c = ws.cell(row=1, column=col)
                c.font, c.fill, c.border = hdr, hf, bdr
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Colour rows by Impact_Status
            if sheet_name in ("FRS_Impact", "OQ_Impact"):
                hdr_vals = {ws.cell(row=1, column=c).value: c
                            for c in range(1, ws.max_column + 1)}
                status_col = hdr_vals.get("Impact_Status")
                for row_i in range(2, ws.max_row + 1):
                    status = ws.cell(row=row_i, column=status_col).value if status_col else ""
                    fill_hex = STATUS_COLORS.get(str(status), "FFFFFF")
                    fill = PatternFill("solid", fgColor=fill_hex)
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_i, column=col)
                        cell.fill   = fill
                        cell.border = bdr
                        cell.alignment = Alignment(vertical="top", wrap_text=True)

            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes    = "A2"
            ws.sheet_properties.tabColor = (
                "DC2626" if sheet_name == "FRS_Impact"    else
                "D97706" if sheet_name == "OQ_Impact"     else
                "EA580C" if sheet_name == "Gaps"          else
                "7C3AED" if sheet_name == "Justifications" else "1E3A5F"
            )
            for col in range(1, ws.max_column + 1):
                cl = get_column_letter(col)
                hdr_val = str(ws.cell(1, col).value or "")
                # Justification_String needs extra width for long sentences
                if hdr_val == "Justification_String":
                    ws.column_dimensions[cl].width = 90
                elif hdr_val == "Review_Status":
                    ws.column_dimensions[cl].width = 50
                else:
                    ws.column_dimensions[cl].width = min(
                        max(14, max(
                            (len(str(ws.cell(r, col).value or "")) for r in range(1, ws.max_row + 1)),
                            default=14
                        ) + 4), 60
                    )

        # Summary sheet
        s = result["summary"]
        trc_ok     = s.get("trace_coverage_ok", False)
        trc_pct    = s.get("trace_coverage_pct", 0)
        orphan_cnt = s.get("orphan_oq_count", 0)
        trc_label  = (f"Yes — {trc_pct}% of OQ tests linked in trace matrix"
                      if trc_ok else
                      f"No — {trc_pct}% linked, {orphan_cnt} orphan OQ test(s) not in trace")

        summary_data = {
            "Metric": [
                "Total Changes Detected",
                "FRS — Must Update",           "FRS — Needs Review",
                "FRS — Obsolete",              "FRS — New Required",
                "OQ — Must Update",            "OQ — Needs Review",
                "OQ — Obsolete",               "OQ — New Required",
                "─── Trace Coverage ───",
                "Trace Coverage Verified",
                "Orphan OQ Tests (not in trace)",
                "─── Confidence Guide ───",
                "High Confidence",             "Medium Confidence",
                "Low Confidence",
                "─── Risk Category Guide ───",
                "GxP_Critical",                "Data_Integrity",
                "Business",                    "Cosmetic",
            ],
            "Count / Value": [
                s["total_changes"],
                s["frs_must_update"],  s["frs_needs_review"],
                s["frs_obsolete"],     s["frs_new"],
                s["oq_must_update"],   s["oq_needs_review"],
                s["oq_obsolete"],      s["oq_new"],
                "",
                trc_label,
                orphan_cnt,
                "",
                "Direct unambiguous match — act immediately",
                "Clear semantic link — engineering judgement required",
                "Weak/inferred — human must verify before acting",
                "",
                "Patient safety, e-records, audit trail, e-signatures",
                "Data accuracy, completeness, retention (ALCOA+)",
                "Operational workflows, non-GxP functionality",
                "UI labels, formatting, display only",
            ],
            "Action": [
                "Review full change table",
                "Update before next validation cycle", "Human review required",
                "Mark retired in document register",   "Generate new FRS rows — see Justifications sheet",
                "Re-execute tests after update",       "Re-verify before sign-off",
                "Retire from test suite",              "Generate new OQ tests — see Justifications sheet",
                "", "", "",                # Trace Coverage section (3 rows)
                "", "", "", "",            # Confidence Guide section (4 rows)
                "", "", "", "", "",        # Risk Category section (5 rows: header + 4 values)
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        ws_s = wb["Summary"]
        ws_s.sheet_properties.tabColor = "059669"

        # Style the summary sheet — highlight Trace Coverage Verified row
        for row_i in range(2, ws_s.max_row + 1):
            cell_val = str(ws_s.cell(row=row_i, column=1).value or "")
            if cell_val == "Trace Coverage Verified":
                fill_hex = "D1FAE5" if trc_ok else "FEE2E2"
                for col_i in range(1, 4):
                    ws_s.cell(row=row_i, column=col_i).fill = PatternFill("solid", fgColor=fill_hex)
                    ws_s.cell(row=row_i, column=col_i).font = Font(bold=True)
            elif cell_val.startswith("───"):
                for col_i in range(1, 4):
                    ws_s.cell(row=row_i, column=col_i).fill = PatternFill("solid", fgColor="1E293B")
                    ws_s.cell(row=row_i, column=col_i).font = Font(bold=True, color="FFFFFF")

        for col_i in range(1, 4):
            cl = get_column_letter(col_i)
            ws_s.column_dimensions[cl].width = 45 if col_i == 2 else 30

    return output.getvalue()


def show_change_impact(user: str, role: str, model_id: str):
    """Render the Change Impact Analysis main panel."""
    st.title("🔍 Change Impact Analysis")
    st.markdown(
        "<p style='color:#94a3b8;margin-top:-12px;'>Identify which existing FRS requirements "
        "and OQ test cases are affected by a system change, version upgrade, or configuration update.</p>",
        unsafe_allow_html=True
    )

    ck = st.session_state.get("cia_key_n", 0)

    # ── Step 1 — Change Specification ──────────────────────────────────────
    st.markdown("### Step 1 — Upload Change Specification")
    st.caption("Change Request, Release Notes, Configuration Change Notice, or mini-URS (PDF)")
    chg_widget = st.file_uploader(
        "Change Spec", type="pdf",
        key=f"cia_chg_{ck}", label_visibility="collapsed"
    )
    if chg_widget:
        ok, msg = validate_upload(chg_widget)
        if not ok:
            st.error(msg)
            st.session_state["cia_change_spec_bytes"] = None
            st.session_state["cia_change_spec_name"]  = None
        else:
            st.session_state["cia_change_spec_bytes"] = chg_widget.getvalue()
            st.session_state["cia_change_spec_name"]  = chg_widget.name
            st.success(f"✅ **Change Spec** — {chg_widget.name} loaded "
                       f"({len(chg_widget.getvalue())//1024} KB)")
    elif chg_widget is None:
        st.session_state["cia_change_spec_bytes"] = None
        st.session_state["cia_change_spec_name"]  = None

    st.markdown("---")

    # ── Step 2 — Existing Validated Documents ──────────────────────────────
    st.markdown("### Step 2 — Upload Existing Validated Documents")
    st.caption(
        "The Traceability Matrix is the **prerequisite** — it carries your URS requirement IDs "
        "and their links to FRS and OQ. The URS document itself is not required."
    )

    col_frs, col_oq = st.columns(2)
    col_trc, col_note = st.columns(2)

    with col_frs:
        st.caption("📄 Approved FRS (PDF)")
        frs_widget = st.file_uploader(
            "FRS", type=["pdf"],
            key=f"cia_frs_{ck}", label_visibility="collapsed"
        )
        if frs_widget:
            ok, msg = validate_upload(frs_widget)
            if not ok:
                st.error(msg)
                st.session_state["cia_frs_bytes"] = None
                st.session_state["cia_frs_name"]  = None
            else:
                st.session_state["cia_frs_bytes"] = frs_widget.getvalue()
                st.session_state["cia_frs_name"]  = frs_widget.name
                st.success(f"✅ FRS loaded — {frs_widget.name}")
        elif frs_widget is None:
            st.session_state["cia_frs_bytes"] = None
            st.session_state["cia_frs_name"]  = None

    with col_oq:
        st.caption("📊 OQ Test Cases (.xlsx or .csv from test engine)")
        oq_widget = st.file_uploader(
            "OQ", type=["xlsx", "xls", "csv"],
            key=f"cia_oq_{ck}", label_visibility="collapsed"
        )
        if oq_widget:
            oq_bytes = oq_widget.getvalue()
            oq_df_check = _load_tabular(oq_bytes, oq_widget.name)
            _, ok, msg = _validate_cia_slot(oq_bytes, oq_widget.name, "OQ", "OQ Test Cases")
            if not ok:
                st.error(msg)
                st.session_state["cia_oq_bytes"] = None
                st.session_state["cia_oq_name"]  = None
            else:
                st.markdown(msg)
                st.session_state["cia_oq_bytes"] = oq_bytes
                st.session_state["cia_oq_name"]  = oq_widget.name
        elif oq_widget is None:
            st.session_state["cia_oq_bytes"] = None
            st.session_state["cia_oq_name"]  = None

    with col_trc:
        st.caption("📊 Traceability Matrix (.xlsx or .csv)")
        trc_widget = st.file_uploader(
            "Trace", type=["xlsx", "xls", "csv"],
            key=f"cia_trc_{ck}", label_visibility="collapsed"
        )
        if trc_widget:
            trc_bytes = trc_widget.getvalue()
            _, ok, msg = _validate_cia_slot(trc_bytes, trc_widget.name,
                                            "Traceability", "Traceability Matrix")
            if not ok:
                st.error(msg)
                st.session_state["cia_trace_bytes"] = None
                st.session_state["cia_trace_name"]  = None
            else:
                st.markdown(msg)
                st.session_state["cia_trace_bytes"] = trc_bytes
                st.session_state["cia_trace_name"]  = trc_widget.name
        elif trc_widget is None:
            st.session_state["cia_trace_bytes"] = None
            st.session_state["cia_trace_name"]  = None

    with col_note:
        st.caption("ℹ️ Prerequisites")
        st.markdown(
            "<p style='color:#64748b;font-size:0.82rem;'>"
            "<b style='color:#e2e8f0;'>Traceability Matrix is required</b> — it contains your "
            "URS requirement IDs (URS-NNN) already linked to FRS and OQ rows. "
            "A separate URS upload is not needed.<br><br>"
            "FRS: approved baseline PDF from your document management system.<br><br>"
            "OQ and Traceability: export directly from your test management tool "
            "(Veeva Vault, TestRail, Jira, HP ALM, etc.).</p>",
            unsafe_allow_html=True
        )

    # ── Cross-slot consistency check ────────────────────────────────────────
    oq_bytes  = st.session_state.get("cia_oq_bytes")
    trc_bytes = st.session_state.get("cia_trace_bytes")
    if oq_bytes and trc_bytes:
        oq_df_v  = _load_tabular(oq_bytes,  st.session_state.get("cia_oq_name", ""))
        trc_df_v = _load_tabular(trc_bytes, st.session_state.get("cia_trace_name", ""))
        if not oq_df_v.empty and not trc_df_v.empty:
            oq_ids_set  = set(oq_df_v.iloc[:, 0].astype(str).str.strip())
            # Find OQ-like column in trace
            trc_oq_col  = next((c for c in trc_df_v.columns
                                if "test_id" in c.lower() or c.lower().startswith("oq")), None)
            if trc_oq_col:
                trc_oq_ids = set(trc_df_v[trc_oq_col].astype(str).str.strip())
                shared     = oq_ids_set & trc_oq_ids
                if len(shared) == 0 and len(oq_ids_set) > 0:
                    st.warning(
                        "⚠️ **Version mismatch** — no OQ IDs in the traceability matrix match "
                        "the uploaded OQ file. These files may be from different validation cycles. "
                        "Verify both files are from the same approved baseline."
                    )

    st.markdown("---")

    # ── Run button ──────────────────────────────────────────────────────────
    all_ready = all([
        st.session_state.get("cia_change_spec_bytes"),
        st.session_state.get("cia_frs_bytes"),
        st.session_state.get("cia_oq_bytes"),
        st.session_state.get("cia_trace_bytes"),
    ])

    _es, run_col, _es2 = st.columns([3, 4, 3])
    with run_col:
        run_cia = st.button(
            "🔍 Run Change Impact Analysis",
            key="run_cia_btn",
            disabled=not all_ready,
            use_container_width=True,
            type="primary"
        )

    if not all_ready:
        missing = [
            label for label, key in [
                ("Change Specification", "cia_change_spec_bytes"),
                ("FRS",                  "cia_frs_bytes"),
                ("OQ Test Cases",        "cia_oq_bytes"),
                ("Traceability Matrix",  "cia_trace_bytes"),
            ] if not st.session_state.get(key)
        ]
        st.info(f"📋 Still needed: {', '.join(missing)}")

    if run_cia:
        oq_df_run  = _load_tabular(
            st.session_state["cia_oq_bytes"],
            st.session_state["cia_oq_name"]
        )
        trc_df_run = _load_tabular(
            st.session_state["cia_trace_bytes"],
            st.session_state["cia_trace_name"]
        )
        progress_bar = st.progress(0)
        status_text  = st.empty()

        with st.status("🔍 Change Impact Analysis Pipeline", expanded=True) as cia_status:
            try:
                st.write("📄 Step 1: Extracting change specification...")
                st.write("🗺️ Step 2: Mapping changes to existing FRS requirements...")
                st.write("🧪 Step 3: Propagating impact through traceability matrix to OQ tests...")
                st.write("📊 Step 4: Compiling impact report...")

                result = run_cia_analysis(
                    change_spec_bytes = st.session_state["cia_change_spec_bytes"],
                    frs_bytes         = st.session_state["cia_frs_bytes"],
                    oq_df             = oq_df_run,
                    trace_df          = trc_df_run,
                    model_id          = model_id,
                    status_widget     = status_text,
                    progress_widget   = progress_bar,
                )
                cia_status.update(
                    label=f"✅ Complete — {result['summary']['total_changes']} changes, "
                          f"{result['summary']['frs_must_update'] + result['summary']['oq_must_update']} "
                          f"items require immediate action",
                    state="complete", expanded=False
                )
                log_audit(user, "CIA_COMPLETE", "CHANGE_IMPACT",
                          new_value=f"{result['summary']['total_changes']} changes detected",
                          reason=f"Model: {st.session_state.selected_model}")
                st.session_state["cia_result"] = result

            except Exception as e:
                cia_status.update(label="❌ Analysis failed", state="error")
                st.error(f"❌ {e}")
                log_audit(user, "CIA_ERROR", "CHANGE_IMPACT", reason=str(e)[:300])

        progress_bar.empty()
        status_text.empty()

    # ── Display results ──────────────────────────────────────────────────────
    cia_res = st.session_state.get("cia_result")
    if cia_res:
        s = cia_res["summary"]
        trc_ok     = s.get("trace_coverage_ok", False)
        trc_pct    = s.get("trace_coverage_pct", 0)
        orphan_cnt = s.get("orphan_oq_count", 0)

        # Hero metrics — 5 tiles
        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("📋 Changes Detected",    s["total_changes"])
        m2.metric("🔴 FRS Must Update",     s["frs_must_update"])
        m3.metric("🔴 OQ Must Update",      s["oq_must_update"])
        m4.metric("🔵 New Items Required",  s["frs_new"] + s["oq_new"])
        m5.metric("🔗 Trace Coverage",      f"{trc_pct}%",
                  delta=f"{orphan_cnt} orphan OQ" if orphan_cnt > 0 else "intact",
                  delta_color="inverse" if orphan_cnt > 0 else "off")

        # Colour-coded hero cards
        _c1, _c2, _c3, _c4 = st.columns(4)
        for col, icon, label, frs_v, oq_v, color, bg in [
            (_c1, "🔴", "Must Update",   s["frs_must_update"],  s["oq_must_update"],  "#dc2626", "#1a0505"),
            (_c2, "🟡", "Needs Review",  s["frs_needs_review"], s["oq_needs_review"], "#d97706", "#1a1000"),
            (_c3, "🔵", "New Required",  s["frs_new"],          s["oq_new"],          "#2563eb", "#0f1a2e"),
        ]:
            col.markdown(f"""
<div style="background:{bg};border:2px solid {color};border-radius:10px;
            padding:14px 18px;text-align:center;font-family:'Inter',sans-serif;">
  <p style="margin:0;color:#94a3b8;font-size:0.7rem;letter-spacing:2px;
            text-transform:uppercase;">{label}</p>
  <p style="margin:4px 0 2px;font-size:1.6rem;font-weight:800;color:{color};">
    FRS: {frs_v} &nbsp;|&nbsp; OQ: {oq_v}</p>
  <span style="font-size:1.5rem;">{icon}</span>
</div>""", unsafe_allow_html=True)

        # Trace Coverage Verified card
        _trc_color = "#059669" if trc_ok else "#dc2626"
        _trc_bg    = "#052019"  if trc_ok else "#1a0505"
        _trc_icon  = "🟢" if trc_ok else "🔴"
        _trc_detail = (
            f"All but {orphan_cnt} OQ test(s) are linked in the traceability matrix. "
            "Repair orphan links before signing off the impact report."
            if orphan_cnt > 0 else
            "All OQ tests appear in the traceability matrix. Validation structure is intact."
        )
        _c4.markdown(f"""
<div style="background:{_trc_bg};border:2px solid {_trc_color};border-radius:10px;
            padding:14px 18px;text-align:center;font-family:'Inter',sans-serif;">
  <p style="margin:0;color:#94a3b8;font-size:0.7rem;letter-spacing:2px;
            text-transform:uppercase;">Trace Coverage</p>
  <p style="margin:4px 0 2px;font-size:1.6rem;font-weight:800;color:{_trc_color};">
    {trc_pct}% {_trc_icon}</p>
  <p style="margin:0;color:#94a3b8;font-size:0.72rem;">
    {'✅ Verified' if trc_ok else f'⚠️ {orphan_cnt} orphan OQ test(s)'}</p>
</div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # Previews
        with st.expander("📋 Changes Extracted from Spec", expanded=True):
            st.dataframe(cia_res["chg_df"], use_container_width=True)

        STATUS_DISPLAY = {
            "Must_Update":  "🔴",
            "Needs_Review": "🟡",
            "Obsolete":     "⚫",
            "New_Required": "🔵",
        }

        with st.expander("📐 FRS Impact", expanded=True):
            frs_imp = cia_res["frs_impact_df"].copy()
            if not frs_imp.empty and "Impact_Status" in frs_imp.columns:
                frs_imp["Status"] = frs_imp["Impact_Status"].map(
                    lambda x: f"{STATUS_DISPLAY.get(x, '')} {x}"
                )
            st.dataframe(frs_imp, use_container_width=True)

        with st.expander("🧪 OQ Impact", expanded=True):
            oq_imp = cia_res["oq_impact_df"].copy()
            if not oq_imp.empty and "Impact_Status" in oq_imp.columns:
                oq_imp["Status"] = oq_imp["Impact_Status"].map(
                    lambda x: f"{STATUS_DISPLAY.get(x, '')} {x}"
                )
            st.dataframe(oq_imp, use_container_width=True)

        cia_gap_df = cia_res.get("cia_gap_df", pd.DataFrame())
        with st.expander(
            f"⚠️ Gaps — Missing Test Coverage ({len(cia_gap_df)} item(s))",
            expanded=len(cia_gap_df) > 0
        ):
            if not cia_gap_df.empty:
                st.dataframe(cia_gap_df, use_container_width=True)
            else:
                st.success("✅ No gaps detected — all changes have existing test coverage.")

        # ── Pass 3: GxP Justification Strings ────────────────────────────────
        just_df = cia_res.get("justification_df", pd.DataFrame())
        n_just  = len(just_df) if not just_df.empty else 0
        with st.expander(
            f"✍️ Change Control Justification Strings ({n_just} item(s))",
            expanded=n_just > 0
        ):
            if just_df.empty or n_just == 0:
                st.info("No justification strings were generated. "
                        "Re-run the analysis or draft manually.")
            else:
                # Watermark banner
                st.markdown("""
<div style="background:#1e1b4b;border:2px solid #7c3aed;border-radius:8px;
            padding:12px 18px;margin-bottom:16px;">
  <p style="margin:0;color:#c4b5fd;font-size:0.8rem;font-weight:600;">
    ⚠️ AI-PROPOSED DRAFT — Human review and attestation required before use in any
    regulatory submission, Change Control record, or Impact Assessment.
    The Validation Engineer is responsible for verifying the accuracy of each
    string against the source documents before copying to a Change Control form.
  </p>
</div>""", unsafe_allow_html=True)

                # Acknowledgement checkbox — creates an implicit review record
                ack = st.checkbox(
                    "I have reviewed these justification strings against the source "
                    "Change Specification and FRS/OQ documents and confirm they are "
                    "accurate before use.",
                    key="cia_just_ack"
                )

                st.markdown("<br>", unsafe_allow_html=True)

                # Display each justification as a copy-friendly card
                for _, jrow in just_df.iterrows():
                    doc_id    = str(jrow.get("Document_ID",        "—"))
                    doc_type  = str(jrow.get("Document_Type",      "—"))
                    status    = str(jrow.get("Impact_Status",      "—"))
                    just_str  = str(jrow.get("Justification_String",""))

                    status_color = {
                        "Must_Update":  "#dc2626",
                        "New_Required": "#2563eb",
                        "Obsolete":     "#6b7280",
                    }.get(status, "#d97706")

                    st.markdown(f"""
<div style="background:#0f172a;border-left:3px solid {status_color};
            border-radius:6px;padding:14px 18px;margin-bottom:10px;">
  <div style="display:flex;gap:10px;align-items:center;margin-bottom:8px;">
    <span style="background:{status_color}22;border:1px solid {status_color}66;
           color:{status_color};padding:2px 8px;border-radius:4px;
           font-size:0.72rem;font-weight:600;">{doc_type} · {doc_id}</span>
    <span style="background:#1e293b;color:#94a3b8;padding:2px 8px;border-radius:4px;
           font-size:0.72rem;">{status}</span>
  </div>
  <p style="margin:0;color:#e2e8f0;font-size:0.88rem;line-height:1.6;
            font-style:italic;">"{just_str}"</p>
</div>""", unsafe_allow_html=True)

                if not ack:
                    st.caption("☝️ Check the box above to confirm review before using "
                               "these strings in any formal document.")

        # Download
        st.markdown("---")
        xlsx_bytes = build_cia_excel(
            cia_res, user,
            st.session_state.get("cia_change_spec_name", "change_spec"),
            st.session_state.selected_model
        )
        dl1, _sp, clear_c = st.columns([5, 2, 2])
        with dl1:
            st.download_button(
                label="📥 Download Change Impact Report (.xlsx)",
                data=xlsx_bytes,
                file_name=f"CIA_{st.session_state.get('cia_change_spec_name','report').replace('.pdf','')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="cia_download_btn",
                use_container_width=True,
            )
        with clear_c:
            if st.button("🔄 New Analysis", key="cia_clear_btn", use_container_width=True):
                for k in ["cia_change_spec_bytes","cia_change_spec_name","cia_frs_bytes",
                          "cia_frs_name","cia_oq_bytes","cia_oq_name","cia_trace_bytes",
                          "cia_trace_name","cia_result"]:
                    st.session_state[k] = None
                st.session_state["cia_key_n"] = st.session_state.get("cia_key_n", 0) + 1
                st.rerun()




# =============================================================================
# 10c. PERIODIC REVIEW — MODULE 1: AUDIT TRAIL INTELLIGENCE
# =============================================================================

# ── Scoring constants ─────────────────────────────────────────────────────────
_AT_ADMIN_KW    = ["admin","administrator","sysadmin","system","service",
                   "root","superuser","power user","dba","sa"]
_AT_DELETE_KW   = ["delete","del","remove","purge","void","cancel","retire"]
_AT_MODIFY_KW   = ["modify","update","edit","change","amend","correct",
                   "override","revise","write"]
_AT_CREATE_KW   = ["create","add","insert","new","submit","approve",
                   "release","publish"]
_AT_SENSITIVE   = ["batch record","audit trail","electronic signature","esig",
                   "test result","clinical","raw data","master data",
                   "configuration","user account","role","permission",
                   "gxp","quality record","change control","deviation",
                   "capa","oos","oos result"]
_AT_AUDIT_CTRL  = ["audit trail","audit log","logging","log enabled","log disabled",
                   "audit enabled","audit disabled","configuration change","system setting"]
_AT_BIZ_START   = 7
_AT_BIZ_END     = 20
_AT_WEEKENDS    = {5, 6}
_AT_VEL_WINDOW  = 60    # minutes
_AT_VEL_THRESH  = 5     # same user+action in window = anomaly
_AT_TOP_N       = 20

# US Federal Holidays — fully dynamic observed-date calculation.
# Fixed holidays shift to Friday if they fall on Saturday, Monday if Sunday.
# Floating holidays are computed from weekday + week-of-month rules — these
# are inherently year-agnostic and require no date tables.
import datetime as _dt_mod

_AT_US_FIXED_HOLIDAYS = [
    (1,  1,  "New Year's Day"),
    (6,  19, "Juneteenth"),
    (7,  4,  "Independence Day"),
    (11, 11, "Veterans Day"),
    (12, 25, "Christmas Day"),
]

def _us_observed_date(year: int, month: int, day: int) -> _dt_mod.date:
    """
    Return the federally observed date for a fixed holiday.
    US OPM rule: Saturday → preceding Friday; Sunday → following Monday.
    """
    actual = _dt_mod.date(year, month, day)
    wd = actual.weekday()           # Mon=0 … Sun=6
    if wd == 5:                     # Saturday → Friday
        return actual - _dt_mod.timedelta(days=1)
    if wd == 6:                     # Sunday → Monday
        return actual + _dt_mod.timedelta(days=1)
    return actual


def _is_us_federal_holiday(ts: pd.Timestamp) -> tuple:
    """
    Returns (bool, holiday_name). Covers all 11 US Federal Holidays.

    Fixed holidays (New Year's, Juneteenth, Independence Day, Veterans Day,
    Christmas) are resolved to their exact observed date for the given year
    using proper OPM Saturday→Friday / Sunday→Monday shifting.

    Floating holidays (MLK, Presidents, Memorial, Labor, Columbus,
    Thanksgiving) are detected from weekday + week-of-month arithmetic —
    fully dynamic for any year, no lookup tables required.

    Edge case: when 1 Jan falls on Saturday, the observed holiday is
    31 Dec of the prior year — handled explicitly.
    """
    if pd.isnull(ts):
        return False, ""
    ts          = pd.Timestamp(ts)
    check_date  = ts.date()
    year        = check_date.year

    # ── Fixed holidays with correct observed-date shifting ────────────────
    for month, day, name in _AT_US_FIXED_HOLIDAYS:
        try:
            observed = _us_observed_date(year, month, day)
        except ValueError:
            continue
        if check_date == observed:
            return True, name

    # Edge case: New Year's Day on Saturday → observed Dec 31 prior year
    try:
        ny_observed = _us_observed_date(year + 1, 1, 1)
        if ny_observed == _dt_mod.date(year, 12, 31) and check_date == ny_observed:
            return True, "New Year's Day (observed)"
    except ValueError:
        pass

    # ── Floating holidays (weekday + week-of-month arithmetic) ────────────
    m  = ts.month
    d  = ts.day
    wd = ts.weekday()                   # Mon=0 … Sun=6
    wk = (d - 1) // 7 + 1              # 1-indexed week within month

    if m == 1  and wd == 0 and wk == 3:       return True, "MLK Day"
    if m == 2  and wd == 0 and wk == 3:       return True, "Presidents Day"
    if m == 5  and wd == 0 and wk >= 4:       return True, "Memorial Day"
    if m == 9  and wd == 0 and wk == 1:       return True, "Labor Day"
    if m == 10 and wd == 0 and wk == 2:       return True, "Columbus Day"
    if m == 11 and wd == 3 and wk == 4:       return True, "Thanksgiving"

    return False, ""

_AT_REQUIRED_COLS = {
    "timestamp":   "Timestamp / Date-Time of the event",
    "user_id":     "User ID / Username who performed the action",
    "action_type": "Action / Event Type (e.g. Create, Modify, Delete, Insert)",
    "record_id":   "Record ID / Object ID affected (optional)",
    "record_type": "Record Type / Table Name (e.g. RESULTS, BATCH, SAMPLE_DATA)",
    "role":        "User Role / Permission Level (e.g. Admin, DBA, Analyst)",
    "comments":    "Comments / Rationale / Change Reason field (optional — Rule 1)",
    "new_value":   "New Value / Changed Value (optional — Rule 4 drift detection)",
}

# Vague rationale terms that trigger Rule 1
_AT_VAGUE_TERMS = {"fixed","update","updated","error","changed","change","test",
                   "misc","other","n/a","na","correction","corrected","edit","edited",
                   "modified","mod","ok","done","see above","as per","per request"}

# Tables that are GxP-critical for Rule 1, 3, 4
_AT_GXP_TABLES  = ["results","result","batch","batch_release","sample_data",
                   "sample","test_result","audit_trail","electronic_signature",
                   "quality_record","raw_data"]


def _at_temporal_score(ts) -> float:
    try:
        ts = pd.Timestamp(ts)
    except Exception:
        return 3.0
    if pd.isnull(ts):
        return 3.0
    score = 0.0
    # Weekend
    if ts.weekday() in _AT_WEEKENDS:
        score += 5.0
    # Outside business hours
    if ts.hour < _AT_BIZ_START or ts.hour >= _AT_BIZ_END:
        score += 4.0
    # Deep night 00:00–05:00
    if 0 <= ts.hour < 5:
        score += 1.0
    # US Federal Holiday — shadow activity window
    is_holiday, _ = _is_us_federal_holiday(ts)
    if is_holiday:
        score += 4.0   # same weight as off-hours; holiday + off-hours = 8+ alone
    return min(score, 10.0)


def _at_velocity_scores(df: pd.DataFrame) -> pd.Series:
    scores = pd.Series(0.0, index=df.index)
    if not all(c in df.columns for c in ["timestamp_parsed","user_id","action_type"]):
        return scores
    df_s   = df.sort_values("timestamp_parsed").copy()
    df_s   = df_s[df_s["timestamp_parsed"].notna()].copy()
    if df_s.empty:
        return scores
    ts_arr = df_s["timestamp_parsed"].tolist()   # list of pd.Timestamp
    us_arr = df_s["user_id"].astype(str).tolist()
    ac_arr = df_s["action_type"].astype(str).str.upper().tolist()
    ix_arr = df_s.index.tolist()
    for i in range(len(df_s)):
        count = 0
        for j in range(max(0, i - 200), min(len(df_s), i + 200)):
            if j == i:
                continue
            try:
                diff_mins = abs((ts_arr[j] - ts_arr[i]).total_seconds() / 60)
            except Exception:
                continue
            if diff_mins <= _AT_VEL_WINDOW:
                if us_arr[j] == us_arr[i] and ac_arr[j] == ac_arr[i]:
                    count += 1
        if count >= _AT_VEL_THRESH:
            scores.at[ix_arr[i]] = min(count / _AT_VEL_THRESH * 3.5, 10.0)
    return scores


def _at_gap_scores(df: pd.DataFrame) -> pd.Series:
    scores = pd.Series(0.0, index=df.index)
    if "timestamp_parsed" not in df.columns:
        return scores
    ts   = df["timestamp_parsed"].sort_values()
    prev = ts.shift(1)
    gap  = (ts - prev).dt.total_seconds() / 3600
    scores.loc[gap[gap > 2].index] = 7.0
    return scores


def _at_del_recreate_scores(df: pd.DataFrame) -> pd.Series:
    scores = pd.Series(0.0, index=df.index)
    needed = ["record_id","user_id","action_type","timestamp_parsed"]
    if not all(c in df.columns for c in needed):
        return scores
    df2 = df[df["record_id"].astype(str).str.strip() != ""].copy()
    df2["_del"] = df2["action_type"].astype(str).str.lower().apply(
        lambda x: any(k in x for k in _AT_DELETE_KW))
    df2["_cre"] = df2["action_type"].astype(str).str.lower().apply(
        lambda x: any(k in x for k in _AT_CREATE_KW))
    dels = df2[df2["_del"]]
    cres = df2[df2["_cre"]]
    for _, dr in dels.iterrows():
        if pd.isnull(dr["timestamp_parsed"]):
            continue
        match = cres[
            (cres["record_id"] == dr["record_id"]) &
            (cres["user_id"]   == dr["user_id"]) &
            (cres["timestamp_parsed"] >= dr["timestamp_parsed"]) &
            (cres["timestamp_parsed"] <= dr["timestamp_parsed"] + pd.Timedelta(hours=4))
        ]
        if not match.empty:
            di = df2[(df2["record_id"]==dr["record_id"]) &
                     (df2["user_id"]==dr["user_id"]) &
                     (df2["_del"]) &
                     (df2["timestamp_parsed"]==dr["timestamp_parsed"])].index
            scores.loc[di]          = 9.0
            scores.loc[match.index] = 9.0
    return scores


def at_score_events(df: pd.DataFrame) -> pd.DataFrame:
    """
    Score every event across the original 6 dimensions PLUS
    4 named AI Skill rules from the GxP Operational Anomaly Detection spec.
    Returns sorted DataFrame with individual dimension scores, rule flags,
    rationale strings, and composite Risk_Score.
    """
    df = df.copy()

    # ── Timestamp parsing ─────────────────────────────────────────────────────
    if "timestamp" in df.columns:
        df["timestamp_parsed"] = pd.to_datetime(
            df["timestamp"], errors="coerce", infer_datetime_format=True)
    else:
        df["timestamp_parsed"] = pd.NaT

    # ── Original 6 dimensions ─────────────────────────────────────────────────
    df["score_temporal"]     = df["timestamp_parsed"].apply(_at_temporal_score)
    df["score_velocity"]     = _at_velocity_scores(df)
    df["score_gap"]          = _at_gap_scores(df)
    df["score_del_recreate"] = _at_del_recreate_scores(df)

    def _priv(row):
        role = str(row.get("role","")).lower()
        rec  = str(row.get("record_type","")).lower()
        act  = str(row.get("action_type","")).lower()
        if any(k in role for k in _AT_ADMIN_KW) and any(k in rec for k in _AT_SENSITIVE):
            return 8.0
        if any(k in role for k in _AT_ADMIN_KW) and any(
                k in act for k in _AT_MODIFY_KW + _AT_DELETE_KW):
            return 7.0
        return 0.0
    df["score_privilege"] = df.apply(_priv, axis=1)

    # Read-only actions on any record type — never a data integrity finding
    _AT_READ_ONLY = {"select", "read", "view", "query", "search",
                     "select_audit", "read_audit", "list", "export",
                     "report", "print"}

    def _rec(row):
        rec = str(row.get("record_type","")).lower()
        act = str(row.get("action_type","")).lower()
        # Read-only actions are never a Record integrity finding regardless
        # of which table they touch — a QA reviewer selecting audit trail
        # records is expected behaviour, not a critical integrity event.
        if act in _AT_READ_ONLY:
            return 0.0
        combined = act + " " + rec
        # Direct record_type name check — any modification to a table named
        # audit_trail is a critical integrity event.
        if "audit_trail" in rec or "audit trail" in rec:
            return 10.0
        if any(k in combined for k in _AT_AUDIT_CTRL):
            return 10.0
        if any(k in rec for k in _AT_SENSITIVE) and any(k in act for k in _AT_DELETE_KW):
            return 8.0
        return 0.0
    df["score_record"] = df.apply(_rec, axis=1)

    # ── Rule 1 — Vague Rationale (Compliance Gap) ─────────────────────────────
    # Target: UPDATE on RESULTS or BATCH table with <3-word or non-descriptive comment
    # Risk: High
    def _rule1(row):
        act  = str(row.get("action_type","")).upper()
        tbl  = str(row.get("record_type","")).upper()
        cmt  = str(row.get("comments","")).strip().lower()
        tbl_hit = any(t in tbl for t in ["RESULTS","RESULT","BATCH"])
        if not tbl_hit:
            return 0.0, ""
        if "UPDATE" not in act and "MODIFY" not in act and "EDIT" not in act:
            return 0.0, ""
        # Blank or missing comment — always flag
        if not cmt or cmt in ("", "nan", "none", "-", "—"):
            return 8.0, (
                "No change reason was recorded for this data modification. "
                "Every update to a GxP record requires a documented justification "
                "explaining what changed and why (21 CFR Part 211.68, ALCOA+)."
            )
        # Contains a known vague/non-descriptive term — flag
        words     = [w for w in cmt.split() if len(w) > 1]
        has_vague = any(vague in cmt for vague in _AT_VAGUE_TERMS)
        # Only flag short comments if they ALSO contain a vague term
        # (avoids false positives on legitimate brief annotations like "Outlier test")
        if has_vague:
            matched = [v for v in _AT_VAGUE_TERMS if v in cmt]
            return 7.0, (
                f"The change reason recorded ('{cmt}') uses non-descriptive language "
                f"({', '.join(matched)}) that does not explain what changed or why. "
                "GxP data modifications require a specific, scientifically justified "
                "rationale per 21 CFR Part 211.68."
            )
        # Very short AND no meaningful content (1 word only, not scientific)
        if len(words) < 2:
            return 6.0, (
                f"The change reason ('{cmt}') is too brief to constitute adequate "
                "documentation for a GxP record modification. A minimum of a clear, "
                "specific explanation is required per ALCOA+ Attributable principle."
            )
        return 0.0, ""

    r1_scores   = []
    r1_rationale = []
    for _, row in df.iterrows():
        s, r = _rule1(row)
        r1_scores.append(s)
        r1_rationale.append(r)
    df["score_rule1_vague_rationale"] = r1_scores
    df["rule1_rationale"]             = r1_rationale

    # ── Rule 2 — Contemporaneous Burst (ALCOA Gap) ────────────────────────────
    # Target: >10 RESULT_INSERT actions within 15-minute window per user
    # Risk: Medium
    r2_scores    = pd.Series(0.0, index=df.index)
    r2_rationale = pd.Series("", index=df.index)
    if "timestamp_parsed" in df.columns and "user_id" in df.columns:
        df_s   = df.sort_values("timestamp_parsed").copy()
        df_s   = df_s[df_s["timestamp_parsed"].notna()].copy()
        ts_arr = df_s["timestamp_parsed"].tolist()
        us_arr = df_s["user_id"].astype(str).tolist()
        ac_arr = df_s["action_type"].astype(str).str.upper().tolist()
        ix_arr = df_s.index.tolist()
        insert_kw = ["INSERT","RESULT_INSERT","CREATE","ADD"]
        for i in range(len(df_s)):
            if not any(kw in ac_arr[i] for kw in insert_kw):
                continue
            count = 0
            for j in range(max(0,i-200), min(len(df_s),i+200)):
                if j == i:
                    continue
                try:
                    diff_mins = abs((ts_arr[j]-ts_arr[i]).total_seconds() / 60)
                except Exception:
                    continue
                if diff_mins <= 15:
                    if us_arr[j]==us_arr[i] and any(kw in ac_arr[j] for kw in insert_kw):
                        count += 1
            if count > 10:
                r2_scores.at[ix_arr[i]]    = 6.0
                r2_rationale.at[ix_arr[i]] = (
                    f"Rule 2 — Contemporaneous Burst [MEDIUM]: {count+1} INSERT actions "
                    f"by user '{us_arr[i]}' within 15 minutes. Exceeds the 10-action "
                    "threshold indicating batch processing from memory or paper scraps "
                    "rather than real-time entry, which is inconsistent with the ALCOA+ Contemporaneous principle (FDA Data Integrity Guidance, 2018)."
                )
    df["score_rule2_burst"]    = r2_scores
    df["rule2_rationale"]      = r2_rationale

    # ── Rule 3 — Admin/GxP Conflict (SoD Gap) ────────────────────────────────
    # Target: Admin or DBA performing INSERT or UPDATE on SAMPLE_DATA or BATCH_RELEASE
    # Risk: Critical
    def _rule3(row):
        role = str(row.get("role","")).upper()
        act  = str(row.get("action_type","")).upper()
        tbl  = str(row.get("record_type","")).upper()
        role_hit = any(r in role for r in ["ADMIN","DBA","ADMINISTRATOR","SYSADMIN"])
        act_hit  = any(a in act  for a in ["INSERT","UPDATE","CREATE","MODIFY"])
        tbl_hit  = any(t in tbl  for t in ["SAMPLE_DATA","SAMPLE","BATCH_RELEASE",
                                            "BATCH","RESULTS","RESULT"])
        if role_hit and act_hit and tbl_hit:
            return 10.0, (
                f"Rule 3 — Admin/GxP Conflict [CRITICAL]: Role '{row.get('role','')}' "
                f"performed {row.get('action_type','')} on production GxP table "
                f"'{row.get('record_type','')}'. Admins must maintain system configuration "
                "only — direct modification of production data by an administrative account may indicate a Segregation of Duties gap inconsistent with data integrity expectations under 21 CFR Part 11 §11.10(d) and FDA Data Integrity Guidance (2018)."
            )
        return 0.0, ""

    r3_scores    = []
    r3_rationale = []
    for _, row in df.iterrows():
        s, r = _rule3(row)
        r3_scores.append(s)
        r3_rationale.append(r)
    df["score_rule3_admin_conflict"] = r3_scores
    df["rule3_rationale"]            = r3_rationale

    # ── Rule 4 — Change Control Drift (Validation Gap) ────────────────────────
    # Target: new_value column present + deviation from expected patterns
    # Risk: High
    # Note: without an uploaded Change Request PDF, we detect numeric outliers
    # and flag any new_value that differs from the modal value for that record type
    def _rule4_scores(df: pd.DataFrame) -> tuple:
        scores    = pd.Series(0.0, index=df.index)
        rationale = pd.Series("", index=df.index)
        if "new_value" not in df.columns or "record_type" not in df.columns:
            return scores, rationale
        # Group by record_type, find modal new_value; flag deviations
        for rec_type, grp in df.groupby("record_type"):
            if len(grp) < 3:
                continue
            vals = grp["new_value"].astype(str).str.strip()
            # Try numeric deviation detection
            try:
                numeric_vals = pd.to_numeric(vals, errors="coerce").dropna()
                if len(numeric_vals) >= 3:
                    mean = numeric_vals.mean()
                    std  = numeric_vals.std()
                    if std > 0:
                        for idx in grp.index:
                            try:
                                v = float(df.at[idx,"new_value"])
                                z = abs(v - mean) / std
                                if z > 3.0:   # >3 std dev from mean for this record type
                                    scores.at[idx] = 8.0
                                    rationale.at[idx] = (
                                        f"Rule 4 — Change Control Drift [HIGH]: "
                                        f"new_value '{v}' deviates {z:.1f} standard deviations "
                                        f"from the expected range for '{rec_type}' records "
                                        f"(mean={mean:.2f}, σ={std:.2f}). "
                                        "May indicate manual override of a validated setpoint "
                                        "without Change Control per 21 CFR Part 820.70(b)."
                                    )
                            except (ValueError, TypeError):
                                pass
            except Exception:
                pass
        return scores, rationale

    r4_sc, r4_rat              = _rule4_scores(df)
    df["score_rule4_drift"]    = r4_sc
    df["rule4_rationale"]      = r4_rat

    # ── Rule 5 — Failed Login → Data Manipulation (Credential Abuse) ─────────
    # Target: 3+ failed login events followed by a successful login, then a
    #         DELETE or MODIFY on a GxP record within 30 minutes.
    # Risk: Critical. Pattern indicates brute-force or credential stuffing
    #       preceding unauthorised data manipulation.
    # Requires: action_type column with LOGIN_FAILED / AUTHENTICATION_FAILED
    #           and a subsequent successful login + data action in the same file.
    r5_scores    = pd.Series(0.0, index=df.index)
    r5_rationale = pd.Series("", index=df.index)
    if "timestamp_parsed" in df.columns and "user_id" in df.columns:
        df_s   = df.sort_values("timestamp_parsed").copy()
        df_s   = df_s[df_s["timestamp_parsed"].notna()].copy()
        ts_lst = df_s["timestamp_parsed"].tolist()
        us_lst = df_s["user_id"].astype(str).tolist()
        ac_lst = df_s["action_type"].astype(str).str.upper().tolist()
        ix_lst = df_s.index.tolist()

        failed_kw  = ["LOGIN_FAILED","AUTH_FAILED","AUTHENTICATION_FAILED",
                      "FAILED_LOGIN","LOGIN_FAILURE","LOGON_FAILURE","FAILED_LOGON"]
        success_kw = ["LOGIN","LOGON","AUTHENTICATION_SUCCESS","LOGIN_SUCCESS"]
        manip_kw   = ["DELETE","UPDATE","MODIFY","MODIFY_RESULT","RESULT_UPDATE",
                      "AMEND","OVERRIDE","REVISE","INSERT","RESULT_INSERT"]
        manip_tbl  = ["results","result","batch","batch_release","sample_data",
                      "sample","test_result","raw_data","quality_record"]

        for i in range(len(df_s)):
            # Must be a successful login
            if not any(kw in ac_lst[i] for kw in success_kw):
                continue
            usr = us_lst[i]
            t_login = ts_lst[i]

            # Count failed logins by same user in the 2 hours before this login
            failed_count = 0
            for j in range(max(0, i-300), i):
                if us_lst[j] != usr:
                    continue
                if not any(kw in ac_lst[j] for kw in failed_kw):
                    continue
                try:
                    mins_before = (t_login - ts_lst[j]).total_seconds() / 60
                    if 0 < mins_before <= 120:
                        failed_count += 1
                except Exception:
                    continue

            if failed_count < 3:
                continue

            # Now look for a GxP data manipulation within 30 min after the login
            for k in range(i+1, min(len(df_s), i+500)):
                if us_lst[k] != usr:
                    continue
                try:
                    mins_after = (ts_lst[k] - t_login).total_seconds() / 60
                except Exception:
                    continue
                if mins_after > 30:
                    break
                act_k = ac_lst[k]
                rec_k = str(df_s.iloc[k].get("record_type","")).lower() \
                        if hasattr(df_s.iloc[k], "get") else ""
                rec_k = df_s["record_type"].iloc[k].lower() \
                        if "record_type" in df_s.columns else ""
                if any(kw in act_k for kw in manip_kw):
                    # Flag both the login event and the manipulation event
                    rationale_text = (
                        f"Rule 5 — Failed Login → Data Manipulation [CRITICAL]: "
                        f"User '{usr}' had {failed_count} failed login attempt(s) "
                        f"in the 120 minutes preceding successful login at {t_login}. "
                        f"Within 30 minutes of login, action '{df_s['action_type'].iloc[k]}' "
                        f"was performed on '{df_s['record_type'].iloc[k] if 'record_type' in df_s.columns else 'GxP record'}'. "
                        f"This sequence indicates potential brute-force access followed by "
                        f"This sequence may indicate unauthorised data access and manipulation, raising concerns regarding data originality and attributability inconsistent with 21 CFR Part 11 §11.300 and FDA Data Integrity Guidance (2018)."
                    )
                    r5_scores.at[ix_lst[i]] = 10.0
                    r5_scores.at[ix_lst[k]] = 10.0
                    r5_rationale.at[ix_lst[i]] = rationale_text
                    r5_rationale.at[ix_lst[k]] = rationale_text
                    break   # one manipulation is enough to flag

    df["score_rule5_failed_login"] = r5_scores
    df["rule5_rationale"]          = r5_rationale

    # ── Rule 12 — Timestamp Reversal ──────────────────────────────────────────
    # Approval/release timestamp precedes creation timestamp on the same record.
    # Risk: Critical. Audit trail alone — no extra file needed.
    r12_scores    = pd.Series(0.0, index=df.index)
    r12_rationale = pd.Series("", index=df.index)
    if all(c in df.columns for c in ["record_id","action_type","timestamp_parsed"]):
        create_kw  = {"insert","create","add","result_insert","new"}
        approve_kw = {"approve","release","authorise","authorize","sign","submit",
                      "batch_release","approve_result"}
        valid = df[df["record_id"].astype(str).str.strip().ne("") &
                   df["timestamp_parsed"].notna()].copy()
        valid["_rid"] = valid["record_id"].astype(str).str.strip()
        for rid, grp in valid.groupby("_rid"):
            if len(grp) < 2:
                continue
            acts     = grp["action_type"].astype(str).str.lower()
            creates  = grp[acts.isin(create_kw)]
            approves = grp[acts.isin(approve_kw)]
            if creates.empty or approves.empty:
                continue
            t_create  = creates["timestamp_parsed"].min()
            t_approve = approves["timestamp_parsed"].min()
            if pd.isnull(t_create) or pd.isnull(t_approve):
                continue
            if t_approve < t_create:
                for idx in approves.index:
                    r12_scores.at[idx]    = 10.0
                    r12_rationale.at[idx] = (
                        f"Rule 12 — Timestamp Reversal [CRITICAL]: "
                        f"Record '{rid}' was approved/released at {t_approve} "
                        f"which is before its creation timestamp at {t_create}. "
                        "This is chronologically impossible in a correctly functioning "
                        "system and indicates clock manipulation or direct database "
                        "alteration. (21 CFR Part 11 §11.10(e), ALCOA+ Contemporaneous)"
                    )
    df["score_rule12_timestamp_reversal"] = r12_scores
    df["rule12_rationale"]                = r12_rationale

    # ── Rule 13 — Service / Shared Account GxP Action ─────────────────────────
    # Non-personal accounts (service, shared, automated) performing GxP data ops.
    # Risk: Critical for GxP data actions, High otherwise.
    _NONPERSONAL_PREFIXES = (
        "svc_","service_","shr_","share_","shared_","share.",
        "adm_","admin_","tec_","tech_","technical_",
        "interface_","int_","batch_","sys_","system_",
        "robot_","auto_","automation_","script_","api_",
        "sa_","dba_","root","daemon","guest","test_",
    )
    _GXP_ACTIONS_13 = {"insert","update","modify","delete","create","result_insert",
                       "amend","approve","release","override"}

    def _rule13(row) -> tuple:
        uid = str(row.get("user_id","")).lower().strip()
        act = str(row.get("action_type","")).lower().strip()
        rec = str(row.get("record_type","")).lower().strip()
        if not any(uid.startswith(p) for p in _NONPERSONAL_PREFIXES):
            return 0.0, ""
        is_gxp_action = any(kw in act for kw in _GXP_ACTIONS_13)
        is_gxp_rec    = any(kw in rec for kw in _AT_SENSITIVE)
        if is_gxp_action and is_gxp_rec:
            return 10.0, (
                f"Rule 13 — Service/Shared Account GxP Action [CRITICAL]: "
                f"Account '{row.get('user_id','')}' is a non-personal account "
                f"(service, shared, or automated) that performed "
                f"'{row.get('action_type','')}' on '{row.get('record_type','')}' "
                f"record '{row.get('record_id','')}'. Non-personal accounts cannot "
                "be attributed to a single individual, which is inconsistent with "
                "21 CFR Part 11 §11.300 and ALCOA+ Attributable principle."
            )
        if is_gxp_action:
            return 7.0, (
                f"Rule 13 — Service/Shared Account Action [HIGH]: "
                f"Non-personal account '{row.get('user_id','')}' performed "
                f"'{row.get('action_type','')}'. Verify this action was authorised "
                "and that an individual can be identified as responsible "
                "(21 CFR Part 11 §11.300)."
            )
        return 0.0, ""

    r13_res = [_rule13(row) for _, row in df.iterrows()]
    df["score_rule13_service_account"] = [x[0] for x in r13_res]
    df["rule13_rationale"]             = [x[1] for x in r13_res]

    # ── Rule 14 — Dormant Account Sudden Activity ──────────────────────────────
    # User with ≥90-day gap in activity re-activates and performs GxP action.
    # Risk: High. Audit trail alone — no extra file needed.
    _DORMANT_DAYS      = 90
    _DORMANT_MIN_PRIOR = 3

    r14_scores    = pd.Series(0.0, index=df.index)
    r14_rationale = pd.Series("", index=df.index)
    if "timestamp_parsed" in df.columns and "user_id" in df.columns:
        df_s14 = df.sort_values("timestamp_parsed").copy()
        df_s14 = df_s14[df_s14["timestamp_parsed"].notna()].copy()
        for uid, ugrp in df_s14.groupby("user_id"):
            if len(ugrp) < _DORMANT_MIN_PRIOR + 1:
                continue
            ts_list  = ugrp["timestamp_parsed"].tolist()
            idx_list = ugrp.index.tolist()
            for i in range(1, len(ts_list)):
                try:
                    gap_days = (ts_list[i] - ts_list[i-1]).total_seconds() / 86400
                except Exception:
                    continue
                if gap_days < _DORMANT_DAYS:
                    continue
                act = str(df_s14.at[idx_list[i], "action_type"]).lower()
                rec = str(df_s14.at[idx_list[i], "record_type"]).lower()
                is_gxp = (
                    any(kw in rec for kw in _AT_SENSITIVE) or
                    any(kw in act for kw in {"update","modify","delete",
                                             "insert","approve","release"})
                )
                if is_gxp:
                    gap_disp = f"{int(gap_days)} days"
                    r14_scores.at[idx_list[i]]    = 8.0
                    r14_rationale.at[idx_list[i]] = (
                        f"Rule 14 — Dormant Account Sudden Activity [HIGH]: "
                        f"User '{uid}' had no recorded activity for {gap_disp} "
                        f"(last seen {ts_list[i-1].date()}, "
                        f"re-activated {ts_list[i].date()}). "
                        f"This account then performed "
                        f"'{df_s14.at[idx_list[i],'action_type']}' on "
                        f"'{df_s14.at[idx_list[i],'record_type']}'. "
                        "Dormant accounts must be deactivated or formally "
                        "re-authorised. Verify current employment status and "
                        "access approval (21 CFR Part 11 §11.10(d))."
                    )
                    break   # flag only first re-activation per user
    df["score_rule14_dormant_account"] = r14_scores
    df["rule14_rationale"]             = r14_rationale

    # ── Rule 15 — Suspicious Action Sequence (UPDATE → DELETE → INSERT) ───────
    # Detects the 3-step data manipulation pattern:
    #   1. User UPDATEs a record (modification is logged)
    #   2. User DELETEs the same record (removes the logged update)
    #   3. User INSERTs the same record_id (recreates with altered values)
    # All three steps by same user, same record_id, within 30 minutes.
    # This extends Rule 6 (delete-recreate) by detecting the preceding UPDATE
    # that motivates the deletion — the hallmark of deliberate log evasion.
    # Risk: Critical. No extra file needed.
    _SEQ_WINDOW_MINS = 30
    r15_scores    = pd.Series(0.0, index=df.index)
    r15_rationale = pd.Series("", index=df.index)
    r15_chain_key = pd.Series("", index=df.index)  # for Event Chain ID
    if all(c in df.columns for c in ["record_id","user_id","action_type",
                                       "timestamp_parsed"]):
        df_s15 = df.sort_values("timestamp_parsed").copy()
        df_s15 = df_s15[df_s15["timestamp_parsed"].notna() &
                        df_s15["record_id"].astype(str).str.strip().ne("")].copy()
        df_s15["_rid"] = df_s15["record_id"].astype(str).str.strip()

        chain_counter = [0]
        for rid, grp in df_s15.groupby("_rid"):
            if len(grp) < 3:
                continue
            events = grp.sort_values("timestamp_parsed")
            acts   = events["action_type"].astype(str).str.upper().tolist()
            tss    = events["timestamp_parsed"].tolist()
            usrs   = events["user_id"].astype(str).tolist()
            idxs   = events.index.tolist()

            for i in range(len(events) - 2):
                # Look for UPDATE/MODIFY → DELETE → INSERT in order
                is_upd = any(k in acts[i] for k in
                             ["UPDATE","MODIFY","EDIT","AMEND"])
                is_del = any(k in acts[i+1] for k in
                             ["DELETE","DEL","REMOVE","PURGE"])
                is_ins = any(k in acts[i+2] for k in
                             ["INSERT","CREATE","ADD","RESULT_INSERT"])
                if not (is_upd and is_del and is_ins):
                    continue
                # Same user all three steps
                if not (usrs[i] == usrs[i+1] == usrs[i+2]):
                    continue
                # Within time window
                try:
                    span = (tss[i+2] - tss[i]).total_seconds() / 60
                except Exception:
                    continue
                if span > _SEQ_WINDOW_MINS:
                    continue

                chain_counter[0] += 1
                chain_id = f"EC-{chain_counter[0]:03d}"
                rationale = (
                    f"Rule 15 — Suspicious Action Sequence [CRITICAL]: "
                    f"User '{usrs[i]}' performed UPDATE → DELETE → INSERT on record "
                    f"'{rid}' within {span:.0f} minutes "
                    f"({tss[i].strftime('%H:%M')} to {tss[i+2].strftime('%H:%M')}). "
                    "This three-step sequence — modify, delete, recreate — is the "
                    "primary method for altering a locked GxP record while obscuring "
                    "the original modification from the audit trail. "
                    "(21 CFR Part 11 §11.10(e), ALCOA+ Original)"
                )
                for j, idx in enumerate([idxs[i], idxs[i+1], idxs[i+2]]):
                    r15_scores.at[idx]    = 10.0
                    r15_rationale.at[idx] = rationale
                    r15_chain_key.at[idx] = chain_id
                break   # one sequence per record is enough

    df["score_rule15_suspicious_sequence"] = r15_scores
    df["rule15_rationale"]                 = r15_rationale

    # ── Rule 16 — First-Time Behavior Detection ────────────────────────────────
    # Detects when a user performs an action_type they have never performed before
    # in the uploaded audit trail history. High-prior-event users with a sudden
    # new high-risk action are a meaningful insider risk signal.
    #
    # Scoring tiers:
    #   ≥50 prior events + first-time DELETE/APPROVE on GxP table → 9.0
    #   ≥20 prior events + first-time DELETE/APPROVE any table    → 8.0
    #   ≥20 prior events + first-time any action                  → 6.0
    #   ≥5  prior events + first-time high-risk action            → 5.0
    #   <5  prior events (too little history — skip)              → 0.0
    #
    # Rationale always states prior event count so reviewer can judge
    # whether "first time" is statistically meaningful.

    _HIGH_RISK_FIRST_ACTIONS = {
        "delete","del","purge",
        "approve","release","authorise","authorize","sign",
        "override","batch_release","approve_result",
    }
    _MIN_PRIOR_EVENTS_16 = 5

    r16_scores    = pd.Series(0.0, index=df.index)
    r16_rationale = pd.Series("", index=df.index)

    if "user_id" in df.columns and "action_type" in df.columns:
        df_s16 = df.copy()
        if "timestamp_parsed" in df_s16.columns:
            df_s16 = df_s16.sort_values("timestamp_parsed").reset_index(drop=False)
            orig_idx = df_s16["index"].tolist()
        else:
            df_s16 = df_s16.reset_index(drop=False)
            orig_idx = df_s16["index"].tolist()

        df_s16["_uid"] = df_s16["user_id"].astype(str)
        df_s16["_act"] = df_s16["action_type"].astype(str).str.lower().str.strip()
        df_s16["_rec"] = df_s16["record_type"].astype(str).str.lower() \
                         if "record_type" in df_s16.columns \
                         else pd.Series([""] * len(df_s16))

        for uid, ugrp in df_s16.groupby("_uid"):
            if len(ugrp) < _MIN_PRIOR_EVENTS_16 + 1:
                continue

            seen_acts: set = set()
            for pos, (_, urow) in enumerate(ugrp.iterrows()):
                orig = urow["index"]
                act  = urow["_act"]
                rec  = urow["_rec"]
                raw_act = str(df.at[orig, "action_type"]) \
                          if orig in df.index else act
                raw_rec = str(df.at[orig, "record_type"]) \
                          if ("record_type" in df.columns and orig in df.index) \
                          else rec

                if pos < _MIN_PRIOR_EVENTS_16:
                    seen_acts.add(act)
                    continue

                prior = pos  # events seen before this position
                is_new = act not in seen_acts

                if not is_new:
                    seen_acts.add(act)
                    continue

                is_gxp  = any(kw in rec for kw in _AT_SENSITIVE)
                is_hira  = any(kw in act for kw in _HIGH_RISK_FIRST_ACTIONS)

                score = 0.0
                if prior >= 50 and is_hira and is_gxp:
                    score = 9.0
                elif prior >= 20 and is_hira:
                    score = 8.0
                elif prior >= _MIN_PRIOR_EVENTS_16 and is_hira:
                    score = 5.0
                # Note: first-time non-high-risk actions (SELECT, VIEW, READ, etc.)
                # do not score independently — they require is_hira to fire.
                # This prevents alert fatigue from routine read operations.

                if score > 0:
                    conf = (
                        "High confidence" if prior >= 50 else
                        "Moderate confidence" if prior >= 20 else
                        "Low confidence — limited prior history"
                    )
                    r16_scores.at[orig]    = score
                    r16_rationale.at[orig] = (
                        f"Rule 16 — First-Time Behavior [HIGH]: "
                        f"User '{uid}' performed '{raw_act}' on '{raw_rec}' "
                        f"for the first time after {prior} prior recorded events in "
                        f"this audit trail. ({conf}) "
                        "A sudden new action type from an established user is an "
                        "insider risk signal — especially high-risk actions like "
                        "delete or approve that were never performed before. "
                        "Verify this action was within the user's approved access rights at the time "
                        "and obtain documented authorisation if not already on file "
                        "(21 CFR Part 11 §11.10(d), ALCOA+ Attributable)."
                    )

                seen_acts.add(act)

    df["score_rule16_first_time_behavior"] = r16_scores
    df["rule16_rationale"]                 = r16_rationale

    # ── Event Chain ID — group related events from Rules 5, 6, 15 ─────────────
    # Gives reviewer a shared identifier to filter and see complete event stories.
    # Format: EC-NNN where NNN increments per chain found in the dataset.
    chain_id_col = pd.Series("", index=df.index)
    chain_id_col[r15_chain_key != ""] = r15_chain_key[r15_chain_key != ""]

    # Rule 5 chains — LOGIN_FAILED sequence → same user → DELETE/UPDATE
    # Assign chain IDs to the login and manipulation events together
    if "score_rule5_failed_login" in df.columns:
        r5_flagged = df[df["score_rule5_failed_login"] >= 8].copy()
        for uid, ugrp in r5_flagged.groupby("user_id"):
            if len(ugrp) == 0:
                continue
            chain_counter[0] += 1
            cid = f"EC-{chain_counter[0]:03d}"
            # Also flag the LOGIN_FAILED rows for same user within 2hr window
            if "timestamp_parsed" in df.columns:
                for idx in ugrp.index:
                    t_flag = df.at[idx, "timestamp_parsed"]
                    if pd.isnull(t_flag):
                        continue
                    related = df[
                        (df["user_id"].astype(str) == str(uid)) &
                        (df["timestamp_parsed"].notna()) &
                        ((df["timestamp_parsed"] - t_flag).abs() <=
                         pd.Timedelta(hours=2))
                    ].index
                    for ridx in related:
                        if chain_id_col.at[ridx] == "":
                            chain_id_col.at[ridx] = cid

    # Rule 6 chains — DELETE → INSERT same record
    if "score_del_recreate" in df.columns:
        r6_flagged = df[df["score_del_recreate"] >= 9].copy()
        if not r6_flagged.empty and "record_id" in df.columns:
            for rid in r6_flagged["record_id"].astype(str).unique():
                related_idx = df[
                    (df["record_id"].astype(str) == rid) &
                    (df["score_del_recreate"] >= 9)
                ].index
                if len(related_idx) > 0:
                    existing = chain_id_col.loc[related_idx]
                    existing_ids = existing[existing != ""].tolist()
                    if existing_ids:
                        cid = existing_ids[0]
                    else:
                        chain_counter[0] += 1
                        cid = f"EC-{chain_counter[0]:03d}"
                    for idx in related_idx:
                        if chain_id_col.at[idx] == "":
                            chain_id_col.at[idx] = cid

    df["Event_Chain_ID"] = chain_id_col

    # ── Sequence_Context — natural language description of chain membership ───
    # Replaces EC-NNN identifiers in all reviewer-facing output.
    # Events not in a chain get an empty string (displayed as "—" in Excel).
    def _sequence_context(row):
        cid     = str(row.get("Event_Chain_ID","")).strip()
        primary = str(row.get("Primary_Rule","")).lower()
        act     = str(row.get("action_type","")).lower()
        rec_id  = str(row.get("record_id","")).strip()
        rec_type= str(row.get("record_type","")).strip()
        if not cid or cid in ("","None","nan","—"):
            return ""
        # Rule 5 chain: failed-login sequence
        if "rule 5" in primary or "failed login" in primary:
            if any(k in act for k in ["login","logon"]):
                return ("Part of failed-login sequence — this login followed "
                        "repeated failed attempts by the same user")
            return ("Part of failed-login sequence — data action occurred "
                    "within 30 minutes of repeated failed logins by this user")
        # Rule 6 chain: delete-recreate
        if "rule 6" in primary or "delete and recreate" in primary:
            ref = f" ({rec_id})" if rec_id and rec_id not in ("","nan","—") else f" ({rec_type})"
            if "delete" in act:
                return f"Part of delete-recreate sequence — original record{ref} deleted here"
            return f"Part of delete-recreate sequence — record{ref} recreated here after deletion"
        # Rule 15 chain: update-delete-insert
        if "rule 15" in primary or "suspicious" in primary:
            return ("Part of Update→Delete→Insert sequence on the same record — "
                    "see related events for the complete sequence")
        # Generic fallback for any other chain type
        return "Part of a multi-event sequence — review related events together"

    # Sequence_Context depends on Primary_Rule so it must be computed after it
    # We temporarily compute primary_rule inline here for sequencing, then
    # the master table block below will also set Primary_Rule on df.
    # We use a forward reference — chain_id_col is already set above.
    # Primary_Rule will be set by the master table; we read from chain_id_col directly.
    # Simple approach: derive context after master table populates Primary_Rule.
    # Flag for post-master-table computation:
    df["_needs_seq_ctx"] = chain_id_col.ne("").astype(int)

    # ── Composite Risk Score ──────────────────────────────────────────────────
    # Original 6 dimensions + 4 named rules, weighted
    weights = {
        "score_temporal":               0.06,
        "score_velocity":               0.07,
        "score_privilege":              0.09,
        "score_record":                 0.08,
        "score_del_recreate":           0.08,
        "score_gap":                    0.06,
        "score_rule1_vague_rationale":  0.07,
        "score_rule2_burst":            0.07,
        "score_rule3_admin_conflict":   0.10,
        "score_rule4_drift":            0.06,
        "score_rule5_failed_login":     0.08,
        "score_rule12_timestamp_reversal": 0.09,  # Critical — impossible in valid system
        "score_rule13_service_account":    0.09,  # Critical — attribution violation
        "score_rule14_dormant_account":    0.07,  # High — access control gap
        "score_rule15_suspicious_sequence":0.09,  # Critical — deliberate log evasion
        "score_rule16_first_time_behavior": 0.07,  # High — insider risk signal
    }
    df["Risk_Score"] = sum(df[c]*w for c,w in weights.items()).round(2)

    def _tier(s):
        t_crit = float(st.session_state.get("at_thresh_critical", 7.0))
        t_high = float(st.session_state.get("at_thresh_high",     5.0))
        t_med  = float(st.session_state.get("at_thresh_medium",   3.0))
        if s >= t_crit: return "Critical"
        if s >= t_high: return "High"
        if s >= t_med:  return "Medium"
        return "Low"
    df["Risk_Tier"] = df["Risk_Score"].apply(_tier)

    # ── Named rule tier overrides — named rules always win over composite ─────
    # A Rule 3 (Admin/GxP) or Rule 5 (Failed Login) at 10/10 must be Critical
    # regardless of what the composite score produces.
    # A Rule 1 (Vague) or Rule 4 (Drift) at ≥7 must be at least High.
    def _apply_tier_override(row):
        tier = row["Risk_Tier"]
        if float(row.get("score_rule3_admin_conflict",      0)) >= 8: return "Critical"
        if float(row.get("score_rule5_failed_login",        0)) >= 8: return "Critical"
        if float(row.get("score_del_recreate",              0)) >= 9: return "Critical"
        if float(row.get("score_record",                    0)) >= 10: return "Critical"
        if float(row.get("score_rule12_timestamp_reversal", 0)) >= 9: return "Critical"
        if float(row.get("score_rule13_service_account",    0)) >= 9: return "Critical"
        if float(row.get("score_rule15_suspicious_sequence",0)) >= 9: return "Critical"
        if float(row.get("score_rule16_first_time_behavior", 0)) >= 8 \
                and tier == "Low": return "High"
        if float(row.get("score_rule14_dormant_account",    0)) >= 8 and tier == "Low":
            return "High"
        if float(row.get("score_rule13_service_account",    0)) >= 6 and tier in ("Low","Medium"):
            return "High"   # Rule 13 High (non-GxP-record) must surface as at least High
        if float(row.get("score_privilege",                 0)) >= 7 and tier == "Low":
            return "High"   # Rule 8 privileged user must surface as at least High
        if float(row.get("score_rule1_vague_rationale",     0)) >= 7 and tier == "Low":
            return "High"
        if float(row.get("score_rule4_drift",               0)) >= 7 and tier == "Low":
            return "High"
        if float(row.get("score_rule2_burst",               0)) >= 6 and tier == "Low":
            return "Medium"
        return tier
    df["Risk_Tier"] = df.apply(_apply_tier_override, axis=1)

    # ── Deduplicate burst events — keep one representative per user+action burst
    # so the Top 20 isn't dominated by 11 identical Rule 2 rows
    df["_burst_key"] = (
        df["user_id"].astype(str) + "|" +
        df["action_type"].astype(str) + "|" +
        df["score_rule2_burst"].astype(str)
    )
    # Mark all but the highest-scoring row in each burst group as duplicates
    burst_mask = df["score_rule2_burst"] > 0
    if burst_mask.any():
        df["_burst_rank"] = df.groupby("_burst_key")["Risk_Score"].rank(
            method="first", ascending=False)
        df["_is_burst_dup"] = burst_mask & (df["_burst_rank"] > 1)
    else:
        df["_is_burst_dup"] = False

    # ── Triggered Rules summary column ───────────────────────────────────────
    # Lists which named rules fired for each event — useful for the Excel output
    def _triggered(row):
        fired = []
        # ── Named AI Skill Rules ──────────────────────────────────────────────
        if row.get("score_rule1_vague_rationale", 0) > 0:
            fired.append("Rule 1 — Vague Rationale [HIGH]")
        if row.get("score_rule2_burst", 0) > 0:
            fired.append("Rule 2 — Contemporaneous Burst [MEDIUM]")
        if row.get("score_rule3_admin_conflict", 0) > 0:
            fired.append("Rule 3 — Admin/GxP Conflict [CRITICAL]")
        if row.get("score_rule4_drift", 0) > 0:
            fired.append("Rule 4 — Change Control Drift [HIGH]")
        if row.get("score_rule5_failed_login", 0) > 0:
            fired.append("Rule 5 — Failed Login → Data Manipulation [CRITICAL]")
        if float(row.get("score_del_recreate", 0)) >= 9:
            fired.append("Rule 6 — Delete and Recreate Pattern [CRITICAL]")
        # ── Dimension-based rules — only fire when score is meaningful ────────
        if float(row.get("score_record", 0)) >= 10:
            fired.append("Rule 7 — Audit Trail Integrity Event [CRITICAL]")
        elif float(row.get("score_record", 0)) >= 8:
            fired.append("Rule 7 — Sensitive Record Deletion [HIGH]")
        if float(row.get("score_privilege", 0)) >= 7:
            fired.append("Rule 8 — Privileged User on GxP Data [HIGH]")
        if float(row.get("score_velocity", 0)) >= 3.5:
            fired.append("Rule 9 — High-Volume Activity Burst [MEDIUM]")
        if float(row.get("score_gap", 0)) >= 7:
            fired.append("Rule 10 — Audit Trail Timestamp Gap [HIGH]")
        if float(row.get("score_temporal", 0)) >= 9:
            fired.append("Rule 11 — Off-Hours Activity [HIGH]")
        elif float(row.get("score_temporal", 0)) >= 5:
            fired.append("Rule 11 — Off-Hours Activity [MEDIUM]")
        # ── Federal Holiday ───────────────────────────────────────────────────
        try:
            is_hol, hol_name = _is_us_federal_holiday(
                pd.Timestamp(str(row.get("timestamp",""))))
            if is_hol:
                fired.append(f"Rule 11 — Federal Holiday Activity ({hol_name})")
        except Exception:
            pass
        # ── Rules 12–14 ───────────────────────────────────────────────────────
        if float(row.get("score_rule12_timestamp_reversal", 0)) >= 9:
            fired.append("Rule 12 — Timestamp Reversal [CRITICAL]")
        if float(row.get("score_rule13_service_account", 0)) >= 9:
            fired.append("Rule 13 — Service/Shared Account GxP Action [CRITICAL]")
        elif float(row.get("score_rule13_service_account", 0)) >= 6:
            fired.append("Rule 13 — Service/Shared Account Action [HIGH]")
        if float(row.get("score_rule14_dormant_account", 0)) >= 7:
            fired.append("Rule 14 — Dormant Account Sudden Activity [HIGH]")
        if float(row.get("score_rule15_suspicious_sequence", 0)) >= 9:
            fired.append("Rule 15 — Suspicious Action Sequence [CRITICAL]")
        if float(row.get("score_rule16_first_time_behavior", 0)) >= 5:
            fired.append("Rule 16 — First-Time Behavior [HIGH]")
        return "; ".join(fired) if fired else ""
    df["Triggered_Rules"] = df.apply(_triggered, axis=1)

    # ══════════════════════════════════════════════════════════════════════════
    # MASTER RULE TABLE — single source of truth for ALL derived fields.
    #
    # Every field that depends on "which rule is primary" consults THIS table
    # in THIS order. No separate priority lists, no separate dicts.
    #
    # Columns:
    #   score_col    — the DataFrame column to test
    #   threshold    — minimum score for this entry to fire
    #   label        — human-readable rule name (used in Primary_Rule)
    #   evidence     — "High" | "Medium" | "Low" (used in Evidence_Strength)
    #   reg_basis    — citation-only string (used in Regulatory_Basis / Why It Matters)
    #   action       — procedural instruction (used in Action_Required)
    # ══════════════════════════════════════════════════════════════════════════
    _MASTER = [
        # ── TIER 1: Structural / Attribution failures ─────────────────────────
        (
            "score_rule15_suspicious_sequence", 9,
            "Rule 15 — Suspicious Action Sequence [CRITICAL]",
            "High",
            ("ALCOA+ Original principle; 21 CFR Part 11 §11.10(e) — "
             "GxP records must reflect what was originally observed and must not be "
             "altered after the fact without a complete, auditable change history."),
            ("Retrieve all three events in this sequence — the UPDATE, DELETE, and "
             "INSERT on the same record. Compare values before and after to identify "
             "what changed. Obtain a written explanation from the user for each step. "
             "If the deletion circumvented the locked record workflow, "
             "initiate a formal data integrity investigation."),
        ),
        (
            "score_rule12_timestamp_reversal", 9,
            "Rule 12 — Timestamp Reversal [CRITICAL]",
            "High",
            ("21 CFR Part 11 §11.10(e); FDA Data Integrity Guidance (2018) — "
             "Audit trail timestamps must be sequentially consistent; "
             "an approval timestamp cannot precede a creation timestamp."),
            ("Retrieve the full audit trail for this record and compare all creation "
             "and approval timestamps. Determine whether a server clock error, system "
             "migration, or direct database alteration caused the reversal. "
             "Document all findings formally."),
        ),
        (
            "score_rule13_service_account", 9,
            "Rule 13 — Service/Shared Account GxP Action [CRITICAL]",
            "High",
            ("21 CFR Part 11 §11.300; ALCOA+ Attributable principle — "
             "Each GxP action must be traceable to a single, identified individual; "
             "non-personal accounts cannot satisfy this requirement."),
            ("Identify the individual responsible for authorising this service or "
             "shared account to perform this action. Obtain written justification "
             "for use of a non-personal account on GxP data. Assess whether the "
             "data modified can be attributed to a specific individual as required."),
        ),
        (
            "score_rule5_failed_login", 8,
            "Rule 5 — Failed Login → Data Manipulation [CRITICAL]",
            "High",
            ("ALCOA+ Original and Attributable principles; FDA Data Integrity Guidance (2018) — "
             "GxP systems must ensure that only authorised individuals can access "
             "and modify controlled records."),
            ("Initiate a data integrity investigation immediately. Obtain all failed "
             "login records for this user. Cross-reference every data change made "
             "in the 30 minutes following login against source documents. "
             "If manipulation is confirmed without authorisation, raise a Critical "
             "non-conformance."),
        ),
        (
            "score_rule3_admin_conflict", 8,
            "Rule 3 — Admin/GxP Conflict [CRITICAL]",
            "High",
            ("21 CFR Part 11 §11.10(d); Segregation of Duties principle — "
             "Administrative accounts are authorised for system configuration only "
             "and must not directly create or modify GxP production records."),
            ("Obtain documented business justification for this administrative action "
             "on production data. Verify whether an Emergency Access Request was "
             "approved prior to this action. If no documented authorisation exists, "
             "assess impact on GxP record integrity and raise a non-conformance."),
        ),
        (
            "score_del_recreate", 9,
            "Rule 6 — Delete and Recreate Pattern [CRITICAL]",
            "High",
            ("ALCOA+ Original principle; 21 CFR Part 11 §11.10(e) — "
             "GxP data must not be altered by deleting and recreating records "
             "with different values. A DELETE followed by INSERT on the same "
             "record ID replaces the original data with potentially altered "
             "values, breaking the traceability chain the audit trail requires."),
            ("Retrieve both the original deleted record and the recreated record. "
             "Compare all field values for discrepancies. Obtain a retrospective "
             "written explanation from the user. If the change cannot be justified, "
             "initiate a data integrity investigation."),
        ),
        (
            "score_record", 10,
            "Rule 7 — Audit Trail Integrity Event [CRITICAL]",
            "High",
            ("21 CFR Part 11 §11.10(e); EU Annex 11 Clause 9 — "
             "Audit trail systems must be protected from modification; "
             "any change to audit trail configuration is a critical integrity event."),
            ("This action affected the audit trail system itself. Retrieve the full "
             "system configuration log and determine the scope of the change. "
             "Assess whether any events may have been suppressed during the affected "
             "period. Document all findings and escalate immediately."),
        ),
        # ── TIER 2: High-risk patterns ────────────────────────────────────────
        (
            "score_rule4_drift", 7,
            "Rule 4 — Change Control Drift [HIGH]",
            "Medium",
            ("21 CFR Part 820.70(b); FDA Data Integrity Guidance (2018) — "
             "Validated system parameters must remain within approved specifications; "
             "deviations require documented Change Control authorisation."),
            ("Cross-reference the recorded value against the approved specification "
             "or validated setpoint. Verify whether a formal Change Control was "
             "authorised before this value was applied. If not, assess impact on "
             "the validated state."),
        ),
        (
            "score_rule1_vague_rationale", 7,
            "Rule 1 — Vague Rationale [HIGH]",
            "Medium",
            ("ALCOA+ Attributable and Legible principles; FDA Data Integrity Guidance (2018) — "
             "GxP data changes must be accompanied by a specific, scientifically "
             "justified rationale attributable to the performing individual."),
            ("Obtain a retrospective written amendment from the analyst explaining "
             "what changed and why. Assess whether the undocumented change could "
             "affect the quality or disposition of the associated batch or result."),
        ),
        (
            "score_rule16_first_time_behavior", 8,
            "Rule 16 — First-Time Behavior [HIGH]",
            "Low",
            ("21 CFR Part 11 §11.10(d); ALCOA+ Attributable principle — "
             "Each GxP action must be within the performing user's approved, "
             "documented access rights at the time the action was taken."),
            ("Verify this action was within the user's approved access rights at the "
             "time it was performed. Obtain written confirmation from the user's "
             "supervisor that this action type was within their authorised role. "
             "If the user was not authorised, assess all affected records and raise "
             "a non-conformance."),
        ),
        (
            "score_rule14_dormant_account", 8,
            "Rule 14 — Dormant Account Sudden Activity [HIGH]",
            "Medium",
            ("21 CFR Part 11 §11.10(d) — "
             "User access rights must be reviewed periodically and must be formally "
             "re-authorised following extended periods of inactivity."),
            ("Verify the current employment and access authorisation status of this "
             "user. Confirm whether access was formally reviewed and re-approved "
             "before this re-activation. If no re-authorisation record exists, "
             "assess this event for data integrity impact."),
        ),
        (
            "score_record", 8,
            "Rule 7 — Sensitive Record Deletion [HIGH]",
            "High",
            ("21 CFR Part 11 §11.10(e); ALCOA+ Original principle — "
             "Deletions of GxP records must be fully justified, authorised, "
             "and traceable in the audit trail."),
            ("Review the event against the source document for that record. "
             "Obtain written justification from the performing user. "
             "If no authorisation can be demonstrated, escalate to a formal "
             "non-conformance."),
        ),
        (
            "score_privilege", 7,
            "Rule 8 — Privileged User on GxP Data [HIGH]",
            "Medium",
            ("21 CFR Part 11 §11.10(d) — "
             "Privileged accounts must be restricted to their authorised purpose; "
             "direct modification of production GxP records is outside that scope."),
            ("Verify that the use of this privileged account was consistent with its "
             "authorised purpose. Administrative accounts must be used for system "
             "configuration only and must not directly create or modify production "
             "records."),
        ),
        # ── TIER 3: Statistical / Behavioural ────────────────────────────────
        (
            "score_rule2_burst", 6,
            "Rule 2 — Contemporaneous Burst [MEDIUM]",
            "Low",
            ("ALCOA+ Contemporaneous principle; FDA Data Integrity Guidance (2018) — "
             "GxP data entries must be recorded at the time of the activity, "
             "not retrospectively from memory or paper records."),
            ("Verify that contemporaneous source data exists — instrument printouts "
             "or laboratory worksheets — confirming each entry was recorded in real "
             "time. If entries were transcribed after the fact, investigate as a "
             "data integrity concern."),
        ),
        (
            "score_gap", 7,
            "Rule 10 — Audit Trail Timestamp Gap [HIGH]",
            "Low",
            ("21 CFR Part 11 §11.10(e) — "
             "Audit trails must be continuous and computer-generated; "
             "gaps in coverage must be explained and documented."),
            ("Investigate whether the audit trail was suspended during the gap "
             "period. Document findings. If no legitimate explanation exists, "
             "treat this as a critical audit trail integrity finding."),
        ),
        (
            "score_velocity", 3.5,
            "Rule 9 — High-Volume Activity Burst [MEDIUM]",
            "Low",
            ("ALCOA+ Contemporaneous principle; FDA Data Integrity Guidance (2018) — "
             "High-volume data entry in a short window must be supported by "
             "contemporaneous source records confirming real-time recording."),
            ("Verify that contemporaneous source data exists to confirm each entry "
             "was recorded in real time. If entries appear to have been made in bulk "
             "or retrospectively, assess whether data integrity has been compromised."),
        ),
        (
            "score_temporal", 5,
            "Rule 11 — Off-Hours/Holiday Activity [MEDIUM]",
            "Low",
            ("21 CFR Part 11 §11.10(e); FDA Data Integrity Guidance (2018) — "
             "Activity on GxP systems outside approved working hours must be "
             "covered by a documented maintenance window or approved overtime record."),
            ("Obtain documented business justification for this activity outside "
             "normal working hours. Verify whether an approved overtime record or "
             "scheduled maintenance window covers this period."),
        ),
        (
            "score_rule13_service_account", 6,
            "Rule 13 — Service/Shared Account Action [HIGH]",
            "Medium",
            ("21 CFR Part 11 §11.300; ALCOA+ Attributable principle — "
             "Each GxP action must be traceable to a single, identified individual; "
             "non-personal accounts cannot satisfy this requirement."),
            ("Verify the individual responsible for this service or shared account "
             "action. Obtain written justification for the use of a non-personal "
             "account and confirm the action was authorised."),
        ),
        (
            "score_rule16_first_time_behavior", 5,
            "Rule 16 — First-Time Behavior [HIGH]",
            "Low",
            ("21 CFR Part 11 §11.10(d); ALCOA+ Attributable principle — "
             "Each GxP action must be within the performing user's approved, "
             "documented access rights at the time the action was taken."),
            ("Verify this action was within the user's approved access rights. "
             "If the user was not authorised to perform this action type, "
             "assess all affected records and raise a non-conformance."),
        ),
    ]

    # ── Derive all rule-dependent fields from the master table ────────────────
    # One loop, one priority order, guaranteed consistency across all fields.

    def _master_lookup(row):
        """
        Walk _MASTER in order. Return the first entry whose score_col >= threshold.
        Returns the full tuple so callers can extract any field without re-walking.
        Returns None if no rule fires.
        """
        for entry in _MASTER:
            score_col, threshold = entry[0], entry[1]
            if float(row.get(score_col, 0)) >= threshold:
                return entry
        return None

    def _primary_rule(row):
        m = _master_lookup(row)
        return m[2] if m else "Composite risk score — no single rule dominant"

    def _evidence_strength(row):
        m = _master_lookup(row)
        return m[3] if m else "Low"

    def _reg_basis(row):
        m = _master_lookup(row)
        return m[4] if m else (
            "No named data integrity risk indicator detected at a significant level.")

    def _action_req(row):
        m = _master_lookup(row)
        return m[5] if m else (
            "Review this event against source documentation and obtain a written "
            "justification from the performing user if the reason for the action "
            "is not already documented.")

    def _supporting_signals(row):
        primary = _primary_rule(row)
        clean = lambda s: (s.replace(" [CRITICAL]","").replace(" [HIGH]","")
                            .replace(" [MEDIUM]","").replace(" [LOW]",""))
        primary_clean = clean(primary)
        all_rules = [r.strip() for r in str(row.get("Triggered_Rules","")).split(";")
                     if r.strip()]
        supporting = [clean(r) for r in all_rules
                      if clean(r) != primary_clean and r]
        return "; ".join(supporting) if supporting else "—"

    df["Primary_Rule"]       = df.apply(_primary_rule, axis=1)
    df["Supporting_Signals"] = df.apply(_supporting_signals, axis=1)
    df["Evidence_Strength"]  = df.apply(_evidence_strength, axis=1)
    df["Regulatory_Basis"]   = df.apply(_reg_basis, axis=1)
    df["Action_Required"]    = df.apply(_action_req, axis=1)

    # ── Sequence_Context — computed after Primary_Rule is available ───────────
    df["Sequence_Context"] = df.apply(_sequence_context, axis=1)
    # Primary Rule = the named rule that drove the tier classification.
    # Derived from the same priority order as _apply_tier_override so the two
    # are always consistent. Supporting Signals = all remaining triggered rules.
    _RULE_PRIORITY = [
        ("score_rule15_suspicious_sequence", 9,  "Rule 15 — Suspicious Action Sequence [CRITICAL]"),
        ("score_rule12_timestamp_reversal",  9,  "Rule 12 — Timestamp Reversal [CRITICAL]"),
        ("score_rule13_service_account",     9,  "Rule 13 — Service/Shared Account GxP Action [CRITICAL]"),
        ("score_rule5_failed_login",         8,  "Rule 5 — Failed Login → Data Manipulation [CRITICAL]"),
        ("score_rule3_admin_conflict",       8,  "Rule 3 — Admin/GxP Conflict [CRITICAL]"),
        ("score_del_recreate",               9,  "Rule 6 — Delete and Recreate Pattern [CRITICAL]"),
        ("score_record",                     10, "Rule 7 — Audit Trail Integrity Event [CRITICAL]"),
        ("score_rule4_drift",                7,  "Rule 4 — Change Control Drift [HIGH]"),
        ("score_rule1_vague_rationale",      7,  "Rule 1 — Vague Rationale [HIGH]"),
        ("score_rule16_first_time_behavior", 8,  "Rule 16 — First-Time Behavior [HIGH]"),
        ("score_rule14_dormant_account",     8,  "Rule 14 — Dormant Account Sudden Activity [HIGH]"),
        ("score_rule2_burst",                6,  "Rule 2 — Contemporaneous Burst [MEDIUM]"),
        ("score_privilege",                  7,  "Rule 8 — Privileged User on GxP Data [HIGH]"),
        ("score_record",                     8,  "Rule 7 — Sensitive Record Deletion [HIGH]"),
        ("score_velocity",                   3.5,"Rule 9 — High-Volume Activity Burst [MEDIUM]"),
        ("score_gap",                        7,  "Rule 10 — Audit Trail Timestamp Gap [HIGH]"),
        ("score_temporal",                   5,  "Rule 11 — Off-Hours/Holiday Activity [MEDIUM]"),
    ]

    def _primary_rule(row):
        for score_col, threshold, label in _RULE_PRIORITY:
            if float(row.get(score_col, 0)) >= threshold:
                return label
        # Fallback: highest scoring dimension
        return "Composite risk score — no single rule dominant"

    def _supporting_signals(row):
        primary = _primary_rule(row)
        all_rules = [r.strip() for r in str(row.get("Triggered_Rules","")).split(";")
                     if r.strip()]
        # Remove severity labels for cleaner display
        clean = lambda s: s.replace(" [CRITICAL]","").replace(" [HIGH]","").replace(" [MEDIUM]","").replace(" [LOW]","")
        primary_clean = clean(primary)
        supporting = [clean(r) for r in all_rules
                      if clean(r) != primary_clean and r]
        return "; ".join(supporting) if supporting else "—"

    # ── Combined rationale — includes dimension rationales ────────────────────
    def _dim_rationale(row) -> str:
        """Build plain-English rationale for dimension-based findings."""
        parts = []
        rec_s = float(row.get("score_record", 0))
        pri_s = float(row.get("score_privilege", 0))
        vel_s = float(row.get("score_velocity", 0))
        gap_s = float(row.get("score_gap", 0))
        tmp_s = float(row.get("score_temporal", 0))
        usr   = str(row.get("user_id","Unknown"))
        act   = str(row.get("action_type","Unknown"))
        rec   = str(row.get("record_type","Unknown"))
        ts    = str(row.get("timestamp","Unknown"))

        if rec_s >= 10:
            parts.append(
                f"Rule 7 — Audit Trail Integrity Event [CRITICAL]: "
                f"Action '{act}' was performed on audit trail configuration or records. "
                "Any modification to the audit trail system is a critical data integrity finding "
                "requiring immediate investigation (21 CFR Part 11 §11.10(e))."
            )
        elif rec_s >= 8:
            parts.append(
                f"Rule 7 — Sensitive Record Deletion [HIGH]: "
                f"A deletion was performed on a GxP-critical record type ('{rec}'). "
                "Deletions of GxP records must be fully justified and authorised "
                "(21 CFR Part 11 §11.10(e), ALCOA+ Original)."
            )
        if pri_s >= 7:
            parts.append(
                f"Rule 8 — Privileged User on GxP Data [HIGH]: "
                f"User '{usr}' holds an administrative or privileged role and performed "
                f"'{act}' on '{rec}'. Privileged accounts must be restricted to system "
                "configuration only and must not directly modify production data "
                "(21 CFR Part 11 §11.10(d))."
            )
        if vel_s >= 3.5:
            parts.append(
                f"Rule 9 — High-Volume Activity Burst [MEDIUM]: "
                f"User '{usr}' performed the same action repeatedly in a short time window. "
                "This pattern may indicate automated or retrospective data entry rather than "
                "real-time recording, which may be inconsistent with the ALCOA+ Contemporaneous principle as described in FDA Data Integrity Guidance (2018)."
            )
        if gap_s >= 7:
            parts.append(
                f"Rule 10 — Audit Trail Timestamp Gap [HIGH]: "
                f"A gap of more than 2 hours was detected in the audit trail before this event at {ts}. "
                "Continuous audit trail coverage is required — gaps may indicate logging was "
                "suspended during that period (21 CFR Part 11 §11.10(e))."
            )
        if tmp_s >= 5:
            try:
                is_hol, hol_name = _is_us_federal_holiday(pd.Timestamp(ts))
            except Exception:
                is_hol, hol_name = False, ""
            if is_hol:
                parts.append(
                    f"Rule 11 — Federal Holiday Activity [{hol_name}]: "
                    f"User '{usr}' performed '{act}' on a US Federal Holiday. "
                    "Activity on scheduled non-working days requires documented business "
                    "justification and is a classic indicator of unauthorised shadow activity."
                )
            else:
                parts.append(
                    f"Rule 11 — Off-Hours Activity: "
                    f"User '{usr}' performed '{act}' at {ts}, outside normal business hours. "
                    "Off-hours activity on GxP records must be justified by an approved "
                    "overtime record or maintenance window."
                )
        return " | ".join(parts)

    def _combined_rat(row):
        parts = [r for r in [
            row.get("rule1_rationale",""),
            row.get("rule2_rationale",""),
            row.get("rule3_rationale",""),
            row.get("rule4_rationale",""),
            row.get("rule5_rationale",""),
            row.get("rule12_rationale",""),
            row.get("rule13_rationale",""),
            row.get("rule14_rationale",""),
            row.get("rule15_rationale",""),
            row.get("rule16_rationale",""),
            _dim_rationale(row),
        ] if r]
        return " | ".join(parts)
    df["Rule_Rationale"] = df.apply(_combined_rat, axis=1)

    # ── Suggested Disposition — fully deterministic 7-tier engine ────────────
    # Single source of truth: _MASTER table determines Primary_Rule, and the
    # same priority logic drives tier selection here. No separate lists.
    #
    # Comment-gate rule: comment presence downgrades Escalate → Investigate
    # ONLY for documentation-gap rules (R1, R4, R8, Rr-deletion, off-hours).
    # It NEVER downgrades structural findings (R3, R5, R6, R12, R13, R15, Rr≥10).
    #
    # Gap timestamp rule: is_biz defaults to False on parse failure —
    # unknown timestamps never auto-escalate; they route to Investigate.
    #
    # Decision table (strict priority order — first match wins):
    # ┌──────────────────────────────────────────────┬──────────────────────────┐
    # │ Condition                                    │ Disposition              │
    # ├──────────────────────────────────────────────┼──────────────────────────┤
    # │ R15≥9  Update→Delete→Insert sequence        │ Escalate to CAPA         │
    # │ R12≥9  Approval before creation timestamp   │ Escalate to CAPA         │
    # │ R13≥9  Service/shared account on GxP record │ Escalate to CAPA         │
    # │ R5≥8   Failed logins → GxP data action      │ Escalate to CAPA         │
    # │ R3≥8   Admin account on production GxP data │ Escalate to CAPA         │
    # │ R6≥9   Delete then recreate same record      │ Escalate to CAPA         │
    # │ Rr≥10  Audit trail config modified           │ Escalate to CAPA         │
    # │ Rr≥8 + no comment   GxP record deletion     │ Escalate to CAPA         │
    # │ Rr≥8 + comment      GxP record deletion     │ Investigate              │
    # │ Rg≥7 + business hrs  Audit trail gap        │ Escalate to CAPA         │
    # │ Rg≥7 + other/unknown Audit trail gap        │ Investigate              │
    # │ R4≥7 + no comment   Value drift             │ Escalate to CAPA         │
    # │ R4≥7 + comment      Value drift             │ Investigate              │
    # │ Rp≥7 + no comment   Privileged user         │ Escalate to CAPA         │
    # │ Rp≥7 + comment      Privileged user         │ Investigate              │
    # │ R1≥8 + no comment   Blank GxP change reason │ Escalate to CAPA         │
    # │ R1≥6  Vague/insufficient change reason      │ Amendment Required       │
    # │ R2≥6  Contemporaneous burst                 │ Investigate              │
    # │ R14≥7 Dormant account re-activation         │ Investigate              │
    # │ R16≥5 First-time high-risk action           │ Investigate              │
    # │ Rt≥9 + comment      Deep off-hours          │ Document Rationale       │
    # │ Rt≥9 + no comment   Deep off-hours          │ Investigate              │
    # │ Rt≥5 + comment      Off-hours               │ Document Rationale       │
    # │ Rt≥5 + no comment   Off-hours               │ Investigate              │
    # │ R13≥6 Service account (non-GxP record)      │ Investigate              │
    # │ named_max≥7.0  Any high-level named rule     │ Investigate              │
    # │ default        No significant named rule     │ No Action Required       │
    # └──────────────────────────────────────────────┴──────────────────────────┘

    def _suggested_disposition(row) -> tuple:
        """Return (disposition_label, rationale_text) — reviewer-facing language only."""
        r3  = float(row.get("score_rule3_admin_conflict",      0))
        r5  = float(row.get("score_rule5_failed_login",        0))
        r1  = float(row.get("score_rule1_vague_rationale",     0))
        r4  = float(row.get("score_rule4_drift",               0))
        r6  = float(row.get("score_del_recreate",              0))
        rg  = float(row.get("score_gap",                       0))
        rr  = float(row.get("score_record",                    0))
        rp  = float(row.get("score_privilege",                 0))
        rt  = float(row.get("score_temporal",                  0))
        r12 = float(row.get("score_rule12_timestamp_reversal",  0))
        r13 = float(row.get("score_rule13_service_account",     0))
        r14 = float(row.get("score_rule14_dormant_account",     0))
        r15 = float(row.get("score_rule15_suspicious_sequence", 0))
        r16 = float(row.get("score_rule16_first_time_behavior", 0))
        r2  = float(row.get("score_rule2_burst",                0))
        cmt     = str(row.get("comments","")).lower().strip()
        has_cmt = bool(cmt and cmt not in ("nan","none","-","—",""))

        # TIER 1 — Structural: always Escalate regardless of documentation
        if r15 >= 9:
            return ("Escalate to CAPA",
                    "Update, Delete, and re-Insert were performed on the same record "
                    "within 30 minutes — this sequence is the primary method for "
                    "altering locked GxP records while obscuring the original data.")
        if r12 >= 9:
            return ("Escalate to CAPA",
                    "Approval or release timestamp precedes the creation timestamp "
                    "on the same record — this is chronologically impossible in a "
                    "correctly functioning system and requires immediate investigation.")
        if r13 >= 9:
            return ("Escalate to CAPA",
                    "A service or shared account directly modified a GxP record — "
                    "this action cannot be attributed to a named individual as required.")
        if r5 >= 8:
            return ("Escalate to CAPA",
                    "Three or more failed login attempts preceded a GxP data action "
                    "within 30 minutes — this sequence requires investigation for "
                    "potential unauthorised access.")
        if r3 >= 8:
            return ("Escalate to CAPA",
                    "An administrative account directly modified production GxP data — "
                    "administrative accounts are authorised for system configuration "
                    "only, not direct data modification.")
        if r6 >= 9:
            return ("Escalate to CAPA",
                    "The same record was deleted then recreated — this pattern may "
                    "replace original GxP data with altered values, breaking the "
                    "traceability chain.")
        if rr >= 10:
            return ("Escalate to CAPA",
                    "The audit trail system itself was modified — any change to "
                    "audit trail configuration requires immediate investigation.")

        # TIER 2 — Destructive/high-risk: documentation present downgrades to Investigate
        if rr >= 8:
            if has_cmt:
                return ("Investigate — Verify Source Data",
                        "A GxP-sensitive record was deleted with a comment on file — "
                        "verify the comment constitutes adequate justification and "
                        "the deletion was formally authorised.")
            return ("Escalate to CAPA",
                    "A GxP-sensitive record was deleted with no documented "
                    "justification — authorisation must be established before "
                    "this event can be closed.")

        if rg >= 7:
            try:
                gap_ts = pd.Timestamp(str(row.get("timestamp", "")))
                is_biz = (not pd.isnull(gap_ts)
                          and gap_ts.weekday() < 5
                          and _AT_BIZ_START <= gap_ts.hour < _AT_BIZ_END)
            except Exception:
                is_biz = False
            if is_biz:
                return ("Escalate to CAPA",
                        "An unexplained gap in audit trail coverage occurred during "
                        "normal business hours — continuous logging is required and "
                        "any pause during working hours requires investigation.")
            return ("Investigate — Verify Source Data",
                    "A gap in audit trail coverage was detected — verify whether "
                    "this aligns with an approved maintenance window or scheduled "
                    "system downtime.")

        if r4 >= 7:
            if has_cmt:
                return ("Investigate — Verify Source Data",
                        "The recorded value is significantly outside the expected range "
                        "for this record type, with a comment on file — verify the "
                        "comment references an approved Change Control or specification.")
            return ("Escalate to CAPA",
                    "The recorded value is significantly outside the expected range "
                    "with no documented justification — verify against the approved "
                    "specification before this event can be closed.")

        if rp >= 7:
            if has_cmt:
                return ("Investigate — Verify Source Data",
                        "A privileged account acted on GxP data with a comment on file — "
                        "verify the comment constitutes adequate business justification "
                        "and that an Emergency Access Request was approved if required.")
            return ("Escalate to CAPA",
                    "A privileged account modified GxP production data with no "
                    "documented justification — authorisation must be established.")

        # TIER 3 — Documentation gap
        if r1 >= 8 and not has_cmt:
            return ("Escalate to CAPA",
                    "A GxP data modification was recorded with no change reason — "
                    "retrospective justification is required before this event "
                    "can be closed.")
        if r1 >= 6:
            return ("Amendment Required",
                    "The change reason recorded is insufficient or uses "
                    "non-descriptive language — a retrospective written amendment "
                    "from the analyst is required.")

        # TIER 4 — Statistical/behavioural
        # Note: r14 checked BEFORE rg to prevent gap score from overriding
        # dormant account as the disposition driver for long-inactivity events.
        if r14 >= 7:
            return ("Investigate — Verify Source Data",
                    "This account had no recorded activity for 90 or more days "
                    "before this GxP action — verify current employment status "
                    "and confirm access was formally re-authorised.")
        if r2 >= 6:
            return ("Investigate — Verify Source Data",
                    "More than ten data entries were recorded by the same user "
                    "within 15 minutes — verify that contemporaneous source "
                    "records exist for each entry.")
        if r16 >= 5:
            return ("Investigate — Verify Source Data",
                    "This user performed a high-risk action type for the first "
                    "time in their recorded history — verify this was within "
                    "their approved access rights at the time.")

        # TIER 5 — Temporal: all four combinations explicit
        if rt >= 9:
            return (
                "Document Rationale" if has_cmt else "Investigate — Verify Source Data",
                ("Activity at an unusually late or early hour was detected with a "
                 "comment on file — confirm a corresponding approved overtime record "
                 "or maintenance window covers this period.")
                if has_cmt else
                ("Activity at an unusually late or early hour was detected with no "
                 "documented reason — obtain business justification before closing "
                 "this finding.")
            )
        if rt >= 5:
            return (
                "Document Rationale" if has_cmt else "Investigate — Verify Source Data",
                ("Off-hours activity was detected with a comment on file — confirm "
                 "a corresponding approved overtime record or maintenance window "
                 "covers this period.")
                if has_cmt else
                ("Off-hours activity was detected with no documented reason — "
                 "obtain business justification before closing this finding.")
            )

        # TIER 6 — Service account, non-GxP record
        if r13 >= 6:
            return ("Investigate — Verify Source Data",
                    "A service or shared account performed this action — verify "
                    "that a responsible individual can be identified and that "
                    "the action was authorised.")

        # TIER 7 — Hard gate
        named_max = max(r16,r15,r12,r13,r5,r3,r6,rg,rr,r4,r1,r2,rp,rt,r14)
        if named_max >= 7.0:
            return ("Investigate — Verify Source Data",
                    "A risk indicator was detected that warrants documented "
                    "reviewer investigation before this event can be closed.")

        return ("No Action Required",
                "No significant risk indicator was detected — a brief review "
                "and documented disposition is sufficient.")

    sugg_disp = []
    sugg_rat  = []
    for _, row in df.iterrows():
        d, r = _suggested_disposition(row)
        sugg_disp.append(d)
        sugg_rat.append(r)
    df["Suggested_Disposition"]          = sugg_disp
    df["Suggested_Disposition_Rationale"] = sugg_rat

    return df.sort_values("Risk_Score", ascending=False).reset_index(drop=True)


def _at_deterministic_justification(row: dict) -> str:
    """
    Builds a single factual sentence for the System Narrative column.
    States only observable log data: who, what, which record, when, comment.
    No regulatory language. No action instructions. No inference.
    Used as fallback when LLM is unavailable — output is format-identical.
    """
    user     = str(row.get("user_id",     "unknown user"))
    action   = str(row.get("action_type", "unknown action"))
    rec_type = str(row.get("record_type", ""))
    rec_id   = str(row.get("record_id",   ""))
    ts       = str(row.get("timestamp",   "unknown time"))
    cmt      = str(row.get("comments",    "")).strip()

    # Build record reference
    if rec_id and rec_id not in ("", "nan", "—", "None"):
        record_ref = f"{rec_type}/{rec_id}".strip("/")
    elif rec_type and rec_type not in ("", "nan", "—"):
        record_ref = rec_type
    else:
        record_ref = "an unspecified record"

    # Build comment clause
    cmt_clean = cmt.strip() if cmt and cmt not in ("nan", "none", "-", "—", "") else ""
    if cmt_clean:
        comment_clause = f"; comment on file reads '{cmt_clean[:60]}'"
    else:
        comment_clause = "; no change reason was recorded"

    # Build chain context from natural language Sequence_Context field
    seq_ctx      = str(row.get("Sequence_Context", "")).strip()
    chain_clause = f" — {seq_ctx}" if seq_ctx else ""

    return f"{user} performed {action} on {record_ref} at {ts}{comment_clause}{chain_clause}."


def at_generate_justifications(top_df: pd.DataFrame, model_id: str) -> pd.DataFrame:
    """
    Generates a 3-sentence GxP justification for each escalated event.
    Primary path: LLM via litellm.
    Fallback: deterministic Python justification — same format, no visible error.
    The client never sees an error string in the output regardless of LLM status.
    """
    justifications = []
    total = len(top_df)

    for rank, (_, row) in enumerate(top_df.iterrows(), 1):
        text = None

        # ── Primary: LLM ──────────────────────────────────────────────────────
        try:
            from litellm import completion as _comp
            triggered = str(row.get("Triggered_Rules","None"))
            rule_rat  = str(row.get("Rule_Rationale","")).split(" | ")[0][:250]
            sugg_disp = str(row.get("Suggested_Disposition",""))
            record_id = str(row.get("record_id","unknown record"))
            rec_type  = str(row.get("record_type","unknown type"))
            usr       = str(row.get("user_id","unknown user"))
            act       = str(row.get("action_type","unknown action"))
            cmt       = str(row.get("comments","none provided"))
            ts        = str(row.get("timestamp","unknown time"))

            seq_ctx   = str(row.get("Sequence_Context","")).strip()
            chain_ctx = f"  Sequence context: {seq_ctx}\n" if seq_ctx else ""

            # Determine primary rule to anchor the narrative
            all_rules = [r.strip() for r in triggered.split(";") if r.strip()]
            primary_rule = all_rules[0] if all_rules else "Risk indicator detected"

            prompt = f"""You are writing the System Narrative column in a GxP audit trail review table.

YOUR ONLY JOB: Write exactly ONE sentence stating the observable facts from the log.

HARD RULES — no exceptions:
1. State ONLY what the log record shows: who, what action, on which record, at what time, what comment was recorded.
2. If Sequence context is provided below, you MUST include it in your sentence — it is factual log context.
3. Do NOT use any regulatory language: no "may indicate", "raises a concern", "is inconsistent with", "warrants", "ALCOA", "21 CFR", "data integrity".
4. Do NOT recommend any action: no "verify", "confirm", "obtain", "investigate", "escalate".
5. Do NOT infer intent or motivation.
6. ONE sentence. Max 40 words.

Log data:
  User: {usr}
  Action: {act}
  Record type: {rec_type}
  Record ID: {record_id}
  Timestamp: {ts}
  Comment on file: "{cmt}"
{chain_ctx}
CORRECT example: "analyst_x performed DELETE on RESULTS/RES-042 at 02:14 on 15-Mar-2026; no change reason was recorded."
CORRECT example: "svc_batch executed APPROVE on BATCH_RELEASE/BR-117 at 09:32 on 22-Jan-2026; comment reads 'release ok'."
WRONG example: "This action may indicate unauthorised access to a GxP record." [contains regulatory language]
WRONG example: "Verify the user's access rights before closing this finding." [contains action instruction]

Write only the one sentence. No labels, no preamble, no explanation."""

            resp = _comp(
                model=model_id, stream=False, temperature=0.05, max_tokens=80,
                messages=[
                    {"role": "system", "content":
                     "You write one-sentence factual log summaries for pharmaceutical QA tables. "
                     "State only observable facts: who, what, which record, when, what comment. "
                     "No regulatory language. No action instructions. One sentence, max 30 words."},
                    {"role": "user", "content": prompt}
                ]
            )
            candidate = resp.choices[0].message.content.strip()
            if len(candidate) > 40 and "error" not in candidate.lower()[:30]:
                text = candidate

        except Exception:
            pass   # silently fall through to deterministic fallback

        # ── Fallback: deterministic Python justification ───────────────────────
        if not text:
            text = _at_deterministic_justification(row.to_dict())

        justifications.append(text)

    top_df = top_df.copy()
    top_df["AI_Justification"] = justifications  # internal key unchanged for compatibility
    return top_df


def at_build_excel(top_df, scored_df, system_name, r_start, r_end, fname) -> bytes:
    """
    Build a clean, professional evidence workbook for QA reviewers and auditors.
    White background, dark text, colour only on Risk Level cells.
    Three sheets: Cover & Summary | Events for Review | Full Audit Log
    """
    from openpyxl import Workbook
    output = io.BytesIO()
    wb     = Workbook()

    # ── Colour palette — professional, printable ──────────────────────────────
    C_HEADER_BG  = "1E3A5F"   # dark navy for header rows
    C_HEADER_FG  = "FFFFFF"
    C_SECTION_BG = "EBF3FB"   # very light blue for section dividers
    C_SECTION_FG = "1E3A5F"
    C_LABEL_FG   = "374151"   # dark grey for labels
    C_VALUE_FG   = "111827"   # near-black for values
    C_ALT_ROW    = "F9FAFB"   # very light grey alternating rows
    C_WHITE      = "FFFFFF"

    # Risk tier colours — background / text (light pastels, readable when printed)
    TIER_BG = {
        "Critical": "FECACA",   # soft red
        "High":     "FED7AA",   # soft orange
        "Medium":   "FEF08A",   # soft yellow
        "Low":      "DCFCE7",   # soft green
    }
    TIER_FG = {
        "Critical": "991B1B",
        "High":     "9A3412",
        "Medium":   "854D0E",
        "Low":      "166534",
    }

    thin  = Side(style="thin",  color="D1D5DB")
    thick = Side(style="medium", color="1E3A5F")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    bdr_t = Border(left=thick, right=thick, top=thick, bottom=thick)

    def _hdr_font(size=10, bold=True, color=C_HEADER_FG):
        return Font(bold=bold, color=color, name="Calibri", size=size)
    def _body_font(size=10, bold=False, color=C_VALUE_FG):
        return Font(bold=bold, color=color, name="Calibri", size=size)
    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    t_crit = float(st.session_state.get("at_thresh_critical", 7.0))
    t_high = float(st.session_state.get("at_thresh_high",     5.0))
    t_med  = float(st.session_state.get("at_thresh_medium",   3.0))

    total     = len(scored_df)
    n_esc     = len(top_df)
    n_crit    = int((scored_df["Risk_Tier"]=="Critical").sum())
    n_high    = int((scored_df["Risk_Tier"]=="High").sum())
    n_med     = int((scored_df["Risk_Tier"]=="Medium").sum())
    n_low     = int((scored_df["Risk_Tier"]=="Low").sum())
    pct_clear = round((total-n_esc)/total*100,1) if total>0 else 0

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Cover & Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "1E3A5F"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 52
    ws.sheet_view.showGridLines = False

    row = 1
    # Title banner
    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1,
                value="Audit Trail Review — Evidence Package")
    c.font      = Font(bold=True, color=C_HEADER_FG, name="Calibri", size=14)
    c.fill      = _fill(C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(f"A{row}:B{row}")
    c = ws.cell(row=row, column=1,
                value="Periodic Review — Section 9.1.6 Technical Review of Audit Trail")
    c.font      = Font(bold=False, color="94A3B8", name="Calibri", size=10)
    c.fill      = _fill(C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18
    row += 2

    # ── Critical Findings Box — shown FIRST so auditor sees it immediately ────
    critical_events = top_df[
        top_df["Risk_Tier"].astype(str) == "Critical"
    ].copy() if not top_df.empty else pd.DataFrame()

    if not critical_events.empty:
        # Red alert header
        ws.merge_cells(f"A{row}:B{row}")
        alert_hdr = ws.cell(
            row=row, column=1,
            value=f"⚠  CRITICAL FINDINGS REQUIRING IMMEDIATE ACTION  "
                  f"({len(critical_events)} event{'s' if len(critical_events)>1 else ''})"
        )
        alert_hdr.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        alert_hdr.fill      = _fill("991B1B")
        alert_hdr.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

        # Column headers — no chain ID column
        for ci, (lbl, w) in enumerate([
            ("User", 20), ("Rule Triggered", 38),
            ("Record / Table", 26), ("Date & Time", 22)
        ], 1):
            c_hdr = ws.cell(row=row, column=ci, value=lbl)
            c_hdr.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
            c_hdr.fill      = _fill("7F1D1D")
            c_hdr.border    = bdr
            c_hdr.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 16

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 38
        ws.column_dimensions["C"].width = 26
        ws.column_dimensions["D"].width = 22
        row += 1

        # Group chain-related events visually — show sequence context in plain language
        seen_chains = {}  # chain_id → first row written
        for _, ev in critical_events.iterrows():
            chain = str(ev.get("Event_Chain_ID","")).strip()
            usr   = str(ev.get("user_id","—"))
            rule  = str(ev.get("Primary_Rule","—"))
            rule  = rule.replace(" [CRITICAL]","").replace(" [HIGH]","").strip()
            rec   = str(ev.get("record_id","—"))
            if rec in ("","nan","—"):
                rec = str(ev.get("record_type","—"))
            ts    = str(ev.get("timestamp","—"))
            seq   = str(ev.get("Sequence_Context","")).strip()

            # For chain events, prefix with ↳ after first member to show grouping
            if chain and chain not in ("","None","nan"):
                if chain not in seen_chains:
                    seen_chains[chain] = row
                    user_label = usr
                else:
                    user_label = f"  ↳ {usr}"  # indent to show sequence membership
            else:
                user_label = usr

            row_data = [user_label, rule, rec, ts]
            for ci, val in enumerate(row_data, 1):
                c_data = ws.cell(row=row, column=ci, value=val)
                c_data.font      = Font(bold=(ci==2), color="7F1D1D",
                                        name="Calibri", size=9)
                c_data.fill      = _fill("FEF2F2")
                c_data.border    = bdr
                c_data.alignment = Alignment(vertical="center", wrap_text=True)
            ws.row_dimensions[row].height = 16
            row += 1

            # If this is part of a sequence, add a light-grey context row
            if seq and chain and chain in seen_chains:
                ws.merge_cells(f"A{row}:D{row}")
                ctx_cell = ws.cell(row=row, column=1,
                                   value=f"   ⟳ {seq}")
                ctx_cell.font      = Font(italic=True, color="6B7280",
                                          name="Calibri", size=8)
                ctx_cell.fill      = _fill("FFF7F7")
                ctx_cell.border    = bdr
                ctx_cell.alignment = Alignment(vertical="center", wrap_text=True)
                ws.row_dimensions[row].height = 13
                row += 1

        row += 1   # spacer after critical box

    elif n_crit == 0 and n_high == 0:
        # Green all-clear box
        ws.merge_cells(f"A{row}:B{row}")
        ok_cell = ws.cell(row=row, column=1,
                          value="✓  NO CRITICAL OR HIGH-RISK EVENTS IDENTIFIED")
        ok_cell.font      = Font(bold=True, color="166534", name="Calibri", size=11)
        ok_cell.fill      = _fill("DCFCE7")
        ok_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 2

    # Reset column widths back to standard after critical box
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 52

    def _summary_row(label, value, bold_val=False, section=False, tier=None):
        nonlocal row
        c1 = ws.cell(row=row, column=1, value=label)
        c2 = ws.cell(row=row, column=2, value=value)
        if section:
            c1.font = Font(bold=True, color=C_SECTION_FG, name="Calibri", size=10)
            c2.font = Font(bold=True, color=C_SECTION_FG, name="Calibri", size=10)
            c1.fill = c2.fill = _fill(C_SECTION_BG)
        else:
            c1.font = _body_font(color=C_LABEL_FG)
            c2.font = _body_font(bold=bold_val, color=C_VALUE_FG)
        if tier:
            c2.fill = _fill(TIER_BG.get(tier, C_WHITE))
            c2.font = Font(bold=True, color=TIER_FG.get(tier, C_VALUE_FG),
                           name="Calibri", size=10)
        for c in (c1, c2):
            c.border    = bdr
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 18
        row += 1

    _summary_row("REVIEW INFORMATION", "", section=True)
    _summary_row("System Name",     system_name or "(not entered)", bold_val=True)
    _summary_row("Review Period",   f"{r_start}  →  {r_end}")
    _summary_row("Source File",     fname)
    _summary_row("Analysis Date",   str(datetime.date.today()))
    _summary_row("Regulatory Basis","21 CFR Part 11 §11.10(e)  |  EU Annex 11, Clause 9")

    # Out-of-period scope note — added to Summary sheet if events outside review period exist
    try:
        _r_s = pd.to_datetime(r_start, dayfirst=True, errors="coerce")
        _r_e = pd.to_datetime(r_end,   dayfirst=True, errors="coerce")
        if not pd.isnull(_r_s) and not pd.isnull(_r_e) and "timestamp_parsed" in scored_df.columns:
            _ts      = scored_df["timestamp_parsed"].dropna()
            _before  = int((_ts < _r_s).sum())
            _after   = int((_ts > _r_e).sum())
            if _before > 0 or _after > 0:
                _parts = []
                if _before: _parts.append(f"{_before:,} event(s) pre-date {r_start}")
                if _after:  _parts.append(f"{_after:,} event(s) post-date {r_end}")
                _summary_row(
                    "⚠ Dataset Scope Note",
                    f"The uploaded file contains events outside the defined review period "
                    f"({'; '.join(_parts)}). These events are included in the analysis "
                    f"because the full dataset was provided. Reviewers must confirm whether "
                    f"out-of-period events are within scope before signing this report.",
                    bold_val=True
                )
    except Exception:
        pass
    row += 1

    _summary_row("RESULTS AT A GLANCE", "", section=True)
    _summary_row("Total Records Reviewed",   f"{total:,}")
    _summary_row("Records Auto-Cleared",     f"{total-n_esc:,}  ({pct_clear}%)")
    _summary_row("Records Requiring Review", f"{n_esc}",         bold_val=True)
    row += 1

    _summary_row("ESCALATED EVENTS — BREAKDOWN", "", section=True)
    # These counts reflect the Events for Review sheet only (escalated events).
    # The Full Audit Log sheet contains tier counts for all 1,000 scored events.
    n_esc_crit = int((top_df["Risk_Tier"] == "Critical").sum()) if not top_df.empty else 0
    n_esc_high = int((top_df["Risk_Tier"] == "High").sum())     if not top_df.empty else 0
    n_esc_med  = int((top_df["Risk_Tier"] == "Medium").sum())   if not top_df.empty else 0
    _summary_row("Critical — Escalated for immediate action",
                 f"{n_esc_crit} of {n_esc} escalated events", tier="Critical")
    _summary_row("High — Escalated for investigation",
                 f"{n_esc_high} of {n_esc} escalated events", tier="High")
    _summary_row("Medium — Escalated for review",
                 f"{n_esc_med} of {n_esc} escalated events",  tier="Medium")
    _summary_row("Full dataset tier distribution",
                 f"See Full Audit Log sheet — Risk Level column")
    row += 1

    row += 1

    # ── Review Narrative — human-readable story of findings ───────────────────
    def _build_narrative(top_df: pd.DataFrame, scored_df: pd.DataFrame,
                         sys_name: str, r_start: str, r_end: str,
                         n_crit: int, n_high: int, n_med: int,
                         n_esc: int, total: int) -> str:
        """
        Build a 4–6 sentence narrative summary of the audit trail findings.
        Uses only escalated-event counts — full-dataset tier counts are NOT
        referenced here because they represent a different population from
        the escalated events and would create an irreconcilable contradiction
        on the same page. Full-dataset distribution is in the Full Audit Log.
        """
        sentences = []

        # Derive counts from the escalated set (top_df) only
        n_esc_crit = int((top_df["Risk_Tier"] == "Critical").sum()) if not top_df.empty else 0
        n_esc_high = int((top_df["Risk_Tier"] == "High").sum())     if not top_df.empty else 0
        n_esc_med  = int((top_df["Risk_Tier"] == "Medium").sum())   if not top_df.empty else 0

        # ── Sentence 1: Period, volume, and escalated summary ─────────────────
        _missing = "(review period dates not specified)"
        period_str = (
            f"from {r_start} to {r_end}"
            if r_start and r_end
               and r_start != _missing and r_end != _missing
            else _missing
        )
        if n_esc == 0:
            sentences.append(
                f"The audit trail for {sys_name} was reviewed {period_str}. "
                f"Of {total:,} recorded events, none met the escalation threshold. "
                "No findings require reviewer action."
            )
        elif n_esc_crit == 0 and n_esc_high == 0:
            sentences.append(
                f"The audit trail for {sys_name} was reviewed {period_str}. "
                f"Of {total:,} recorded events, {n_esc} were escalated for review — "
                f"none Critical or High. "
                f"Full dataset tier distribution is available in the Full Audit Log sheet."
            )
        else:
            esc_parts = []
            if n_esc_crit > 0:
                esc_parts.append(f"{n_esc_crit} Critical")
            if n_esc_high > 0:
                esc_parts.append(f"{n_esc_high} High")
            if n_esc_med > 0:
                esc_parts.append(f"{n_esc_med} Medium")
            esc_str = ", ".join(esc_parts)
            sentences.append(
                f"The audit trail for {sys_name} was reviewed {period_str}. "
                f"Of {total:,} recorded events, {n_esc} were escalated for review "
                f"({esc_str}). "
                f"Full dataset tier distribution is available in the Full Audit Log sheet."
            )

        # ── Sentences 2–4: Describe the most significant findings ─────────────
        if top_df.empty:
            sentences.append("No significant findings were detected.")
        else:
            # Group by rule type to avoid repeating same finding
            rule_groups: dict = {}
            for _, ev in top_df.iterrows():
                triggered = str(ev.get("Triggered_Rules",""))
                tier      = str(ev.get("Risk_Tier",""))
                if tier not in ("Critical","High"):
                    continue
                for part in triggered.split(";"):
                    part = part.strip()
                    if not part:
                        continue
                    key = part.split("—")[0].strip() if "—" in part else part
                    if key not in rule_groups:
                        rule_groups[key] = []
                    rule_groups[key].append(ev)

            # Build one sentence per unique finding type (max 4)
            # seen_rules deduplicates by rule label; seen_sentences deduplicates
            # by output text — catches cases where two events share a rule and
            # produce an identical aggregated sentence (e.g. two Rule 1 events).
            finding_sentences = []
            seen_rules     = set()
            seen_sentences = set()

            for _, ev in top_df.iterrows():
                if len(finding_sentences) >= 4:
                    break
                triggered = str(ev.get("Triggered_Rules",""))
                tier      = str(ev.get("Risk_Tier","Low"))
                rat       = str(ev.get("Rule_Rationale",""))
                usr       = str(ev.get("user_id","an unidentified user"))
                act       = str(ev.get("action_type","an action"))
                rec_id    = str(ev.get("record_id",""))
                rec_type  = str(ev.get("record_type","a record"))
                chain_id  = str(ev.get("Event_Chain_ID",""))

                # Pick the primary rule for this event
                primary_rule = ""
                for part in triggered.split(";"):
                    p = part.strip()
                    if p and p not in seen_rules:
                        primary_rule = p
                        break
                if not primary_rule or primary_rule in seen_rules:
                    continue
                seen_rules.add(primary_rule)

                rec_ref = (f" on record {rec_id}" if rec_id and
                           rec_id not in ("","nan","—") else "")
                chain_ref = (f" (event chain {chain_id})" if chain_id else "")

                # Map rule to narrative sentence template
                r = primary_rule.lower()
                if "rule 3" in r or "admin" in r and "conflict" in r:
                    finding_sentences.append(
                        f"A system administrator ({usr}) directly modified "
                        f"production {rec_type} data{rec_ref}, which is inconsistent "
                        "with Segregation of Duties expectations under 21 CFR Part 11 §11.10(d)."
                    )
                elif "rule 5" in r or "failed login" in r:
                    finding_sentences.append(
                        f"A potential unauthorised access event was detected — "
                        f"user '{usr}' had repeated failed login attempts followed "
                        f"by a data modification{rec_ref}{chain_ref}."
                    )
                elif "rule 15" in r or "suspicious" in r and "sequence" in r:
                    finding_sentences.append(
                        f"A data manipulation sequence was detected: "
                        f"user '{usr}' modified, deleted, then recreated "
                        f"record {rec_id}{chain_ref}, suggesting an attempt "
                        "to alter a locked GxP record."
                    )
                elif "rule 12" in r or "timestamp reversal" in r:
                    finding_sentences.append(
                        f"A chronological impossibility was detected — "
                        f"record {rec_id} shows an approval timestamp "
                        "before its creation, indicating clock manipulation "
                        "or direct database alteration."
                    )
                elif "rule 13" in r or "service" in r and "account" in r:
                    finding_sentences.append(
                        f"Non-personal account '{usr}' performed a "
                        f"{act} action on {rec_type}{rec_ref}. "
                        "This action cannot be attributed to a specific individual."
                    )
                elif "rule 1" in r or "vague" in r:
                    # Count total Rule 1 findings
                    r1_count = sum(
                        1 for _, e in top_df.iterrows()
                        if "Rule 1" in str(e.get("Triggered_Rules",""))
                    )
                    if r1_count > 1:
                        finding_sentences.append(
                            f"{r1_count} instances of missing or inadequate change "
                            "rationale were observed across multiple records, "
                            "indicating deficiencies in data attribution practices."
                        )
                    else:
                        finding_sentences.append(
                            f"One instance of missing or inadequate change rationale "
                            f"was observed on {rec_type}{rec_ref}."
                        )
                elif "rule 6" in r or "delete and recreate" in r:
                    finding_sentences.append(
                        f"A delete-and-recreate pattern was identified on "
                        f"record {rec_id} by user '{usr}', a known method "
                        "for circumventing locked record controls."
                    )
                elif "rule 16" in r or "first-time" in r:
                    finding_sentences.append(
                        f"User '{usr}' performed a '{act}' action for the "
                        "first time in their recorded history, representing "
                        "an unexpected behaviour change requiring investigation."
                    )
                elif "rule 2" in r or "burst" in r or "contemporaneous" in r:
                    r2_count = sum(
                        1 for _, e in top_df.iterrows()
                        if "Rule 2" in str(e.get("Triggered_Rules",""))
                    )
                    finding_sentences.append(
                        f"{r2_count if r2_count > 1 else 'An'} instance"
                        f"{'s' if r2_count > 1 else ''} of high-volume "
                        "data entry in a short time window were detected, "
                        "raising concerns about contemporaneous recording practices."
                    )
                elif "off-hours" in r or "rule 11" in r or "holiday" in r:
                    finding_sentences.append(
                        f"Activity outside normal business hours was detected "
                        f"by user '{usr}' on {rec_type}{rec_ref}, "
                        "with no documented business justification on file."
                    )
                else:
                    continue

            # Deduplicate by sentence text before extending
            sentences.extend(
                s for s in finding_sentences
                if s not in seen_sentences
                and not seen_sentences.add(s)
            )

        # ── Final sentence: Overall risk assessment (escalated counts only) ──
        if n_esc_crit >= 2:
            overall = (
                "Overall, the system presents a high risk to data integrity "
                "and requires immediate corrective actions, particularly in "
                "access control and audit trail integrity."
            )
        elif n_esc_crit == 1:
            overall = (
                "Overall, the system presents a moderate-to-high risk to data "
                "integrity. The critical finding identified requires prompt "
                "investigation and formal non-conformance documentation."
            )
        elif n_esc_high >= 3:
            overall = (
                "Overall, the system presents a moderate risk to data integrity. "
                "Multiple high-risk findings indicate systemic issues in "
                "documentation practices and access control that require "
                "corrective action."
            )
        elif n_esc_high >= 1 or n_esc_med >= 3:
            overall = (
                "Overall, the system presents a low-to-moderate risk to data "
                "integrity. The findings identified are manageable but require "
                "documented review and corrective action where applicable."
            )
        else:
            overall = (
                "Overall, the system presents a low risk to data integrity "
                "for the review period. No critical or high-risk findings "
                "were identified that require immediate action."
            )
        sentences.append(overall)

        return "\n\n".join(sentences)

    narrative = _build_narrative(
        top_df, scored_df, system_name or "the reviewed system",
        r_start, r_end, n_crit, n_high, n_med, n_esc, total
    )

    _summary_row("REVIEW NARRATIVE", "", section=True)
    ws.merge_cells(f"A{row}:B{row}")
    narr_cell = ws.cell(row=row, column=1, value=narrative)
    narr_cell.font      = _body_font(color=C_VALUE_FG)
    narr_cell.fill      = _fill("EFF6FF")   # very light blue — distinct from white
    narr_cell.border    = bdr
    narr_cell.alignment = Alignment(vertical="top", wrap_text=True)
    # Height scales with content — approx 18pt per sentence, min 90
    n_sentences = narrative.count("\n\n") + 1
    ws.row_dimensions[row].height = max(90, n_sentences * 22)
    row += 2

    _summary_row("REVIEWER STATEMENT", "", section=True)
    ws.merge_cells(f"A{row}:B{row}")
    stmt = ws.cell(row=row, column=1, value=(
        f"This audit trail review was performed using rule-based anomaly detection with system-generated narrative summaries. "
        f"All risk classifications are derived from deterministic rules; narrative text does not influence risk scoring or tier assignment. "
        f"{total:,} records were reviewed across the period {r_start} to {r_end}. "
        f"{pct_clear}% of records were automatically cleared as low risk. "
        f"The {n_esc} records requiring review are documented in the 'Events for Review' sheet. "
        f"Each finding has been independently reviewed and dispositioned by the undersigned reviewer. "
        f"The 'System-Proposed Disposition' column is informational only; the reviewer's determination "
        f"in the 'Reviewer Decision' column reflects independent human judgement and is the authoritative record. "
        f"Regulatory basis: 21 CFR Part 11 §11.10(e) and EU Annex 11 Clause 9.\n\n"
        f"Timezone note: Off-hours scoring (Rule 11) assumes all timestamps in the uploaded "
        f"file are in the same timezone. If your system exports timestamps in UTC or a "
        f"non-local timezone, off-hours flags should be interpreted accordingly. "
        f"Reviewers should verify local shift patterns before dispositing temporal anomalies."
    ))
    stmt.font      = _body_font(color=C_VALUE_FG)
    stmt.fill      = _fill(C_ALT_ROW)
    stmt.border    = bdr
    stmt.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 72
    row += 2

    _summary_row("Reviewer Name",      "")
    _summary_row("Reviewer Title",     "")
    _summary_row("Date of Review",     "")
    _summary_row("Reviewer Signature", "")

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Events for Review (QA-friendly, no technical columns)
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Events for Review")
    ws2.sheet_properties.tabColor = "DC2626"
    ws2.sheet_view.showGridLines  = False

    # Director-facing column set — clean, scannable, no technical noise.
    # Risk Score removed: composite score and tier are intentionally decoupled
    # (named rule overrides force tier regardless of score — showing both
    # creates confusion). Tier alone is the reviewer-facing classification.
    # Supporting_Signals and Triggered_Rules are in Full Audit Log only.
    reviewer_cols = [
        ("No.",                            "Rank",              5),
        ("Risk Level",                     "Risk_Tier",         11),
        ("Evidence Strength",              "Evidence_Strength", 10),
        ("Date & Time",                    "timestamp",         19),
        ("User",                           "user_id",           16),
        ("Action",                         "action_type",       18),
        ("Record",                         "record_type",       16),
        ("Record ID",                      "record_id",         14),
        ("Change Reason",                  "comments",          28),
        ("Primary Rule",                   "Primary_Rule",      34),
        ("Related Sequence",               "Sequence_Context",  28),
        ("Why It Matters",                 "Regulatory_Basis",  50),
        ("What Happened",                  "AI_Justification",  46),
        ("Recommended Action",             "Action_Required",   46),
        ("System-Proposed Disposition\n(see Decision Basis — reviewer must\n"
         "make independent determination)",
                                           "Suggested_Disposition", 30),
        ("Decision Basis",                 "Suggested_Disposition_Rationale", 40),
        ("Reviewer Decision\n(independent — tick one)",
                                           "Reviewer_Disposition", 34),
        ("Reviewer Notes",                 "Reviewer_Notes",    28),
    ]

    top_out = top_df.copy().reset_index(drop=True)
    top_out.insert(0, "Rank", range(1, len(top_out)+1))
    top_out["Reviewer_Disposition"] = "☐ Justified     ☐ Escalate to CAPA     ☐ False Positive"
    top_out["Reviewer_Notes"]       = ""

    # Suggested Disposition cell colours
    SUGG_FILL = {
        "Escalate to CAPA":              ("FEE2E2", "991B1B"),
        "Justified — Amendment Required":("FEF3C7", "92400E"),
        "Investigate — Verify Source Data":("DBEAFE","1E40AF"),
        "Justified — Document Rationale":("F0FDF4","166534"),
        "Justified — No Action Required":("F9FAFB","374151"),
    }

    # Header row
    for ci, (hdr_label, _, col_w) in enumerate(reviewer_cols, 1):
        c = ws2.cell(row=1, column=ci, value=hdr_label)
        c.font      = _hdr_font(size=10)
        c.fill      = _fill(C_HEADER_BG)
        c.border    = bdr
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        ws2.column_dimensions[get_column_letter(ci)].width = col_w
    ws2.row_dimensions[1].height = 30

    # Data rows
    for ri, (_, drow) in enumerate(top_out.iterrows(), 2):
        tier    = str(drow.get("Risk_Tier", "Low"))
        alt_bg  = C_ALT_ROW if ri % 2 == 0 else C_WHITE
        for ci, (_, data_col, _) in enumerate(reviewer_cols, 1):
            val = drow.get(data_col, "")
            if pd.isnull(val):
                val = ""
            # Clean up Issues Found — remove score labels for readability
            if data_col == "Triggered_Rules" and val:
                val = str(val).replace(" [HIGH]","").replace(
                    " [MEDIUM]","").replace(" [CRITICAL]","").replace(
                    " [LOW]","")
            # Regulatory_Basis is already one sentence — truncate only as safety net
            if data_col == "Regulatory_Basis" and val:
                val = str(val)[:500]
            if isinstance(val, float) and not pd.isnull(val):
                val = round(val, 2)
            c = ws2.cell(row=ri, column=ci, value=val)
            c.border    = bdr
            c.alignment = Alignment(vertical="top", wrap_text=True)

            # Risk Level column gets tier colour; rest get alternating white/grey
            if data_col == "Risk_Tier":
                c.fill = _fill(TIER_BG.get(tier, C_WHITE))
                c.font = Font(bold=True,
                              color=TIER_FG.get(tier, C_VALUE_FG),
                              name="Calibri", size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif data_col == "Suggested_Disposition":
                sugg_bg, sugg_fg = SUGG_FILL.get(str(val), ("F9FAFB","374151"))
                c.fill      = _fill(sugg_bg)
                c.font      = Font(bold=True, color=sugg_fg,
                                   name="Calibri", size=9)
                c.alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
            elif data_col == "Suggested_Disposition_Rationale":
                c.fill = _fill("FFFBEB")
                c.font = _body_font(color="78350F", size=9)
            elif data_col in ("Reviewer_Disposition", "Reviewer_Notes"):
                c.fill = _fill(C_ALT_ROW)
                c.font = _body_font(color=C_LABEL_FG)
            else:
                c.fill = _fill(alt_bg)
                c.font = _body_font(color=C_VALUE_FG,
                                    bold=(data_col == "Rank"))
        ws2.row_dimensions[ri].height = 60

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(reviewer_cols))}1"
    ws2.freeze_panes    = "A2"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — Full Audit Log (all events, reviewer-friendly columns only)
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Full Audit Log")
    ws3.sheet_properties.tabColor = "374151"
    ws3.sheet_view.showGridLines  = False

    log_cols = [
        ("Date & Time",      "timestamp",          20),
        ("User",             "user_id",            16),
        ("Action",           "action_type",        18),
        ("Record Type",      "record_type",        18),
        ("Record ID",        "record_id",          16),
        ("Change Reason",    "comments",           28),
        ("Risk Level",       "Risk_Tier",          11),
        ("Evidence",         "Evidence_Strength",  10),
        ("Primary Rule",     "Primary_Rule",       34),
        ("All Rules Fired",  "Triggered_Rules",    38),
        ("Related Sequence",  "Sequence_Context",   22),
    ]

    # Strip internal columns — keep only the log_cols fields
    keep_fields = [f for _, f, _ in log_cols]
    log_df = scored_df[[c for c in keep_fields if c in scored_df.columns]].copy()

    for ci, (hdr_label, _, col_w) in enumerate(log_cols, 1):
        c = ws3.cell(row=1, column=ci, value=hdr_label)
        c.font      = _hdr_font(size=10)
        c.fill      = _fill(C_HEADER_BG)
        c.border    = bdr
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws3.column_dimensions[get_column_letter(ci)].width = col_w
    ws3.row_dimensions[1].height = 24

    for ri, (_, drow) in enumerate(log_df.iterrows(), 2):
        tier   = str(drow.get("Risk_Tier", "Low"))
        alt_bg = C_ALT_ROW if ri % 2 == 0 else C_WHITE
        for ci, (_, data_col, _) in enumerate(log_cols, 1):
            val = drow.get(data_col, "")
            if pd.isnull(val):
                val = ""
            if data_col == "Triggered_Rules":
                val = str(val).replace(" [HIGH]","").replace(
                    " [MEDIUM]","").replace(" [CRITICAL]","")
                if not val or val.strip() == "":
                    val = "No anomaly detected"
            if data_col == "Event_Chain_ID":
                if not val or str(val).strip() in ("", "nan"):
                    val = "None"
            if isinstance(val, float) and not pd.isnull(val):
                val = round(val, 2)
            c = ws3.cell(row=ri, column=ci, value=val)
            c.border    = bdr
            c.alignment = Alignment(vertical="center", wrap_text=False)
            c.font      = _body_font(color=C_VALUE_FG, size=9)
            if data_col == "Risk_Tier" and tier in ("Critical","High"):
                c.fill = _fill(TIER_BG.get(tier, C_WHITE))
                c.font = Font(bold=True, color=TIER_FG.get(tier, C_VALUE_FG),
                              name="Calibri", size=9)
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.fill = _fill(alt_bg)
        ws3.row_dimensions[ri].height = 15

    ws3.auto_filter.ref = f"A1:{get_column_letter(len(log_cols))}1"
    ws3.freeze_panes    = "A2"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4 — Detection Logic Reference (full 14-rule specification)
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Detection Logic")
    ws4.sheet_properties.tabColor = "374151"
    ws4.sheet_view.showGridLines  = False
    ws4.column_dimensions["A"].width = 120

    # Pull the detection logic text from the session — built in show_audit_trail
    detection_text = st.session_state.get("at_detection_logic_text", "")
    if not detection_text:
        detection_text = (
            "Detection Logic Reference not available. "
            "Please re-run the analysis to populate this sheet."
        )

    # Title row
    t4 = ws4.cell(row=1, column=1,
                  value="Audit Trail Intelligence — Detection Logic Reference")
    t4.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
    t4.fill      = _fill(C_HEADER_BG)
    t4.alignment = Alignment(horizontal="left", vertical="center")
    ws4.row_dimensions[1].height = 26

    # Write each line of the detection logic as a row
    row_num = 2
    for line in detection_text.split("\n"):
        c = ws4.cell(row=row_num, column=1, value=line)
        # Style headings (lines starting with ##)
        if line.startswith("## ") or line.startswith("### "):
            c.font = Font(bold=True, color=C_HEADER_BG, name="Calibri", size=11)
            c.fill = _fill(C_SECTION_BG)
            ws4.row_dimensions[row_num].height = 18
        elif line.startswith("# "):
            c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
            c.fill = _fill(C_HEADER_BG)
            ws4.row_dimensions[row_num].height = 20
        elif line.startswith("---"):
            c.font = Font(color="D1D5DB", name="Calibri", size=9)
            ws4.row_dimensions[row_num].height = 8
        elif line.startswith("**") or line.startswith("| "):
            c.font = Font(bold=True, color=C_VALUE_FG, name="Calibri", size=10)
            ws4.row_dimensions[row_num].height = 15
        else:
            c.font = Font(color=C_LABEL_FG, name="Calibri", size=10)
            ws4.row_dimensions[row_num].height = 15
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        row_num += 1

    wb.save(output)
    return output.getvalue()


def show_periodic_review(user: str, role: str, model_id: str):
    """
    Periodic Review landing page — shows 3 module cards.
    Clicking a live module opens it; coming-soon modules show a placeholder.
    """
    active = st.session_state.get("pr_active_module")

    # ── If a sub-module is open, show it ─────────────────────────────────────
    if active == "audit_trail":
        show_audit_trail(user, role, model_id)
        return

    if active in ("access_review", "report_drafter"):
        st.markdown("<br>", unsafe_allow_html=True)
        label = "User Access Review Intelligence" if active == "access_review" \
                else "Periodic Review Report Drafter"
        st.title(f"🚧 {label}")
        st.info("This module is coming soon. It will be available in the next release.")
        return

    # ── Landing page ──────────────────────────────────────────────────────────
    st.title("📋 Periodic Review")
    st.markdown(
        "<p style='color:#94a3b8;margin-top:-12px;'>Select a module below. "
        "Each module covers a mandatory section of your Periodic Review Report "
        "per 21 CFR Part 11, EU Annex 11, and your internal SOP(s).</p>",
        unsafe_allow_html=True
    )
    st.markdown("<br>", unsafe_allow_html=True)

    # ── Module cards ──────────────────────────────────────────────────────────
    modules = [
        {
            "key":     "audit_trail",
            "title":   "Audit Trail Review",
            "section": "21 CFR Part 11 §11.10(e) · EU Annex 11 Cl. 9",
            "desc":    (
                "Upload your audit trail log file to run the 16-rule detection engine. "
                "Escalates the 20 highest-risk events with a evidence package for "
                "your Periodic Review Report."
            ),
            "status":  "live",
            "btn_label": "Launch →",
            "color":   "#0284c7",
            "bg":      "#0c1f36",
            "border":  "#1e3a5f",
        },
        {
            "key":     "access_review",
            "title":   "User Access Review",
            "section": "21 CFR Part 11 §11.300",
            "desc":    (
                "Upload your user access list to flag dormant accounts, admin roles "
                "on non-admin functions, accounts active before training completion, "
                "and shared account fingerprinting."
            ),
            "status":  "coming_soon",
            "btn_label": "Coming Soon",
            "color":   "#475569",
            "bg":      "#0a1628",
            "border":  "#1e293b",
        },
        {
            "key":     "report_drafter",
            "title":   "Periodic Review Report Drafter",
            "section": "All sections · Full report output",
            "desc":    (
                "Upload your Periodic Review SOP and a model Periodic Report to "
                "auto-draft a complete report in your company terminology — "
                "Word document ready for approval."
            ),
            "status":  "coming_soon",
            "btn_label": "Coming Soon",
            "color":   "#475569",
            "bg":      "#0a1628",
            "border":  "#1e293b",
        },
    ]

    # ── All 3 cards in one HTML grid — guarantees equal height ───────────────
    cards_html = ""
    for mod in modules:
        live         = mod["status"] == "live"
        accent_color = mod["color"] if live else "#d2d2d7"
        badge_cls    = "pr-card-badge-live" if live else "pr-card-badge-soon"
        badge_text   = "Live" if live else "Coming Soon"
        title_color  = "#1d1d1f" if live else "#a1a1a6"
        desc_color   = "#3d3d3f" if live else "#a1a1a6"
        cards_html += f"""
  <div class="pr-card">
    <div class="pr-card-accent" style="background:{accent_color};"></div>
    <span class="pr-card-badge {badge_cls}">{badge_text}</span>
    <p class="pr-card-title" style="color:{title_color};">{mod['title']}</p>
    <p class="pr-card-ref">{mod['section']}</p>
    <p class="pr-card-desc" style="color:{desc_color};">{mod['desc']}</p>
  </div>"""

    st.markdown(f"""
<style>
.pr-grid {{
    display: grid;
    grid-template-columns: 1fr 1fr 1fr;
    gap: 16px;
    margin-bottom: 12px;
}}
.pr-card {{
    background: #ffffff;
    border: 1px solid #d2d2d7;
    border-radius: 14px;
    padding: 24px 22px 20px 22px;
    font-family: 'Inter', -apple-system, sans-serif;
    display: flex;
    flex-direction: column;
    box-shadow: 0 1px 4px rgba(0,0,0,0.06);
    transition: box-shadow 0.18s ease;
}}
.pr-card:hover {{ box-shadow: 0 4px 16px rgba(0,0,0,0.10); }}
.pr-card-accent {{ height: 3px; border-radius: 2px; margin-bottom: 18px; flex-shrink:0; }}
.pr-card-badge {{
    display: inline-block; font-size: 0.62rem; font-weight: 600;
    letter-spacing: 1.5px; text-transform: uppercase;
    padding: 2px 8px; border-radius: 20px; margin-bottom: 10px;
    flex-shrink: 0;
}}
.pr-card-badge-live {{ background: #e8f9f0; color: #1a7f4b; border: 1px solid #a3e4c1; }}
.pr-card-badge-soon {{ background: #f5f5f7; color: #a1a1a6; border: 1px solid #d2d2d7; }}
.pr-card-title {{ font-size: 1.05rem; font-weight: 700; margin: 0 0 4px 0; flex-shrink:0; }}
.pr-card-ref   {{ font-size: 0.69rem; color: #a1a1a6; font-family: 'Courier New', monospace;
                  margin: 0 0 12px 0; flex-shrink:0; }}
.pr-card-desc  {{ font-size: 0.81rem; line-height: 1.55; margin: 0; flex: 1; }}
</style>
<div class="pr-grid">{cards_html}</div>
""", unsafe_allow_html=True)

    # ── Button row — separate st.columns so Streamlit click handling works ───
    btn_col1, btn_col2, btn_col3 = st.columns(3)
    for btn_col, mod in zip([btn_col1, btn_col2, btn_col3], modules):
        live = mod["status"] == "live"
        with btn_col:
            if live:
                if st.button(mod["btn_label"], key=f"pr_open_{mod['key']}",
                             type="primary", use_container_width=True):
                    st.session_state["pr_active_module"] = mod["key"]
                    st.rerun()
            else:
                st.button(mod["btn_label"], key=f"pr_open_{mod['key']}",
                          disabled=True, use_container_width=True)


def show_audit_trail(user: str, role: str, model_id: str):
    """Render Periodic Review — Module 1: Audit Trail Intelligence."""
    st.title("🔍 Audit Trail Review Intelligence")
    st.markdown(
        "<p style='color:#94a3b8;margin-top:-12px;'>"
        "Periodic Review: Audit Trail Review Intelligence — Reduce audit log entries "
        "to the 20 highest-risk events, with documented statistical justification "
        "for the Periodic Review Report.</p>",
        unsafe_allow_html=True
    )

    # ── System metadata (always shown) ───────────────────────────────────────
    mc1, mc2, mc3 = st.columns(3)
    with mc1:
        st.session_state["at_system_name"] = st.text_input(
            "System Name", value=st.session_state.get("at_system_name",""),
            placeholder="e.g. DocuSign Part 11", key="at_sysname")

    # ── Previous quarter defaults for date pickers ───────────────────────────
    import datetime as _dt_q
    def _prev_quarter_dates():
        today = _dt_q.date.today()
        q = (today.month - 1) // 3 + 1
        if q == 1:
            return (_dt_q.date(today.year - 1, 10, 1),
                    _dt_q.date(today.year - 1, 12, 31))
        elif q == 2:
            return (_dt_q.date(today.year, 1, 1),
                    _dt_q.date(today.year, 3, 31))
        elif q == 3:
            return (_dt_q.date(today.year, 4, 1),
                    _dt_q.date(today.year, 6, 30))
        else:
            return (_dt_q.date(today.year, 7, 1),
                    _dt_q.date(today.year, 9, 30))

    _pq_start, _pq_end = _prev_quarter_dates()

    def _parse_date_input(key_date, key_str, label, fallback):
        """
        Render a date_input calendar picker. Defaults to `fallback` (previous
        quarter start or end) when no value has been stored yet.
        Parses any pre-stored string back to a date object on rerender.
        """
        import datetime as _dt
        stored = st.session_state.get(key_str, "")
        default_val = fallback
        if stored:
            for fmt in ("%d-%b-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d%m%Y", "%Y%m%d"):
                try:
                    default_val = _dt.datetime.strptime(stored.strip(), fmt).date()
                    break
                except ValueError:
                    continue
        picked = st.date_input(
            label,
            value=default_val,
            format="DD/MM/YYYY",
            key=key_date,
        )
        if picked:
            formatted = picked.strftime("%d-%b-%Y")
            st.session_state[key_str] = formatted
            return formatted
        return stored

    with mc2:
        _parse_date_input("at_rstart_picker", "at_review_start",
                          "Review Period Start", _pq_start)
    with mc3:
        _parse_date_input("at_rend_picker",   "at_review_end",
                          "Review Period End",   _pq_end)

    st.markdown("---")

    # ── STEP 1: Upload ────────────────────────────────────────────────────────
    if not st.session_state.get("at_mapping_done"):
        st.markdown("### Step 1 — Upload Audit Trail Export")
        st.caption(
            "CSV or Excel export from any GxP system — Veeva Vault, SAP, "
            "MasterControl, LIMS, or any custom system. No integration required."
        )

        # ── Detection Logic Reference Page ────────────────────────────────────
        DETECTION_LOGIC = """# Audit Trail Intelligence — Detection Logic Reference
## GxP Operational Anomaly Detection Engine v1.0

This document describes every detection rule implemented in the engine.
It supports Computer System Validation (CSV) of this tool and enables
QA reviewers to understand the basis for every finding produced.

**Total rules implemented: 16**
Rules 1–5: Named AI Skill rules (targeted, specific violations)
Rules 6–11: Dimension-based rules (pattern and behavioural anomalies)
Rules 12–16: Advanced integrity rules (no second file required)

---

## Named AI Skill Rules

### Rule 1 — Vague Rationale [Risk: HIGH]
**Target:** UPDATE or MODIFY on RESULTS or BATCH record types.
**Trigger:** Change reason field is blank, OR contains a non-descriptive term:
fixed, error, update, changed, correction, misc, ok, done, see above, as per,
per request, modified, edit, n/a.
**Not triggered by:** Short but specific comments — the rule requires a recognised
vague term, not just brevity. "pH outlier" does NOT trigger this rule.
**Regulatory basis:** 21 CFR Part 211.68; ALCOA+ Attributable and Legible.
**Threshold justification:** Vague rationale terms leave GxP data changes
unattributable and untraceable, making them impossible to evaluate during an audit.

---

### Rule 2 — Contemporaneous Burst [Risk: MEDIUM]
**Target:** INSERT, RESULT_INSERT, CREATE, ADD actions by any single user.
**Trigger:** More than 10 such actions by the same user within any 15-minute window.
**Regulatory basis:** ALCOA+ Contemporaneous; EU Annex 11 Clause 9.
**Threshold justification:** 10 entries in 15 minutes = one every 90 seconds.
Physiologically possible but statistically unusual for complex measurements.
Threshold is conservative to minimise false positives on high-throughput workflows.
**Deduplication note:** When this rule fires across multiple rows from the same
burst, only one representative event appears in the Top 20 escalated list.
All events remain visible in the Full Audit Log sheet.

---

### Rule 3 — Admin / GxP Conflict [Risk: CRITICAL]
**Target:** Users with roles containing Admin, DBA, Administrator, Sysadmin.
**Trigger:** That user performs INSERT, UPDATE, CREATE, or MODIFY on:
SAMPLE_DATA, BATCH_RELEASE, BATCH, RESULTS, or RESULT.
**Regulatory basis:** 21 CFR Part 11 §11.10(d); Segregation of Duties.
**Why Critical:** Admin accounts are authorised for system configuration only.
Direct modification of production records bypasses the standard review workflow,
creating an uncontrolled change pathway with no independent oversight.

---

### Rule 4 — Change Control Drift [Risk: HIGH]
**Target:** Events with a populated new_value column containing numeric data.
**Trigger:** Numeric new_value deviates by more than 3 standard deviations from
the mean of all new_values for the same record_type in the uploaded file.
**Regulatory basis:** 21 CFR Part 820.70(b); validated state requirements.
**Threshold justification:** 3 standard deviations is a statistically robust
outlier threshold. The reference distribution is computed from the uploaded file
itself, making the rule system-agnostic.
**Limitation:** Only numeric new_value data is scored. Status values such as
RELEASED or PENDING are not evaluated by this rule.

---

### Rule 5 — Failed Login → Data Manipulation [Risk: CRITICAL]
**Target:** All users.
**Trigger:** 3+ LOGIN_FAILED / AUTHENTICATION_FAILED events within 120 minutes
before a successful LOGIN, followed by DELETE, UPDATE, MODIFY, or INSERT on a
GxP record within 30 minutes of that login.
**Regulatory basis:** 21 CFR Part 11 §11.300; ALCOA+ Original.
**Threshold justification:** 3 failed attempts is the industry-standard threshold
for flagging a brute-force or credential-stuffing attempt. The 30-minute post-login
window captures immediate post-access manipulation. The 120-minute lookback covers
slow credential attacks.

---

## Dimension-Based Rules

### Rule 6 — Delete and Recreate Pattern [Risk: CRITICAL]
**Target:** All record types with a populated record_id column.
**Trigger:** Same user deletes a record_id and creates a new record with the same
record_id within 4 hours. Score: 9.0 on both the delete and recreate events.
**Regulatory basis:** 21 CFR Part 11 §11.10(e); ALCOA+ Original.
**Why Critical:** GxP systems typically lock approved records to prevent editing.
Deleting and recreating the same record is a known method of modifying locked data
while making the change appear as a new entry.
**4-hour window:** Set to capture same-session recreation while excluding legitimate
archival or system migration workflows that occur over longer periods.

---

### Rule 7 — Audit Trail Integrity Event [Risk: CRITICAL / HIGH / MEDIUM]
**Trigger (Critical — 10.0):** action_type or record_type contains keywords
indicating audit trail configuration was changed: audit trail, audit log,
log enabled, log disabled, audit enabled, audit disabled, configuration change,
system setting.
**Trigger (High — 8.0):** A delete action on a GxP-sensitive record type.
**Note:** Non-destructive read/access actions on GxP-sensitive record types
(SELECT, VIEW, READ) are no longer scored. Only audit trail integrity events
(Critical) and sensitive record deletions (High) produce a Rule 7 score.
This prevents routine system use from appearing as false positives.
**GxP-sensitive record types include:** batch record, audit trail, electronic
signature, test result, clinical, raw data, master data, configuration, user
account, role, permission, quality record, change control, deviation, CAPA, OOS.
**Regulatory basis:** 21 CFR Part 11 §11.10(e).

---

### Rule 8 — Privileged User on GxP Data [Risk: HIGH]
**Trigger:** A user with an admin-keyword role (admin, sysadmin, DBA, root,
service, superuser, power user) performs a modify or delete on a GxP-sensitive
record. Score: 8.0 if record is sensitive; 7.0 otherwise.
**Distinction from Rule 3:** Rule 3 fires on specific high-risk table names
(BATCH_RELEASE, SAMPLE_DATA etc.). Rule 8 fires on record type sensitivity
keywords and covers a broader range of tables.
**Regulatory basis:** 21 CFR Part 11 §11.10(d); Segregation of Duties.

---

### Rule 9 — High-Volume Activity Burst [Risk: MEDIUM]
**Trigger:** Same user performs the same action_type 5+ times within 60 minutes.
**Distinction from Rule 2:** Rule 2 is INSERT-specific with a 15-minute window
and a threshold of 10. Rule 9 applies to any action type with a 60-minute window
and a threshold of 5. Both can fire independently.
**Score:** Proportional to (count ÷ threshold × 3.5), capped at 10.0.
**Regulatory basis:** ALCOA+ Contemporaneous.

---

### Rule 10 — Audit Trail Timestamp Gap [Risk: HIGH]
**Trigger:** A gap of more than 2 hours between consecutive audit trail entries
(sorted by timestamp). Score: 7.0 fixed, applied to the first event after the gap.
**Regulatory basis:** 21 CFR Part 11 §11.10(e) — audit trail completeness.
**Important note:** Gaps may occur legitimately overnight or during approved
maintenance windows. The reviewer must verify whether the gap aligns with a
scheduled downtime before escalating.

---

### Rule 11 — Off-Hours and Federal Holiday Activity [Risk: HIGH / MEDIUM]
**Business hours definition:** Monday–Friday 07:00–20:00 local time.
**Trigger (Weekend — +5.0):** Event falls on Saturday or Sunday.
**Trigger (Federal Holiday — +4.0):** Event falls on a US Federal Holiday.
Holidays detected: New Year's Day, MLK Day, Presidents Day, Memorial Day,
Juneteenth, Independence Day, Labor Day, Columbus Day, Veterans Day,
Thanksgiving, Christmas Day.
**Trigger (Outside business hours — +4.0):** Event before 07:00 or after 20:00
on a weekday.
**Trigger (Deep night — additional +1.0):** Event between 00:00 and 04:59.
**Score cap:** 10.0 maximum regardless of how many triggers combine.
**Regulatory basis:** 21 CFR Part 11 §11.10(e). Off-hours activity is a risk
indicator requiring justification, not a violation per se.
**Timezone assumption:** This rule assumes all timestamps in the uploaded file are
in the same timezone. If your system exports timestamps in UTC or a non-local
timezone, off-hours flags must be interpreted accordingly. For global systems,
reviewers should verify the local shift pattern for the user's site before
dispositing any Rule 11 finding. A flag at 02:00 UTC may be 10:00 local time
for a user in a different region — not an anomaly at all.
*Recommendation:* Add a note in your Periodic Review Report specifying which
timezone the audit trail timestamps represent.

---

## Advanced Integrity Rules (No Second File Required)

### Rule 12 — Timestamp Reversal [Risk: CRITICAL]
**Target:** Records where both a creation action and an approval/release action
exist for the same record_id.
**Trigger:** The approval/release timestamp is earlier than the creation timestamp
for the same record_id. Score: 10.0 on the approval event.
**Creation keywords detected:** insert, create, add, result_insert, new.
**Approval keywords detected:** approve, release, authorise, authorize, sign,
submit, batch_release, approve_result.
**Regulatory basis:** 21 CFR Part 11 §11.10(e); ALCOA+ Contemporaneous (O).
**Why this matters:** Chronological impossibility in a correctly functioning system.
Indicates server clock manipulation, system migration error, or direct database
alteration. No legitimate explanation exists for an approval preceding creation.
**Requirements:** record_id column must be populated and timestamp column must
be parseable for this rule to fire.

---

### Rule 13 — Service / Shared Account GxP Action [Risk: CRITICAL / HIGH]
**Target:** User accounts whose username begins with a non-personal prefix.
**Detected prefixes:** svc_, service_, shr_, share_, shared_, share., adm_,
admin_, tec_, tech_, technical_, interface_, int_, batch_, sys_, system_,
robot_, auto_, automation_, script_, api_, sa_, dba_, root, daemon, guest, test_
**Trigger (Critical — 10.0):** Non-personal account performs a GxP data action
(insert, update, modify, delete, approve, release) on a GxP-sensitive record type.
**Trigger (High — 7.0):** Non-personal account performs a data action on any table.
**Regulatory basis:** 21 CFR Part 11 §11.300 — electronic signatures must be
unique to one individual; ALCOA+ Attributable.
**Why this matters:** Non-personal accounts cannot be attributed to a specific
individual, making any data they create or modify non-ALCOA+ compliant.
**Configuration note:** Add or remove prefixes from the _NONPERSONAL_PREFIXES
tuple in the engine code to match your organisation's naming conventions.

---



---

### Rule 15 — Suspicious Action Sequence: UPDATE → DELETE → INSERT [Risk: CRITICAL]
**Target:** All record types with a populated record_id column.
**Trigger:** Same user performs UPDATE (or MODIFY/EDIT/AMEND), then DELETE,
then INSERT/CREATE on the same record_id, all within 30 minutes.
Score: 10.0 on all three events. An Event Chain ID (EC-NNN) links all three rows.
**Regulatory basis:** 21 CFR Part 11 §11.10(e); ALCOA+ Original.
**Why Critical:** GxP systems lock approved records to prevent direct editing.
This three-step sequence — modify, delete, recreate — is the primary method
for altering a locked record while making the change appear as a new entry,
obscuring the original modification from the audit trail.
**Distinction from Rule 6 (Delete-Recreate):** Rule 6 detects DELETE → INSERT
(two steps). Rule 15 detects UPDATE → DELETE → INSERT (three steps), which is
more specific — it indicates the user first tried to modify the locked record,
then switched to the delete-recreate method.
**30-minute window:** Captures same-session sequences while excluding coincidental
patterns across different working sessions.

---

## Event Chain IDs

When multiple events are causally linked, the engine assigns a shared Event Chain
ID (format: EC-001, EC-002 etc.) to all related events. This allows reviewers to
filter the Full Audit Log by chain ID and read the complete event story in sequence.

Chains are assigned for:
- **Rule 15** (UPDATE → DELETE → INSERT): all three steps share one chain ID
- **Rule 5** (Failed Login → Manipulation): all login attempts and the triggering
  data action share one chain ID
- **Rule 6** (Delete → Recreate): both events share one chain ID

The Event_Chain_ID column appears in both the Events for Review sheet and the
Full Audit Log sheet in the Excel output.


---

### Rule 16 — First-Time Behavior Detection [Risk: HIGH]
**Target:** All users with at least 5 prior recorded events in the uploaded file.
**Trigger:** A user performs an action_type they have never performed before
in their recorded audit trail history. Score varies by prior history and action risk:

| Prior events | Action type | Score |
|---|---|---|
| ≥50 | First-time DELETE/APPROVE on GxP table | 9.0 |
| ≥20 | First-time DELETE/APPROVE on any table | 8.0 |
| ≥20 | First-time any action | 6.0 |
| ≥5  | First-time high-risk action | 5.0 |
| <5  | (not enough history — skipped) | 0.0 |

**High-risk action types detected:** delete, del, remove, purge, void, cancel,
approve, release, authorise, authorize, sign, override, batch_release, approve_result.
**Regulatory basis:** 21 CFR Part 11 §11.10(d) access controls; ALCOA+ Attributable.
**Why this matters:** An established user who suddenly performs an action they
have never done before — especially a high-risk action like delete or approve —
is a meaningful insider risk signal. This pattern may indicate:
- Unauthorised escalation of access or responsibilities
- Credential sharing (someone else used the account)
- A data integrity incident being covered up

**Confidence labelling:** The rationale always states the number of prior events
observed, so the reviewer can judge statistical significance. A user with 200 prior
events performing their first-ever delete is far more anomalous than a user with 6.

**Important limitation:** This rule requires sufficient history in the uploaded file.
Short-period exports (under 30 days) will produce many "first-time" flags that
are simply normal activity not captured in the window. For reliable detection,
upload audit trails covering at least 3 months.

**Minimum prior events:** 5. Users with fewer than 5 recorded events are skipped
to exclude newly created accounts where any action is technically a "first time."

### Rule 14 — Dormant Account Sudden Activity [Risk: HIGH]
**Target:** All users with at least 4 events in the uploaded audit trail.
**Trigger:** A user has no activity for 90 or more consecutive days, then
performs a data action on a GxP-sensitive record type or performs
update/modify/delete/insert/approve/release on any record. Score: 8.0.
**Only the first re-activation event per user is flagged** to avoid filling
the Top 20 with multiple events from the same dormant account.
**Regulatory basis:** 21 CFR Part 11 §11.10(d) — access controls; access review
best practice requiring deactivation of accounts inactive for 90+ days.
**Limitation:** Requires the uploaded audit trail to cover at least 90+ days of
history for reliable detection. Short-period extracts will not trigger this rule.
**Minimum event count:** A user must have at least 3 prior events before the gap
to be considered an established account (excludes newly created accounts).

---

## Composite Scoring

Each event receives scores across all 14 rules. A weighted composite (0–10):

| Rule | Dimension | Weight |
|------|-----------|--------|
| 11 | Temporal / Holiday | 6% |
| 9  | Velocity burst | 7% |
| 8  | Privilege | 9% |
| 7  | Record sensitivity | 8% |
| 6  | Delete-Recreate | 8% |
| 10 | Timestamp gap | 6% |
| 1  | Vague Rationale | 7% |
| 2  | Contemporaneous Burst | 7% |
| 3  | Admin/GxP Conflict | 10% |
| 4  | Change Control Drift | 6% |
| 5  | Failed Login | 8% |
| 12 | Timestamp Reversal | 9% |
| 13 | Service Account | 9% |
| 14 | Dormant Account | 7% |
| 15 | Suspicious Sequence | 9% |
| 16 | First-Time Behavior | 7% |
| | **Total** | **123%** |

*Note: Weights are normalised internally — the sum does not need to equal 100%.*

**Named rule tier overrides:** Rules 3, 5, 12, and 13 at score ≥9 always produce
a Critical tier, regardless of composite score. Rule 6 at ≥9 and Rule 7 at 10.0
also produce Critical. This ensures named Critical violations are never
downgraded by a low composite score.

---

## Suggested Disposition Logic

| Condition | Suggested Disposition |
|---|---|
| Rule 12 fired | Escalate to CAPA |
| Rule 13 fired (Critical) | Escalate to CAPA |
| Rule 5 fired | Escalate to CAPA |
| Rule 3 fired | Escalate to CAPA |
| Rule 6 fired | Escalate to CAPA |
| Rule 7 (audit ctrl) fired | Escalate to CAPA |
| Rule 10 (gap) during business hours | Escalate to CAPA |
| Rule 10 (gap) outside business hours | Investigate — Verify Source Data |
| Rule 4 fired | Escalate to CAPA |
| Rule 1 — blank comment | Escalate to CAPA |
| Rule 1 — vague term | Justified — Amendment Required |
| Rule 14 fired | Investigate — Verify Source Data |
| Rule 2 burst | Investigate — Verify Source Data |
| Off-hours with comment | Justified — Document Rationale |
| Off-hours, no comment | Escalate to CAPA |
| No rule triggered significantly | Justified — No Action Required |

---

## Review Narrative (Summary Sheet)

Following the detection analysis, the Summary sheet includes a **Review Narrative**
section — a plain-English paragraph summarising the key findings of the review.
This section is generated deterministically from the scored data with no LLM
involvement, making it fully reproducible.

The narrative covers:
1. **Volume statement** — total events reviewed, period covered, events escalated
2. **Key findings** — one sentence per significant rule triggered (max 4 findings),
   naming the user, record, and action type involved
3. **Overall risk assessment** — concludes with one of five risk levels:
   - High risk — 2+ Critical findings
   - Moderate-to-high risk — 1 Critical finding
   - Moderate risk — 3+ High findings
   - Low-to-moderate risk — 1–2 High findings
   - Low risk — no Critical or High findings

The narrative is intended to be copied into the introduction of the Periodic Review
Report Section 9.1.6 or used as a management summary.

---

---

## Not Yet Implemented — Planned Rules

These checks were designed and scoped but not yet built into the engine.
They are documented here for transparency and future CSV planning.

### Planned Rule 16 — Shared Account / Same User, Two Locations
**What it detects:** The same user_id performing actions from two different
IP addresses within 30 minutes — physically impossible for one person.
**Why not yet built:** Requires an IP address or workstation column in the
audit trail export. Many systems do not include this by default. Will be
activated automatically when the column mapper detects an IP column.
**Regulatory basis:** 21 CFR Part 11 §11.300.

### Planned Rule 17 — Bulk Export Before Departure
**What it detects:** A user exporting or printing large volumes of records
(SELECT, EXPORT, PRINT actions) close to their account deactivation date.
**Why not yet built:** Requires an HR termination date as a second input file.
Cannot be determined from the audit trail alone.
**Regulatory basis:** Data protection and confidentiality requirements.

### Planned Rule 18 — Activity During Approved Maintenance Window
**What it detects:** User activity during a scheduled system maintenance
window, indicating either the maintenance did not occur as planned, or
unauthorised activity occurred during the maintenance period.
**Why not yet built:** Requires the approved maintenance schedule (from the
change control system) as a second input file.
**Regulatory basis:** 21 CFR Part 11 §11.10(e).

### Planned Rule 19 — Account Created and Used Same Day
**What it detects:** A new user account that is created and then used to
perform GxP data actions on the same day, before training records could
have been completed.
**Why not yet built:** Requires a user account creation log or training
completion record as a second input file to confirm training status.
**Regulatory basis:** 21 CFR Part 11 §11.10(i) — training.

---

## Validation Evidence

This document constitutes part of the functional specification for the
Audit Trail Intelligence module. For Computer System Validation purposes,
the following test cases should be executed against the sample CSV template
(downloadable from the tool's upload screen) to verify each rule fires correctly:

| Test Case | Expected Rule | Expected Tier |
|---|---|---|
| UPDATE RESULTS with comment "fixed" | Rule 1 | High |
| 12 RESULT_INSERT in 12 minutes, same user | Rule 2 | Medium |
| admin_sys INSERT BATCH_RELEASE | Rule 3 | Critical |
| new_value = 147.3 when mean ≈ 7.1 | Rule 4 | High |
| 3x LOGIN_FAILED → LOGIN → DELETE, 18 min | Rule 5 | Critical |
| DELETE then INSERT same record_id, 3 min | Rule 6 | Critical |
| UPDATE AUDIT_TRAIL, new_value = DISABLED | Rule 7 | Critical |
| admin_sys DELETE RESULTS | Rule 8 | High |
| Same user, same action, 5+ times in 60 min | Rule 9 | Medium |
| >2 hour gap in timestamps | Rule 10 | High |
| Timestamp 02:14 AM weekday | Rule 11 | High |
| Any event on 2024-07-04 | Rule 11 (Holiday) | Medium-High |
| Approval timestamp before creation, same record | Rule 12 | Critical |
| svc_batch INSERT RESULTS | Rule 13 | Critical |
| Same user, 90+ day gap, then GxP action | Rule 14 | High |
| User with 50+ events performs first-ever DELETE | Rule 16 | High |

*This document was generated by the Audit Trail Intelligence module and
represents the validated detection logic at the time of this release.
Review and approve as part of the Computer System Validation package.*
"""

        def _detection_logic_pdf() -> bytes:
            """Return Detection Logic as a plain-text UTF-8 encoded file."""
            st.session_state["at_detection_logic_text"] = DETECTION_LOGIC
            return DETECTION_LOGIC.encode("utf-8")

        with st.expander(
            "🔬 Detection Logic Reference — How does the engine work? (click to expand)",
            expanded=False
        ):
            st.markdown(
                "<div style='max-height:420px;overflow-y:auto;padding:0 8px;'>",
                unsafe_allow_html=True
            )
            st.session_state["at_detection_logic_text"] = DETECTION_LOGIC
            st.markdown(DETECTION_LOGIC)
            st.markdown("</div>", unsafe_allow_html=True)
            st.download_button(
                "📄 Download Detection Logic Reference (.txt)",
                data=_detection_logic_pdf(),
                file_name="AuditTrail_Detection_Logic_Reference.txt",
                mime="text/plain",
                key="at_detection_logic_download",
                help="Download the full detection logic document for inclusion "
                     "in your Computer System Validation package."
            )

        # ── Column guidance card + sample download ────────────────────────────
        with st.expander("📋 What columns do I need? (click to expand)", expanded=False):
            st.markdown("""
**Required columns (3) — the engine will not run without these:**

| Column | What it contains | Example values |
|---|---|---|
| `timestamp` | Date-time of the event | `2024-03-15 02:14:33` |
| `user_id` | Username who performed the action | `jsmith`, `admin_user` |
| `action_type` | What action was taken | `UPDATE`, `DELETE`, `INSERT`, `LOGIN` |

**Optional columns — unlock additional AI Skill detection rules:**

| Column | What it contains | Unlocks |
|---|---|---|
| `record_type` | Table or record category touched | Rules 1, 3, 4 |
| `role` | User permission level | Rule 3 — Admin/GxP Conflict |
| `record_id` | ID of the record affected | Delete→Recreate detection |
| `comments` | Change reason / rationale text | Rule 1 — Vague Rationale |
| `new_value` | The value that was written | Rule 4 — Change Control Drift |

**Column names don't need to match exactly.** The column mapper in Step 2 lets you
match your system's export column names to the fields above — rename nothing in advance.

**Your system probably exports these under different names, for example:**

| Your system might call it | Maps to |
|---|---|
| `Event DateTime`, `Logged At`, `Date/Time` | `timestamp` |
| `Performed By`, `Modified By`, `Actor` | `user_id` |
| `Event Type`, `Operation`, `Transaction` | `action_type` |
| `Table Name`, `Object`, `Module` | `record_type` |
| `User Role`, `Permission`, `Access Level` | `role` |
| `Record Number`, `Object ID`, `Key` | `record_id` |
| `Reason for Change`, `Justification`, `Note` | `comments` |
| `Changed To`, `New Entry`, `Value After` | `new_value` |
""")

        # ── Generate and offer sample audit log template download (xlsx) ──────
        def _build_sample_xlsx() -> bytes:
            """
            Returns a two-sheet Excel workbook:
              Sheet 1 — Usage:     Plain-English guide to every column and the
                                   detection rules the tool can identify.
              Sheet 2 — Audit Log: 1,000-row realistic GxP audit trail using
                                   the validated labware_comprehensive_test
                                   dataset. Covers all 10 detection scenarios.
            Replace Sheet 2 with your own system export to analyse real data.
            """
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()

            # ── Sheet 1: Usage Instructions ───────────────────────────────────
            ws_usage = wb.active
            ws_usage.title = "Usage Instructions"

            # All white background, black text — clean and readable
            white_fill = PatternFill("solid", fgColor="FFFFFF")
            hdr_font   = Font(bold=True, color="000000", size=12)
            sec_font   = Font(bold=True, color="000000", size=10)
            body_font  = Font(color="000000", size=9)
            key_font   = Font(bold=True, color="000000", size=9)
            thin       = Side(style="thin", color="CCCCCC")
            border     = Border(bottom=thin)

            def _uw(row_num, col_num, value, font=None, fill=None, wrap=False):
                cell = ws_usage.cell(row=row_num, column=col_num, value=value)
                cell.fill = white_fill
                if font:  cell.font  = font
                cell.alignment = Alignment(wrap_text=wrap, vertical="top")
                return cell

            ws_usage.column_dimensions["A"].width = 28
            ws_usage.column_dimensions["B"].width = 72
            ws_usage.sheet_view.showGridLines = False

            r = 1
            _uw(r, 1, "VALINTEL.AI — Audit Trail Review Intelligence",
                Font(bold=True, color="000000", size=13))
            ws_usage.merge_cells(f"A{r}:B{r}")
            ws_usage.row_dimensions[r].height = 26
            r += 1

            _uw(r, 1, "Sample Audit Log Template — Usage Instructions",
                Font(italic=True, color="444444", size=9))
            ws_usage.merge_cells(f"A{r}:B{r}")
            r += 2

            # Column guide
            _uw(r, 1, "COLUMN REFERENCE", sec_font)
            _uw(r, 2, "Description & accepted values", sec_font)
            ws_usage.cell(r, 1).border = Border(bottom=Side(style="medium", color="000000"))
            ws_usage.cell(r, 2).border = Border(bottom=Side(style="medium", color="000000"))
            r += 1

            col_guide = [
                ("timestamp",    "Date/time of the event. Format: YYYY-MM-DD HH:MM:SS  (required)"),
                ("user_id",      "Username or account ID that performed the action  (required)"),
                ("action_type",  "Type of operation, e.g. UPDATE, INSERT, DELETE, LOGIN, LOGIN_FAILED  (required)"),
                ("record_type",  "GxP table or entity affected, e.g. RESULTS, BATCH_RELEASE, AUDIT_TRAIL"),
                ("role",         "System role of the user at time of event, e.g. Analyst, Admin, DBA"),
                ("record_id",    "Unique identifier of the record changed, e.g. RES-001, BAT-2024-001"),
                ("comments",     "Change rationale or reason field as logged by the system"),
                ("new_value",    "Updated value after the change (numeric or text)"),
            ]
            for col, desc in col_guide:
                _uw(r, 1, col,  key_font, wrap=True)
                _uw(r, 2, desc, body_font, wrap=True)
                ws_usage.row_dimensions[r].height = 18
                r += 1

            r += 1
            _uw(r, 1, "DETECTION SCENARIOS IN THIS TEMPLATE", sec_font)
            _uw(r, 2, "Rule triggered — what to look for", sec_font)
            ws_usage.cell(r, 1).border = Border(bottom=Side(style="medium", color="000000"))
            ws_usage.cell(r, 2).border = Border(bottom=Side(style="medium", color="000000"))
            r += 1

            scenarios = [
                ("Rule 1 — Vague Rationale",
                 "4 UPDATE events where comments = 'Error', 'Correction', 'fixed', 'ok'. "
                 "These should surface as High risk in the escalated events list."),
                ("Rule 2 — Contemporaneous Burst",
                 "12 RESULT_INSERT actions by analyst_jones within 2 minutes (23:00–23:01). "
                 "Threshold is >10 inserts in any 15-minute window by the same user."),
                ("Rule 3 — Admin / GxP Conflict",
                 "admin_sys performs INSERT on BATCH_RELEASE — admin accounts should never "
                 "write directly to production GxP tables (segregation of duties)."),
                ("Rule 4 — Change Control Drift",
                 "A new_value of 147.3 is >3 standard deviations above the dataset mean (~7.4). "
                 "Flags statistically anomalous numeric changes."),
                ("Rule 5 — Failed Login → Data Manipulation",
                 "analyst_x has 3 LOGIN_FAILED events at 10:00–10:06, successful login at 10:08, "
                 "then a DELETE on RESULTS (RES-5050) just 7 minutes later."),
                ("Rule 6 — Delete → Recreate Same Record",
                 "analyst_y DELETEs RES-8888 at 18:00 then creates a new entry for RES-8888 "
                 "at 18:15 — a known method of modifying locked GxP records."),
                ("Rule 7 — Audit Trail Integrity Event",
                 "dba_prod performs UPDATE on AUDIT_TRAIL at 13:00 — any modification of the "
                 "audit trail itself is Critical by definition."),
                ("Rule 8 — Privileged User on GxP Data",
                 "admin_sys and dba_prod (Admin role) act directly on GxP tables. "
                 "Complements Rule 3 with broader record-type coverage."),
                ("Off-Hours Activity",
                 "Entry at 02:14 and multiple entries after 20:00. The engine flags activity "
                 "outside standard business hours (08:00–20:00) as elevated risk."),
                ("Weekend + Holiday Activity",
                 "2026-03-28 (Saturday) and 2026-07-04 (Saturday + US federal holiday July 4th). "
                 "Both are flagged — holiday detection is independent of weekend detection."),
                ("Timestamp Gap",
                 "A 5h 46m gap between 02:14 and 08:00, and a 98-day gap before the July 4 entry. "
                 "Gaps > 2 hours during expected activity periods may indicate audit trail tampering."),
            ]
            for name, detail in scenarios:
                _uw(r, 1, name,   key_font, wrap=True)
                _uw(r, 2, detail, body_font, wrap=True)
                ws_usage.row_dimensions[r].height = 30
                r += 1

            r += 1
            _uw(r, 1, "HOW TO USE YOUR OWN DATA", sec_font)
            ws_usage.merge_cells(f"A{r}:B{r}")
            ws_usage.cell(r, 1).border = Border(bottom=Side(style="medium", color="000000"))
            r += 1
            instructions = [
                "1. Export your GxP system's audit trail as CSV or Excel.",
                "2. Delete or replace the data in the 'Audit Log' sheet (keep the header row).",
                "3. Upload the file in the VALINTEL.AI Audit Trail Review module.",
                "4. Use the column mapper to match your system's column names to the required fields.",
                "5. Run the analysis — the engine scores every event and escalates the top 20.",
            ]
            for inst in instructions:
                _uw(r, 1, inst, body_font, wrap=True)
                ws_usage.merge_cells(f"A{r}:B{r}")
                ws_usage.row_dimensions[r].height = 16
                r += 1

            # ── Sheet 2: Audit Log data ───────────────────────────────────────
            ws_data = wb.create_sheet("Audit Log")

            # Embed the validated labware_comprehensive_test dataset
            HEADER = ["timestamp","user_id","action_type","record_type",
                      "role","record_id","comments","new_value"]

            DATA_ROWS = [
                ["2026-03-23 02:14:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1401", "Insomnia entry", "7.5"],
                ["2026-03-23 08:00:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1000", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-23 08:05:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1001", "Standard value entry per SOP-01", "7.05"],
                ["2026-03-23 08:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1002", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-23 08:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1003", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-23 08:20:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1004", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-23 08:25:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1005", "Standard value entry per SOP-01", "7.5"],
                ["2026-03-23 08:30:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1006", "Standard value entry per SOP-01", "7.41"],
                ["2026-03-23 08:35:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1007", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-23 08:40:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1008", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-23 08:45:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1009", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-23 08:55:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1011", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-23 09:00:00", "admin_sys", "INSERT", "BATCH_RELEASE", "Admin", "BATCH-999", "Urgent release", "RELEASED"],
                ["2026-03-23 09:00:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1012", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-23 09:05:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1013", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-23 09:10:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1014", "Standard value entry per SOP-01", "7.5"],
                ["2026-03-23 09:15:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1015", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-23 09:20:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1016", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-23 09:25:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1017", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-23 09:30:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1018", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-23 09:35:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1019", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-23 10:00:00", "analyst_x", "LOGIN_FAILED", "USER_SESSION", "Analyst", "SES-01", "Wrong password", ""],
                ["2026-03-23 10:03:00", "analyst_x", "LOGIN_FAILED", "USER_SESSION", "Analyst", "SES-01", "Wrong password", ""],
                ["2026-03-23 10:05:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1025", "Standard value entry per SOP-01", "7.37"],
                ["2026-03-23 10:06:00", "analyst_x", "LOGIN_FAILED", "USER_SESSION", "Analyst", "SES-01", "Wrong password", ""],
                ["2026-03-23 10:08:00", "analyst_x", "LOGIN", "USER_SESSION", "Analyst", "SES-01", "Success", ""],
                ["2026-03-23 10:10:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1026", "Standard value entry per SOP-01", "7.27"],
                ["2026-03-23 10:15:00", "analyst_x", "DELETE", "RESULTS", "Analyst", "RES-5050", "Cleaning up error", ""],
                ["2026-03-23 10:15:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1027", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-23 10:20:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1028", "Standard value entry per SOP-01", "7.43"],
                ["2026-03-23 10:25:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1029", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-23 10:30:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1030", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-23 10:35:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1031", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-23 10:40:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1032", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-23 10:45:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1033", "Standard value entry per SOP-01", "7.43"],
                ["2026-03-23 10:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1034", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-23 10:55:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1035", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-23 11:00:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1036", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-23 11:05:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1037", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-23 11:10:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1038", "Standard value entry per SOP-01", "7.21"],
                ["2026-03-23 11:15:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1039", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-23 11:20:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1040", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-23 11:25:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1041", "Standard value entry per SOP-01", "7.21"],
                ["2026-03-23 11:30:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1042", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-23 11:35:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1043", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-23 11:40:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1044", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-23 11:45:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1045", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-23 11:50:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1046", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-23 11:55:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1047", "Standard value entry per SOP-01", "7.15"],
                ["2026-03-23 12:00:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1048", "Standard value entry per SOP-01", "7.05"],
                ["2026-03-23 12:05:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1049", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-23 12:15:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1051", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-23 12:20:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1052", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-23 12:25:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1053", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-23 12:30:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1054", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-23 12:35:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1055", "Standard value entry per SOP-01", "7.46"],
                ["2026-03-23 12:40:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1056", "Standard value entry per SOP-01", "7.33"],
                ["2026-03-23 12:45:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1057", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-23 12:50:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1058", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-23 12:55:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1059", "Standard value entry per SOP-01", "7.5"],
                ["2026-03-23 13:00:00", "dba_prod", "UPDATE", "AUDIT_TRAIL", "Admin", "SYS-001", "System maintenance", "DISABLED"],
                ["2026-03-23 13:00:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1060", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-23 13:05:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1061", "Standard value entry per SOP-01", "7.06"],
                ["2026-03-23 13:10:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1062", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-23 13:15:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1063", "Standard value entry per SOP-01", "7.38"],
                ["2026-03-23 13:20:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1064", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-23 13:25:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1065", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-23 13:30:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1066", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-23 13:35:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1067", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-23 13:40:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1068", "Standard value entry per SOP-01", "7.06"],
                ["2026-03-23 13:45:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1069", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-23 13:50:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1070", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-23 13:55:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1071", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-23 14:00:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1072", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-23 14:05:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1073", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-23 14:10:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1074", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-23 14:15:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1075", "Standard value entry per SOP-01", "7.27"],
                ["2026-03-23 14:20:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1076", "Standard value entry per SOP-01", "7.2"],
                ["2026-03-23 14:25:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1077", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-23 14:30:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1078", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-23 14:35:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1079", "Standard value entry per SOP-01", "7.21"],
                ["2026-03-23 14:40:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1080", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-23 14:45:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1081", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-23 14:50:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1082", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-23 14:55:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1083", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-23 15:00:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1084", "Standard value entry per SOP-01", "7.06"],
                ["2026-03-23 15:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1085", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-23 15:10:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1086", "Standard value entry per SOP-01", "7.06"],
                ["2026-03-23 15:15:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1087", "Standard value entry per SOP-01", "7.07"],
                ["2026-03-23 15:20:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1088", "Standard value entry per SOP-01", "7.2"],
                ["2026-03-23 15:25:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1089", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-23 15:30:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1090", "Standard value entry per SOP-01", "7.24"],
                ["2026-03-23 15:35:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1091", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-23 15:40:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1092", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-23 15:45:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1093", "Standard value entry per SOP-01", "7.27"],
                ["2026-03-23 15:50:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1094", "Standard value entry per SOP-01", "7.5"],
                ["2026-03-23 15:55:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1095", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-23 16:00:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1096", "Standard value entry per SOP-01", "7.28"],
                ["2026-03-23 16:05:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1097", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-23 16:10:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1098", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-23 16:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1099", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-23 16:25:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1101", "Standard value entry per SOP-01", "7.33"],
                ["2026-03-23 16:30:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1102", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-23 16:35:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1103", "Standard value entry per SOP-01", "7.06"],
                ["2026-03-23 16:40:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1104", "Standard value entry per SOP-01", "7.07"],
                ["2026-03-23 16:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1106", "Standard value entry per SOP-01", "7.24"],
                ["2026-03-23 16:55:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1107", "Standard value entry per SOP-01", "7.21"],
                ["2026-03-23 17:00:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1108", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-23 17:05:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1109", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-23 17:10:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1110", "Standard value entry per SOP-01", "7.06"],
                ["2026-03-23 17:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1111", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-23 17:20:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1112", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-23 17:25:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1113", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-23 17:30:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1114", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-23 17:35:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1115", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-23 17:40:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1116", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-23 17:45:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1117", "Standard value entry per SOP-01", "7.08"],
                ["2026-03-23 17:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1118", "Standard value entry per SOP-01", "7.46"],
                ["2026-03-23 17:55:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1119", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-23 18:00:00", "analyst_y", "DELETE", "RESULTS", "Analyst", "RES-8888", "Error", ""],
                ["2026-03-23 18:00:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1120", "Standard value entry per SOP-01", "7.33"],
                ["2026-03-23 18:05:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1121", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-23 18:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1122", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-23 18:15:00", "analyst_y", "INSERT", "RESULTS", "Analyst", "RES-8888", "Correction", "7.2"],
                ["2026-03-23 18:15:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1123", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-23 18:20:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1124", "Standard value entry per SOP-01", "7.16"],
                ["2026-03-23 18:25:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1125", "Standard value entry per SOP-01", "7.16"],
                ["2026-03-23 18:30:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1126", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-23 18:35:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1127", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-23 18:40:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1128", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-23 18:45:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1129", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-23 18:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1130", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-23 18:55:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1131", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-23 19:00:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1132", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-23 19:05:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1133", "Standard value entry per SOP-01", "7.05"],
                ["2026-03-23 19:10:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1134", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-23 19:15:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1135", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-23 19:20:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1136", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-23 19:25:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1137", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-23 19:30:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1138", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-23 19:35:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1139", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-23 19:40:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1140", "Standard value entry per SOP-01", "7.15"],
                ["2026-03-23 19:45:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1141", "Standard value entry per SOP-01", "7.16"],
                ["2026-03-23 19:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1142", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-23 19:55:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1143", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-23 20:00:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1144", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-23 20:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1145", "Standard value entry per SOP-01", "7.46"],
                ["2026-03-23 20:10:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1146", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-23 20:15:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1147", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-23 20:20:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1148", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-23 20:25:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1149", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-23 20:30:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1150", "", "7.13"],
                ["2026-03-23 20:35:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1151", "fixed", "7.02"],
                ["2026-03-23 20:40:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1152", "ok", "7.16"],
                ["2026-03-23 20:45:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1153", "Standard value entry per SOP-01", "7.33"],
                ["2026-03-23 20:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1154", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-23 20:55:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1155", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-23 21:00:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1156", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-23 21:05:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1157", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-23 21:10:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1158", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-23 21:15:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1159", "Standard value entry per SOP-01", "7.27"],
                ["2026-03-23 21:20:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1160", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-23 21:25:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1161", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-23 21:30:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1162", "Standard value entry per SOP-01", "7.2"],
                ["2026-03-23 21:35:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1163", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-23 21:40:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1164", "Standard value entry per SOP-01", "7.2"],
                ["2026-03-23 21:45:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1165", "Standard value entry per SOP-01", "7.37"],
                ["2026-03-23 21:50:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1166", "Standard value entry per SOP-01", "7.46"],
                ["2026-03-23 21:55:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1167", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-23 22:00:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1168", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-23 22:05:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1169", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-23 22:10:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1170", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-23 22:15:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1171", "Standard value entry per SOP-01", "7.05"],
                ["2026-03-23 22:20:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1172", "Standard value entry per SOP-01", "7.41"],
                ["2026-03-23 22:25:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1173", "Standard value entry per SOP-01", "7.21"],
                ["2026-03-23 22:30:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1174", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-23 22:35:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1175", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-23 22:40:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1176", "Standard value entry per SOP-01", "7.06"],
                ["2026-03-23 22:45:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1177", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-23 22:50:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1178", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-23 22:55:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1179", "Standard value entry per SOP-01", "7.07"],
                ["2026-03-23 23:00:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1180", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-23 23:00:00", "analyst_jones", "RESULT_INSERT", "RESULTS", "Analyst", "RES-BURST-0", "Bulk upload", "7.2"],
                ["2026-03-23 23:00:10", "analyst_jones", "RESULT_INSERT", "RESULTS", "Analyst", "RES-BURST-1", "Bulk upload", "7.2"],
                ["2026-03-23 23:00:20", "analyst_jones", "RESULT_INSERT", "RESULTS", "Analyst", "RES-BURST-2", "Bulk upload", "7.2"],
                ["2026-03-23 23:00:30", "analyst_jones", "RESULT_INSERT", "RESULTS", "Analyst", "RES-BURST-3", "Bulk upload", "7.2"],
                ["2026-03-23 23:00:40", "analyst_jones", "RESULT_INSERT", "RESULTS", "Analyst", "RES-BURST-4", "Bulk upload", "7.2"],
                ["2026-03-23 23:00:50", "analyst_jones", "RESULT_INSERT", "RESULTS", "Analyst", "RES-BURST-5", "Bulk upload", "7.2"],
                ["2026-03-23 23:01:00", "analyst_jones", "RESULT_INSERT", "RESULTS", "Analyst", "RES-BURST-6", "Bulk upload", "7.2"],
                ["2026-03-23 23:01:10", "analyst_jones", "RESULT_INSERT", "RESULTS", "Analyst", "RES-BURST-7", "Bulk upload", "7.2"],
                ["2026-03-23 23:01:20", "analyst_jones", "RESULT_INSERT", "RESULTS", "Analyst", "RES-BURST-8", "Bulk upload", "7.2"],
                ["2026-03-23 23:01:30", "analyst_jones", "RESULT_INSERT", "RESULTS", "Analyst", "RES-BURST-9", "Bulk upload", "7.2"],
                ["2026-03-23 23:01:40", "analyst_jones", "RESULT_INSERT", "RESULTS", "Analyst", "RES-BURST-10", "Bulk upload", "7.2"],
                ["2026-03-23 23:01:50", "analyst_jones", "RESULT_INSERT", "RESULTS", "Analyst", "RES-BURST-11", "Bulk upload", "7.2"],
                ["2026-03-23 23:05:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1181", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-23 23:10:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1182", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-23 23:15:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1183", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-23 23:20:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1184", "Standard value entry per SOP-01", "7.28"],
                ["2026-03-23 23:25:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1185", "Standard value entry per SOP-01", "7.07"],
                ["2026-03-23 23:30:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1186", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-23 23:30:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1400", "Late night check", "7.32"],
                ["2026-03-23 23:35:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1187", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-23 23:40:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1188", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-23 23:45:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1189", "Standard value entry per SOP-01", "7.33"],
                ["2026-03-23 23:50:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1190", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-23 23:55:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1191", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-24 00:00:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1192", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-24 00:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1193", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-24 00:10:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1194", "Standard value entry per SOP-01", "7.37"],
                ["2026-03-24 00:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1195", "Standard value entry per SOP-01", "7.47"],
                ["2026-03-24 00:20:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1196", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-24 00:25:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1197", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-24 00:30:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1198", "Standard value entry per SOP-01", "7.41"],
                ["2026-03-24 00:35:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1199", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-24 00:40:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1200", "Outlier test", "147.3"],
                ["2026-03-24 00:45:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1201", "Standard value entry per SOP-01", "7.0"],
                ["2026-03-24 00:50:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1202", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-24 00:55:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1203", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-24 01:00:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1204", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-24 01:05:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1205", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-24 01:10:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1206", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-24 01:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1207", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-24 01:20:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1208", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-24 01:25:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1209", "Standard value entry per SOP-01", "7.33"],
                ["2026-03-24 01:30:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1210", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-24 01:35:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1211", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-24 01:40:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1212", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-24 01:45:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1213", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-24 01:50:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1214", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-24 01:55:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1215", "Standard value entry per SOP-01", "7.06"],
                ["2026-03-24 02:00:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1216", "Standard value entry per SOP-01", "7.38"],
                ["2026-03-24 02:05:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1217", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-24 02:10:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1218", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-24 02:15:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1219", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-24 02:20:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1220", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-24 02:25:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1221", "Standard value entry per SOP-01", "7.27"],
                ["2026-03-24 02:30:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1222", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-24 02:35:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1223", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-24 02:40:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1224", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-24 02:45:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1225", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-24 02:50:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1226", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-24 02:55:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1227", "Standard value entry per SOP-01", "7.28"],
                ["2026-03-24 03:00:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1228", "Standard value entry per SOP-01", "7.07"],
                ["2026-03-24 03:05:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1229", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-24 03:10:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1230", "Standard value entry per SOP-01", "7.47"],
                ["2026-03-24 03:15:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1231", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-24 03:20:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1232", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-24 03:25:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1233", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-24 03:30:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1234", "Standard value entry per SOP-01", "7.05"],
                ["2026-03-24 03:35:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1235", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-24 03:40:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1236", "Standard value entry per SOP-01", "7.0"],
                ["2026-03-24 03:45:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1237", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-24 03:50:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1238", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-24 03:55:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1239", "Standard value entry per SOP-01", "7.07"],
                ["2026-03-24 04:00:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1240", "Standard value entry per SOP-01", "7.28"],
                ["2026-03-24 04:05:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1241", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-24 04:10:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1242", "Standard value entry per SOP-01", "7.39"],
                ["2026-03-24 04:15:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1243", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-24 04:20:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1244", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-24 04:25:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1245", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-24 04:30:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1246", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-24 04:35:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1247", "Standard value entry per SOP-01", "7.5"],
                ["2026-03-24 04:40:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1248", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-24 04:45:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1249", "Standard value entry per SOP-01", "7.28"],
                ["2026-03-24 04:50:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1250", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-24 04:55:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1251", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-24 05:00:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1252", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-24 05:05:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1253", "Standard value entry per SOP-01", "7.5"],
                ["2026-03-24 05:10:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1254", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-24 05:15:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1255", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-24 05:20:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1256", "Standard value entry per SOP-01", "7.15"],
                ["2026-03-24 05:25:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1257", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-24 05:30:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1258", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-24 05:35:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1259", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-24 05:40:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1260", "Standard value entry per SOP-01", "7.46"],
                ["2026-03-24 05:45:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1261", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-24 05:50:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1262", "Standard value entry per SOP-01", "7.16"],
                ["2026-03-24 05:55:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1263", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-24 06:00:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1264", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-24 06:05:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1265", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-24 06:10:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1266", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-24 06:15:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1267", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-24 06:20:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1268", "Standard value entry per SOP-01", "7.37"],
                ["2026-03-24 06:25:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1269", "Standard value entry per SOP-01", "7.15"],
                ["2026-03-24 06:30:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1270", "Standard value entry per SOP-01", "7.15"],
                ["2026-03-24 06:35:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1271", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-24 06:40:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1272", "Standard value entry per SOP-01", "7.43"],
                ["2026-03-24 06:45:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1273", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-24 06:50:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1274", "Standard value entry per SOP-01", "7.41"],
                ["2026-03-24 06:55:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1275", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-24 07:00:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1276", "Standard value entry per SOP-01", "7.37"],
                ["2026-03-24 07:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1277", "Standard value entry per SOP-01", "7.37"],
                ["2026-03-24 07:10:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1278", "Standard value entry per SOP-01", "7.24"],
                ["2026-03-24 07:15:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1279", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-24 07:20:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1280", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-24 07:25:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1281", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-24 07:30:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1282", "Standard value entry per SOP-01", "7.47"],
                ["2026-03-24 07:35:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1283", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-24 07:40:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1284", "Standard value entry per SOP-01", "7.06"],
                ["2026-03-24 07:45:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1285", "Standard value entry per SOP-01", "7.2"],
                ["2026-03-24 07:50:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1286", "Standard value entry per SOP-01", "7.41"],
                ["2026-03-24 07:55:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1287", "Standard value entry per SOP-01", "7.2"],
                ["2026-03-24 08:00:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1288", "Standard value entry per SOP-01", "7.38"],
                ["2026-03-24 08:05:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1289", "Standard value entry per SOP-01", "7.41"],
                ["2026-03-24 08:10:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1290", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-24 08:15:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1291", "Standard value entry per SOP-01", "7.46"],
                ["2026-03-24 08:20:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1292", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-24 08:25:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1293", "Standard value entry per SOP-01", "7.08"],
                ["2026-03-24 08:30:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1294", "Standard value entry per SOP-01", "7.41"],
                ["2026-03-24 08:35:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1295", "Standard value entry per SOP-01", "7.0"],
                ["2026-03-24 08:40:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1296", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-24 08:45:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1297", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-24 08:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1298", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-24 08:55:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1299", "Standard value entry per SOP-01", "7.39"],
                ["2026-03-24 10:00:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1312", "Standard value entry per SOP-01", "7.37"],
                ["2026-03-24 10:05:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1313", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-24 10:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1314", "Standard value entry per SOP-01", "7.47"],
                ["2026-03-24 10:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1315", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-24 10:20:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1316", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-24 10:25:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1317", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-24 10:30:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1318", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-24 10:35:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1319", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-24 10:40:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1320", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-24 10:45:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1321", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-24 10:50:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1322", "Standard value entry per SOP-01", "7.33"],
                ["2026-03-24 10:55:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1323", "Standard value entry per SOP-01", "7.38"],
                ["2026-03-24 11:00:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1324", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-24 11:05:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1325", "Standard value entry per SOP-01", "7.07"],
                ["2026-03-24 11:10:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1326", "Standard value entry per SOP-01", "7.5"],
                ["2026-03-24 11:15:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1327", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-24 11:20:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1328", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-24 11:25:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1329", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-24 11:30:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1330", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-24 11:35:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1331", "Standard value entry per SOP-01", "7.28"],
                ["2026-03-24 11:40:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1332", "Standard value entry per SOP-01", "7.0"],
                ["2026-03-24 11:45:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1333", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-24 11:50:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1334", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-24 11:55:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1335", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-24 12:00:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1336", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-24 12:05:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1337", "Standard value entry per SOP-01", "7.47"],
                ["2026-03-24 12:10:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1338", "Standard value entry per SOP-01", "7.15"],
                ["2026-03-24 12:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1339", "Standard value entry per SOP-01", "7.41"],
                ["2026-03-24 12:20:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1340", "Standard value entry per SOP-01", "7.15"],
                ["2026-03-24 12:25:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1341", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-24 12:30:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1342", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-24 12:35:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1343", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-24 12:40:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1344", "Standard value entry per SOP-01", "7.27"],
                ["2026-03-24 12:45:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1345", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-24 12:50:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1346", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-24 12:55:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1347", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-24 13:00:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1348", "Standard value entry per SOP-01", "7.38"],
                ["2026-03-24 13:05:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1349", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-24 13:10:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1350", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-24 13:15:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1351", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-24 13:20:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1352", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-24 13:25:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1353", "Standard value entry per SOP-01", "7.37"],
                ["2026-03-24 13:30:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1354", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-24 13:35:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1355", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-24 13:40:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1356", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-24 13:45:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1357", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-24 13:50:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1358", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-24 13:55:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1359", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-24 14:00:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1360", "Standard value entry per SOP-01", "7.06"],
                ["2026-03-24 14:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1361", "Standard value entry per SOP-01", "7.06"],
                ["2026-03-24 14:10:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1362", "Standard value entry per SOP-01", "7.5"],
                ["2026-03-24 14:15:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1363", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-24 14:20:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1364", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-24 14:25:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1365", "Standard value entry per SOP-01", "7.39"],
                ["2026-03-24 14:30:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1366", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-24 14:35:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1367", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-24 14:40:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1368", "Standard value entry per SOP-01", "7.08"],
                ["2026-03-24 14:45:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1369", "Standard value entry per SOP-01", "7.24"],
                ["2026-03-24 14:50:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1370", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-24 14:55:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1371", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-24 15:00:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1372", "Standard value entry per SOP-01", "7.2"],
                ["2026-03-24 15:05:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1373", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-24 15:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1374", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-24 15:15:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1375", "Standard value entry per SOP-01", "7.08"],
                ["2026-03-24 15:20:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1376", "Standard value entry per SOP-01", "7.43"],
                ["2026-03-24 15:25:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1377", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-24 15:30:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1378", "Standard value entry per SOP-01", "7.28"],
                ["2026-03-24 15:35:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1379", "Standard value entry per SOP-01", "7.28"],
                ["2026-03-24 15:40:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1380", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-24 15:45:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1381", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-24 15:50:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1382", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-24 15:55:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1383", "Standard value entry per SOP-01", "7.38"],
                ["2026-03-24 16:00:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1384", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-24 16:05:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1385", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-24 16:10:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1386", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-24 16:15:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1387", "Standard value entry per SOP-01", "7.43"],
                ["2026-03-24 16:20:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1388", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-24 16:25:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1389", "Standard value entry per SOP-01", "7.28"],
                ["2026-03-24 16:30:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1390", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-24 16:35:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1391", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-24 16:40:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1392", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-24 16:45:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1393", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-24 16:50:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1394", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-24 16:55:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1395", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-24 17:00:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1396", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-24 17:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1397", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-24 17:10:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1398", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-24 17:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1399", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-24 17:30:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1402", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-24 17:35:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1403", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-24 17:40:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1404", "Standard value entry per SOP-01", "7.21"],
                ["2026-03-24 17:45:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1405", "Standard value entry per SOP-01", "7.16"],
                ["2026-03-24 17:50:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1406", "Standard value entry per SOP-01", "7.39"],
                ["2026-03-24 17:55:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1407", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-24 18:00:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1408", "Standard value entry per SOP-01", "7.08"],
                ["2026-03-24 18:05:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1409", "Standard value entry per SOP-01", "7.33"],
                ["2026-03-24 18:10:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1410", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-24 18:15:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1411", "Standard value entry per SOP-01", "7.38"],
                ["2026-03-24 18:20:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1412", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-24 18:25:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1413", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-24 18:30:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1414", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-24 18:35:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1415", "Standard value entry per SOP-01", "7.21"],
                ["2026-03-24 18:40:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1416", "Standard value entry per SOP-01", "7.07"],
                ["2026-03-24 18:45:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1417", "Standard value entry per SOP-01", "7.08"],
                ["2026-03-24 18:50:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1418", "Standard value entry per SOP-01", "7.21"],
                ["2026-03-24 18:55:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1419", "Standard value entry per SOP-01", "7.37"],
                ["2026-03-24 19:00:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1420", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-24 19:05:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1421", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-24 19:10:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1422", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-24 19:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1423", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-24 19:20:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1424", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-24 19:25:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1425", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-24 19:30:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1426", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-24 19:35:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1427", "Standard value entry per SOP-01", "7.05"],
                ["2026-03-24 19:40:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1428", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-24 19:45:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1429", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-24 19:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1430", "Standard value entry per SOP-01", "7.43"],
                ["2026-03-24 19:55:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1431", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-24 20:00:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1432", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-24 20:05:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1433", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-24 20:10:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1434", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-24 20:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1435", "Standard value entry per SOP-01", "7.27"],
                ["2026-03-24 20:20:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1436", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-24 20:25:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1437", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-24 20:30:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1438", "Standard value entry per SOP-01", "7.39"],
                ["2026-03-24 20:35:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1439", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-24 20:40:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1440", "Standard value entry per SOP-01", "7.16"],
                ["2026-03-24 20:45:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1441", "Standard value entry per SOP-01", "7.07"],
                ["2026-03-24 20:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1442", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-24 20:55:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1443", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-24 21:00:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1444", "Standard value entry per SOP-01", "7.0"],
                ["2026-03-24 21:05:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1445", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-24 21:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1446", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-24 21:15:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1447", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-24 21:20:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1448", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-24 21:25:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1449", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-24 21:30:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1450", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-24 21:35:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1451", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-24 21:40:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1452", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-24 21:45:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1453", "Standard value entry per SOP-01", "7.43"],
                ["2026-03-24 21:50:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1454", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-24 21:55:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1455", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-24 22:00:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1456", "Standard value entry per SOP-01", "7.06"],
                ["2026-03-24 22:05:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1457", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-24 22:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1458", "Standard value entry per SOP-01", "7.0"],
                ["2026-03-24 22:15:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1459", "Standard value entry per SOP-01", "7.21"],
                ["2026-03-24 22:20:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1460", "Standard value entry per SOP-01", "7.21"],
                ["2026-03-24 22:25:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1461", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-24 22:30:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1462", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-24 22:35:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1463", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-24 22:40:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1464", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-24 22:45:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1465", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-24 22:50:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1466", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-24 22:55:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1467", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-24 23:00:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1468", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-24 23:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1469", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-24 23:10:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1470", "Standard value entry per SOP-01", "7.39"],
                ["2026-03-24 23:15:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1471", "Standard value entry per SOP-01", "7.21"],
                ["2026-03-24 23:20:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1472", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-24 23:25:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1473", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-24 23:30:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1474", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-24 23:35:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1475", "Standard value entry per SOP-01", "7.15"],
                ["2026-03-24 23:40:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1476", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-24 23:45:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1477", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-24 23:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1478", "Standard value entry per SOP-01", "7.24"],
                ["2026-03-24 23:55:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1479", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-25 00:00:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1480", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-25 00:05:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1481", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-25 00:10:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1482", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-25 00:15:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1483", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-25 00:20:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1484", "Standard value entry per SOP-01", "7.24"],
                ["2026-03-25 00:25:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1485", "Standard value entry per SOP-01", "7.05"],
                ["2026-03-25 00:30:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1486", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-25 00:35:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1487", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-25 00:40:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1488", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-25 00:45:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1489", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-25 00:50:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1490", "Standard value entry per SOP-01", "7.05"],
                ["2026-03-25 00:55:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1491", "Standard value entry per SOP-01", "7.46"],
                ["2026-03-25 01:00:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1492", "Standard value entry per SOP-01", "7.46"],
                ["2026-03-25 01:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1493", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-25 01:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1494", "Standard value entry per SOP-01", "7.39"],
                ["2026-03-25 01:15:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1495", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-25 01:20:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1496", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-25 01:25:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1497", "Standard value entry per SOP-01", "7.33"],
                ["2026-03-25 01:30:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1498", "Standard value entry per SOP-01", "7.33"],
                ["2026-03-25 01:35:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1499", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-25 01:45:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1501", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-25 01:50:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1502", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-25 01:55:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1503", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-25 02:00:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1504", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-25 02:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1505", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-25 02:10:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1506", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-25 02:15:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1507", "Standard value entry per SOP-01", "7.41"],
                ["2026-03-25 02:20:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1508", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-25 02:25:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1509", "Standard value entry per SOP-01", "7.47"],
                ["2026-03-25 02:30:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1510", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-25 02:35:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1511", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-25 02:40:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1512", "Standard value entry per SOP-01", "7.07"],
                ["2026-03-25 02:45:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1513", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-25 02:50:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1514", "Standard value entry per SOP-01", "7.46"],
                ["2026-03-25 02:55:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1515", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-25 03:00:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1516", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-25 03:05:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1517", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-25 03:10:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1518", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-25 03:15:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1519", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-25 03:20:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1520", "Standard value entry per SOP-01", "7.2"],
                ["2026-03-25 03:25:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1521", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-25 03:30:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1522", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-25 03:35:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1523", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-25 03:40:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1524", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-25 03:45:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1525", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-25 03:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1526", "Standard value entry per SOP-01", "7.07"],
                ["2026-03-25 03:55:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1527", "Standard value entry per SOP-01", "7.08"],
                ["2026-03-25 04:00:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1528", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-25 04:05:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1529", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-25 04:10:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1530", "Standard value entry per SOP-01", "7.27"],
                ["2026-03-25 04:15:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1531", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-25 04:20:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1532", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-25 04:25:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1533", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-25 04:30:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1534", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-25 04:35:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1535", "Standard value entry per SOP-01", "7.08"],
                ["2026-03-25 04:40:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1536", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-25 04:45:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1537", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-25 04:50:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1538", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-25 04:55:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1539", "Standard value entry per SOP-01", "7.38"],
                ["2026-03-25 05:00:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1540", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-25 05:05:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1541", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-25 05:10:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1542", "Standard value entry per SOP-01", "7.08"],
                ["2026-03-25 05:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1543", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-25 05:20:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1544", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-25 05:25:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1545", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-25 05:30:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1546", "Standard value entry per SOP-01", "7.47"],
                ["2026-03-25 05:35:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1547", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-25 05:40:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1548", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-25 05:45:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1549", "Standard value entry per SOP-01", "7.46"],
                ["2026-03-25 05:50:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1550", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-25 05:55:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1551", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-25 06:00:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1552", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-25 06:05:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1553", "Standard value entry per SOP-01", "7.39"],
                ["2026-03-25 06:10:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1554", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-25 06:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1555", "Standard value entry per SOP-01", "7.38"],
                ["2026-03-25 06:20:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1556", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-25 06:25:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1557", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-25 06:30:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1558", "Standard value entry per SOP-01", "7.28"],
                ["2026-03-25 06:35:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1559", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-25 06:40:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1560", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-25 06:45:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1561", "Standard value entry per SOP-01", "7.16"],
                ["2026-03-25 06:50:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1562", "Standard value entry per SOP-01", "7.2"],
                ["2026-03-25 06:55:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1563", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-25 07:00:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1564", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-25 07:05:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1565", "Standard value entry per SOP-01", "7.38"],
                ["2026-03-25 07:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1566", "Standard value entry per SOP-01", "7.28"],
                ["2026-03-25 07:15:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1567", "Standard value entry per SOP-01", "7.15"],
                ["2026-03-25 07:20:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1568", "Standard value entry per SOP-01", "7.27"],
                ["2026-03-25 07:25:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1569", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-25 07:30:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1570", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-25 07:35:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1571", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-25 07:40:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1572", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-25 07:45:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1573", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-25 07:50:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1574", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-25 07:55:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1575", "Standard value entry per SOP-01", "7.41"],
                ["2026-03-25 08:00:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1576", "Standard value entry per SOP-01", "7.21"],
                ["2026-03-25 08:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1577", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-25 08:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1578", "Standard value entry per SOP-01", "7.05"],
                ["2026-03-25 08:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1579", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-25 08:20:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1580", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-25 08:25:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1581", "Standard value entry per SOP-01", "7.39"],
                ["2026-03-25 08:30:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1582", "Standard value entry per SOP-01", "7.46"],
                ["2026-03-25 08:35:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1583", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-25 08:40:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1584", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-25 08:45:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1585", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-25 08:50:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1586", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-25 08:55:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1587", "Standard value entry per SOP-01", "7.08"],
                ["2026-03-25 09:00:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1588", "Standard value entry per SOP-01", "7.05"],
                ["2026-03-25 09:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1589", "Standard value entry per SOP-01", "7.07"],
                ["2026-03-25 09:10:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1590", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-25 09:15:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1591", "Standard value entry per SOP-01", "7.16"],
                ["2026-03-25 09:20:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1592", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-25 09:25:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1593", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-25 09:30:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1594", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-25 09:35:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1595", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-25 09:40:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1596", "Standard value entry per SOP-01", "7.41"],
                ["2026-03-25 09:45:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1597", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-25 09:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1598", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-25 09:55:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1599", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-25 10:05:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1601", "Standard value entry per SOP-01", "7.05"],
                ["2026-03-25 10:10:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1602", "Standard value entry per SOP-01", "7.5"],
                ["2026-03-25 10:15:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1603", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-25 10:20:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1604", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-25 10:25:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1605", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-25 10:30:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1606", "Standard value entry per SOP-01", "7.33"],
                ["2026-03-25 10:35:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1607", "Standard value entry per SOP-01", "7.15"],
                ["2026-03-25 10:40:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1608", "Standard value entry per SOP-01", "7.46"],
                ["2026-03-25 10:45:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1609", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-25 10:50:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1610", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-25 10:55:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1611", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-25 11:00:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1612", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-25 11:05:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1613", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-25 11:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1614", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-25 11:15:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1615", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-25 11:20:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1616", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-25 11:25:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1617", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-25 11:30:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1618", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-25 11:35:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1619", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-25 11:40:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1620", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-25 11:45:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1621", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-25 11:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1622", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-25 11:55:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1623", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-25 12:00:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1624", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-25 12:05:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1625", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-25 12:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1626", "Standard value entry per SOP-01", "7.16"],
                ["2026-03-25 12:15:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1627", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-25 12:20:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1628", "Standard value entry per SOP-01", "7.47"],
                ["2026-03-25 12:25:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1629", "Standard value entry per SOP-01", "7.38"],
                ["2026-03-25 12:30:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1630", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-25 12:35:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1631", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-25 12:40:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1632", "Standard value entry per SOP-01", "7.2"],
                ["2026-03-25 12:45:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1633", "Standard value entry per SOP-01", "7.2"],
                ["2026-03-25 12:50:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1634", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-25 12:55:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1635", "Standard value entry per SOP-01", "7.41"],
                ["2026-03-25 13:00:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1636", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-25 13:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1637", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-25 13:10:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1638", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-25 13:15:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1639", "Standard value entry per SOP-01", "7.21"],
                ["2026-03-25 13:20:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1640", "Standard value entry per SOP-01", "7.47"],
                ["2026-03-25 13:25:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1641", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-25 13:30:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1642", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-25 13:35:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1643", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-25 13:40:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1644", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-25 13:45:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1645", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-25 13:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1646", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-25 13:55:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1647", "Standard value entry per SOP-01", "7.24"],
                ["2026-03-25 14:00:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1648", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-25 14:05:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1649", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-25 14:10:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1650", "Standard value entry per SOP-01", "7.07"],
                ["2026-03-25 14:15:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1651", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-25 14:20:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1652", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-25 14:25:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1653", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-25 14:30:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1654", "Standard value entry per SOP-01", "7.39"],
                ["2026-03-25 14:35:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1655", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-25 14:40:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1656", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-25 14:45:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1657", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-25 14:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1658", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-25 14:55:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1659", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-25 15:00:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1660", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-25 15:05:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1661", "Standard value entry per SOP-01", "7.28"],
                ["2026-03-25 15:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1662", "Standard value entry per SOP-01", "7.16"],
                ["2026-03-25 15:15:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1663", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-25 15:20:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1664", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-25 15:25:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1665", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-25 15:30:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1666", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-25 15:35:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1667", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-25 15:40:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1668", "Standard value entry per SOP-01", "7.43"],
                ["2026-03-25 15:45:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1669", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-25 15:50:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1670", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-25 15:55:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1671", "Standard value entry per SOP-01", "7.33"],
                ["2026-03-25 16:00:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1672", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-25 16:05:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1673", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-25 16:10:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1674", "Standard value entry per SOP-01", "7.21"],
                ["2026-03-25 16:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1675", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-25 16:20:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1676", "Standard value entry per SOP-01", "7.2"],
                ["2026-03-25 16:25:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1677", "Standard value entry per SOP-01", "7.39"],
                ["2026-03-25 16:30:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1678", "Standard value entry per SOP-01", "7.06"],
                ["2026-03-25 16:35:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1679", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-25 16:40:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1680", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-25 16:45:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1681", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-25 16:50:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1682", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-25 16:55:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1683", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-25 17:00:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1684", "Standard value entry per SOP-01", "7.16"],
                ["2026-03-25 17:00:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1700", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-25 17:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1685", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-25 17:10:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1686", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-25 17:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1687", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-25 17:20:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1688", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-25 17:25:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1689", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-25 17:30:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1690", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-25 17:35:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1691", "Standard value entry per SOP-01", "7.16"],
                ["2026-03-25 17:40:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1692", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-25 17:45:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1693", "Standard value entry per SOP-01", "7.15"],
                ["2026-03-25 17:50:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1694", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-25 17:55:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1695", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-25 18:00:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1696", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-25 18:05:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1697", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-25 18:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1698", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-25 18:15:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1699", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-25 18:30:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1702", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-25 18:35:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1703", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-25 18:40:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1704", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-25 18:45:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1705", "Standard value entry per SOP-01", "7.06"],
                ["2026-03-25 18:50:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1706", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-25 18:55:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1707", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-25 19:00:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1708", "Standard value entry per SOP-01", "7.27"],
                ["2026-03-25 19:05:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1709", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-25 19:10:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1710", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-25 19:15:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1711", "Standard value entry per SOP-01", "7.2"],
                ["2026-03-25 19:20:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1712", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-25 19:25:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1713", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-25 19:30:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1714", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-25 19:35:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1715", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-25 19:40:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1716", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-25 19:45:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1717", "Standard value entry per SOP-01", "7.24"],
                ["2026-03-25 19:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1718", "Standard value entry per SOP-01", "7.39"],
                ["2026-03-25 19:55:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1719", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-25 20:00:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1701", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-25 20:00:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1720", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-25 20:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1721", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-25 20:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1722", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-25 20:15:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1723", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-25 20:20:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1724", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-25 20:25:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1725", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-25 20:30:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1726", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-25 20:35:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1727", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-25 20:40:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1728", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-25 20:45:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1729", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-25 20:50:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1730", "Standard value entry per SOP-01", "7.06"],
                ["2026-03-25 20:55:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1731", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-25 21:00:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1732", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-25 21:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1733", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-25 21:10:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1734", "Standard value entry per SOP-01", "7.38"],
                ["2026-03-25 21:15:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1735", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-25 21:20:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1736", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-25 21:25:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1737", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-25 21:30:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1738", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-25 21:35:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1739", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-25 21:40:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1740", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-25 21:45:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1741", "Standard value entry per SOP-01", "7.08"],
                ["2026-03-25 21:50:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1742", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-25 21:55:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1743", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-25 22:00:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1744", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-25 22:05:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1745", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-25 22:10:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1746", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-25 22:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1747", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-25 22:20:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1748", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-25 22:25:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1749", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-25 22:30:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1750", "Standard value entry per SOP-01", "7.25"],
                ["2026-03-25 22:35:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1751", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-25 22:40:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1752", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-25 22:45:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1753", "Standard value entry per SOP-01", "7.24"],
                ["2026-03-25 22:50:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1754", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-25 22:55:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1755", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-25 23:00:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1756", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-25 23:05:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1757", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-25 23:10:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1758", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-25 23:15:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1759", "Standard value entry per SOP-01", "7.37"],
                ["2026-03-25 23:20:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1760", "Standard value entry per SOP-01", "7.05"],
                ["2026-03-25 23:25:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1761", "Standard value entry per SOP-01", "7.21"],
                ["2026-03-25 23:30:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1762", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-25 23:35:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1763", "Standard value entry per SOP-01", "7.33"],
                ["2026-03-25 23:40:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1764", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-25 23:45:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1765", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-25 23:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1766", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-25 23:55:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1767", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-26 00:00:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1768", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-26 00:05:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1769", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-26 00:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1770", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-26 00:15:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1771", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-26 00:20:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1772", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-26 00:25:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1773", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-26 00:30:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1774", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-26 00:35:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1775", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-26 00:40:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1776", "Standard value entry per SOP-01", "7.28"],
                ["2026-03-26 00:45:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1777", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-26 00:50:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1778", "Standard value entry per SOP-01", "7.47"],
                ["2026-03-26 00:55:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1779", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-26 01:00:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1780", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-26 01:05:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1781", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-26 01:10:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1782", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-26 01:15:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1783", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-26 01:20:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1784", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-26 01:25:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1785", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-26 01:30:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1786", "Standard value entry per SOP-01", "7.47"],
                ["2026-03-26 01:35:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1787", "Standard value entry per SOP-01", "7.33"],
                ["2026-03-26 01:40:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1788", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-26 01:45:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1789", "Standard value entry per SOP-01", "7.28"],
                ["2026-03-26 01:50:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1790", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-26 01:55:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1791", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-26 02:00:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1792", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-26 02:05:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1793", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-26 02:10:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1794", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-26 02:15:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1795", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-26 02:20:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1796", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-26 02:25:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1797", "Standard value entry per SOP-01", "7.37"],
                ["2026-03-26 02:30:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1798", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-26 02:35:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1799", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-26 02:45:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1801", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-26 02:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1802", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-26 02:55:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1803", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-26 03:00:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1804", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-26 03:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1805", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-26 03:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1806", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-26 03:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1807", "Standard value entry per SOP-01", "7.5"],
                ["2026-03-26 03:20:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1808", "Standard value entry per SOP-01", "7.21"],
                ["2026-03-26 03:25:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1809", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-26 03:30:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1810", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-26 03:35:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1811", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-26 03:40:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1812", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-26 03:45:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1813", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-26 03:50:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1814", "Standard value entry per SOP-01", "7.16"],
                ["2026-03-26 03:55:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1815", "Standard value entry per SOP-01", "7.21"],
                ["2026-03-26 04:00:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1816", "Standard value entry per SOP-01", "7.33"],
                ["2026-03-26 04:05:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1817", "Standard value entry per SOP-01", "7.33"],
                ["2026-03-26 04:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1818", "Standard value entry per SOP-01", "7.39"],
                ["2026-03-26 04:15:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1819", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-26 04:20:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1820", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-26 04:25:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1821", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-26 04:30:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1822", "Standard value entry per SOP-01", "7.39"],
                ["2026-03-26 04:35:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1823", "Standard value entry per SOP-01", "7.37"],
                ["2026-03-26 04:40:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1824", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-26 04:45:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1825", "Standard value entry per SOP-01", "7.37"],
                ["2026-03-26 04:50:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1826", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-26 04:55:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1827", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-26 05:00:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1828", "Standard value entry per SOP-01", "7.2"],
                ["2026-03-26 05:05:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1829", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-26 05:10:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1830", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-26 05:15:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1831", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-26 05:20:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1832", "Standard value entry per SOP-01", "7.27"],
                ["2026-03-26 05:25:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1833", "Standard value entry per SOP-01", "7.38"],
                ["2026-03-26 05:30:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1834", "Standard value entry per SOP-01", "7.46"],
                ["2026-03-26 05:35:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1835", "Standard value entry per SOP-01", "7.34"],
                ["2026-03-26 05:40:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1836", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-26 05:45:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1837", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-26 05:50:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1838", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-26 05:55:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1839", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-26 06:00:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1840", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-26 06:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1841", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-26 06:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1842", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-26 06:15:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1843", "Standard value entry per SOP-01", "7.47"],
                ["2026-03-26 06:20:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1844", "Standard value entry per SOP-01", "7.24"],
                ["2026-03-26 06:25:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1845", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-26 06:30:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1846", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-26 06:35:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1847", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-26 06:40:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1848", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-26 06:45:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1849", "Standard value entry per SOP-01", "7.48"],
                ["2026-03-26 06:50:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1850", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-26 06:55:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1851", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-26 07:00:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1852", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-26 07:05:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1853", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-26 07:10:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1854", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-26 07:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1855", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-26 07:20:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1856", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-26 07:25:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1857", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-26 07:30:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1858", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-26 07:35:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1859", "Standard value entry per SOP-01", "7.41"],
                ["2026-03-26 07:40:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1860", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-26 07:45:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1861", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-26 07:50:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1862", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-26 07:55:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1863", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-26 08:00:00", "analyst_jones", "SELECT", "AUDIT_TRAIL", "Analyst", "", "Reviewing logs", ""],
                ["2026-03-26 08:00:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1864", "Standard value entry per SOP-01", "7.37"],
                ["2026-03-26 08:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1865", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-26 08:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1866", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-26 08:15:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1867", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-26 08:20:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1868", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-26 08:25:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1869", "Standard value entry per SOP-01", "7.47"],
                ["2026-03-26 08:30:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1870", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-26 08:35:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1871", "Standard value entry per SOP-01", "7.0"],
                ["2026-03-26 08:40:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1872", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-26 08:45:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1873", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-26 08:50:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1874", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-26 08:55:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1875", "Standard value entry per SOP-01", "7.06"],
                ["2026-03-26 09:00:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1876", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-26 09:05:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1877", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-26 09:10:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1878", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-26 09:15:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1879", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-26 09:20:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1880", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-26 09:25:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1881", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-26 09:30:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1882", "Standard value entry per SOP-01", "7.0"],
                ["2026-03-26 09:35:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1883", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-26 09:40:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1884", "Standard value entry per SOP-01", "7.37"],
                ["2026-03-26 09:45:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1885", "Standard value entry per SOP-01", "7.16"],
                ["2026-03-26 09:50:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1886", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-26 09:55:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1887", "Standard value entry per SOP-01", "7.38"],
                ["2026-03-26 10:00:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1888", "Standard value entry per SOP-01", "7.46"],
                ["2026-03-26 10:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1889", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-26 10:10:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1890", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-26 10:15:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1891", "Standard value entry per SOP-01", "7.2"],
                ["2026-03-26 10:20:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1892", "Standard value entry per SOP-01", "7.47"],
                ["2026-03-26 10:25:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1893", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-26 10:30:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1894", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-26 10:35:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1895", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-26 10:40:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1896", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-26 10:45:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1897", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-26 10:50:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1898", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-26 10:55:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1899", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-26 11:00:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1900", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-26 11:05:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1901", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-26 11:10:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1902", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-26 11:15:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1903", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-26 11:20:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1904", "Standard value entry per SOP-01", "7.15"],
                ["2026-03-26 11:25:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1905", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-26 11:30:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1906", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-26 11:35:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1907", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-26 11:40:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1908", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-26 11:45:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1909", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-26 11:50:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1910", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-26 11:55:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1911", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-26 12:00:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1912", "Standard value entry per SOP-01", "7.47"],
                ["2026-03-26 12:05:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1913", "Standard value entry per SOP-01", "7.28"],
                ["2026-03-26 12:10:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1914", "Standard value entry per SOP-01", "7.24"],
                ["2026-03-26 12:15:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1915", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-26 12:20:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1916", "Standard value entry per SOP-01", "7.28"],
                ["2026-03-26 12:25:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1917", "Standard value entry per SOP-01", "7.2"],
                ["2026-03-26 12:30:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1918", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-26 12:35:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1919", "Standard value entry per SOP-01", "7.2"],
                ["2026-03-26 12:40:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1920", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-26 12:45:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1921", "Standard value entry per SOP-01", "7.46"],
                ["2026-03-26 12:50:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1922", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-26 12:55:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1923", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-26 13:00:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1924", "Standard value entry per SOP-01", "7.47"],
                ["2026-03-26 13:05:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1925", "Standard value entry per SOP-01", "7.5"],
                ["2026-03-26 13:10:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1926", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-26 13:15:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1927", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-26 13:20:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1928", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-26 13:25:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1929", "Standard value entry per SOP-01", "7.38"],
                ["2026-03-26 13:30:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1930", "Standard value entry per SOP-01", "7.38"],
                ["2026-03-26 13:35:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1931", "Standard value entry per SOP-01", "7.06"],
                ["2026-03-26 13:40:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1932", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-26 13:45:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1933", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-26 13:50:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1934", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-26 13:55:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1935", "Standard value entry per SOP-01", "7.12"],
                ["2026-03-26 14:00:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1936", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-26 14:05:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1937", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-26 14:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1938", "Standard value entry per SOP-01", "7.49"],
                ["2026-03-26 14:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1939", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-26 14:20:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1940", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-26 14:25:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1941", "Standard value entry per SOP-01", "7.08"],
                ["2026-03-26 14:30:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1942", "Standard value entry per SOP-01", "7.03"],
                ["2026-03-26 14:35:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1943", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-26 14:40:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1944", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-26 14:45:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1945", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-26 14:50:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1946", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-26 14:55:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1947", "Standard value entry per SOP-01", "7.08"],
                ["2026-03-26 15:00:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1948", "Standard value entry per SOP-01", "7.41"],
                ["2026-03-26 15:05:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1949", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-26 15:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1950", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-26 15:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1951", "Standard value entry per SOP-01", "7.19"],
                ["2026-03-26 15:20:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1952", "Standard value entry per SOP-01", "7.42"],
                ["2026-03-26 15:25:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1953", "Standard value entry per SOP-01", "7.16"],
                ["2026-03-26 15:30:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1954", "Standard value entry per SOP-01", "7.02"],
                ["2026-03-26 15:35:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1955", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-26 15:40:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1956", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-26 15:45:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1957", "Standard value entry per SOP-01", "7.37"],
                ["2026-03-26 15:50:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1958", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-26 15:55:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1959", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-26 16:00:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1960", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-26 16:05:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1961", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-26 16:10:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1962", "Standard value entry per SOP-01", "7.37"],
                ["2026-03-26 16:15:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1963", "Standard value entry per SOP-01", "7.0"],
                ["2026-03-26 16:20:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1964", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-26 16:25:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1965", "Standard value entry per SOP-01", "7.07"],
                ["2026-03-26 16:30:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1966", "Standard value entry per SOP-01", "7.07"],
                ["2026-03-26 16:35:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1967", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-26 16:40:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1968", "Standard value entry per SOP-01", "7.22"],
                ["2026-03-26 16:45:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1969", "Standard value entry per SOP-01", "7.1"],
                ["2026-03-26 16:50:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1970", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-26 16:55:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1971", "Standard value entry per SOP-01", "7.3"],
                ["2026-03-26 17:00:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1972", "Standard value entry per SOP-01", "7.26"],
                ["2026-03-26 17:05:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1973", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-26 17:10:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1974", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-26 17:15:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1975", "Standard value entry per SOP-01", "7.01"],
                ["2026-03-26 17:20:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1976", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-26 17:25:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1977", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-26 17:30:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1978", "Standard value entry per SOP-01", "7.27"],
                ["2026-03-26 17:35:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1979", "Standard value entry per SOP-01", "7.18"],
                ["2026-03-26 17:40:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1980", "Standard value entry per SOP-01", "7.36"],
                ["2026-03-26 17:45:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1981", "Standard value entry per SOP-01", "7.4"],
                ["2026-03-26 17:50:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1982", "Standard value entry per SOP-01", "7.35"],
                ["2026-03-26 17:55:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1983", "Standard value entry per SOP-01", "7.45"],
                ["2026-03-26 18:00:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1984", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-26 18:05:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1985", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-26 18:10:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1986", "Standard value entry per SOP-01", "7.14"],
                ["2026-03-26 18:15:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1987", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-26 18:20:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1988", "Standard value entry per SOP-01", "7.11"],
                ["2026-03-26 18:25:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1989", "Standard value entry per SOP-01", "7.09"],
                ["2026-03-26 18:30:00", "analyst_z", "UPDATE", "RESULTS", "Analyst", "RES-1990", "Standard value entry per SOP-01", "7.39"],
                ["2026-03-26 18:35:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1991", "Standard value entry per SOP-01", "7.04"],
                ["2026-03-26 18:40:00", "analyst_brown", "UPDATE", "RESULTS", "Analyst", "RES-1992", "Standard value entry per SOP-01", "7.32"],
                ["2026-03-26 18:45:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1993", "Standard value entry per SOP-01", "7.23"],
                ["2026-03-26 18:50:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1994", "Standard value entry per SOP-01", "7.17"],
                ["2026-03-26 18:55:00", "analyst_x", "UPDATE", "RESULTS", "Analyst", "RES-1995", "Standard value entry per SOP-01", "7.13"],
                ["2026-03-26 19:00:00", "analyst_jones", "UPDATE", "RESULTS", "Analyst", "RES-1996", "Standard value entry per SOP-01", "7.31"],
                ["2026-03-26 19:05:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1997", "Standard value entry per SOP-01", "7.29"],
                ["2026-03-26 19:10:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1998", "Standard value entry per SOP-01", "7.24"],
                ["2026-03-26 19:15:00", "jsmith", "UPDATE", "RESULTS", "Analyst", "RES-1999", "Standard value entry per SOP-01", "7.44"],
                ["2026-03-28 10:00:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1600", "Weekend catchup", "7.04"],
                ["2026-07-04 11:00:00", "analyst_y", "UPDATE", "RESULTS", "Analyst", "RES-1500", "Holiday work", "7.18"],
            ]

            # Write header row
            hdr_data_font  = Font(bold=True, color="E2E8F0", size=9)
            hdr_data_fill  = PatternFill("solid", fgColor="0F172A")
            for ci, col_name in enumerate(HEADER, 1):
                cell = ws_data.cell(row=1, column=ci, value=col_name)
                cell.font  = hdr_data_font
                cell.fill  = hdr_data_fill
                cell.alignment = Alignment(horizontal="center")

            # Write data rows
            data_font = Font(color="CBD5E1", size=9)
            alt_fill  = PatternFill("solid", fgColor="0A1628")
            for ri, row_vals in enumerate(DATA_ROWS, 2):
                fill = alt_fill if ri % 2 == 0 else PatternFill("solid", fgColor="0F172A")
                for ci, val in enumerate(row_vals, 1):
                    cell = ws_data.cell(row=ri, column=ci, value=val)
                    cell.font  = data_font
                    cell.fill  = fill

            # Column widths
            col_widths = [22, 18, 18, 18, 12, 16, 36, 12]
            for ci, w in enumerate(col_widths, 1):
                ws_data.column_dimensions[
                    openpyxl.utils.get_column_letter(ci)].width = w

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        st.download_button(
            label="📥 Download Sample Audit Log Template",
            data=_build_sample_xlsx(),
            file_name="valintel_sample_audit_log_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="at_sample_download",
            help=(
                "1,000-row validated GxP audit trail covering all 10 detection "
                "scenarios. Sheet 1 (Usage) explains every column and rule. "
                "Sheet 2 (Audit Log) contains the data — replace with your own "
                "system export to analyse real audit trails."
            ),
        )

        st.markdown("<br>", unsafe_allow_html=True)

        ck = st.session_state.get("at_key_n", 0)
        uploaded = st.file_uploader(
            "Drag and drop your Audit Log file here (CSV or Excel)",
            type=["csv","xlsx","xls"],
            key=f"at_upload_{ck}"
        )
        if uploaded:
            try:
                raw = uploaded.getvalue()
                if uploaded.name.lower().endswith(".csv"):
                    df = pd.read_csv(io.BytesIO(raw), dtype=str,
                                     low_memory=False).fillna("")
                else:
                    # Smart sheet selection: skip "Usage Instructions", prefer
                    # "Audit Log" sheet if present, otherwise use first data sheet
                    xl       = pd.ExcelFile(io.BytesIO(raw))
                    sheets   = xl.sheet_names
                    SKIP     = {"usage instructions", "usage", "instructions",
                                "readme", "read me", "guide"}
                    data_sheets = [s for s in sheets
                                   if s.strip().lower() not in SKIP]
                    # Prefer a sheet whose name contains "audit" or "log"
                    preferred = [s for s in data_sheets
                                 if any(kw in s.lower()
                                        for kw in ("audit", "log", "data", "trail"))]
                    sheet_to_use = (preferred or data_sheets or sheets)[0]
                    df = pd.read_excel(io.BytesIO(raw), sheet_name=sheet_to_use,
                                       dtype=str).fillna("")
                    if len(sheets) > 1:
                        st.caption(f"📋 Reading sheet: **{sheet_to_use}**"
                                   + (f"  ·  skipped: {', '.join(s for s in sheets if s != sheet_to_use)}"
                                      if len(sheets) > 1 else ""))
                st.session_state["at_raw_df"]   = df
                st.session_state["at_file_name"] = uploaded.name
                st.success(f"✅ **{uploaded.name}** — "
                           f"**{len(df):,} rows** × **{len(df.columns)} columns**")
                with st.expander("Preview (first 10 rows)", expanded=False):
                    st.dataframe(df.head(10), use_container_width=True)
            except Exception as e:
                st.error(f"⛔ Could not read file: {e}")

        # ── Column mapper ─────────────────────────────────────────────────────
        if st.session_state.get("at_raw_df") is not None:
            st.markdown("### Step 2 — Map Your Columns")
            st.caption("Match your file's column names to the required fields. "
                       "★ = required.")
            df      = st.session_state["at_raw_df"]
            avail   = ["(not in file)"] + list(df.columns)
            mapping = {}
            cols3   = st.columns(3)
            for i,(field,desc) in enumerate(_AT_REQUIRED_COLS.items()):
                req  = field in ("timestamp","user_id","action_type")
                auto = "(not in file)"
                for col in df.columns:
                    cl = col.lower().replace(" ","_").replace("-","_")
                    if field in cl or cl in field:
                        auto = col; break
                with cols3[i%3]:
                    st.caption(desc)
                    mapping[field] = st.selectbox(
                        f"{'★ ' if req else ''}{field.replace('_',' ').title()}",
                        avail,
                        index=avail.index(auto) if auto in avail else 0,
                        key=f"at_map_{field}"
                    )

            req_ok = all(
                mapping.get(f,"(not in file)") != "(not in file)"
                for f in ("timestamp","user_id","action_type")
            )
            if not req_ok:
                st.warning("⚠️ Map the three required fields to continue.")
            else:
                _, bc, _ = st.columns([3,4,3])
                with bc:
                    if st.button("✅ Confirm Mapping & Continue",
                                 type="primary", use_container_width=True,
                                 key="at_confirm_map"):
                        rename = {v:k for k,v in mapping.items()
                                  if v != "(not in file)"}
                        mdf = df.rename(columns=rename)
                        for c in _AT_REQUIRED_COLS:
                            if c not in mdf.columns:
                                mdf[c] = ""
                        st.session_state["at_mapped_df"]   = mdf
                        st.session_state["at_column_map"]  = mapping
                        st.session_state["at_mapping_done"] = True
                        st.rerun()

    # ── STEP 2: Run analysis ──────────────────────────────────────────────────
    elif not st.session_state.get("at_analysis_done"):
        df = st.session_state["at_mapped_df"]
        st.success(f"✅ Mapping confirmed — **{len(df):,} events** ready")

        st.markdown("### Step 3 — Run Analysis")
        st.markdown("""
<div style="display:inline-flex;align-items:center;gap:8px;
            background:#f0fdf4;border:1px solid #86efac;border-radius:8px;
            padding:8px 16px;margin-bottom:16px;">
  <span style="color:#16a34a;font-size:0.88rem;">●</span>
  <span style="color:#15803d;font-size:0.82rem;font-weight:600;">
    16-Rule Detection Engine Ready
  </span>
</div>""", unsafe_allow_html=True)

        _, rc2, _ = st.columns([2,6,2])
        with rc2:
            run = st.button(
                f"🚀 Analyse {len(df):,} Events → Generate Top {_AT_TOP_N} Risk Report",
                type="primary", use_container_width=True, key="at_run_btn"
            )

        st.markdown("<br>", unsafe_allow_html=True)
        _, na2_col, _ = st.columns([3, 4, 3])
        with na2_col:
            if st.button("🔄 Start New Analysis", key="at_reset_btn_pre",
                         use_container_width=True):
                for k in ["at_raw_df","at_mapped_df","at_scored_df","at_top20_df",
                          "at_file_name","at_mapping_done","at_analysis_done","at_total_events"]:
                    st.session_state[k] = _defaults.get(k)
                st.session_state["at_key_n"] = st.session_state.get("at_key_n",0) + 1
                st.rerun()

        if run:
            prog   = st.progress(0)
            status = st.empty()
            with st.status("🔍 Audit Trail Analysis", expanded=True) as atstat:
                st.write("📊 Step 1: Parsing timestamps...")
                prog.progress(0.15)
                scored = at_score_events(df)

                st.write(f"⚡ Step 2: Scoring {len(scored):,} events across 16 rules...")
                prog.progress(0.45)
                # ── Select Top 20 with two filters ────────────────────────────
                # Filter 1: Remove burst duplicates (already built)
                # Filter 2: Named rule gate — events only enter Top 20 if at
                #   least one named rule (Rules 1–15) fired above minimum score.
                #   Pure dimension-score events with no named rule are noise
                #   and erode reviewer trust in the tool.
                # ── High-signal rules only qualify events for Top N ─────────
                # Rule 11 (temporal) and Rule 2 (burst) are excluded from the gate
                # — they can boost composite score but cannot independently qualify
                # an event. This prevents off-hours and burst floods in rows 13–20.
                _NAMED_RULE_COLS = [
                    "score_rule1_vague_rationale",
                    # score_rule2_burst excluded — burst alone is not a qualifier
                    "score_rule3_admin_conflict",
                    "score_rule4_drift",
                    "score_rule5_failed_login",
                    "score_del_recreate",               # Rule 6
                    "score_record",                     # Rule 7
                    "score_privilege",                  # Rule 8
                    # score_temporal excluded — off-hours alone is not a qualifier
                    "score_rule12_timestamp_reversal",
                    "score_rule13_service_account",
                    "score_rule14_dormant_account",
                    "score_rule15_suspicious_sequence",
                    "score_rule16_first_time_behavior",
                ]
                _GATE_THRESHOLD = 7.0  # raised from 6.0 — reduces low-signal noise

                def _has_named_rule(row):
                    return any(
                        float(row.get(c, 0)) >= _GATE_THRESHOLD
                        for c in _NAMED_RULE_COLS
                        if c in row.index
                    )

                # Deduplicate burst events (Rule 2)
                non_dup = scored[~scored.get("_is_burst_dup",
                                 pd.Series(False, index=scored.index))]

                # Deduplicate Rule 11 (off-hours) — keep only the highest-scoring
                # off-hours event per user to prevent temporal floods
                _TEMPORAL_KEY = non_dup["user_id"].astype(str) + "||temporal"
                temporal_mask = non_dup["score_temporal"] > 0
                if temporal_mask.any():
                    non_dup = non_dup.copy()
                    non_dup["_temporal_rank"] = non_dup.groupby(
                        _TEMPORAL_KEY)["score_temporal"].rank(
                        method="first", ascending=False)
                    non_dup = non_dup[
                        (non_dup["score_temporal"] == 0) |
                        (non_dup["_temporal_rank"] <= 1)
                    ]

                has_rule  = non_dup.apply(_has_named_rule, axis=1)
                qualified = non_dup[has_rule].head(_AT_TOP_N)
                # No fill padding — if fewer than TOP_N genuine findings exist,
                # show only genuine findings. A shorter honest report is better
                # than a padded one with composite-score filler.
                top20 = qualified.copy()

                st.write(f"✍️ Step 3: Generating system narratives for top {_AT_TOP_N} events...")
                prog.progress(0.65)
                top20  = at_generate_justifications(top20, model_id)

                prog.progress(0.90)
                st.write("📋 Step 4: Building evidence package...")
                st.session_state["at_scored_df"]     = scored
                st.session_state["at_top20_df"]      = top20
                st.session_state["at_total_events"]  = len(scored)
                st.session_state["at_analysis_done"] = True

                n_crit = int((scored["Risk_Tier"]=="Critical").sum())
                prog.progress(1.0)
                log_audit(user,"AT_ANALYSIS_COMPLETE","AUDIT_TRAIL",
                          new_value=f"{len(scored)} events, {n_crit} critical",
                          reason=f"System: {st.session_state.get('at_system_name','?')}")
                atstat.update(
                    label=f"✅ {len(scored):,} events analysed — "
                          f"{_AT_TOP_N} escalated — {n_crit} critical",
                    state="complete", expanded=False
                )
            prog.empty(); status.empty()
            st.rerun()

    # ── STEP 3: Results ───────────────────────────────────────────────────────
    else:
        scored = st.session_state["at_scored_df"]
        top20  = st.session_state["at_top20_df"]
        n_total = st.session_state["at_total_events"]
        n_esc   = len(top20)
        n_crit  = int((scored["Risk_Tier"]=="Critical").sum())
        n_high  = int((scored["Risk_Tier"]=="High").sum())
        n_med   = int((scored["Risk_Tier"]=="Medium").sum())
        n_low   = int((scored["Risk_Tier"]=="Low").sum())
        pct_clr = round((n_total-n_esc)/n_total*100,1) if n_total>0 else 0

        # ── Hero banner ───────────────────────────────────────────────────────
        st.markdown(f"""
<div style="background:#0f172a;border:2px solid #38bdf8;border-radius:10px;
            padding:18px 24px;margin-bottom:12px;">
  <p style="margin:0;color:#475569;font-size:0.68rem;letter-spacing:3px;
            text-transform:uppercase;font-family:'Inter',sans-serif;">
    Audit Trail Review Complete — {st.session_state.get('at_system_name','System')}</p>
  <div style="display:flex;align-items:baseline;gap:20px;margin:8px 0 4px;">
    <p style="margin:0;font-size:2.4rem;font-weight:800;color:#4ade80;
              line-height:1;font-family:'Inter',sans-serif;">{pct_clr}%
      <span style="font-size:1rem;font-weight:500;color:#4ade80;"> auto-cleared</span>
    </p>
    <p style="margin:0;font-size:1.1rem;font-weight:600;color:#38bdf8;">
      {n_esc} of {n_total:,} events escalated for human review
    </p>
  </div>
  <p style="margin:4px 0 0;font-size:0.8rem;color:#475569;font-style:italic;">
    Only events where a named detection rule fired above threshold are escalated —
    dimension scores alone cannot trigger an escalation.
    &nbsp;·&nbsp; {st.session_state.get('at_file_name','')}
  </p>
</div>""", unsafe_allow_html=True)

        # ── Metrics ───────────────────────────────────────────────────────────
        c1,c2,c3,c4,c5 = st.columns(5)
        for col,val,label,color in [
            (c1, f"{n_total:,}", "Total Events",  "#38bdf8"),
            (c2, str(n_crit),   "🔴 Critical",    "#dc2626"),
            (c3, str(n_high),   "🟠 High",        "#ea580c"),
            (c4, str(n_med),    "🟡 Medium",      "#d97706"),
            (c5, f"{pct_clr}%", "✅ Auto-Cleared","#4ade80"),
        ]:
            col.metric(label, val)

        # ── Out-of-period events warning ─────────────────────────────────────
        # If the user entered review period dates, check whether the uploaded
        # file contains events outside that range and warn clearly.
        _r_start_str = st.session_state.get("at_review_start","").strip()
        _r_end_str   = st.session_state.get("at_review_end","").strip()
        _missing_str = "(review period dates not specified)"
        if (_r_start_str and _r_end_str
                and _r_start_str != _missing_str
                and _r_end_str   != _missing_str
                and "timestamp_parsed" in scored.columns):
            try:
                import datetime as _dtp
                _r_s = pd.to_datetime(_r_start_str, dayfirst=True, errors="coerce")
                _r_e = pd.to_datetime(_r_end_str,   dayfirst=True, errors="coerce")
                if not pd.isnull(_r_s) and not pd.isnull(_r_e):
                    _ts  = scored["timestamp_parsed"].dropna()
                    _before = int((_ts < _r_s).sum())
                    _after  = int((_ts > _r_e).sum())
                    if _before > 0 or _after > 0:
                        _parts = []
                        if _before: _parts.append(f"{_before:,} event(s) before {_r_start_str}")
                        if _after:  _parts.append(f"{_after:,} event(s) after {_r_end_str}")
                        st.warning(
                            f"⚠️ **Dataset scope note:** The uploaded file contains events "
                            f"outside the defined review period ({' and '.join(_parts)}). "
                            "These are included in the analysis as the full dataset was provided. "
                            "If only the defined review period should be assessed, filter your "
                            "system export before uploading."
                        )
            except Exception:
                pass

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Risk Distribution ─────────────────────────────────────────────────
        st.markdown("### Risk Distribution — Full Dataset")
        st.dataframe(pd.DataFrame({
            "Risk Tier":  ["Critical","High","Medium","Low"],
            "Count":      [n_crit,n_high,n_med,n_low],
            "% of Total": [round(v/n_total*100,1) for v in [n_crit,n_high,n_med,n_low]],
            "Escalated":  ["Yes","Yes" if n_high>0 else "No","No","No"],
        }), use_container_width=True, hide_index=True)

        # ── Download ──────────────────────────────────────────────────────────
        st.markdown("---")
        sys_name = st.session_state.get("at_system_name","").strip()
        r_start  = st.session_state.get("at_review_start","").strip()
        r_end    = st.session_state.get("at_review_end","").strip()

        if not sys_name:
            st.warning("⚠️ Please enter a **System Name** above before downloading. "
                       "The system name is required for the evidence package header.")
        else:
            if not r_start or not r_end:
                st.warning(
                    "⚠️ **Review Period dates not entered.** "
                    "The narrative in the Summary sheet will read "
                    "*'(review period dates not specified)'* instead of actual dates. "
                    "Enter Start and End dates above to include them in the report."
                )
            xlsx = at_build_excel(
                top20, scored,
                sys_name,
                r_start  or "(review period dates not specified)",
                r_end    or "(review period dates not specified)",
                st.session_state.get("at_file_name",""),
            )
            dl_c, inf_c = st.columns([4,5])
            with dl_c:
                st.download_button(
                    "📥 Download Evidence Package (.xlsx)",
                    data=xlsx,
                    file_name=(f"AuditTrail_{sys_name.replace(' ','_')}"
                               f"_{datetime.date.today()}.xlsx"),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="at_download_btn", use_container_width=True,
                )
            with inf_c:
                st.markdown(
                    "<p style='color:#475569;font-size:0.8rem;padding-top:8px;'>"
                    "4 sheets: <b style='color:#e2e8f0;'>Summary</b> · "
                    "<b style='color:#e2e8f0;'>Events for Review</b> · "
                    "<b style='color:#e2e8f0;'>Full Audit Log</b> · "
                    "<b style='color:#e2e8f0;'>Detection Logic</b>"
                    "<br>Attach to Periodic Review Report Section 9.1.6.</p>",
                    unsafe_allow_html=True
                )

        st.markdown(f"""
<div style="background:#0f172a;border:1px solid #1e293b;border-radius:8px;
            padding:14px 20px;margin-top:14px;font-size:0.8rem;">
  <b style="color:#94a3b8;">Content for your Periodic Review Report:</b><br>
  <i style="color:#cbd5e1;">
  "System-assisted audit trail review identified the {n_esc} highest-risk events from
  {n_total:,} total entries using a 16-rule anomaly detection engine. {pct_clr}% of
  events were auto-cleared as low risk. All {n_esc} escalated events are available
  for human review and have been dispositioned by the undersigned as documented in
  the attached Appendix. Complies with 21 CFR Part 11 §11.10(e) and EU Annex 11 Clause 9."
  </i>
</div>""", unsafe_allow_html=True)

        # ── New Analysis button ───────────────────────────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)
        _, na_col, _ = st.columns([3, 4, 3])
        with na_col:
            if st.button("🔄 Start New Analysis", key="at_reset_btn",
                         use_container_width=True):
                for k in ["at_raw_df","at_mapped_df","at_scored_df","at_top20_df",
                          "at_file_name","at_mapping_done","at_analysis_done","at_total_events"]:
                    st.session_state[k] = _defaults.get(k)
                st.session_state["at_key_n"] = st.session_state.get("at_key_n",0) + 1
                st.rerun()

        # ── Top 20 Events — collapsed by default ──────────────────────────────
        st.markdown("---")
        st.markdown(f"### Top {_AT_TOP_N} Highest-Risk Events")
        st.caption("Click any event to expand the full detail. "
                   "All events are collapsed by default to keep the page clean.")

        tier_colors = {
            "Critical":"#dc2626","High":"#ea580c",
            "Medium":"#d97706","Low":"#4ade80"
        }
        tier_icons = {
            "Critical":"🔴","High":"🟠","Medium":"🟡","Low":"🟢"
        }

        # ── Low-risk grouping — collapse repetitive same-disposition entries ────
        # Events sharing the same Primary_Rule + Risk_Tier + Suggested_Disposition
        # where disposition is Justified (low urgency) are grouped into a single
        # summary card so the reviewer sees intelligent aggregation, not repetition.
        _COLLAPSIBLE_DISPOSITIONS = {
            "No Action Required",
            "Document Rationale",
        }
        _COLLAPSIBLE_TIERS = {"Medium", "Low"}

        def _group_key(r):
            disp = str(r.get("Suggested_Disposition",""))
            tier = str(r.get("Risk_Tier",""))
            prim = str(r.get("Primary_Rule","")).split("[")[0].strip()
            if disp in _COLLAPSIBLE_DISPOSITIONS and tier in _COLLAPSIBLE_TIERS:
                return f"{prim}||{tier}||{disp}"
            return None   # non-collapsible — display individually

        # Build display list: individual rows or grouped summaries
        display_items = []   # each item: ("single", rank, row) or ("group", [rows])
        group_buckets = {}   # key → [rows]
        group_ranks   = {}   # key → [ranks]

        for rank, (_, row) in enumerate(top20.iterrows(), 1):
            key = _group_key(row)
            if key:
                group_buckets.setdefault(key, []).append(row)
                group_ranks.setdefault(key, []).append(rank)
            else:
                display_items.append(("single", rank, row))

        # Insert group summaries at the position of the first member's rank
        # so ordering roughly matches the original risk-sorted list
        group_insert = []
        for key, rows in group_buckets.items():
            first_rank = group_ranks[key][0]
            group_insert.append((first_rank, "group", rows, group_ranks[key]))
        group_insert.sort(key=lambda x: x[0])

        for first_rank, _, rows, ranks in group_insert:
            display_items.append(("group", first_rank, rows, ranks))

        # Sort final display list by first rank in group / individual rank
        display_items.sort(key=lambda x: x[1])

        for item in display_items:
          if item[0] == "group":
            _, first_rank, rows, ranks = item
            grp_tier  = str(rows[0].get("Risk_Tier","Medium"))
            grp_icon  = tier_icons.get(grp_tier,"🟡")
            grp_bc    = tier_colors.get(grp_tier,"#d97706")
            grp_prim  = str(rows[0].get("Primary_Rule","")).replace(" [MEDIUM]","").replace(" [HIGH]","").replace(" [CRITICAL]","")
            grp_disp  = str(rows[0].get("Suggested_Disposition",""))
            grp_users = ", ".join(sorted({str(r.get("user_id","")) for r in rows}))
            grp_label = (
                f"{grp_icon} Events #{ranks[0]}–#{ranks[-1]} · {len(rows)} similar entries "
                f"· {grp_tier} · {grp_prim}"
            )
            with st.expander(grp_label, expanded=False):
                st.markdown(f"""
<div style="background:#0f172a;border-left:3px solid {grp_bc};border-radius:6px;
            padding:14px 18px;">
  <p style="color:{grp_bc};font-weight:700;margin:0 0 6px;">
    {len(rows)} similar entries reviewed collectively</p>
  <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;font-size:0.8rem;
              margin-bottom:10px;">
    <div><span style="color:#475569;">Primary Rule: </span>
         <span style="color:#e2e8f0;">{grp_prim}</span></div>
    <div><span style="color:#475569;">Risk Tier: </span>
         <span style="color:#e2e8f0;">{grp_tier}</span></div>
    <div><span style="color:#475569;">Disposition: </span>
         <span style="color:#4ade80;">{grp_disp}</span></div>
    <div><span style="color:#475569;">Users involved: </span>
         <span style="color:#e2e8f0;">{grp_users}</span></div>
    <div><span style="color:#475569;">Event ranks: </span>
         <span style="color:#e2e8f0;">#{ranks[0]} – #{ranks[-1]}</span></div>
  </div>
  <p style="color:#475569;font-size:0.78rem;margin:0;font-style:italic;">
    These entries share the same primary rule finding and disposition.
    Reviewed collectively — all {len(rows)} events are recorded individually in the
    downloaded evidence package with full detail.</p>
</div>""", unsafe_allow_html=True)
          else:
            # individual event
            _, rank, row = item
            tier         = str(row.get("Risk_Tier","Medium"))
            score        = float(row.get("Risk_Score",0))
            bc           = tier_colors.get(tier,"#d97706")
            icon         = tier_icons.get(tier,"🟡")
            user_id      = str(row.get("user_id","—"))
            action       = str(row.get("action_type","—"))
            triggered    = str(row.get("Triggered_Rules",""))
            rule_rat     = str(row.get("Rule_Rationale",""))
            reg_basis    = str(row.get("Regulatory_Basis",""))
            action_req   = str(row.get("Action_Required",""))
            chain_id     = str(row.get("Event_Chain_ID",""))
            primary_r    = str(row.get("Primary_Rule","")).replace(" [CRITICAL]","").replace(" [HIGH]","").replace(" [MEDIUM]","")
            supporting   = str(row.get("Supporting_Signals",""))
            evidence_str = str(row.get("Evidence_Strength","Low"))
            ev_color     = {"High":"#16a34a","Medium":"#d97706","Low":"#475569"}.get(evidence_str,"#475569")

            seq_ctx_ui     = str(row.get("Sequence_Context","")).strip()
            seq_suffix     = f"  ·  {seq_ctx_ui}" if seq_ctx_ui else ""
            expander_label = (
                f"{icon} Event #{rank} · {tier} · {evidence_str} Evidence"
                f"  |  {user_id}  ·  {action}  ·  {primary_r}{seq_suffix}"
            )


            with st.expander(expander_label, expanded=False):

                badges_html = ""
                if triggered:
                    rule_badge_colors = {
                        "Rule 1": ("#7c3aed","#ede9fe"),
                        "Rule 2": ("#0369a1","#dbeafe"),
                        "Rule 3": ("#dc2626","#fee2e2"),
                        "Rule 4": ("#d97706","#fef3c7"),
                        "Rule 5": ("#b91c1c","#fee2e2"),
                        "Holida": ("#b45309","#fef3c7"),
                    }
                    for rule_label in triggered.split("; "):
                        key = rule_label[:6]
                        fg, bg2 = rule_badge_colors.get(key, ("#64748b","#1e293b"))
                        badges_html += (
                            f'<span style="background:{bg2};color:{fg};border:1px solid {fg}44;'
                            f'padding:2px 8px;border-radius:4px;font-size:0.68rem;'
                            f'margin-right:4px;margin-bottom:4px;display:inline-block;">'
                            f'{rule_label}</span>'
                        )

                chain_badge = ""  # removed — chain context in sequence field

                sugg_disp = str(row.get("Suggested_Disposition",""))
                sugg_rat  = str(row.get("Suggested_Disposition_Rationale",""))
                disp_color = {
                    "Escalate to CAPA":               "#dc2626",
                    "Investigate — Verify Source Data":"#2563eb",
                    "Justified — Amendment Required":  "#d97706",
                    "Justified — Document Rationale":  "#16a34a",
                    "Justified — No Action Required":  "#475569",
                }.get(sugg_disp, "#475569")

                st.markdown(f"""
<div style="background:#0f172a;border-left:3px solid {bc};border-radius:6px;
            padding:14px 18px;">
  <div style="display:flex;justify-content:space-between;margin-bottom:10px;">
    <span style="color:{bc};font-weight:700;">Event #{rank}
      <span style="background:{bc}22;border:1px solid {bc}44;color:{bc};
             padding:2px 8px;border-radius:4px;font-size:0.7rem;margin-left:8px;">
        {tier}</span>
      <span style="background:{ev_color}22;border:1px solid {ev_color}44;color:{ev_color};
             padding:2px 8px;border-radius:4px;font-size:0.68rem;margin-left:6px;">
        {evidence_str} Evidence</span></span>
    <span style="color:#334155;font-size:0.72rem;">{str(row.get('timestamp',''))}</span>
  </div>
  <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:6px;
              font-size:0.8rem;margin-bottom:10px;">
    <div><span style="color:#475569;">User: </span>
         <span style="color:#e2e8f0;">{user_id}</span></div>
    <div><span style="color:#475569;">Action: </span>
         <span style="color:#e2e8f0;">{action}</span></div>
    <div><span style="color:#475569;">Role: </span>
         <span style="color:#e2e8f0;">{row.get('role','—')}</span></div>
    <div><span style="color:#475569;">Record ID: </span>
         <span style="color:#e2e8f0;">{row.get('record_id','—')}</span></div>
    <div><span style="color:#475569;">Record Type: </span>
         <span style="color:#e2e8f0;">{row.get('record_type','—')}</span></div>
    <div><span style="color:#475569;">Comment: </span>
         <span style="color="{'#fbbf24' if row.get('score_rule1_vague_rationale',0)>0 else '#e2e8f0'}">
           {str(row.get('comments','—'))[:60]}</span></div>
  </div>
  <!-- Primary Rule — dominant badge -->
  <div style="background:#1e1b4b;border:1.5px solid {bc};border-radius:6px;
              padding:8px 14px;margin-bottom:8px;display:flex;align-items:center;gap:10px;">
    <span style="color:#a5b4fc;font-size:0.65rem;text-transform:uppercase;
                 letter-spacing:1.5px;white-space:nowrap;">Primary Rule</span>
    <span style="color:#e2e8f0;font-size:0.82rem;font-weight:600;">{primary_r}</span>
  </div>
  {f'<div style="margin-bottom:8px;"><span style="color:#334155;font-size:0.67rem;text-transform:uppercase;letter-spacing:1px;">Supporting signals: </span><span style="color:#475569;font-size:0.75rem;">{supporting}</span></div>' if supporting and supporting != "—" else ""}
  {f'<div style="margin-bottom:8px;">{badges_html}</div>' if badges_html else ''}
  {f'''<div style="background:#1a0f2e;border:1px solid #7c3aed44;border-radius:4px;
              padding:10px 14px;margin-bottom:8px;">
    <p style="color:#7c3aed;font-size:0.67rem;text-transform:uppercase;
              letter-spacing:1px;margin:0 0 5px;font-weight:700;">&#9312; Why It Matters &mdash; Regulatory Risk</p>
    <p style="color:#c4b5fd;font-size:0.81rem;line-height:1.5;margin:0 0 6px;">{reg_basis}</p>
    <details style="margin:0;"><summary style="color:#475569;font-size:0.7rem;cursor:pointer;">Full technical detail &#9658;</summary>
      <p style="color:#94a3b8;font-size:0.77rem;line-height:1.4;margin:6px 0 0;">{rule_rat[:400]}</p></details>
  </div>''' if reg_basis or rule_rat else ''}
  {f'''<div style="background:#0c1a2e;border:1px solid {disp_color}44;border-radius:4px;
              padding:10px 14px;margin-bottom:8px;">
    <p style="color:{disp_color};font-size:0.67rem;text-transform:uppercase;
              letter-spacing:1px;margin:0 0 5px;font-weight:700;">&#9313; Suggested Disposition &mdash; {sugg_disp}</p>
    <p style="color:#94a3b8;font-size:0.79rem;line-height:1.4;margin:0;">{sugg_rat}</p>
  </div>''' if sugg_disp else ''}
  {f'''<div style="background:#0f1f12;border:1px solid #16a34a44;border-radius:4px;
              padding:10px 14px;margin-bottom:8px;">
    <p style="color:#16a34a;font-size:0.67rem;text-transform:uppercase;
              letter-spacing:1px;margin:0 0 5px;font-weight:700;">&#9314; Action Required</p>
    <p style="color:#86efac;font-size:0.79rem;line-height:1.4;margin:0;">{action_req}</p>
  </div>''' if action_req else ''}
  <div style="margin-top:4px;">
    <p style="color:#334155;font-size:0.67rem;text-transform:uppercase;
             letter-spacing:1px;margin:0 0 5px;">&#9315; What Happened</p>
    <p style="color:#475569;font-size:0.73rem;font-style:italic;margin:0 0 4px;">
      One-sentence log summary. All interpretation is in the fields above.</p>
    <p style="color:#cbd5e1;font-size:0.79rem;line-height:1.5;margin:0;">
      {str(row.get('AI_Justification',''))}</p>
  </div>
</div>""", unsafe_allow_html=True)


# =============================================================================
# 11. LOGIN
# =============================================================================

def show_login():
    # Disable browser autocomplete / password-save on all password inputs
    _inject_password_security()

    # Remove the sticky End Session button injected by show_app() into the
    # parent document body — Streamlit cannot clean up manually-injected DOM
    # elements on rerun, so we do it explicitly here on every login page render.
    _st_components.html("""
    <script>
    (function() {
        var btn = window.parent.document.getElementById('sticky-terminate-btn');
        if (btn) btn.parentNode.removeChild(btn);
    })();
    </script>
    """, height=0)

    left_space, center_content, right_space = st.columns([3, 4, 3])
    with center_content:
        st.markdown(
            '<div class="top-banner"><p class="banner-text-inner">GxP Validation — CSV Accelerator</p></div>',
            unsafe_allow_html=True
        )
        st.markdown("<br>", unsafe_allow_html=True)
        u = st.text_input("Professional Identity", placeholder="Username", label_visibility="collapsed",
                          key="login_username_field")
        p = st.text_input("Security Token", type="password", placeholder="Password",
                          label_visibility="collapsed", key="login_password_field")
        st.markdown("<br>", unsafe_allow_html=True)
        b_left, b_center, b_right = st.columns([1, 2, 1])
        with b_center:
            if st.button("Initialize Secure Session", key="login_btn", use_container_width=True):
                u_clean = sanitize_input(u, max_length=64)
                p_clean = sanitize_input(p, max_length=256)
                success, err_msg = authenticate_user(u_clean, p_clean)
                if success:
                    st.session_state.user_name     = u_clean
                    st.session_state.user_role     = get_user_role(u_clean)
                    st.session_state.authenticated = True
                    st.session_state.last_activity = datetime.datetime.utcnow()
                    log_audit(u_clean, "LOGIN_SUCCESS", "SESSION",
                              new_value=f"Role: {st.session_state.user_role}")
                    st.rerun()
                else:
                    st.error(err_msg or "Invalid credentials.")

        # ── Branding footer — small, italic, grey ───────────────────────────
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown(
            "<p style='text-align:center; font-style:italic; font-size:0.78rem; color:#94a3b8;'>"
          # ── MANUAL EDIT v29-custom — DO NOT OVERWRITE ──────────────
            "LLM-Powered GxP Validation</p>"
            "<p style='text-align:center; font-size:0.72rem; color:#b0bec5; margin-top:-6px;'>"
            "AI-driven validation accelerator that generates FRS, OQ tests, and Traceability matrices from URS and system documentation, while detecting validation gaps. "
            "</p>",
            unsafe_allow_html=True
          # ── END MANUAL EDIT ────────────────────────────────────────
        )


# =============================================================================
# 12. MAIN APPLICATION
# =============================================================================

def show_app():
    # Disable browser autocomplete / password-save on all password inputs (incl. admin panel)
    _inject_password_security()

    # ── Centralised logout handler — checked BEFORE any widget renders ────────
    # Button handlers only set this flag and call st.rerun(). The actual clear
    # happens here on the very next run, before Streamlit draws a single widget.
    # This guarantees a clean single-click logout with no mid-render state flush.
    if st.session_state.get("_logout_requested"):
        _logout_user = st.session_state.get("user_name", "unknown")
        _logout_reason = st.session_state.get("_logout_reason", "User terminated session")
        log_audit(_logout_user, "LOGOUT", "SESSION", reason=_logout_reason)
        st.session_state.clear()
        st.rerun()

    # Session timeout enforcement
    if not check_session_timeout():
        user = st.session_state.get("user_name", "unknown")
        log_audit(user, "SESSION_TIMEOUT", "SESSION",
                  reason=f"Inactivity exceeded {SESSION_TIMEOUT_MINUTES} min")
        st.session_state.clear()   # wipe everything — no stale results on re-login
        st.warning("⏱️ Session expired due to inactivity. Please log in again.")
        st.rerun()

    touch_session()

    user = st.session_state.get("user_name", "unknown")
    role = st.session_state.get("user_role", "Validator")

    # ── Sidebar ──
    with st.sidebar:
        st.markdown('<p class="sb-title">VALINTEL.AI — Validation Intelligence</p>', unsafe_allow_html=True)
        st.markdown(
            '<p style="color:#94a3b8;font-size:0.68rem;margin:-6px 0 4px;'
            'letter-spacing:0.5px;font-family:\'IBM Plex Mono\',monospace;">'
            'Build v60 · Audit Trail Module</p>',
            unsafe_allow_html=True)
        st.divider()
        st.markdown('<p class="sb-sub">🔧 Analysis Mode</p>', unsafe_allow_html=True)
        _modes = [
            "New Validation",
            "Change Impact Analysis",
            "Periodic Review",
            "Delta Generation (coming soon)",
        ]
        app_mode = st.radio(
            "Mode", _modes,
            index=_modes.index(st.session_state.get("app_mode","New Validation"))
                  if st.session_state.get("app_mode","New Validation") in _modes else 0,
            label_visibility="collapsed",
            key="app_mode_radio",
        )
        st.session_state["app_mode"] = app_mode
        st.divider()
        st.markdown('<p class="sb-sub">🤖 AI Model</p>', unsafe_allow_html=True)

        engine_name = st.radio(
            "Model",
            list(MODELS.keys()),
            index=list(MODELS.keys()).index(st.session_state.selected_model),
            label_visibility="collapsed",
            key="model_radio"
        )
      
        # ── MANUAL EDIT v29-custom — DO NOT OVERWRITE ──────────────
        st.session_state.selected_model = engine_name
        # ── END MANUAL EDIT ────────────────────────────────────────

        st.divider()
        st.markdown(f'<p class="sidebar-stats">Operator: {user} &nbsp;|&nbsp; Role: {role}</p>', unsafe_allow_html=True)

        if st.button("Terminate Session", key="terminate_sidebar", use_container_width=True):
            st.session_state["_logout_requested"] = True
            st.session_state["_logout_reason"]    = "User clicked Terminate Session"
            st.rerun()

        if role == "Admin":
            with st.expander("👤 User Management", expanded=False):
                st.markdown('<p class="sidebar-stats">Create New User</p>', unsafe_allow_html=True)
                new_u = st.text_input("New Username", key="new_username_input",
                                      label_visibility="collapsed", placeholder="New Username")
                new_p = st.text_input("New Password", type="password", key="new_password_input",
                                      label_visibility="collapsed", placeholder="New Password")
                new_r = st.selectbox("New Role", ROLES, key="new_role_select",
                                     label_visibility="collapsed")
                if st.button("➕ Create User", key="create_user_btn"):
                    new_u_clean = sanitize_input(new_u, max_length=64)
                    new_p_clean = sanitize_input(new_p, max_length=256)
                    if new_u_clean and new_p_clean:
                        if len(new_p_clean) < 8:
                            st.warning("⚠️ Password must be at least 8 characters.")
                        else:
                            create_user(new_u_clean, new_p_clean, new_r)
                            log_audit(user, "USER_CREATED", "USER",
                                      new_value=f"{new_u_clean} ({new_r})",
                                      reason=f"Created by Admin: {user}")
                            st.success(f"User '{new_u_clean}' created with role: {new_r}.")
                    else:
                        st.warning("Username and password are required.")

    # ── Top action bar — Back to Periodic Review (left) + End Session (right) ──
    # The back button only appears when inside a Periodic Review sub-module.
    # Both buttons sit in the same row so they align at the same visual level.
    _in_pr_submodule = (
        st.session_state.get("app_mode") == "Periodic Review"
        and st.session_state.get("pr_active_module") is not None
    )
    if _in_pr_submodule:
        _back_col, _spacer_col, _end_col = st.columns([5, 4, 3])
        with _back_col:
            if st.button("← Back to Periodic Review", key="pr_back_btn",
                         use_container_width=True):
                st.session_state["pr_active_module"] = None
                st.rerun()
        with _end_col:
            if st.button("⏹ End Session", key="terminate_hidden_trigger",
                         use_container_width=True):
                st.session_state["_logout_requested"] = True
                st.session_state["_logout_reason"]    = "User clicked End Session (sticky button)"
                st.rerun()
    else:
        _es_space, _es_col = st.columns([11, 3])
        with _es_col:
            if st.button("⏹ End Session", key="terminate_hidden_trigger"):
                st.session_state["_logout_requested"] = True
                st.session_state["_logout_reason"]    = "User clicked End Session (sticky button)"
                st.rerun()

    # ── Sticky End Session: only inject when page has results (scrollable) ──
    # Using Python session state is the only reliable way to gate this —
    # JS DOM scroll measurements inside an iframe are defeated by Streamlit's
    # rerun cycle. When results are present the page is always scrollable;
    # when no results are loaded the page fits the viewport.
    _has_results = bool(
        st.session_state.get("last_result") or
        st.session_state.get("esig_pending")
    )
    if _has_results:
        _st_components.html("""
    <script>
    (function() {
        var DOC = window.parent.document;
        var old = DOC.getElementById('sticky-terminate-btn');
        if (old) old.parentNode.removeChild(old);
        var btn = DOC.createElement('button');
        btn.id = 'sticky-terminate-btn';
        btn.innerHTML = '&#9209; End Session';
        Object.assign(btn.style, {
            position:     'fixed',
            top:          '58px',
            right:        '20px',
            zIndex:       '2147483647',
            background:   '#dc2626',
            color:        'white',
            border:       'none',
            borderRadius: '8px',
            padding:      '6px 16px',
            fontSize:     '0.82rem',
            fontWeight:   '600',
            cursor:       'pointer',
            boxShadow:    '0 2px 8px rgba(220,38,38,0.4)',
            fontFamily:   'inherit'
        });
        btn.onmouseover = function(){ this.style.background = '#b91c1c'; };
        btn.onmouseout  = function(){ this.style.background = '#dc2626'; };
        btn.onclick = function() {
            var all = DOC.querySelectorAll('button');
            for (var i = 0; i < all.length; i++) {
                if (all[i].innerText && all[i].innerText.trim().indexOf('\u23f9') === 0) {
                    all[i].click(); return;
                }
            }
        };
        DOC.body.appendChild(btn);
    })();
    </script>
        """, height=0)
    else:
        # No results — remove any stale sticky button (e.g. after New Analysis)
        _st_components.html("""
    <script>
    (function() {
        var old = window.parent.document.getElementById('sticky-terminate-btn');
        if (old) old.parentNode.removeChild(old);
    })();
    </script>
        """, height=0)

    # IP capture on every authenticated load
    _capture_client_ip()

    # ── Mode routing ──────────────────────────────────────────────────────────
    _mode = st.session_state.get("app_mode", "New Validation")

    if _mode == "Change Impact Analysis":
        show_change_impact(user, role, MODELS[st.session_state.selected_model])
        return

    if _mode == "Periodic Review":
        show_periodic_review(user, role, MODELS[st.session_state.selected_model])
        return

    if _mode == "Delta Generation (coming soon)":
        st.title("Delta Generation")
        st.info("🚧 Coming soon. Select another mode in the sidebar to continue.")
        return

    # ── Main area — New Validation ────────────────────────────────────────────
    st.title("Auto-Generate Validation Package")

    # ── Restore active job from DB on page load ───────────────────────────────
    # If the user closed the browser during a run and logged back in,
    # restore their most recent active or completed job automatically.
    if not st.session_state.get("active_job_id"):
        try:
            _conn = db_connect()
            _restore = _conn.execute(
                """SELECT job_id, status FROM jobs
                   WHERE user = ?
                   AND status IN ('queued','running','complete')
                   ORDER BY created_at DESC LIMIT 1""",
                (user,)
            ).fetchone()
            _conn.close()
            if _restore:
                _rjob_id, _rjob_status = _restore
                # Only restore if job is recent (< 24 hours old)
                _conn2 = db_connect()
                _rjob_age = _conn2.execute(
                    "SELECT created_at FROM jobs WHERE job_id = ?", (_rjob_id,)
                ).fetchone()
                _conn2.close()
                if _rjob_age:
                    import datetime as _dt_restore
                    _created = _dt_restore.datetime.fromisoformat(_rjob_age[0])
                    _age_hrs = (_dt_restore.datetime.utcnow() - _created).total_seconds() / 3600
                    if _age_hrs < 24:
                        st.session_state["active_job_id"] = _rjob_id
                        # Restart worker in case server was rebooted
                        if _rjob_status in ("queued", "running"):
                            ensure_worker_running()
        except Exception:
            pass  # non-fatal — user just won't see restored job

    # Dynamic key allows the file-uploader widget to be fully reset after a
    # completed run. Incrementing uploader_key_n forces Streamlit to mount
    # a brand-new widget instance, which clears any retained file state.
    uploader_key = f"main_sop_uploader_{st.session_state.uploader_key_n}"
    sop_widget = st.file_uploader(
        "Upload URS (The 'What')", type="pdf", key=uploader_key
    )

    if sop_widget is not None:
        is_valid_upload, upload_err = validate_upload(sop_widget)
        if not is_valid_upload:
            st.error(upload_err)
            st.session_state.sop_file_bytes = None
            st.session_state.sop_file_name  = None
        else:
            raw_bytes = sop_widget.getvalue()
            if raw_bytes and b'%PDF' in raw_bytes[:1024]:
                new_file = (st.session_state.sop_file_name != sop_widget.name)
                if new_file:
                    # New file — run Stage 0+1 heuristic immediately (free, instant)
                    st.session_state.pop("last_result", None)
                    st.session_state.pop("doc_validation_msg", None)
                    try:
                        pages       = extract_pages(raw_bytes)
                        full_text   = "\n".join(pages) if pages else ""
                        sample      = "\n\n".join(pages[:5]).lower() if pages else ""  # extended from 2 to 5 pages
                        pos_hits    = [p for p in _URS_POSITIVE if re.search(p, sample, re.IGNORECASE)]
                        neg_hits    = [p for p in _URS_NEGATIVE if re.search(p, sample, re.IGNORECASE)]
                        shall_count = len(re.findall(r'\b(shall|must)\b', full_text, re.IGNORECASE))

                        if len(full_text.strip()) < 300:
                            st.session_state["doc_validation_msg"] = (
                                "error",
                                "⛔ **Document rejected:** too little extractable text. "
                                "The PDF may be image-only or corrupt. Upload a text-based URS."
                            )
                            st.session_state.sop_file_bytes = None
                            st.session_state.sop_file_name  = None
                        elif neg_hits:
                            matched = [p.replace(r'\b','').replace('\\','') for p in neg_hits[:3]]
                            st.session_state["doc_validation_msg"] = (
                                "error",
                                f"⛔ **Document rejected:** non-URS content detected "
                                f"({', '.join(matched)}). Upload a URS, SRS, or SOP."
                            )
                            st.session_state.sop_file_bytes = None
                            st.session_state.sop_file_name  = None
                        elif shall_count < 2:
                            st.session_state["doc_validation_msg"] = (
                                "error",
                                f"⛔ **Document rejected:** only {shall_count} 'shall'/'must' "
                                f"statement(s) found. A URS must contain requirement statements. "
                                f"Upload a valid User Requirements Specification."
                            )
                            st.session_state.sop_file_bytes = None
                            st.session_state.sop_file_name  = None
                        elif len(pos_hits) < 3:
                            st.session_state["doc_validation_msg"] = (
                                "warning",
                                f"⚠️ **Low URS signal** ({len(pos_hits)} indicator(s), "
                                f"{shall_count} shall/must). "
                                f"The AI will perform a deeper content check at Run Analysis."
                            )
                            st.session_state.sop_file_bytes = raw_bytes
                            st.session_state.sop_file_name  = sop_widget.name
                        else:
                            st.session_state["doc_validation_msg"] = (
                                "success",
                                f"✅ **Pre-screen passed** — {len(pos_hits)} URS indicator(s), "
                                f"{shall_count} requirement statement(s). "
                                f"AI deep-check runs at analysis time."
                            )
                            st.session_state.sop_file_bytes = raw_bytes
                            st.session_state.sop_file_name  = sop_widget.name
                    except Exception:
                        st.session_state.sop_file_bytes = raw_bytes
                        st.session_state.sop_file_name  = sop_widget.name
                else:
                    # Same file retained (e.g. model change rerun)
                    if st.session_state.sop_file_bytes is None:
                        st.session_state.sop_file_bytes = raw_bytes
            else:
                st.error("⚠️ Uploaded file does not appear to be a valid PDF. Please try again.")
                st.session_state.sop_file_bytes = None
                st.session_state.sop_file_name  = None
    else:
        st.session_state.sop_file_bytes = None
        st.session_state.sop_file_name  = None
        st.session_state.pop("last_result", None)
        st.session_state.pop("doc_validation_msg", None)

    # Show document validation banner
    val_msg = st.session_state.get("doc_validation_msg")
    if val_msg:
        level, msg = val_msg
        if   level == "error":   st.error(msg)
        elif level == "warning": st.warning(msg)
        elif level == "success": st.success(msg)



    is_ready = st.session_state.sop_file_bytes is not None

    # Retained-file banner — only shown when file is retained but uploader widget is empty
    if is_ready and sop_widget is None and st.session_state.sop_file_name:
        st.info(
            f"📎 Retained: **{st.session_state.sop_file_name}** — model change did not clear the file."
        )

    # ── MANUAL EDIT v29-custom — DO NOT OVERWRITE ──────────────
    st.markdown("<br>", unsafe_allow_html=True)
    # ── END MANUAL EDIT ────────────────────────────────────────

    # ── System Document uploader (New Validation only) ────────────────────────
    st.markdown(
        "<p style='font-size:0.85rem;font-weight:600;color:#475569;margin-bottom:4px;'>"
        "📂 System Context Document <span style='font-weight:400;color:#94a3b8;'>"
        "(optional) — User Guide, SOP, or Instruction Manual</span></p>",
        unsafe_allow_html=True
    )
    sys_up_key = f"sidebar_sys_uploader_{st.session_state.sys_uploader_key_n}"
    sidebar_sys = st.file_uploader(
        "SysContext", type="pdf", key=sys_up_key, label_visibility="collapsed"
    )
    if sidebar_sys is not None:
        raw = sidebar_sys.getvalue()
        if raw and b'%PDF' in raw[:1024]:
            _sys_pages = []
            try:
                import pdfplumber as _plb
                with _plb.open(io.BytesIO(raw)) as _pdf:
                    for _pg in _pdf.pages[:4]:
                        _t = _pg.extract_text() or ""
                        if _t.strip():
                            _sys_pages.append(_t.lower())
            except Exception:
                pass
            _sys_sample = " ".join(_sys_pages)

            _SYS_POSITIVE = [
                r'\bclick\b', r'\bnavigate\b', r'\bselect\b', r'\bdashboard\b',
                r'\bscreen\b', r'\bbutton\b', r'\bmenu\b', r'\bfield\b',
                r'\bworkflow\b', r'\bprocedure\b', r'\bconfigure\b',
                r'\binstall\b', r'\blog.?in\b', r'\buser guide\b',
                r'\bmanual\b', r'\bsop\b', r'\binstruction\b',
                r'\bstep\s+\d\b', r'\bmodule\b', r'\btab\b', r'\bform\b',
                r'\bsystem\b', r'\bapplication\b', r'\bsoftware\b',
            ]
            _SYS_NEGATIVE = [
                r'\bwork experience\b', r'\bemployment history\b',
                r'\bcurriculum vitae\b', r'\b\bcv\b\b',
                r'\beducation\b.*\buniversity\b', r'\bdegree\b.*\bgraduat',
                r'\bskills\b.*\bproficien', r'\breferences available\b',
                r'\bdate of birth\b', r'\blinkedin\.com\b',
                r'\binvoice\b', r'\bpurchase order\b',
                r'\btotal due\b', r'\bremit payment\b',
                r'\bdear\b.*\bsincerely\b',
            ]
            if _sys_sample:
                _neg_hits = [p for p in _SYS_NEGATIVE
                             if re.search(p, _sys_sample, re.IGNORECASE)]
                _pos_hits = [p for p in _SYS_POSITIVE
                             if re.search(p, _sys_sample, re.IGNORECASE)]
                if _neg_hits:
                    st.error(
                        "⛔ **Document rejected** — this appears to be a personal or "
                        "non-operational document (CV, invoice, letter, etc.). "
                        "Upload a system User Guide, operational SOP, or instruction manual."
                    )
                    st.session_state["sys_context_bytes"] = None
                    st.session_state["sys_context_name"]  = None
                elif len(_pos_hits) < 3:
                    st.warning(
                        f"⚠️ **Low system-doc signal** ({len(_pos_hits)} indicator(s)). "
                        "Expected a User Guide, SOP, or instruction manual with screen "
                        "names, workflow steps, or procedural language. "
                        "The document will be used but may not improve FRS quality."
                    )
                    st.session_state["sys_context_bytes"] = raw
                    st.session_state["sys_context_name"]  = sidebar_sys.name
                else:
                    st.success(f"✅ System document accepted: **{sidebar_sys.name}**")
                    st.session_state["sys_context_bytes"] = raw
                    st.session_state["sys_context_name"]  = sidebar_sys.name
            else:
                st.warning(
                    "⚠️ Could not extract text from this PDF. "
                    "Ensure it is OCR-searchable for best results."
                )
                st.session_state["sys_context_bytes"] = raw
                st.session_state["sys_context_name"]  = sidebar_sys.name
    elif sidebar_sys is None:
        st.session_state["sys_context_bytes"] = None
        st.session_state["sys_context_name"]  = None

    ctx_name = st.session_state.get("sys_context_name")
    if ctx_name and sidebar_sys is None:
        st.info(f"📄 Retained system document: **{ctx_name}**")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Async job status polling ─────────────────────────────────────────────
    # If a job was previously submitted, show its status and poll for completion.
    _active_job = st.session_state.get("active_job_id")
    if _active_job:
        _job = _job_get(_active_job)
        if _job:
            _status = _job.get("status", "unknown")
            _prog   = _job.get("progress", 0)
            _msg    = _job.get("progress_msg", "")

            if _status == "complete":
                st.success(f"✅ Analysis complete — {_msg}")
                _xlsx = _job.get("result_xlsx")
                if _xlsx:
                    import pandas as _pd
                    import io as _io
                    _fname = _job.get("file_name","validation").replace(".pdf","")
                    st.download_button(
                        label="📥 Download Validation Package (.xlsx)",
                        data=_xlsx,
                        file_name=f"Validation_Package_{_fname}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="async_download_btn",
                        use_container_width=True,
                    )
                    # Preview sheets
                    with st.expander("📋 Preview Generated Sheets", expanded=True):
                        for _sheet, _csv_key in [
                            ("URS", "result_urs"), ("FRS", "result_frs"),
                            ("OQ", "result_oq"),   ("Traceability", "result_trace"),
                        ]:
                            _csv = _job.get(_csv_key, "")
                            if _csv and _csv.strip():
                                try:
                                    _df = _pd.read_csv(_io.StringIO(_csv))
                                    st.markdown(f"**{_sheet}** — {len(_df)} rows")
                                    st.dataframe(_df, use_container_width=True)
                                except Exception:
                                    pass
                if st.button("🔄 New Analysis", key="async_new_btn"):
                    st.session_state.pop("active_job_id", None)
                    st.session_state.sop_file_bytes = None
                    st.session_state.sop_file_name  = None
                    st.rerun()

            elif _status == "failed":
                st.error(f"❌ Analysis failed: {_job.get('error_msg','Unknown error')}")
                if st.button("🔄 Retry", key="async_retry_btn"):
                    st.session_state.pop("active_job_id", None)
                    st.rerun()

            else:
                # queued or running — show live progress and auto-refresh
                _label = "⏳ Queued — starting shortly..." if _status == "queued" else f"🔬 Running..."
                st.info(_label)
                st.progress(max(_prog, 2) / 100)
                if _msg:
                    st.caption(_msg)
                _jid_short = _active_job[:8]
                st.caption(f"Job reference: **{_active_job}**  |  You can safely close this tab and return later.")
                _time_mod.sleep(3)
                st.rerun()
            # Do not render the upload/button UI while a job is active
            return

    if st.button("🚀 Run Analysis", key="run_analysis_btn", disabled=not is_ready):
        file_bytes = st.session_state.sop_file_bytes
        file_name  = st.session_state.sop_file_name or "unknown.pdf"
        model_id   = MODELS[st.session_state.selected_model]
        sys_ctx    = st.session_state.get("sys_context_bytes", None)

        # ── Stage 2: LLM document pre-flight ─────────────────────────────────
        with st.spinner("🔍 Validating document type..."):
            is_valid_doc, validation_msg = validate_urs_document(file_bytes, model_id)

        if not is_valid_doc:
            st.error(validation_msg)
            log_audit(user, "DOCUMENT_REJECTED", "URS_FILE",
                      new_value=file_name,
                      reason=f"Document validation failed: {validation_msg[:300]}")
            st.session_state.sop_file_bytes = None
            st.session_state.sop_file_name  = None
            st.session_state["doc_validation_msg"] = (
                "error",
                f"⛔ **Document rejected** — not a valid URS/SRS/SOP. "
                f"Please upload a requirements specification document."
            )
            st.stop()

        log_audit(user, "ANALYSIS_INITIATED", "URS_FILE",
                  new_value=file_name,
                  reason=(
                      f"Model: {st.session_state.selected_model} | Prompt: {PROMPT_VERSION} | "
                      f"Temp: {TEMPERATURE} | Validation: {validation_msg[:100]}"
                      + (f" | Guide: {st.session_state.get('sys_context_name','')}"
                         if st.session_state.get("sys_context_name") else "")
                  ))

        # ── Submit async job ──────────────────────────────────────────────────
        _job_id = submit_job(
            user         = user,
            file_bytes   = file_bytes,
            file_name    = file_name,
            model_id     = model_id,
            sys_ctx_bytes = sys_ctx,
            sys_ctx_name = st.session_state.get("sys_context_name",""),
        )
        st.session_state["active_job_id"] = _job_id
        log_audit(user, "JOB_SUBMITTED", "ASYNC_QUEUE",
                  new_value=_job_id, reason=f"file={file_name}")
        st.rerun()   # immediately show the polling UI

        progress_bar = st.progress(0)
        status_text  = st.empty()

        # ── Chain-of-thought status container — wraps the real work ──────────
        # st.status stays "running" while the with block executes, collapses to
        # "complete" on success or "error" on exception. Each st.write() call
        # appends a live step log that the user can read during the wait.
        with st.status("🔍 Running GxP Validation Pipeline...", expanded=True) as cot_status:
          try:
            # ── Parser Quality Indicator ─────────────────────────────────────
            # Warn early if the document has low text density (many images,
            # scanned pages) so the user knows to expect lower extraction quality.
            st.write("📑 Parsing URS document structure and page layout...")
            _pq_pages = []
            try:
                import pdfplumber as _plumber
                with _plumber.open(io.BytesIO(file_bytes)) as _pdf:
                    _total_pg   = len(_pdf.pages)
                    _image_pg   = sum(1 for pg in _pdf.pages
                                      if len(pg.images or []) > 0
                                      and len((pg.extract_text() or "").strip()) < 100)
                    _text_chars = sum(len((pg.extract_text() or "").strip())
                                      for pg in _pdf.pages)
                _avg_density = _text_chars / max(_total_pg, 1)
                if _image_pg > _total_pg * 0.4 or _avg_density < 200:
                    st.warning(
                        f"⚠️ **Parser Quality Warning** — {_image_pg}/{_total_pg} pages appear "
                        f"image-heavy or have low text density (avg {int(_avg_density)} chars/page). "
                        "Ensure the URS is OCR-searchable for 100% extraction accuracy. "
                        "Scanned PDFs without OCR will produce incomplete results."
                    )
                else:
                    st.write(f"✅ Document quality check passed — "
                             f"{_total_pg} pages, avg {int(_avg_density)} chars/page")
            except Exception:
                st.write("📑 Document quality check skipped (pdfplumber unavailable)")

            st.write("🔒 Applying GxP document integrity checks (ALCOA+)...")
            if st.session_state.get("sys_context_name"):
                st.write(f"📖 Loading System User Guide: "
                         f"{st.session_state.get('sys_context_name')} ...")

            # ── Step 1: Two-pass AI analysis ─────────────────────────────────
            st.write(f"🔬 Pass 1 — Extracting URS requirements using "
                     f"{st.session_state.selected_model}...")
            urs_df, frs_df, oq_df, trace_df, gap_df = run_segmented_analysis(
                file_bytes, model_id, progress_bar, status_text, sys_ctx
            )

            st.write("🏗️ Pass 2 — Mapping Functional Requirements to system architecture...")
            st.write("🧪 Pass 2 — Risk-adjusted OQ test cases (High ≥3 | Med ≥2 | Low ≥1)...")
            st.write("📊 Building traceability matrix (URS → FRS → OQ)...")
            if st.session_state.get("sys_context_name"):
                st.write("🔀 Pass 3 — Bidirectional URS ↔ User Guide cross-source gap analysis...")

            # ── Guard: if the AI returned nothing useful, stop cleanly ────────
            if urs_df.empty and frs_df.empty:
                progress_bar.empty()
                status_text.empty()
                cot_status.update(label="❌ Pipeline aborted — no output", state="error")
                st.error(
                    "⚠️ No requirements were extracted. This usually means:\n"
                    "- **API quota exceeded** — check your API key billing/limits\n"
                    "- **Rate limit** — wait a minute and try again\n"
                    "- **Model unavailable** — try a different AI Model\n\n"
                    "The error detail is shown above."
                )
                log_audit(user, "ANALYSIS_ABORTED", "URS_FILE",
                          reason="Empty AI output — possible rate limit or quota error")
                return

            # ── URS Accountability Check ───────────────────────────────────────
            if not urs_df.empty and "Req_ID" in urs_df.columns:
                urs_ids_all  = set(urs_df["Req_ID"].dropna().astype(str).str.strip())
                frs_urs_refs = set()
                if not frs_df.empty and "Source_URS_Ref" in frs_df.columns:
                    frs_urs_refs = set(frs_df["Source_URS_Ref"].dropna().astype(str).str.strip())
                uncovered_urs = urs_ids_all - frs_urs_refs
                if uncovered_urs:
                    log_audit(user, "URS_FRS_GAP_DETECTED", "URS_FILE",
                              new_value=f"{len(uncovered_urs)} URS IDs missing FRS",
                              reason=f"IDs: {', '.join(sorted(uncovered_urs)[:20])}")

            # ── Step 2: Deterministic validation ──────────────────────────────
            st.write("🛡️ Running deterministic validation rules R0–R6...")
            status_text.text("🔍 Running deterministic validation rules R1–R5...")
            gap_df, det_df = run_deterministic_validation(frs_df, oq_df, gap_df, urs_df)
            for _df in [gap_df, det_df, trace_df]:
                _df.fillna("N/A", inplace=True)
                _df.replace("", "N/A", inplace=True)

            # ── Step 3: Persist documents ──────────────────────────────────────
            id_urs   = save_document("URS_Extraction", urs_df.to_csv(index=False),  user, file_name)
            id_frs   = save_document("FRS",            frs_df.to_csv(index=False),  user, file_name)
            id_oq    = save_document("OQ",             oq_df.to_csv(index=False),   user, file_name)
            id_trace = save_document("Traceability",   trace_df.to_csv(index=False),user, file_name)
            id_gap   = save_document("Gap_Analysis",   gap_df.to_csv(index=False),  user, file_name)
            id_det   = save_document("Det_Validation", det_df.to_csv(index=False),  user, file_name)
            doc_ids  = (f"URS:{id_urs}, FRS:{id_frs}, OQ:{id_oq}, "
                        f"Trace:{id_trace}, Gap:{id_gap}, Det:{id_det}")

            # ── Step 4: AI generation log ──────────────────────────────────────
            log_ai_generation(
                user, st.session_state.selected_model,
                PROMPT_VERSION, TEMPERATURE, file_name,
                document_ids_used=doc_ids
            )

            # ── Step 5: Audit entries ──────────────────────────────────────────
            log_audit(user, "URS_EXTRACTED",           f"URS doc:{id_urs}",
                      new_value=f"{len(urs_df)} structured requirements")
            log_audit(user, "FRS_GENERATED",           f"FRS doc:{id_frs}",
                      new_value=f"{len(frs_df)} requirements")
            log_audit(user, "OQ_GENERATED",            f"OQ doc:{id_oq}",
                      new_value=f"{len(oq_df)} test cases")
            log_audit(user, "TRACEABILITY_GENERATED",  f"Trace doc:{id_trace}",
                      new_value=f"{len(trace_df)} rows")
            log_audit(user, "GAP_ANALYSIS_GENERATED",  f"Gap doc:{id_gap}",
                      new_value=f"{len(gap_df)} AI gaps")
            log_audit(user, "DET_VALIDATION_RUN",      f"Det doc:{id_det}",
                      new_value=f"{len(det_df)} deterministic issues (R1-R5)")

            # ── Step 6: Confidence summary ─────────────────────────────────────
            frs_review = 0
            oq_review  = 0
            if not frs_df.empty and "Confidence_Flag" in frs_df.columns:
                frs_review = int(frs_df["Confidence_Flag"].astype(str).str.contains("Review").sum())
            if not oq_df.empty and "Confidence_Flag" in oq_df.columns:
                oq_review = int(oq_df["Confidence_Flag"].astype(str).str.contains("Review").sum())

            # ── Step 7: Build audit log and dashboard ──────────────────────────
            ver_frs = get_next_doc_version("FRS") - 1
            ver_oq  = get_next_doc_version("OQ")  - 1
            audit_df     = build_audit_log_sheet(
                user, file_name, st.session_state.selected_model,
                frs_df, oq_df, gap_df, det_df, ver_frs, ver_oq, doc_ids,
                sys_context_name=st.session_state.get("sys_context_name") or ""
            )
            dashboard_df = build_dashboard_sheet(
                frs_df, oq_df, gap_df, det_df, trace_df, file_name,
                st.session_state.selected_model
            )

            gap_sheet_included = not gap_df.empty
            dataframes = {
                "Dashboard":      dashboard_df,
                "URS_Extraction": urs_df,
                "FRS":            frs_df,
                "OQ":             oq_df,
                "Traceability":   trace_df,
                "Det_Validation": det_df,
                "Audit_Log":      audit_df,
            }
            if gap_sheet_included:
                dataframes = {
                    "Dashboard":      dashboard_df,
                    "URS_Extraction": urs_df,
                    "FRS":            frs_df,
                    "OQ":             oq_df,
                    "Traceability":   trace_df,
                    "Gap_Analysis":   gap_df,
                    "Det_Validation": det_df,
                    "Audit_Log":      audit_df,
                }

            status_text.text("📊 Building validation package — this may take 20–30 seconds...")
            st.write("📋 Compiling signed validation workbook...")
            xlsx_bytes_presig = build_styled_excel(
                dataframes,
                user=user,
                file_name=file_name,
                model_name=st.session_state.selected_model,
                sys_context_name=st.session_state.get("sys_context_name") or "",
                dashboard_df=dashboard_df,
            )

            doc_hash = hashlib.sha256(xlsx_bytes_presig).hexdigest()
            log_audit(user, "WORKBOOK_BUILT", "VALIDATION_PACKAGE",
                      new_value=f"Validation_Package_{datetime.date.today()}.xlsx",
                      reason=f"doc_ids={doc_ids} | hash={doc_hash[:16]}...")

            status_text.empty()
            progress_bar.empty()

            covered = partial_cov = 0
            if not trace_df.empty and "Coverage_Status" in trace_df.columns:
                covered     = int((trace_df["Coverage_Status"] == "Covered").sum())
                partial_cov = int((trace_df["Coverage_Status"] == "Partial").sum())
            total_reqs = len(frs_df)
            has_tests  = covered + partial_cov
            cov_pct    = round(has_tests / total_reqs * 100, 1) if total_reqs > 0 else 0.0

            cot_status.update(
                label=f"✅ Pipeline complete — {len(urs_df)} requirements, "
                      f"{len(oq_df)} tests, {len(gap_df)+len(det_df)} issues",
                state="complete", expanded=False
            )

            st.session_state["esig_pending"] = {
                "xlsx_bytes_presig": xlsx_bytes_presig,
                "doc_hash":          doc_hash,
                "dataframes":        dataframes,
                "frs_review":        frs_review,
                "oq_review":         oq_review,
                "total_reqs":        len(frs_df),
                "total_tests":       len(oq_df),
                "total_urs":         len(urs_df),
                "covered":           covered,
                "cov_pct":           cov_pct,
                "gap_count":         len(gap_df),
                "det_count":         len(det_df),
                "non_testable_count": int(
                    det_df[det_df["Gap_Type"].astype(str) == "Non_Testable_Requirement"].shape[0]
                    if not det_df.empty and "Gap_Type" in det_df.columns else 0
                ),
                "file_name":         file_name,
                "model_name":        st.session_state.selected_model,
                "doc_ids":           doc_ids,
            }
            st.rerun()

          except Exception as e:
            err_msg = str(e)
            log_audit(user, "ANALYSIS_ERROR", "URS_FILE", reason=err_msg[:500])
            cot_status.update(label="❌ Pipeline failed — see error below", state="error")
            if "Pass 1" in err_msg or "Pass 2" in err_msg or "ALCOA+" in err_msg or "segment" in err_msg.lower():
                st.error(f"🛑 **GxP Fail-Stop Protocol Activated**\n\n{err_msg}")
            else:
                st.error(f"❌ Engineering Error: {err_msg}")
                import traceback
                st.error(traceback.format_exc())

    # ── Determine what to display: signed result or pending preview ──────────
    # Preview is shown from either state so the user sees tables immediately.
    # The e-sig form only appears inline when they click a download button.
    _display = st.session_state.get("last_result") or st.session_state.get("esig_pending")

    if _display:
        r            = _display
        is_signed    = "last_result" in st.session_state
        pending      = st.session_state.get("esig_pending")

        # ── Signature confirmation banner (only after signing) ────────────────
        if is_signed and r.get("sig_id"):
            st.success(
                f"✅ Electronically signed — SIG-{r['sig_id']:06d} | "
                f"{r['sig_meaning']} | {r['sig_timestamp'][:19]} UTC"
            )

        # ── Validation Package Summary ─────────────────────────────────────────
        fully_covered = r.get("covered", 0)
        total_reqs    = r["total_reqs"]
        cov_pct       = r["cov_pct"]
        gap_total     = r["gap_count"] + r["det_count"]
        cov_status    = "✅ PASS" if cov_pct >= 80 else ("⚠️ REVIEW" if cov_pct >= 60 else "❌ FAIL")

        st.markdown(f"""
<div style="background:#0f172a;border-radius:12px;padding:20px 28px;margin-bottom:18px;
            border-left:5px solid #2563eb;font-family:'Inter',sans-serif;">
  <p style="color:#94a3b8;font-size:0.75rem;letter-spacing:3px;text-transform:uppercase;
            margin:0 0 10px 0;">Validation Package Summary</p>
  <hr style="border:none;border-top:1px solid #1e293b;margin:0 0 14px 0;">
  <table style="width:100%;border-collapse:collapse;color:white;font-size:0.92rem;">
    <tr>
      <td style="padding:4px 0;color:#94a3b8;width:55%;">📋 Requirements extracted (URS)</td>
      <td style="padding:4px 0;font-weight:700;color:#e2e8f0;">{r["total_urs"]}</td>
    </tr>
    <tr>
      <td style="padding:4px 0;color:#94a3b8;">📐 FRS requirements generated</td>
      <td style="padding:4px 0;font-weight:700;color:#e2e8f0;">{total_reqs}</td>
    </tr>
    <tr>
      <td style="padding:4px 0;color:#94a3b8;">🧪 OQ test cases generated</td>
      <td style="padding:4px 0;font-weight:700;color:#e2e8f0;">{r["total_tests"]}</td>
    </tr>
    <tr>
      <td style="padding:4px 0;color:#94a3b8;">✅ Fully covered requirements</td>
      <td style="padding:4px 0;font-weight:700;color:#4ade80;">{fully_covered}</td>
    </tr>
    <tr>
      <td style="padding:4px 0;color:#94a3b8;">📊 Traceability coverage</td>
      <td style="padding:4px 0;font-weight:700;
                 color:{'#4ade80' if cov_pct >= 80 else ('#facc15' if cov_pct >= 60 else '#f87171')};">
        {cov_pct}% &nbsp; {cov_status}
      </td>
    </tr>
    <tr>
      <td style="padding:4px 0;color:#94a3b8;">⚠️ Gaps detected (AI + deterministic)</td>
      <td style="padding:4px 0;font-weight:700;
                 color:{'#f87171' if gap_total > 0 else '#4ade80'};">
        {gap_total}
      </td>
    </tr>
  </table>
</div>
""", unsafe_allow_html=True)

        col1, col2, col3, col4, col5, col6 = st.columns(6)
        col1.metric("📄 URS Requirements", r["total_urs"])
        col2.metric("📋 FRS Requirements", r["total_reqs"])
        col3.metric("🧪 OQ Test Cases",    r["total_tests"])
        col4.metric("📊 Coverage",          f"{r['cov_pct']}%")
        col5.metric("⚠️ Issues (AI+Det)",   r["gap_count"] + r["det_count"])
        col6.metric("🚫 Non-Testable",
                    f"{round(r.get('non_testable_count',0)/r['total_urs']*100,1) if r['total_urs']>0 else 0}%",
                    delta=f"{r.get('non_testable_count',0)} reqs",
                    delta_color="inverse")

        # ── Hero metric: Unmitigated GxP Risks ───────────────────────────────
        # Counts High-Risk FRS items with zero OQ test coverage — the single
        # most important compliance signal. Red = active regulatory exposure.
        _unmitigated = 0
        _dfs = r.get("dataframes", {})
        _frs_df = _dfs.get("FRS", pd.DataFrame())
        _oq_df  = _dfs.get("OQ",  pd.DataFrame())
        if not _frs_df.empty and "Risk" in _frs_df.columns:
            _high_risk_ids = set(
                _frs_df[_frs_df["Risk"].astype(str).str.lower() == "high"]
                .get("ID", pd.Series(dtype=str)).astype(str).str.strip()
            )
            if _high_risk_ids:
                if not _oq_df.empty and "Requirement_Link" in _oq_df.columns:
                    _tested_ids = set(_oq_df["Requirement_Link"].astype(str).str.strip())
                    _unmitigated = len(_high_risk_ids - _tested_ids)
                else:
                    _unmitigated = len(_high_risk_ids)

        _hero_color  = "#dc2626" if _unmitigated > 0 else "#059669"
        _hero_icon   = "🔴" if _unmitigated > 0 else "🟢"
        _hero_label  = "CRITICAL — Regulatory Exposure" if _unmitigated > 0 else "All High-Risk Requirements Covered"
        _hero_detail = (
            f"{_unmitigated} High-Risk FRS item(s) have zero OQ test coverage. "
            "Per GAMP 5, high-risk requirements require ≥3 test cases. "
            "This package will fail a regulatory audit as-is — add OQ tests before signing."
            if _unmitigated > 0 else
            "All High-Risk FRS requirements have OQ test coverage. "
            "No unmitigated regulatory exposure detected."
        )
        st.markdown(f"""
<div style="background:{'#1a0505' if _unmitigated > 0 else '#052019'};
            border:2px solid {_hero_color};border-radius:12px;
            padding:16px 24px;margin:12px 0 8px 0;
            font-family:'Inter',sans-serif;">
  <div style="display:flex;align-items:center;gap:12px;margin-bottom:6px;">
    <span style="font-size:2rem;">{_hero_icon}</span>
    <div>
      <p style="margin:0;color:#94a3b8;font-size:0.72rem;letter-spacing:2px;
                text-transform:uppercase;">Unmitigated GxP Risks</p>
      <p style="margin:0;font-size:2rem;font-weight:800;color:{_hero_color};
                line-height:1;">{_unmitigated}</p>
    </div>
    <div style="margin-left:16px;border-left:1px solid #334155;padding-left:16px;">
      <p style="margin:0;color:{'#fca5a5' if _unmitigated > 0 else '#6ee7b7'};
                font-size:0.85rem;font-weight:600;">{_hero_label}</p>
      <p style="margin:4px 0 0 0;color:#94a3b8;font-size:0.76rem;">{_hero_detail}</p>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

        # ── Non-Testable % hero card ──────────────────────────────────────────
        _nt_count = r.get("non_testable_count", 0)
        _nt_pct   = round(_nt_count / r["total_urs"] * 100, 1) if r["total_urs"] > 0 else 0.0
        _nt_color = "#dc2626" if _nt_pct >= 20 else ("#d97706" if _nt_pct >= 10 else "#059669")
        _nt_bg    = "#1a0505"  if _nt_pct >= 20 else ("#1a1000" if _nt_pct >= 10 else "#052019")
        _nt_status = (
            "CRITICAL — URS requires significant rewriting before validation"
            if _nt_pct >= 20 else
            "WARNING — Some requirements need measurable acceptance criteria"
            if _nt_pct >= 10 else
            "PASS — URS is sufficiently testable for validation"
        )
        _nt_detail = (
            f"{_nt_count} of {r['total_urs']} requirements contain ambiguous or non-testable language "
            f"(e.g. 'user-friendly', 'fast', 'should'). "
            "These cannot be objectively validated — a direct 21 CFR Part 11 compliance risk. "
            "See Det_Validation tab → Rule R3d for specific terms and remediation guidance."
            if _nt_count > 0 else
            "All URS requirements contain measurable, testable language. No ambiguous terms detected."
        )
        st.markdown(f"""
<div style="background:{_nt_bg};border:2px solid {_nt_color};border-radius:12px;
            padding:16px 24px;margin:8px 0 12px 0;font-family:'Inter',sans-serif;">
  <div style="display:flex;align-items:center;gap:12px;">
    <span style="font-size:2rem;">{'🔴' if _nt_pct >= 20 else ('🟡' if _nt_pct >= 10 else '🟢')}</span>
    <div>
      <p style="margin:0;color:#94a3b8;font-size:0.72rem;letter-spacing:2px;
                text-transform:uppercase;">Non-Testable Requirements</p>
      <p style="margin:0;font-size:2rem;font-weight:800;color:{_nt_color};line-height:1;">
        {_nt_pct}%
        <span style="font-size:1rem;font-weight:400;color:#64748b;margin-left:8px;">
          ({_nt_count} of {r['total_urs']} requirements)
        </span>
      </p>
    </div>
    <div style="margin-left:16px;border-left:1px solid #334155;padding-left:16px;">
      <p style="margin:0;font-size:0.85rem;font-weight:600;
                color:{'#fca5a5' if _nt_pct >= 20 else ('#fde68a' if _nt_pct >= 10 else '#6ee7b7')};">
        {_nt_status}</p>
      <p style="margin:4px 0 0 0;color:#94a3b8;font-size:0.76rem;">{_nt_detail}</p>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

        if r["frs_review"] > 0 or r["oq_review"] > 0:
            st.warning(
                f"🔶 Confidence Review Required: **{r['frs_review']}** FRS row(s) and "
                f"**{r['oq_review']}** OQ row(s) have confidence < 0.70 — "
                "check Confidence_Flag columns in FRS and OQ sheets."
            )
        if r["det_count"] > 0:
            st.warning(
                f"🔍 Deterministic Validation: **{r['det_count']}** issue(s) (R0–R5) "
                "— see Det_Validation tab."
            )
        if r["gap_count"] > 0:
            st.warning(
                f"⚠️ Gap Analysis: **{r['gap_count']}** AI-detected gap(s) "
                "— see Gap_Analysis tab in the workbook."
            )
        elif r["det_count"] == 0:
            st.success("✅ No AI gaps and no deterministic issues. Review Traceability for partial coverage details.")

        with st.expander("📋 Preview Generated Sheets", expanded=True):
            for sheet_name, df in r["dataframes"].items():
                st.markdown(f"**{sheet_name}** — {len(df)} rows")
                st.dataframe(df, use_container_width=True)

        # ── Download / Sign buttons ────────────────────────────────────────────
        st.markdown("---")

        if is_signed:
            # ── Already signed: real download buttons + New Analysis far right ─
            dl1, dl2, _spacer, clear_col = st.columns([5, 5, 1, 2])
            with dl1:
                st.download_button(
                    label="📥 Download Signed Workbook (.xlsx)",
                    data=r["xlsx_bytes"],
                    file_name=f"Validation_Package_{r['file_name'].replace('.pdf','')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_xlsx_btn",
                    use_container_width=True,
                )
            with dl2:
                st.download_button(
                    label="📄 Download Signed Summary (.pdf)",
                    data=r["pdf_bytes"],
                    file_name=f"Validation_Summary_{r['file_name'].replace('.pdf','')}.pdf",
                    mime="application/pdf",
                    key="download_pdf_btn",
                    use_container_width=True,
                )
            # _spacer is empty — pushes New Analysis to the far right
            with clear_col:
                if st.button("🔄 New Analysis", key="clear_results_btn",
                             use_container_width=True,
                             help="Clear results and upload a new URS document"):
                    st.session_state["uploader_key_n"]     = st.session_state.get("uploader_key_n", 0) + 1
                    st.session_state["sys_uploader_key_n"] = st.session_state.get("sys_uploader_key_n", 0) + 1
                    st.session_state.sop_file_bytes  = None
                    st.session_state.sop_file_name   = None
                    st.session_state["sys_context_bytes"] = None
                    st.session_state["sys_context_name"]  = None
                    st.session_state.pop("last_result",        None)
                    st.session_state.pop("esig_pending",       None)
                    st.session_state.pop("show_esig_form",     None)
                    st.session_state.pop("esig_target",        None)
                    st.session_state.pop("doc_validation_msg", None)
                    log_audit(user, "NEW_ANALYSIS_STARTED", "SESSION",
                              reason="User cleared previous results and sidebar guide to start a new analysis")
                    st.rerun()

        else:
            # ── Not yet signed: Sign & Download triggers + New Analysis far right
            show_form = st.session_state.get("show_esig_form", False)

            if not show_form:
                st.info(
                    "🔏 **Electronic signature required** (21 CFR Part 11). "
                    "Click a download button below to sign and release your package."
                )
                btn1, btn2, _spacer, clear_col = st.columns([5, 5, 1, 2])
                with btn1:
                    if st.button("🔏 Sign & Download Excel (.xlsx)",
                                 key="trigger_esig_xlsx", use_container_width=True,
                                 type="primary"):
                        st.session_state["show_esig_form"] = True
                        st.session_state["esig_target"]    = "xlsx"
                        st.rerun()
                with btn2:
                    if st.button("🔏 Sign & Download PDF (.pdf)",
                                 key="trigger_esig_pdf", use_container_width=True,
                                 type="primary"):
                        st.session_state["show_esig_form"] = True
                        st.session_state["esig_target"]    = "pdf"
                        st.rerun()
                # _spacer empty — pushes New Analysis to the far right
                with clear_col:
                    if st.button("🔄 New Analysis", key="clear_results_btn",
                                 use_container_width=True,
                                 help="Clear results and upload a new URS document"):
                        st.session_state["uploader_key_n"]     = st.session_state.get("uploader_key_n", 0) + 1
                        st.session_state["sys_uploader_key_n"] = st.session_state.get("sys_uploader_key_n", 0) + 1
                        st.session_state.sop_file_bytes  = None
                        st.session_state.sop_file_name   = None
                        st.session_state["sys_context_bytes"] = None
                        st.session_state["sys_context_name"]  = None
                        st.session_state.pop("last_result",        None)
                        st.session_state.pop("esig_pending",       None)
                        st.session_state.pop("show_esig_form",     None)
                        st.session_state.pop("esig_target",        None)
                        st.session_state.pop("doc_validation_msg", None)
                        log_audit(user, "NEW_ANALYSIS_STARTED", "SESSION",
                                  reason="User cleared previous results to start a new analysis")
                        st.rerun()

            # ── Inline e-sig form (appears below preview when triggered) ──────
            if show_form and pending:
                p = pending
                st.markdown("""
<div class="esig-container">
  <p class="esig-title">🔏 Electronic Signature Required</p>
  <p class="esig-subtitle">
    21 CFR Part 11 §11.200 — Two-component non-biometric signature.<br>
    Re-enter your password to sign and release this validation package.
  </p>
</div>
""", unsafe_allow_html=True)

                with st.form("esig_form", clear_on_submit=False):
                    st.markdown(
                        f'<p class="esig-field-label">Signer</p>'
                        f'<div class="esig-user-display">👤 &nbsp; {user}'
                        f' &nbsp;&nbsp;|&nbsp;&nbsp; {role}</div>',
                        unsafe_allow_html=True
                    )
                    st.markdown(
                        '<p class="esig-field-label">Password (re-enter to verify identity)</p>',
                        unsafe_allow_html=True)
                    esig_password = st.text_input(
                        "Password", type="password",
                        placeholder="Enter your login password",
                        label_visibility="collapsed",
                        key="esig_password_input"
                    )
                    st.markdown('<p class="esig-field-label">Meaning of Signature</p>',
                                unsafe_allow_html=True)
                    esig_meaning = st.selectbox(
                        "Meaning", ESIG_MEANINGS, index=0,
                        label_visibility="collapsed",
                        key="esig_meaning_select"
                    )
                    st.markdown(
                        '<p class="esig-warning">⚠️ By submitting this signature you confirm '
                        'that the information in this validation package is accurate and '
                        'complete. This action is recorded and cannot be undone.</p>',
                        unsafe_allow_html=True
                    )
                    col_sign, col_cancel = st.columns([2, 1])
                    with col_sign:
                        submitted = st.form_submit_button(
                            "✍️ Sign & Release Package",
                            use_container_width=True, type="primary"
                        )
                    with col_cancel:
                        cancelled = st.form_submit_button("✖ Cancel",
                                                          use_container_width=True)

                if cancelled:
                    st.session_state["show_esig_form"] = False
                    st.session_state.pop("esig_target", None)
                    log_audit(user, "ESIG_CANCELLED", "VALIDATION_PACKAGE",
                              reason="User cancelled e-signature — package not released")
                    st.info("Signature cancelled. Review the preview and click a download "
                            "button when ready to sign.")
                    st.rerun()

                if submitted:
                    if not esig_password:
                        st.error("⛔ Password is required to sign.")
                    else:
                        conn_v = db_connect()
                        stored = conn_v.execute(
                            "SELECT password_hash FROM users WHERE username=?", (user,)
                        ).fetchone()
                        conn_v.close()

                        if not stored or not verify_password(esig_password, stored[0]):
                            log_audit(user, "ESIG_IDENTITY_FAILED", "VALIDATION_PACKAGE",
                                      reason="E-sig identity verification failed — wrong password")
                            st.error(
                                "⛔ Identity verification failed. The password does not match "
                                "your account. This attempt has been recorded in the audit trail."
                            )
                        else:
                            sig_ts = datetime.datetime.utcnow().isoformat()
                            sig_id = log_esignature(
                                user          = user,
                                role          = role,
                                action        = "GENERATED_VALIDATION_PACKAGE",
                                meaning       = esig_meaning,
                                document_hash = p["doc_hash"],
                                document_name = p["file_name"],
                                model_used    = p["model_name"],
                                prompt_ver    = PROMPT_VERSION,
                                ip_address    = st.session_state.get("user_ip", ""),
                                doc_ids       = p["doc_ids"],
                            )
                            log_audit(user, "ESIG_APPLIED", "VALIDATION_PACKAGE",
                                      new_value=f"SIG-{sig_id:06d}",
                                      reason=(f"Meaning: {esig_meaning} | "
                                              f"Hash: {p['doc_hash'][:16]}... | "
                                              f"Model: {p['model_name']}"))

                            # ── Build signed Excel ────────────────────────────
                            import openpyxl
                            wb_final = openpyxl.load_workbook(
                                filename=io.BytesIO(p["xlsx_bytes_presig"])
                            )
                            build_signature_sheet(
                                wb            = wb_final,
                                user          = user,
                                role          = role,
                                meaning       = esig_meaning,
                                document_hash = p["doc_hash"],
                                document_name = p["file_name"],
                                model_used    = p["model_name"],
                                signature_id  = sig_id,
                                timestamp     = sig_ts,
                            )
                            final_buf = io.BytesIO()
                            wb_final.save(final_buf)
                            xlsx_bytes_final = final_buf.getvalue()

                            # ── Build signed PDF ──────────────────────────────
                            pdf_bytes_final = build_pdf_bytes(
                                r             = p,
                                sig_id        = sig_id,
                                sig_meaning   = esig_meaning,
                                sig_timestamp = sig_ts,
                                user          = user,
                                role          = role,
                            )

                            st.session_state["last_result"] = {
                                **p,
                                "xlsx_bytes":    xlsx_bytes_final,
                                "pdf_bytes":     pdf_bytes_final,
                                "sig_id":        sig_id,
                                "sig_meaning":   esig_meaning,
                                "sig_timestamp": sig_ts,
                            }
                            st.session_state.pop("esig_pending",   None)
                            st.session_state.pop("show_esig_form", None)
                            st.session_state.pop("esig_target",    None)
                            st.rerun()


# =============================================================================
# 13. ROUTER
# =============================================================================
if not st.session_state.authenticated:
    show_login()
else:
    show_app()